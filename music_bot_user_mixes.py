#!/usr/bin/env python3
"""
Полностью асинхронный VLMB Music Bot с Яндекс.Музыкой, VK и YouTube
"""

import asyncio
import logging
import os
import io
import aiohttp
import tempfile
import shutil
import re
import time
import html
import math
import signal
import secrets
import json
from collections import OrderedDict
import random
import hashlib
import functools
try:
    import psutil
except Exception:  # pragma: no cover
    psutil = None
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from typing import Dict, List, Optional, Any, Set, Tuple
from io import BytesIO
from dataclasses import dataclass
from contextlib import asynccontextmanager
from pathlib import Path
import redis.asyncio as redis
import aiosqlite
import aiofiles
import aiofiles.os as async_os
import openpyxl
from logging.handlers import RotatingFileHandler
from concurrent.futures import ThreadPoolExecutor

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup, 
    ReplyKeyboardMarkup
)
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler, 
    MessageHandler, ChatMemberHandler, filters, CallbackContext, ContextTypes
)
from telegram.request import HTTPXRequest
from telegram.error import TimedOut, NetworkError, RetryAfter, TelegramError, BadRequest

import config
from services.provider_health import ProviderHealth
from services.provider_router import ProviderRouter, ProviderFailure
from services.search_engine import rank_tracks as _rank_tracks
from services.metrics import MetricsRegistry
from services.download_queue import DownloadQueue
from services.playlist_manager import PlaylistManager
from services.search_scoring import rank_tracks_by_artist as _legacy_rank_tracks_by_artist
from providers import YandexProviderAdapter, VKProviderAdapter, YouTubeProviderAdapter
from storage import InMemoryUserStateStore

# Импорт Яндекс.Музыки
try:
    # sync + async clients (если доступно)
    from yandex_music import Client  # type: ignore
    try:
        from yandex_music import ClientAsync  # type: ignore
    except Exception:
        ClientAsync = None  # type: ignore
    YANDEX_MUSIC_AVAILABLE = True
except ImportError:
    Client = None  # type: ignore
    ClientAsync = None  # type: ignore
    YANDEX_MUSIC_AVAILABLE = False
    print("⚠️ Библиотека yandex-music не установлена. Установите: pip install yandex-music")

# Импорт yt-dlp для поиска и загрузки аудио с YouTube
try:
    import yt_dlp  # type: ignore
    YT_DLP_AVAILABLE = True
except ImportError:
    yt_dlp = None  # type: ignore
    YT_DLP_AVAILABLE = False

# ==================== ИНИЦИАЛИЗАЦИЯ ЛОГГИНГА ====================
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=getattr(logging, config.LOG_LEVEL.upper(), logging.INFO),
    handlers=[logging.StreamHandler()]
)

if config.LOG_FILE:
    file_handler = RotatingFileHandler(
        config.LOG_FILE,
        maxBytes=int(getattr(config, "LOG_MAX_BYTES", 10 * 1024 * 1024) or 10 * 1024 * 1024),
        backupCount=int(getattr(config, "LOG_BACKUP_COUNT", 5) or 5),
        encoding="utf-8",
    )
    file_handler.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
    logging.getLogger().addHandler(file_handler)

# Do not leak bot tokens through full Telegram API URLs in INFO logs.
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
# Never allow HTTP client request URLs (which may contain Telegram bot tokens) into normal logs.
logging.getLogger("httpx").propagate = False
logging.getLogger("httpcore").propagate = False

logger = logging.getLogger(__name__)


def _sqlite_connect(db_path: str, **kwargs):
    """Create SQLite connections with one consistent busy timeout."""
    kwargs.setdefault(
        "timeout",
        float(getattr(config, "DATABASE_TIMEOUT", 20) or 20),
    )
    return aiosqlite.connect(db_path, **kwargs)


@asynccontextmanager
async def _sqlite_connection(db_path: str, **kwargs):
    """Context-managed SQLite connection with a configured busy timeout."""
    conn = await _sqlite_connect(db_path, **kwargs)
    try:
        timeout_ms = max(1, int(float(getattr(config, "DATABASE_TIMEOUT", 20) or 20) * 1000))
        try:
            await conn.execute(f"PRAGMA busy_timeout={timeout_ms}")
        except Exception as exc:
            logger.debug("SQLite busy_timeout setup failed for %s: %s", db_path, exc)
        yield conn
    finally:
        await conn.close()


class _SecretRedactionFilter(logging.Filter):
    """Redact configured API secrets from ordinary log messages."""

    def __init__(self) -> None:
        super().__init__()
        names = ("TELEGRAM_BOT_TOKEN", "YANDEX_TOKEN", "LASTFM_API_KEY", "VK_TOKEN")
        self._secrets = tuple(
            str(getattr(config, name, "") or "")
            for name in names
            if str(getattr(config, name, "") or "")
        )

    def _redact(self, value: Any) -> Any:
        if isinstance(value, str):
            for secret in self._secrets:
                value = value.replace(secret, "<redacted>")
        return value

    def filter(self, record: logging.LogRecord) -> bool:
        record.msg = self._redact(record.msg)
        if isinstance(record.args, tuple):
            record.args = tuple(self._redact(item) for item in record.args)
        elif isinstance(record.args, dict):
            record.args = {key: self._redact(value) for key, value in record.args.items()}
        return True


_secret_filter = _SecretRedactionFilter()
for _handler in logging.getLogger().handlers:
    _handler.addFilter(_secret_filter)

_INSTANCE_LOCK_HANDLE: Optional[Any] = None

def acquire_single_instance_lock() -> bool:
    """Prevent two polling processes from using the same Telegram token."""
    global _INSTANCE_LOCK_HANDLE
    lock_path = str(
        getattr(config, "INSTANCE_LOCK_FILE", "")
        or os.path.join(tempfile.gettempdir(), "vlmb_music_bot.lock")
    )
    try:
        handle = open(lock_path, "a+", encoding="utf-8")
        handle.seek(0)
        if os.name == "nt":
            import msvcrt
            if handle.read(1) == "":
                handle.write("0")
                handle.flush()
            handle.seek(0)
            msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
        else:
            import fcntl
            fcntl.flock(handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        handle.seek(0)
        handle.truncate()
        handle.write(str(os.getpid()))
        handle.flush()
        _INSTANCE_LOCK_HANDLE = handle
        return True
    except (OSError, IOError):
        try:
            handle.close()
        except Exception:
            pass
        return False

def release_single_instance_lock() -> None:
    """Release the polling-process lock during graceful shutdown."""
    global _INSTANCE_LOCK_HANDLE
    handle = _INSTANCE_LOCK_HANDLE
    _INSTANCE_LOCK_HANDLE = None
    if handle is None:
        return
    try:
        if os.name == "nt":
            import msvcrt
            handle.seek(0)
            msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
        else:
            import fcntl
            fcntl.flock(handle.fileno(), fcntl.LOCK_UN)
    except Exception as exc:
        logger.debug("Instance lock release failed: %s", exc)
    finally:
        try:
            handle.close()
        except Exception:
            pass


def _utc_now() -> datetime:
    """Return naive UTC for backward-compatible SQLite ISO timestamps."""
    return datetime.now(timezone.utc).replace(tzinfo=None)

def _process_memory_percent() -> Optional[float]:
    """Return process memory usage percentage when psutil is available."""
    if psutil is None:
        return None
    try:
        return float(psutil.Process().memory_percent())
    except Exception:
        return None

# Search ranking is intentionally limited to performer matching.



def _source_code(value: Any) -> str:
    source = str(value or "vk").strip().lower()
    if source in ("yandex", "yandex_music"):
        return "ym"
    if source in ("youtube", "youtube_music"):
        return "yt"
    return source


def _source_icon(value: Any) -> str:
    return {"ym": "🎵", "vk": "🎶", "yt": "▶️"}.get(_source_code(value), "🎧")


def _source_badge(value: Any) -> str:
    return {"ym": "YM", "vk": "VK", "yt": "YT"}.get(_source_code(value), "MUSIC")


def _source_name(value: Any, *, genitive: bool = False) -> str:
    source = _source_code(value)
    if source == "ym":
        return "Яндекс.Музыки" if genitive else "Яндекс.Музыка"
    if source == "yt":
        return "YouTube"
    if source == "vk":
        return "VK"
    return source.upper() or "источника"


def _search_cache_key(query: str) -> str:
    normalized = _normalize_user_query(query)
    version = str(getattr(config, "SEARCH_CACHE_VERSION", "v1") or "v1")
    source_state = "|".join((
        f"vk={int(bool(getattr(config, 'ENABLE_VK_MUSIC', True)))}",
        f"ym={int(bool(getattr(config, 'ENABLE_YANDEX_MUSIC', True)))}",
        f"yt={int(bool(getattr(config, 'ENABLE_YOUTUBE_MUSIC', True)))}",
        f"priority={getattr(config, 'SOURCE_PRIORITY', 'vk_first')}",
    ))
    raw = f"{source_state}\n{normalized.lower()}"
    digest = hashlib.md5(raw.encode("utf-8", errors="ignore")).hexdigest()
    return f"search:{version}:{digest}"


def _short_id(s: object, n: int = 10) -> str:
    try:
        ss = str(s or "")
    except Exception:
        ss = ""
    if not ss:
        return ""
    return ss[:n] + ("…" if len(ss) > n else "")



@asynccontextmanager
async def sem_guard(sem: asyncio.Semaphore):
    """Safe semaphore guard (no leaked permits on exceptions/returns)."""
    await sem.acquire()
    try:
        yield
    finally:
        sem.release()


# ---- CallbackQuery answer de-duplication (prevents 400 on repeated/late answers) ----
_CALLBACK_ANSWERED: dict = {}
_CALLBACK_ANSWER_TTL_SEC = 70

# ==================== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ====================

# --- GLOBAL PATCH HELPERS (safe to call from anywhere) ---
def _esc(s: str) -> str:
    """HTML-escape helper for Telegram HTML parse_mode."""
    return html.escape(str(s or ""))


def _format_track_duration(value: Any) -> str:
    """Format track duration for UI; keep unknown values explicit and compact."""
    try:
        total_seconds = max(0, int(float(value or 0)))
    except (TypeError, ValueError):
        total_seconds = 0

    if total_seconds <= 0:
        return "--:--"

    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    if hours:
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    return f"{minutes:02d}:{seconds:02d}"


def _track_duration_text(track: Any) -> str:
    """Return a duration prefix for TrackInfo, dictionaries, and compatible objects."""
    duration = 0
    try:
        duration = track.get("duration", 0)
        if not duration:
            duration = track.get("duration_sec", 0)
    except (AttributeError, TypeError):
        duration = getattr(track, "duration_sec", 0) or getattr(track, "duration", 0)
    return _format_track_duration(duration)


def private_main_keyboard():
    """Persistent bottom keyboard for private chats."""
    return ReplyKeyboardMarkup(
        [
            ["🔎 Поиск", "🔥 Чарты"],
            ["🎲 Подборка", "🎧 Похожие"],
            ["❤️ Избранное", "📚 История"],
            ["⚙️ Настройки", "❓ Помощь"],
        ],
        resize_keyboard=True,
        one_time_keyboard=False,
        input_field_placeholder="Напиши запрос или выбери действие…",
    )


_PRIVATE_MENU_ACTIONS = {
    "🔎 Поиск": "search",
    "🔍 Поиск": "search",
    "🔍 Поиск музыки": "search",
    "🔥 Чарты": "charts",
    "📊 Чарты": "charts",
    "🎲 Подборка": "mix",
    "🎲 Подборка по жанру": "mix",
    "🎧 Похожие": "similar",
    "👥 Похожие исполнители": "similar",
    "❤️ Избранное": "favorites",
    "📚 История": "history",
    "⚙️ Настройки": "settings",
    "❓ Помощь": "help",
    "🆘 Помощь": "help",
}

def _private_menu_action(text: str) -> Optional[str]:
    """Return a navigation action for a persistent private-chat menu label."""
    return _PRIVATE_MENU_ACTIONS.get((text or "").strip())


# ==================== КЛАСС ИСКЛЮЧЕНИЙ ====================
class RateLimitExceeded(Exception):
    """Исключение при превышении лимитов запросов."""
    pass


class ProviderError(RuntimeError):
    """Base error for provider-facing operations (reserved for future handlers)."""


class ProviderSearchError(ProviderError):
    """Provider search failed after its configured retry policy."""


class ProviderDownloadError(ProviderError):
    """Provider download failed after its configured retry/fallback policy."""

# ==================== МОДЕЛИ ДАННЫХ ====================
@dataclass
class TrackInfo:
    """Информация о треке из любого источника"""
    idx: int
    track_id: Any
    title: str
    artist: str
    album: str
    duration_sec: int
    source: str = "vk"  # "vk", "ym", "yt"
    vk_url: Optional[str] = None
    ym_download_info: Optional[Any] = None
    vk_key: Optional[str] = None
    youtube_id: Optional[str] = None
    youtube_url: Optional[str] = None
    audio_ext: Optional[str] = None
    track_obj: Optional[Any] = None

    # --- Normalized fields for stable routing / UX ---
    uid: Optional[str] = None           # stable id across group/private (computed)
    bitrate_kbps: Optional[int] = None  # best-effort; may be unknown

    def _compute_uid(self) -> str:
        """Stable UID for routing & favorites/history.

        Prefer provider-specific ids, fallback to hash of main identifying fields.
        """
        try:
            if self.source == 'ym' and self.track_id:
                return f"ym:{self.track_id}"
            if self.source == 'vk':
                if self.vk_key:
                    return f"vk:{self.vk_key}"
                if self.track_id:
                    return f"vkid:{self.track_id}"
            if self.source == 'yt':
                youtube_id = self.youtube_id or self.track_id
                if youtube_id:
                    return f"yt:{youtube_id}"
        except Exception:
            pass
        raw = f"{self.source}|{self.artist}|{self.title}|{self.duration_sec}|{self.vk_url or ''}"
        return "h:" + hashlib.md5(raw.encode('utf-8', errors='ignore')).hexdigest()

    def ensure_uid(self) -> str:
        if not self.uid:
            self.uid = self._compute_uid()
        return self.uid

    # dict-like compatibility (existing code heavily uses .get())
    def get(self, key: str, default: Any = None) -> Any:
        if key == 'url':
            value = self.youtube_url if self.source == 'yt' else self.vk_url
            return value if value is not None else default
        mapping = {
            'idx': 'idx',
            'track_id': 'track_id',
            'title': 'title',
            'artist': 'artist',
            'album': 'album',
            'duration': 'duration_sec',
            'duration_sec': 'duration_sec',
            'source': 'source',
            'vk_url': 'vk_url',
            'vk_key': 'vk_key',
            'youtube_id': 'youtube_id',
            'youtube_url': 'youtube_url',
            'audio_ext': 'audio_ext',
            'uid': 'uid',
            'bitrate_kbps': 'bitrate_kbps',
        }
        attr = mapping.get(key)
        if attr is None:
            return default
        v = getattr(self, attr, default)
        if key == 'uid':
            return self.ensure_uid()
        return v if v is not None else default

    def __getitem__(self, key: str) -> Any:
        v = self.get(key, None)
        if v is None:
            raise KeyError(key)
        return v
    
    def to_dict(self) -> Dict[str, Any]:
        """Конвертация в словарь для кэширования"""
        uid = self.ensure_uid()
        return {
            'idx': self.idx,
            'track_id': str(self.track_id),
            'title': self.title,
            'artist': self.artist,
            'album': self.album,
            'duration_sec': self.duration_sec,
            'source': self.source,
            'vk_url': self.vk_url,
            'vk_key': self.vk_key,
            'youtube_id': self.youtube_id,
            'youtube_url': self.youtube_url,
            'audio_ext': self.audio_ext,
            'uid': uid,
            'bitrate_kbps': self.bitrate_kbps,
            'duration': self.duration_sec
        }
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'TrackInfo':
        """Создание из словаря"""
        ti = cls(
            idx=data['idx'],
            track_id=data['track_id'],
            title=data['title'],
            artist=data['artist'],
            album=data.get('album', ''),
            duration_sec=data['duration_sec'],
            source=data.get('source', 'vk'),
            vk_url=data.get('vk_url'),
            vk_key=data.get('vk_key'),
            youtube_id=data.get('youtube_id'),
            youtube_url=data.get('youtube_url') or (data.get('url') if data.get('source') == 'yt' else None),
            audio_ext=data.get('audio_ext'),
            uid=data.get('uid'),
            bitrate_kbps=data.get('bitrate_kbps')
        )
        ti.ensure_uid()
        return ti


def _track_uid_from_any(track: Any) -> str:
    """Helper: stable UID for dict/TrackInfo."""
    if track is None:
        return ""
    if isinstance(track, TrackInfo):
        return track.ensure_uid()
    try:
        uid = track.get('uid')
        if uid:
            return str(uid)
    except Exception:
        uid = None
    try:
        src = _source_code(track.get('source') or 'vk')
        if src == 'ym' and track.get('track_id'):
            return f"ym:{track.get('track_id')}"
        if src == 'vk' and track.get('vk_key'):
            return f"vk:{track.get('vk_key')}"
        if src == 'yt':
            youtube_id = track.get('youtube_id') or track.get('track_id')
            if youtube_id:
                return f"yt:{youtube_id}"
    except Exception:
        pass
    raw = None
    try:
        raw = f"{track.get('source','')}|{track.get('artist','')}|{track.get('title','')}|{track.get('duration',0)}|{track.get('url','')}"
    except Exception:
        raw = str(track)
    return "h:" + hashlib.md5(str(raw).encode('utf-8', errors='ignore')).hexdigest()




# ==================== TELEGRAM SAFE API WRAPPERS ====================
async def _tg_call_with_retries(call, *args, retries: int = None, base_delay: float = None, **kwargs):
    """Единая обертка для Telegram API: ретраи + backoff + обработка RetryAfter/TimedOut."""
    if retries is None:
        retries = getattr(config, "TELEGRAM_API_RETRIES", 3)
    if base_delay is None:
        base_delay = getattr(config, "TELEGRAM_API_RETRY_BASE_DELAY", 0.7)

    last_exc = None
    for attempt in range(retries + 1):
        # If we retry a request with a file-like object (BytesIO/File), rewind it.
        try:
            for _k in ('audio','document','video','voice','photo','animation'):
                _v = kwargs.get(_k)
                if hasattr(_v, 'seek'):
                    try:
                        _v.seek(0)
                    except Exception:
                        pass
        except Exception:
            pass
        try:
            return await call(*args, **kwargs)
        except RetryAfter as e:
            sleep_for = float(getattr(e, "retry_after", 1.0) or 1.0)
            # Telegram explicitly tells us how long to wait; retrying earlier
            # only causes another flood-control response.
            await asyncio.sleep(max(0.5, sleep_for))
            last_exc = e
        except (TimedOut, NetworkError) as e:
            last_exc = e
            if attempt >= retries:
                break
            await asyncio.sleep(base_delay * (2 ** attempt))
        except TelegramError as e:
            last_exc = e
            break
        except Exception as e:
            last_exc = e
            break
    raise last_exc

async def safe_edit_text(message, text: str, **kwargs):
    """Безопасный edit_text"""
    try:
        return await _tg_call_with_retries(message.edit_text, text, **kwargs)
    except Exception as e:
        logger.warning(f"safe_edit_text fallback due to: {e}")
        try:
            bot = message.get_bot()
            thread_id = getattr(message, 'message_thread_id', None)
            if thread_id is not None and 'message_thread_id' not in kwargs:
                kwargs['message_thread_id'] = thread_id
            return await _tg_call_with_retries(bot.send_message, chat_id=message.chat_id, text=text, **kwargs)
        except Exception:
            return None

async def safe_answer_callback(query, text: str = None, **kwargs):
    """Безопасный ответ на callback_query"""
    if not query:
        return None
    # Avoid answering the same callback_query many times (Telegram will return 400).
    try:
        qid = getattr(query, 'id', None)
        now = time.time()
        if qid:
            ts = _CALLBACK_ANSWERED.get(qid)
            if ts and (now - ts) < _CALLBACK_ANSWER_TTL_SEC:
                return None
            _CALLBACK_ANSWERED[qid] = now
            # small cleanup
            if len(_CALLBACK_ANSWERED) > 2048:
                for k, v in list(_CALLBACK_ANSWERED.items())[:512]:
                    if (now - v) > _CALLBACK_ANSWER_TTL_SEC:
                        _CALLBACK_ANSWERED.pop(k, None)
    except Exception:
        pass
    try:
        return await _tg_call_with_retries(query.answer, text=text, **kwargs)
    except BadRequest as e:
        msg = str(e).lower()
        if "query is too old" in msg or "query_id_invalid" in msg or "query id invalid" in msg:
            logger.info(f"Callback query expired/invalid: {e}")
            return None
        raise
    except TelegramError as e:
        logger.warning(f"Callback answer failed: {e}")
        return None

async def safe_send_audio(bot, **kwargs):
    """Безопасная отправка аудио"""
    return await _tg_call_with_retries(bot.send_audio, **kwargs)

async def safe_edit_message_text(query, text: str, **kwargs):
    """Безопасный edit_message_text для callback query"""
    try:
        return await _tg_call_with_retries(query.edit_message_text, text, **kwargs)
    except Exception as e:
        logger.warning(f"safe_edit_message_text fallback due to: {e}")
        try:
            msg = getattr(query, "message", None)
            bot = query.get_bot() if hasattr(query, "get_bot") else (msg.get_bot() if msg else None)
            if bot and msg:
                thread_id = getattr(msg, "message_thread_id", None)
                if thread_id is not None and "message_thread_id" not in kwargs:
                    kwargs["message_thread_id"] = thread_id
                return await _tg_call_with_retries(bot.send_message, chat_id=msg.chat_id, text=text, **kwargs)
        except Exception:
            return None

async def safe_send_message(bot, chat_id, text, **kwargs):
    """Безопасная отправка сообщения"""
    return await _tg_call_with_retries(bot.send_message, chat_id=chat_id, text=text, **kwargs)

async def safe_delete_message(bot, chat_id, message_id, **kwargs):
    """Безопасное удаление сообщения (игнорирует сетевые/лимитные сбои через ретраи)"""
    return await _tg_call_with_retries(bot.delete_message, chat_id=chat_id, message_id=message_id, **kwargs)

async def safe_send_document(bot, chat_id, document, **kwargs):
    """Безопасная отправка документа"""
    return await _tg_call_with_retries(bot.send_document, chat_id=chat_id, document=document, **kwargs)

# ==================== АСИНХРОННЫЕ ДЕКОРАТОРЫ ====================
def rate_limit(action: str):
    """Декоратор для ограничения частоты запросов"""
    def decorator(func):
        @functools.wraps(func)
        async def wrapper(self, *args, **kwargs):
            user_id = None
            if len(args) > 1:
                user_id = args[1]
            elif 'user_id' in kwargs:
                user_id = kwargs['user_id']
            
            if user_id is not None and not await self.user_manager.check_rate_limit(user_id, action):
                raise RateLimitExceeded(f"Rate limit exceeded for {action}")
            
            return await func(self, *args, **kwargs)
        return wrapper
    return decorator

def validate_query(func):
    """Декоратор для валидации поисковых запросов"""
    @functools.wraps(func)
    async def wrapper(self, query: str, *args, **kwargs):
        user_id = kwargs.get('user_id') or (args[0] if args else None)
        
        if user_id and not await self.user_manager.validate_query(user_id, query):
            logger.warning(f"User {user_id} sent invalid query: {query}")
            return []
        
        return await func(self, query, *args, **kwargs)
    return wrapper

# ==================== ЯНДЕКС.МУЗЫКА МЕНЕДЖЕР ====================
class YandexMusicManager:
    """Асинхронный менеджер для работы с Яндекс.Музыкой"""
    
    def __init__(self):
        self.client = None  # sync client (fallback)
        self.client_async = None  # async client (preferred)
        self.executor: Optional[ThreadPoolExecutor] = None
        self._initialized = False
        self._use_async_client = False

        # cached bot username
        self.bot_username: Optional[str] = None

        # Global YM concurrency limiter (search + download)
        self._semaphore = asyncio.Semaphore(getattr(config, "YANDEX_CONCURRENCY", 4) or 4)
        # Protect sync client from concurrent usage (Client uses requests session; not guaranteed threadsafe)
        self._client_lock = asyncio.Lock()

        # Reuse a shared aiohttp session (set by AsyncMusicBot) to avoid per-download TCP/TLS overhead
        self.http_session: Optional[aiohttp.ClientSession] = None


    async def initialize(self):
        """Инициализация Яндекс.Музыка клиента.

        Предпочтительно используем ClientAsync (без ThreadPoolExecutor).
        Если он недоступен/падает — откатываемся на синхронный Client в отдельном пуле.
        """
        if not YANDEX_MUSIC_AVAILABLE:
            logger.warning("⚠️ Яндекс.Музыка недоступна (библиотека не установлена)")
            return False

        if not getattr(config, "YANDEX_TOKEN", None):
            logger.warning("⚠️ Яндекс.Музыка не настроен (отсутствует токен)")
            return False

        # 1) Try async client first
        try:
            if ClientAsync is not None:
                self.client_async = ClientAsync(getattr(config, "YANDEX_TOKEN"))  # type: ignore
                await asyncio.wait_for(self.client_async.init(), timeout=int(getattr(config, "YM_SEARCH_TIMEOUT", 3)) + 3)
                self.client = None
                self._use_async_client = True
                self._initialized = True
                logger.info("✅ Яндекс.Музыка клиент успешно инициализирован (async)")
                return True
        except Exception as e:
            logger.warning(f"YM async init failed, fallback to sync client: {e}")

        # 2) Fallback to sync client (threadpool)
        try:
            # Sync-client fallback uses a threadpool. Allow it to grow above 4 when explicitly configured,
            # but keep an upper bound to avoid bans/flood and CPU oversubscription.
            max_workers = int(getattr(config, "YANDEX_CONCURRENCY", 4) or 4)
            hard_cap = int(getattr(config, "YM_SYNC_MAX_WORKERS", 12) or 12)
            max_workers = max(1, min(max_workers, max(1, min(hard_cap, 32))))
            self.executor = ThreadPoolExecutor(max_workers=max_workers, thread_name_prefix="ym_")

            loop = asyncio.get_running_loop()
            self.client = await loop.run_in_executor(
                self.executor,
                lambda: Client(getattr(config, "YANDEX_TOKEN")).init()  # type: ignore
            )

            if self.client:
                self.client_async = None
                self._use_async_client = False
                self._initialized = True
                logger.info("✅ Яндекс.Музыка клиент успешно инициализирован (sync)")
                return True

            logger.error("❌ Не удалось инициализировать Яндекс.Музыка клиент")
            return False

        except Exception as e:
            logger.error(f"❌ Ошибка инициализации Яндекс.Музыка: {e}")
            self.client = None
            self.client_async = None
            self._use_async_client = False
            self._initialized = False
            return False

    def _search_sync(self, query: str):
        """Синхронный поиск в Яндекс.Музыке"""
        if not self.client:
            return None
        try:
            return self.client.search(query)
        except Exception as e:
            logger.warning(f"YM search error: {e}")
            return None
    
    def _get_track_sync(self, track_id: int):
        """Синхронное получение трека"""
        if not self.client:
            return None
        try:
            tracks = self.client.tracks([track_id])
            return tracks[0] if tracks else None
        except Exception as e:
            logger.warning(f"YM get track error: {e}")
            return None
    
    async def search_tracks(self, query: str, limit: int = 10) -> List[TrackInfo]:
            """Асинхронный поиск треков в Яндекс.Музыке (устойчивый)."""
            if not self._initialized:
                return []
    
            async with self._semaphore:
                try:
                    # --- Async client path (preferred) ---
                    if self._use_async_client and self.client_async is not None:
                        # sanitize query for YM (telegram users sometimes paste with newlines)
                        query_clean = (query or "").replace("\n", " ").replace("\r", " ").strip()
                        ym_timeout = float(getattr(config, "YM_SEARCH_TIMEOUT", 6))
                        try:
                            search_coro = self.client_async.search(query_clean)
                            search_result = await asyncio.wait_for(search_coro, timeout=ym_timeout)
                        except asyncio.TimeoutError:
                            logger.warning(f"YM search timeout ({ym_timeout}s) for query: {query_clean!r} -> retry with {ym_timeout*2:.1f}s")
                            search_coro = self.client_async.search(query_clean)
                            search_result = await asyncio.wait_for(search_coro, timeout=ym_timeout * 2)
                    else:
                        # --- Sync client fallback (serialized) ---
                        if not self.client or not self.executor:
                            return []
                        loop = asyncio.get_running_loop()
                        async with self._client_lock:
                            search_task = loop.run_in_executor(self.executor, self._search_sync, query)
                            search_result = await asyncio.wait_for(search_task, timeout=getattr(config, "YM_SEARCH_TIMEOUT", 3))
    
                    if not search_result:
                        return []
    
                    tracks: List[TrackInfo] = []
                    track_results = getattr(search_result, 'tracks', None)
                    if track_results:
                        results = getattr(track_results, 'results', [])[:limit]
                        for i, track_short in enumerate(results, 1):
                            try:
                                base = getattr(track_short, 'track', track_short)
                                track_id = getattr(base, 'id', 0)
                                title = getattr(base, 'title', '') or 'Unknown'
    
                                artists = getattr(base, 'artists', []) or []
                                artist_names = []
                                for a in artists:
                                    an = getattr(a, 'name', None)
                                    if an:
                                        artist_names.append(an)
                                artist = ", ".join(artist_names) if artist_names else 'Unknown'
    
                                duration_ms = getattr(base, 'duration_ms', 0) or 0
                                duration = int(duration_ms / 1000) if duration_ms else 0
    
                                # album title (best-effort)
                                album_title = ""
                                try:
                                    albums = getattr(base, "albums", []) or []
                                    if albums:
                                        album_title = (getattr(albums[0], "title", None) or "").strip()
                                except Exception:
                                    album_title = ""

                                tracks.append(TrackInfo(
                                    idx=i,
                                    track_id=str(track_id),
                                    title=title,
                                    artist=artist,
                                    album=album_title,
                                    duration_sec=duration,
                                    source="ym",
                                    vk_url=None,
                                    ym_download_info=None,
                                    vk_key=None,
                                    track_obj=base
                                ))
                            except Exception as e:
                                logger.debug(f"Error parsing YM track: {e}")
                                continue
    
                    return tracks
    
                except asyncio.TimeoutError:
                    logger.warning(f"YM search timeout: {query}")
                    return []
                except Exception as e:
                    logger.error(f"YM search exception: {e}")
                    return []


    async def get_track_download_info(self, track: TrackInfo) -> Optional[Any]:
            """Получение информации для скачивания трека (DownloadInfo)."""
            if not self._initialized or track.source != "ym":
                return None
    
            try:
                if self._use_async_client and self.client_async is not None:
                    tracks_full = await self.client_async.tracks([int(track.track_id)])
                    track_full = tracks_full[0] if tracks_full else None
                    if not track_full:
                        return None
                    download_info_list = await track_full.get_download_info_async(get_direct_links=False)
                else:
                    if not self.client or not self.executor:
                        return None
                    loop = asyncio.get_running_loop()
                    async with self._client_lock:
                        track_full = await loop.run_in_executor(self.executor, self._get_track_sync, track.track_id)
                        if not track_full:
                            return None
                        download_info_list = await loop.run_in_executor(self.executor, track_full.get_download_info)
    
                if not download_info_list:
                    return None
    
                preferred = int(getattr(config, "YM_PREFERRED_MAX_BITRATE_KBPS", 192) or 192)
                best = None
                for di in download_info_list:
                    bitrate = getattr(di, 'bitrate_in_kbps', 0) or 0
                    if best is None:
                        best = di
                        continue
                    best_bitrate = getattr(best, 'bitrate_in_kbps', 0) or 0
                    if bitrate <= preferred:
                        if best_bitrate > preferred or bitrate > best_bitrate:
                            best = di
                    else:
                        if best_bitrate > preferred and bitrate > best_bitrate:
                            best = di
    
                return best or download_info_list[0]
    
            except Exception as e:
                logger.error(f"YM get download info error: {e}")
                return None



    async def download_track(self, track: TrackInfo, file_path: str) -> bool:
            """Скачивание трека из Яндекс.Музыки (устойчивое, без падений)."""
            if not self._initialized or track.source != "ym":
                return False
    
            try:
                download_info = await self.get_track_download_info(track)
                if not download_info:
                    return False
    
                direct_link = None
                try:
                    if self._use_async_client and self.client_async is not None:
                        if hasattr(download_info, "get_direct_link_async"):
                            direct_link = await download_info.get_direct_link_async()
                        else:
                            direct_link = download_info.get_direct_link()
                    else:
                        if not self.executor:
                            return False
                        loop = asyncio.get_running_loop()
                        async with self._client_lock:
                            direct_link = await loop.run_in_executor(self.executor, download_info.get_direct_link)
                except Exception:
                    direct_link = None
    
                if direct_link:
                    session = self.http_session
                    created_session = False
                    if session is None:
                        session = aiohttp.ClientSession()
                        created_session = True
                    try:
                        async with session.get(direct_link) as response:
                            if response.status != 200:
                                return False
                            async with aiofiles.open(file_path, 'wb') as f:
                                async for chunk in response.content.iter_chunked(256 * 1024):
                                    await f.write(chunk)
                        return True
                    finally:
                        if created_session:
                            await session.close()
    
                # Library fallback: download bytes
                try:
                    if self._use_async_client and self.client_async is not None and hasattr(download_info, "download_bytes_async"):
                        data = await download_info.download_bytes_async()
                    elif self.executor is not None and hasattr(download_info, "download_bytes"):
                        loop = asyncio.get_running_loop()
                        async with self._client_lock:
                            data = await loop.run_in_executor(self.executor, download_info.download_bytes)
                    else:
                        return False
    
                    async with aiofiles.open(file_path, 'wb') as f:
                        await f.write(data)
                    return True
                except Exception as e:
                    logger.warning(f"YM download fallback failed: {e}")
                    return False
    
            except Exception as e:
                logger.error(f"YM download exception: {e}")
                return False


    async def download_track_bytes(self, track: TrackInfo) -> Optional[bytes]:
            """Скачать трек из Яндекс.Музыки в память (bytes).

            Цель: ускорить путь YM->Telegram, убрав tempfile->read() round-trip.
            Возвращает bytes или None при ошибке.
            """
            if not self._initialized or track.source != "ym":
                return None

            try:
                download_info = await self.get_track_download_info(track)
                if not download_info:
                    return None

                max_size = int(getattr(config, "MAX_FILE_SIZE", 50 * 1024 * 1024) or (50 * 1024 * 1024))

                # 1) Try direct link (fast, stream in chunks)
                direct_link = None
                try:
                    if self._use_async_client and self.client_async is not None:
                        if hasattr(download_info, "get_direct_link_async"):
                            direct_link = await download_info.get_direct_link_async()
                        else:
                            direct_link = download_info.get_direct_link()
                    else:
                        if not self.executor:
                            direct_link = None
                        else:
                            loop = asyncio.get_running_loop()
                            async with self._client_lock:
                                direct_link = await loop.run_in_executor(self.executor, download_info.get_direct_link)
                except Exception:
                    direct_link = None

                if direct_link:
                    session = self.http_session
                    created_session = False
                    if session is None:
                        session = aiohttp.ClientSession()
                        created_session = True
                    try:
                        buf = BytesIO()
                        total = 0
                        async with session.get(direct_link) as response:
                            if response.status != 200:
                                return None
                            async for chunk in response.content.iter_chunked(256 * 1024):
                                if not chunk:
                                    continue
                                total += len(chunk)
                                if total > max_size:
                                    return None
                                buf.write(chunk)
                        data = buf.getvalue()
                        if not data:
                            return None
                        return data
                    finally:
                        if created_session:
                            await session.close()

                # 2) Library fallback: download bytes
                try:
                    if self._use_async_client and self.client_async is not None and hasattr(download_info, "download_bytes_async"):
                        data = await download_info.download_bytes_async()
                    elif self.executor is not None and hasattr(download_info, "download_bytes"):
                        loop = asyncio.get_running_loop()
                        async with self._client_lock:
                            data = await loop.run_in_executor(self.executor, download_info.download_bytes)
                    else:
                        return None

                    if not data:
                        return None
                    if len(data) > max_size:
                        return None
                    return data
                except Exception as e:
                    logger.warning(f"YM download_bytes failed: {e}")
                    return None

            except Exception as e:
                logger.error(f"YM download_bytes exception: {e}")
                return None


    
    
    
    
    
    

class _YTDLPLogger:
    """Route yt-dlp messages through the bot logger without verbose console output."""

    @staticmethod
    def _clean(message: str) -> str:
        # Signed media URLs may contain short-lived credentials. Never log them.
        return re.sub(r"https?://\S+", "<url>", str(message or ""))

    def debug(self, message: str) -> None:
        if message and not str(message).startswith("[debug]"):
            logger.debug("yt-dlp: %s", self._clean(message))

    def info(self, message: str) -> None:
        logger.debug("yt-dlp: %s", self._clean(message))

    def warning(self, message: str) -> None:
        logger.warning("yt-dlp: %s", self._clean(message))

    def error(self, message: str) -> None:
        logger.error("yt-dlp: %s", self._clean(message))


class YoutubeMusicManager:
    """Bounded async wrapper around the synchronous yt-dlp API."""

    _YOUTUBE_URL_RE = re.compile(
        r"^(?:https?://)?(?:www\.|m\.|music\.)?(?:youtube\.com|youtu\.be)/",
        re.IGNORECASE,
    )

    def __init__(self) -> None:
        self._initialized = False
        self.executor: Optional[ThreadPoolExecutor] = None
        concurrency = max(1, min(int(getattr(config, "YOUTUBE_CONCURRENCY", 2) or 2), 8))
        self._semaphore = asyncio.Semaphore(concurrency)

    async def initialize(self) -> bool:
        # Initialize the provider even when disabled in runtime settings, so an
        # admin can enable it without restarting the bot. Search/download paths
        # still honor ENABLE_YOUTUBE_MUSIC.
        if self._initialized and self.executor is not None:
            return True
        if not YT_DLP_AVAILABLE or yt_dlp is None:
            logger.warning('⚠️ YouTube недоступен: установите pip install -U "yt-dlp[default]"')
            self._initialized = False
            return False
        workers = max(1, min(int(getattr(config, "YOUTUBE_CONCURRENCY", 2) or 2), 8))
        self.executor = ThreadPoolExecutor(max_workers=workers, thread_name_prefix="youtube_")
        self._initialized = True
        logger.info("✅ YouTube/yt-dlp источник инициализирован")
        return True

    def _common_options(self) -> Dict[str, Any]:
        retries = max(0, min(int(getattr(config, "YOUTUBE_DOWNLOAD_RETRIES", 3) or 3), 10))
        options: Dict[str, Any] = {
            "quiet": True,
            "no_warnings": True,
            "logger": _YTDLPLogger(),
            "socket_timeout": float(getattr(config, "YOUTUBE_SEARCH_TIMEOUT", 15) or 15),
            "retries": retries,
            "fragment_retries": retries,
            "extractor_retries": max(0, min(int(getattr(config, "YOUTUBE_SEARCH_RETRIES", 2) or 2), 10)),
            "retry_sleep_functions": {
                "http": lambda attempt: min(8.0, float(getattr(config, "YOUTUBE_RETRY_SLEEP", 1.0) or 1.0) * (2 ** max(0, attempt - 1))),
                "fragment": lambda attempt: min(8.0, float(getattr(config, "YOUTUBE_RETRY_SLEEP", 1.0) or 1.0) * (2 ** max(0, attempt - 1))),
                "extractor": lambda attempt: min(8.0, float(getattr(config, "YOUTUBE_RETRY_SLEEP", 1.0) or 1.0) * (2 ** max(0, attempt - 1))),
            },
        }
        cookie_file = str(getattr(config, "YOUTUBE_COOKIE_FILE", "") or "").strip()
        if cookie_file:
            if os.path.isfile(cookie_file):
                options["cookiefile"] = cookie_file
            else:
                logger.warning("YouTube cookie file not found: %s", cookie_file)
        return options

    @staticmethod
    def _artist_and_title(entry: Dict[str, Any]) -> Tuple[str, str]:
        raw_title = str(entry.get("track") or entry.get("title") or "Неизвестный трек").strip()
        artist = str(
            entry.get("artist")
            or entry.get("creator")
            or entry.get("uploader")
            or entry.get("channel")
            or "YouTube"
        ).strip()
        title = raw_title
        if not entry.get("track"):
            for separator in (" - ", " — ", " – "):
                if separator in raw_title:
                    left, right = raw_title.split(separator, 1)
                    if left.strip() and right.strip():
                        artist = left.strip()
                        title = right.strip()
                    break
        return artist or "YouTube", title or "Неизвестный трек"

    @staticmethod
    def _entry_to_track(entry: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        if not isinstance(entry, dict):
            return None
        video_id = str(entry.get("id") or "").strip()
        webpage_url = str(entry.get("webpage_url") or "").strip()
        if not webpage_url and video_id:
            webpage_url = f"https://www.youtube.com/watch?v={video_id}"
        if not webpage_url:
            raw_url = str(entry.get("url") or "").strip()
            if raw_url.startswith(("http://", "https://")):
                webpage_url = raw_url
        if not webpage_url:
            return None
        artist, title = YoutubeMusicManager._artist_and_title(entry)
        try:
            duration = max(0, int(float(entry.get("duration") or 0)))
        except (TypeError, ValueError):
            duration = 0
        return {
            "url": webpage_url,
            "youtube_url": webpage_url,
            "youtube_id": video_id or None,
            "track_id": video_id or webpage_url,
            "artist": artist,
            "title": title,
            "duration": duration,
            "source": "yt",
            "vk_key": None,
            "audio_ext": None,
        }

    def _search_sync(self, query: str, limit: int) -> List[Dict[str, Any]]:
        if yt_dlp is None:
            return []
        options = self._common_options()
        options.update({
            "skip_download": True,
            "noplaylist": True,
            "extract_flat": "in_playlist",
            "playlistend": limit,
        })
        target = query if self._YOUTUBE_URL_RE.match(query.strip()) else f"ytsearch{limit}:{query}"
        with yt_dlp.YoutubeDL(options) as ydl:  # type: ignore[attr-defined]
            info = ydl.extract_info(target, download=False)
        if not info:
            return []
        entries = info.get("entries") if isinstance(info, dict) else None
        if entries is None:
            entries = [info]
        results: List[Dict[str, Any]] = []
        for entry in entries or []:
            track = self._entry_to_track(entry)
            if track is not None:
                results.append(track)
            if len(results) >= limit:
                break
        return results

    async def search_tracks(self, query: str, limit: int = 10) -> List[Dict[str, Any]]:
        if not self._initialized or not self.executor:
            return []
        query_clean = _normalize_user_query(query)
        if not query_clean:
            return []
        max_limit = max(1, int(getattr(config, "YOUTUBE_MAX_SEARCH_RESULTS", 40) or 40))
        requested = max(1, min(int(limit or 10), max_limit))
        timeout = max(3.0, float(getattr(config, "YOUTUBE_SEARCH_TIMEOUT", 15) or 15))
        loop = asyncio.get_running_loop()
        async with self._semaphore:
            try:
                task = loop.run_in_executor(self.executor, self._search_sync, query_clean, requested)
                return await asyncio.wait_for(task, timeout=timeout)
            except asyncio.TimeoutError:
                logger.warning("YouTube search timeout (%.1fs) for %r", timeout, query_clean)
                return []
            except Exception as exc:
                logger.warning("YouTube search failed for %r: %s", query_clean, exc)
                return []

    @staticmethod
    def _resolve_ffmpeg_location() -> Optional[str]:
        configured = str(getattr(config, "YOUTUBE_FFMPEG_LOCATION", "") or "").strip()
        if configured:
            if os.path.exists(configured) or shutil.which(configured):
                return configured
            logger.warning("Configured ffmpeg location is unavailable: %s", configured)
            return None
        return shutil.which("ffmpeg")

    def _download_sync(self, track: Dict[str, Any]) -> Tuple[bytes, str]:
        temp_dir = tempfile.mkdtemp(prefix="youtube_audio_")
        try:
            return self._download_sync_in_dir(track, temp_dir)
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def _download_sync_in_dir(self, track: Dict[str, Any], temp_dir: str) -> Tuple[bytes, str]:
        if yt_dlp is None:
            raise RuntimeError("yt-dlp не установлен")
        target = str(track.get("youtube_url") or track.get("url") or "").strip()
        video_id = str(track.get("youtube_id") or track.get("track_id") or "").strip()
        if not target and video_id:
            target = f"https://www.youtube.com/watch?v={video_id}"
        if not target:
            raise RuntimeError("YouTube URL недоступен")

        options = self._common_options()
        options.update({
            "format": str(getattr(config, "YOUTUBE_AUDIO_FORMAT", "bestaudio[ext=m4a]/bestaudio/best") or "bestaudio[ext=m4a]/bestaudio/best"),
            "outtmpl": os.path.join(temp_dir, "%(id)s.%(ext)s"),
            "noplaylist": True,
            "nopart": True,
            "overwrites": True,
            "max_filesize": int(getattr(config, "MAX_FILE_SIZE", 50 * 1024 * 1024) or 50 * 1024 * 1024),
            "socket_timeout": float(getattr(config, "YOUTUBE_DOWNLOAD_TIMEOUT", 180) or 180),
        })

        codec = str(getattr(config, "YOUTUBE_AUDIO_CODEC", "mp3") or "").strip().lower()
        ffmpeg_location = self._resolve_ffmpeg_location()
        if codec and ffmpeg_location:
            options["ffmpeg_location"] = ffmpeg_location
            options["postprocessors"] = [{
                "key": "FFmpegExtractAudio",
                "preferredcodec": codec,
                "preferredquality": str(getattr(config, "YOUTUBE_AUDIO_QUALITY", "192") or "192"),
            }]

        with yt_dlp.YoutubeDL(options) as ydl:  # type: ignore[attr-defined]
            info = ydl.extract_info(target, download=True)
            prepared = ydl.prepare_filename(info) if isinstance(info, dict) else ""

        candidates: List[Path] = []
        if prepared:
            candidates.append(Path(prepared))
            if codec and ffmpeg_location:
                candidates.append(Path(prepared).with_suffix(f".{codec}"))
        if isinstance(info, dict):
            for item in info.get("requested_downloads") or []:
                if isinstance(item, dict):
                    filepath = item.get("filepath") or item.get("filename")
                    if filepath:
                        candidates.append(Path(str(filepath)))
        candidates.extend(sorted(Path(temp_dir).glob("*"), key=lambda item: item.stat().st_mtime if item.exists() else 0, reverse=True))

        chosen: Optional[Path] = None
        for candidate in candidates:
            try:
                if candidate.is_file() and candidate.suffix.lower() not in (".part", ".ytdl", ".json"):
                    chosen = candidate
                    break
            except OSError:
                continue
        if chosen is None:
            raise RuntimeError("yt-dlp не создал аудиофайл")

        max_size = int(getattr(config, "MAX_FILE_SIZE", 50 * 1024 * 1024) or 50 * 1024 * 1024)
        size = chosen.stat().st_size
        if size <= 0:
            raise RuntimeError("YouTube вернул пустой аудиофайл")
        if size > max_size:
            raise RuntimeError(f"YouTube аудиофайл слишком большой: {size} байт")
        extension = chosen.suffix.lower().lstrip(".") or (codec if codec else "m4a")
        if not ffmpeg_location and extension not in ("mp3", "m4a", "mp4"):
            raise RuntimeError(
                f"YouTube вернул формат .{extension}; установите ffmpeg для конвертации в MP3"
            )
        data = chosen.read_bytes()
        return data, extension

    async def download_track_bytes(self, track: Dict[str, Any]) -> Tuple[bytes, str]:
        if not self._initialized or not self.executor:
            raise RuntimeError("YouTube источник не инициализирован")
        timeout = max(10.0, float(getattr(config, "YOUTUBE_DOWNLOAD_TIMEOUT", 180) or 180))
        loop = asyncio.get_running_loop()
        async with self._semaphore:
            task = loop.run_in_executor(self.executor, self._download_sync, dict(track))
            try:
                return await asyncio.wait_for(asyncio.shield(task), timeout=timeout)
            except asyncio.TimeoutError as exc:
                # A running worker thread cannot be force-cancelled. It owns and
                # cleans its private temp directory when the yt-dlp call finishes.
                raise RuntimeError(f"Тайм-аут загрузки YouTube ({timeout:.0f} сек.)") from exc

    async def close(self) -> None:
        executor = self.executor
        self.executor = None
        self._initialized = False
        if executor is not None:
            await asyncio.to_thread(executor.shutdown, False, cancel_futures=True)


# ==================== АСИНХРОННЫЙ МЕНЕДЖЕР КЭША ====================
class AsyncCacheManager:
    """Асинхронный менеджер кэширования с Redis"""
    
    def __init__(self):
        self.redis_client = None
        # LRU local cache: key -> (value, expires_at)
        self.local_cache: 'OrderedDict[str, Tuple[Any, float]]' = OrderedDict()
        self._local_max_items = int(getattr(config, 'ASYNC_CACHE_LOCAL_MAX_ITEMS', 5000) or 5000)
        self._hits = 0
        self._misses = 0
        self._redis_hits = 0
        self._redis_misses = 0
        
    async def init_redis(self, redis_url: str = None):
        """Инициализация Redis соединения"""
        if not redis_url or redis_url == "":
            logger.info("Redis отключен (URL не указан)")
            self.redis_client = None
            return
            
        try:
            self.redis_client = redis.from_url(
                redis_url, 
                decode_responses=True,
                max_connections=config.REDIS_MAX_CONNECTIONS,
                socket_keepalive=True,
                retry_on_timeout=True,
                socket_timeout=config.REDIS_TIMEOUT
            )
            await self.redis_client.ping()
            logger.info("Redis connection established with connection pooling")
            
            await self.redis_client.set("vlmb:startup_time", datetime.now().isoformat())
            
        except Exception as e:
            logger.warning(f"Redis not available: {e}")
            self.redis_client = None
    
    async def get(self, key: str) -> Optional[Any]:
        """Асинхронное получение данных из кэша"""
        if not key:
            return None
        
        # Сначала проверяем локальный кэш (LRU + TTL)
        try:
            item = self.local_cache.get(key)
        except Exception:
            item = None
        if item is not None:
            data, expires_at = item
            if time.time() < float(expires_at):
                self._hits += 1
                # refresh LRU
                try:
                    self.local_cache.move_to_end(key)
                except Exception:
                    pass
                return data
            # expired
            try:
                self.local_cache.pop(key, None)
            except Exception:
                pass
        
        # Затем Redis
        if self.redis_client:
            try:
                cached = await self.redis_client.get(key)
                if cached:
                    data = json.loads(cached)
                    ttl = int(getattr(config, 'ASYNC_CACHE_DEFAULT_TTL', getattr(config, 'CACHE_TTL', 3600)) or 3600)
                    self._local_set(key, data, ttl)
                    self._hits += 1
                    self._redis_hits += 1
                    return data
                self._redis_misses += 1
            except Exception as e:
                logger.warning(f"Redis get error: {e}")
        
        self._misses += 1
        return None
    
    async def set(self, key: str, value: Any, ttl: int = None):
        """Асинхронное сохранение данных в кэш"""
        if not key or value is None:
            return

        if ttl is None:
            ttl = int(getattr(config, 'ASYNC_CACHE_DEFAULT_TTL', getattr(config, 'CACHE_TTL', 3600)) or 3600)

        def _make_jsonable(obj: Any):
            # TrackInfo / dataclass-like
            if hasattr(obj, "to_dict") and callable(getattr(obj, "to_dict")):
                try:
                    return obj.to_dict()
                except Exception:
                    pass
            # pydantic-like
            if hasattr(obj, "model_dump") and callable(getattr(obj, "model_dump")):
                try:
                    return obj.model_dump()
                except Exception:
                    pass
            # dict / list / tuple
            if isinstance(obj, dict):
                return {k: _make_jsonable(v) for k, v in obj.items()}
            if isinstance(obj, (list, tuple)):
                return [_make_jsonable(v) for v in obj]
            # basic types
            if isinstance(obj, (str, int, float, bool)) or obj is None:
                return obj
            # fallback to __dict__
            if hasattr(obj, "__dict__"):
                try:
                    return _make_jsonable(dict(obj.__dict__))
                except Exception:
                    pass
            return str(obj)

        try:
            if self.redis_client:
                await self.redis_client.setex(
                    key,
                    ttl,
                    json.dumps(_make_jsonable(value), ensure_ascii=False)
                )
        except Exception as e:
            logger.error(f"Redis set error: {e}")

        self._local_set(key, value, ttl)

    def _local_set(self, key: str, value: Any, ttl: int) -> None:
        """Insert into local LRU cache with TTL and size cap."""
        try:
            expires_at = time.time() + max(1, int(ttl or 1))
            if key in self.local_cache:
                self.local_cache.pop(key, None)
            self.local_cache[key] = (value, float(expires_at))
            self.local_cache.move_to_end(key)
            # evict
            while self._local_max_items and len(self.local_cache) > self._local_max_items:
                self.local_cache.popitem(last=False)
        except Exception as e:
            logger.debug(f"local cache set failed: {e}")
    
    async def delete(self, key: str):
        """Удаление ключа из всех кэшей"""
        if key in self.local_cache:
            del self.local_cache[key]
        
        if self.redis_client:
            try:
                await self.redis_client.delete(key)
            except Exception as e:
                logger.error(f"Redis delete error: {e}")
    
    async def get_cached_results(self, query: str) -> Optional[List[Dict]]:
        """Получение кэшированных результатов поиска"""
        cache_key = _search_cache_key(query)
        return await self.get(cache_key)
    
    async def set_cached_results(self, query: str, results: List[Dict]):
        """Сохранение результатов поиска в кэш"""
        cache_key = _search_cache_key(query)
        await self.set(cache_key, results)
    
    def get_cache_stats(self) -> Dict[str, Any]:
        """Статистика кэша для мониторинга"""
        total = self._hits + self._misses
        redis_total = self._redis_hits + self._redis_misses
        
        return {
            'local_cache_size': len(self.local_cache),
            'hit_rate': self._hits / total if total > 0 else 0,
            'redis_hit_rate': self._redis_hits / redis_total if redis_total > 0 else 0,
            'total_requests': total,
            'hits': self._hits,
            'misses': self._misses,
            'redis_hits': self._redis_hits,
            'redis_misses': self._redis_misses
        }

# ==================== АСИНХРОННЫЙ МЕНЕДЖЕР VK ТОКЕНОВ ====================
class AsyncVKTokenManager:
    """Асинхронный менеджер для управления VK токенами"""
    
    def __init__(self, db_path: str = config.VK_TOKENS_DB_PATH):
        self.db_path = db_path
        self.active_tokens: Dict[str, Dict] = {}
        self._current_token_index = 0
        self._last_usage_flush: Dict[str, float] = {}
    
    async def initialize(self):
        """Инициализация базы данных токенов"""
        await self._init_database()
        await self.load_active_tokens()
    
    async def _init_database(self):
        """Асинхронная инициализация БД токенов"""
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute('''
                CREATE TABLE IF NOT EXISTS vk_tokens (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    token TEXT UNIQUE NOT NULL,
                    added_by INTEGER NOT NULL,
                    added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    is_active INTEGER DEFAULT 1,
                    last_used TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    expires_at TIMESTAMP,
                    description TEXT
                )
            ''')
            
            await conn.execute('CREATE INDEX IF NOT EXISTS idx_tokens_active ON vk_tokens(is_active, last_used)')
            await conn.execute('CREATE INDEX IF NOT EXISTS idx_tokens_expires ON vk_tokens(expires_at)')
            
            await conn.commit()
    
    async def load_active_tokens(self):
        """Загрузка активных токенов из БД"""
        async with _sqlite_connection(self.db_path) as conn:
            cursor = await conn.execute('''
                SELECT token, expires_at, description 
                FROM vk_tokens 
                WHERE is_active = 1 
                ORDER BY last_used
            ''')
            rows = await cursor.fetchall()
            
            self.active_tokens.clear()
            for token, expires_at, description in rows:
                self.active_tokens[token] = {
                    'expires_at': expires_at,
                    'description': description,
                    'is_expired': await self._is_token_expired(expires_at)
                }
            
            logger.info(f"Loaded {len(self.active_tokens)} active VK tokens")
    
    async def _is_token_expired(self, expires_at: str) -> bool:
        """Проверка истечения срока токена"""
        if not expires_at:
            return False
        
        try:
            raw_value = str(expires_at).strip()
            if raw_value.endswith("Z"):
                raw_value = raw_value[:-1] + "+00:00"
            expiry_date = datetime.fromisoformat(raw_value)
            now = datetime.now(expiry_date.tzinfo) if expiry_date.tzinfo else datetime.now()
            return now >= expiry_date
        except (TypeError, ValueError) as exc:
            logger.warning("Invalid VK token expiration timestamp %r: %s", expires_at, exc)
            return False
    
    async def get_next_valid_token(self) -> Optional[str]:
        """Получение следующего валидного токена (round-robin).

        - Основной источник: активные токены из vk_tokens.db
        - Фолбэк: config.VK_TOKEN (если задан)

        Нормализует флаг is_expired, т.к. в БД/кэше он иногда бывает bool/int/str.
        """
        if not self.active_tokens:
            return str(getattr(config, 'VK_TOKEN', '')).strip() or None

        def _is_expired(info: object) -> bool:
            try:
                if not isinstance(info, dict):
                    return False
                v = info.get('is_expired', False)
                if v is None:
                    return False
                if isinstance(v, bool):
                    return v
                if isinstance(v, (int, float)):
                    return bool(v)
                if isinstance(v, str):
                    vv = v.strip().lower()
                    return vv in ('1','true','yes','y','on')
            except Exception:
                return False
            return False

        valid_tokens = [t for t, info in self.active_tokens.items() if not _is_expired(info)]
        if not valid_tokens:
            return None

        # round-robin
        self._current_token_index = (self._current_token_index + 1) % len(valid_tokens)
        selected_token = valid_tokens[self._current_token_index]
        await self.update_token_usage(selected_token)
        return selected_token

    async def update_token_usage(self, token: str):
        """Update token usage at a limited rate to avoid one SQLite write per search."""
        now = time.monotonic()
        interval = float(getattr(config, "VK_TOKEN_USAGE_FLUSH_SECONDS", 60) or 60)
        if now - self._last_usage_flush.get(token, 0.0) < max(1.0, interval):
            return
        self._last_usage_flush[token] = now
        try:
            async with _sqlite_connection(self.db_path) as conn:
                await conn.execute(
                    'UPDATE vk_tokens SET last_used = CURRENT_TIMESTAMP WHERE token = ?',
                    (token,)
                )
                await conn.commit()
        except Exception as e:
            self._last_usage_flush.pop(token, None)
            logger.error(f"Error updating token usage: {e}")
    
    async def add_token(self, token: str, added_by: int, description: str = None, expires_at: str = None) -> bool:
        """Добавление нового токена"""
        try:
            async with _sqlite_connection(self.db_path) as conn:
                await conn.execute(
                    '''
                    INSERT INTO vk_tokens (token, added_by, description, expires_at, is_active)
                    VALUES (?, ?, ?, ?, 1)
                    ON CONFLICT(token) DO UPDATE SET
                        added_by = excluded.added_by,
                        description = excluded.description,
                        expires_at = excluded.expires_at,
                        is_active = 1
                    ''',
                    (token, added_by, description, expires_at)
                )
                await conn.commit()
            
            await self.load_active_tokens()
            return True
        except Exception as e:
            logger.error(f"Error adding token: {e}")
            return False
    
    async def remove_token(self, token: str) -> bool:
        """Удаление токена"""
        try:
            async with _sqlite_connection(self.db_path) as conn:
                await conn.execute(
                    'UPDATE vk_tokens SET is_active = 0 WHERE token = ?',
                    (token,)
                )
                await conn.commit()
            
            await self.load_active_tokens()
            return True
        except Exception as e:
            logger.error(f"Error removing token: {e}")
            return False
    
    async def list_tokens(self) -> List[Dict]:
        """Получение списка всех токенов"""
        try:
            async with _sqlite_connection(self.db_path) as conn:
                cursor = await conn.execute('''
                    SELECT token, added_by, added_at, expires_at, description, is_active,
                           (SELECT COUNT(*) FROM vk_tokens WHERE token = vk_tokens.token) as usage_count
                    FROM vk_tokens
                    ORDER BY is_active DESC, last_used DESC
                ''')
                rows = await cursor.fetchall()
                
                tokens = []
                for row in rows:
                    tokens.append({
                        'token_full': row[0],
                        'token_masked': (row[0][:20] + '...' if len(row[0]) > 20 else row[0]),
                        'added_by': row[1],
                        'added_at': row[2],
                        'expires_at': row[3],
                        'description': row[4],
                        'is_active': bool(row[5]),
                        'is_expired': await self._is_token_expired(row[3]),
                        'usage_count': row[6]
                    })
                
                return tokens
        except Exception as e:
            logger.error(f"Error listing tokens: {e}")
            return []

# ==================== АСИНХРОННЫЙ КЛАСС LAST.FM API ====================
class AsyncLastFM:
    """Асинхронный клиент для Last.fm API"""
    
    def __init__(self, api_key: str):
        self.api_key = api_key
        self.base_url = config.LASTFM_BASE_URL
        self._request_semaphore = asyncio.Semaphore(config.LASTFM_BATCH_SIZE)
        self._last_request_time = 0
    
    async def _make_request(self, session: aiohttp.ClientSession, params: Dict) -> Optional[Dict]:
        """Асинхронный запрос к Last.fm API"""
        if not self.api_key:
            return None
        
        current_time = time.time()
        time_since_last = current_time - self._last_request_time
        if time_since_last < config.LASTFM_REQUEST_DELAY:
            await asyncio.sleep(config.LASTFM_REQUEST_DELAY - time_since_last)
        
        async with self._request_semaphore:
            try:
                self._last_request_time = time.time()
                async with session.get(self.base_url, params=params) as response:
                    if response.status == 200:
                        return await response.json()
                    elif response.status == 429:
                        logger.warning("Last.fm rate limit exceeded")
                        await asyncio.sleep(1)
                        return None
                    else:
                        logger.warning(f"Last.fm API error: {response.status}")
                        return None
            except asyncio.TimeoutError:
                logger.warning("Last.fm API timeout")
                return None
            except Exception as e:
                logger.error(f"Last.fm API error: {e}")
                return None
    
    async def get_top_tracks(self, session: aiohttp.ClientSession, limit: int = config.POPULAR_TRACKS_LIMIT) -> List[Dict[str, Any]]:
        """Получение популярных треков"""
        params = {
            'method': 'chart.gettoptracks',
            'api_key': self.api_key,
            'format': 'json',
            'limit': limit
        }
        
        data = await self._make_request(session, params)
        if data and 'tracks' in data and 'track' in data['tracks']:
            tracks = []
            for track_data in data['tracks']['track']:
                track = {
                    'name': await self.clean_text(track_data.get('name', 'Неизвестный трек')),
                    'artist': await self.clean_text(track_data['artist'].get('name', 'Неизвестный исполнитель')),
                    'playcount': int(track_data.get('playcount', 0)),
                    'listeners': int(track_data.get('listeners', 0)),
                    'url': track_data.get('url', ''),
                    'type': 'global'
                }
                tracks.append(track)
            return tracks
        return []
    
    async def get_top_tracks_by_genre(self, session: aiohttp.ClientSession, genre: str, limit: int = config.POPULAR_TRACKS_LIMIT) -> List[Dict[str, Any]]:
        """Получение треков по жанру"""
        params = {
            'method': 'tag.gettoptracks',
            'tag': genre,
            'api_key': self.api_key,
            'format': 'json',
            'limit': limit
        }
        
        data = await self._make_request(session, params)
        if data and 'tracks' in data and 'track' in data['tracks']:
            tracks = []
            for track_data in data['tracks']['track']:
                track = {
                    'name': await self.clean_text(track_data.get('name', 'Неизвестный трек')),
                    'artist': await self.clean_text(track_data['artist'].get('name', 'Неизвестный исполнитель')),
                    'playcount': int(track_data.get('playcount', 0)),
                    'listeners': int(track_data.get('listeners', 0)),
                    'url': track_data.get('url', ''),
                    'type': 'genre'
                }
                tracks.append(track)
            return tracks
        return []
    
    async def get_similar_artists(self, session: aiohttp.ClientSession, artist: str, limit: int = config.SIMILAR_ARTISTS_LIMIT) -> List[Dict[str, Any]]:
        """Получение похожих исполнителей"""
        params = {
            'method': 'artist.getsimilar',
            'artist': artist,
            'api_key': self.api_key,
            'format': 'json',
            'limit': limit
        }
        
        data = await self._make_request(session, params)
        if data and 'similarartists' in data and 'artist' in data['similarartists']:
            artists = []
            for artist_data in data['similarartists']['artist']:
                if artist_data.get('name'):
                    artist_info = {
                        'name': await self.clean_text(artist_data.get('name')),
                        'match': float(artist_data.get('match', 0)),
                        'url': artist_data.get('url', '')
                    }
                    artists.append(artist_info)
            return artists
        return []
    
    async def _handle_queued_download(self, job):
        """Worker callback used by playlist/batch jobs."""
        payload = dict(job.payload or {})
        update = payload.get("update")
        context = payload.get("context")
        track = payload.get("track")
        if update is None or context is None or not isinstance(track, dict):
            raise RuntimeError("invalid queued download payload")
        async with self.global_download_semaphore:
            return await download_and_send_audio(update, context, update.effective_message, track, update.effective_message.message_id)

    async def clean_text(self, text: str) -> str:
        """Очистка текста (легковесная, не требует executor)"""
        if not text:
            return ""
        try:
            return re.sub(r'<[^>]+>', '', ' '.join(text.split())).strip()
        except Exception:
            return text.strip()



# ==================== DEEZER: ЧАРТЫ И ПОДБОРКИ ====================
# Deezer используется только как источник метаданных популярных треков.
# Аудио берётся из Telegram file_id-кэша, а отсутствующие треки автоматически скачиваются через VK, Яндекс.Музыку или YouTube.
DEEZER_GENRES: "OrderedDict[int, str]" = OrderedDict([
    (0, "Все жанры"),
    (132, "Поп"),
    (116, "Рэп / хип-хоп"),
    (152, "Рок"),
    (464, "Метал"),
    (106, "Электроника"),
    (113, "Танцевальная музыка"),
    (165, "R&B"),
    (85, "Альтернатива"),
    (144, "Регги"),
    (129, "Джаз"),
    (98, "Классика"),
    (153, "Блюз"),
    (169, "Соул и фанк"),
])

DEEZER_GENRE_ALIASES: Dict[str, int] = {
    "global": 0,
    "all": 0,
    "pop": 132,
    "hip-hop": 116,
    "hiphop": 116,
    "rap": 116,
    "rock": 152,
    "metal": 464,
    "electronic": 106,
    "electro": 106,
    "dance": 113,
    "r&b": 165,
    "alternative": 85,
    "reggae": 144,
    "jazz": 129,
    "classical": 98,
    "blues": 153,
    "soul": 169,
    "funk": 169,
}

DIGEST_SCHEDULES: Dict[str, Dict[str, Any]] = {
    # Legacy presets are kept so existing subscriptions continue to work.
    "d09": {"label": "Каждый день в 09:00", "kind": "daily", "hour": 9, "minute": 0},
    "d18": {"label": "Каждый день в 18:00", "kind": "daily", "hour": 18, "minute": 0},
    "d21": {"label": "Каждый день в 21:00", "kind": "daily", "hour": 21, "minute": 0},
    "wd18": {"label": "По будням в 18:00", "kind": "weekdays", "hour": 18, "minute": 0},
    "fr18": {"label": "По пятницам в 18:00", "kind": "weekly", "weekday": 4, "hour": 18, "minute": 0},
    "su12": {"label": "По воскресеньям в 12:00", "kind": "weekly", "weekday": 6, "hour": 12, "minute": 0},
}

DIGEST_WEEKDAYS: Tuple[Tuple[int, str, str], ...] = (
    (0, "Пн", "понедельник"),
    (1, "Вт", "вторник"),
    (2, "Ср", "среда"),
    (3, "Чт", "четверг"),
    (4, "Пт", "пятница"),
    (5, "Сб", "суббота"),
    (6, "Вс", "воскресенье"),
)

_DIGEST_WIZARDS: Dict[Tuple[int, int], Dict[str, Any]] = {}
_DIGEST_SEND_LOCKS: Dict[int, asyncio.Lock] = {}
_DIGEST_SEND_LOCK_LAST_USED: Dict[int, float] = {}
_USER_MIX_WIZARDS: Dict[Tuple[int, int], Dict[str, Any]] = {}
_USER_MIX_RUNNING: Set[int] = set()
_USER_MIX_CHAT_RUNNING: Set[int] = set()


def _user_mix_key(user_id: int, chat_id: int) -> Tuple[int, int]:
    return int(user_id), int(chat_id)


def _user_mix_history_id(user_id: int, chat_id: int) -> int:
    """Stable non-subscription ID for per-user, per-chat mix history."""
    raw = hashlib.blake2b(
        f"{int(user_id)}:{int(chat_id)}".encode("utf-8"),
        digest_size=8,
    ).digest()
    # Keep the value inside SQLite's signed INTEGER range and far above real
    # autoincremented subscription IDs.
    return 2_000_000_000 + (int.from_bytes(raw, "big") % 4_000_000_000_000_000_000)


def _user_mix_no_repeat_days() -> int:
    try:
        return max(1, int(getattr(config, "USER_MIX_NO_REPEAT_DAYS", 7) or 7))
    except Exception:
        return 7


def _digest_send_lock(subscription_id: int) -> asyncio.Lock:
    """Serialize manual and scheduled sends for the same subscription."""
    sid = int(subscription_id)
    lock = _DIGEST_SEND_LOCKS.get(sid)
    if lock is None:
        lock = asyncio.Lock()
        _DIGEST_SEND_LOCKS[sid] = lock
    _DIGEST_SEND_LOCK_LAST_USED[sid] = time.monotonic()
    return lock


def _cleanup_digest_send_locks(max_idle_seconds: Optional[float] = None) -> int:
    """Drop only idle, unlocked per-subscription locks."""
    if max_idle_seconds is None:
        max_idle_seconds = float(getattr(config, "DIGEST_LOCK_TTL", 3600) or 3600)
    cutoff = time.monotonic() - max(60.0, float(max_idle_seconds))
    removed = 0
    for sid, lock in list(_DIGEST_SEND_LOCKS.items()):
        if lock.locked():
            continue
        if _DIGEST_SEND_LOCK_LAST_USED.get(sid, 0.0) >= cutoff:
            continue
        _DIGEST_SEND_LOCKS.pop(sid, None)
        _DIGEST_SEND_LOCK_LAST_USED.pop(sid, None)
        removed += 1
    return removed


def _digest_no_repeat_days() -> int:
    """How long a track is excluded after it was sent."""
    try:
        return max(1, int(getattr(config, "DIGEST_NO_REPEAT_DAYS", 120) or 120))
    except Exception:
        return 120


def _digest_timezone() -> Any:
    name = str(getattr(config, "DIGEST_TIMEZONE", "Europe/Amsterdam") or "Europe/Amsterdam")
    try:
        return ZoneInfo(name)
    except Exception:
        return timezone.utc


def _digest_encode_schedule(days: Set[int], hour: int, minute: int) -> str:
    """Encode a custom weekday/time schedule into the existing TEXT column."""
    normalized = sorted({int(day) for day in days if 0 <= int(day) <= 6})
    if not normalized:
        raise ValueError("At least one weekday is required")
    if not (0 <= int(hour) <= 23 and 0 <= int(minute) <= 59):
        raise ValueError("Invalid digest time")
    return f"c|{','.join(str(day) for day in normalized)}|{int(hour):02d}:{int(minute):02d}"


def _digest_parse_schedule(schedule_key: str) -> Tuple[Set[int], int, int]:
    """Return weekdays, hour and minute for custom and legacy schedules."""
    key = str(schedule_key or "")
    if key.startswith("c|"):
        try:
            _, raw_days, raw_time = key.split("|", 2)
            days = {int(value) for value in raw_days.split(",") if value != ""}
            hour_s, minute_s = raw_time.split(":", 1)
            hour, minute = int(hour_s), int(minute_s)
            if not days or any(day < 0 or day > 6 for day in days):
                raise ValueError
            if not (0 <= hour <= 23 and 0 <= minute <= 59):
                raise ValueError
            return days, hour, minute
        except Exception:
            logger.warning("Invalid custom digest schedule %r; using daily 18:00", schedule_key)
            return set(range(7)), 18, 0

    spec = DIGEST_SCHEDULES.get(key) or DIGEST_SCHEDULES["d18"]
    kind = str(spec.get("kind", "daily"))
    if kind == "daily":
        days = set(range(7))
    elif kind == "weekdays":
        days = set(range(5))
    else:
        days = {int(spec.get("weekday", 4))}
    return days, int(spec.get("hour", 18)), int(spec.get("minute", 0))


def _digest_next_run(schedule_key: str, from_ts: Optional[float] = None) -> float:
    """Return the next UTC timestamp for selected weekdays and local time."""
    days, hour, minute = _digest_parse_schedule(schedule_key)
    tz = _digest_timezone()
    base_ts = time.time() if from_ts is None else float(from_ts)
    now = datetime.fromtimestamp(base_ts, tz=tz)
    for offset in range(8):
        candidate = (now + timedelta(days=offset)).replace(
            hour=hour,
            minute=minute,
            second=0,
            microsecond=0,
        )
        if candidate.weekday() in days and candidate > now:
            return candidate.astimezone(timezone.utc).timestamp()

    # The loop always finds a date within seven days; keep a safe fallback.
    return (now + timedelta(days=1)).astimezone(timezone.utc).timestamp()


def _digest_normalize(value: str) -> str:
    value = (value or "").casefold().replace("ё", "е")
    value = re.sub(r"\s+(?:feat\.?|ft\.?)\s+.*$", "", value, flags=re.IGNORECASE)
    value = re.sub(r"[^\w\s]", " ", value, flags=re.UNICODE)
    return re.sub(r"\s+", " ", value).strip()


def _digest_track_keys(artist: str, title: str) -> List[str]:
    artist_full = _digest_normalize(artist)
    artist_main = _digest_normalize(re.split(r"\s*(?:,|&|/| x )\s*", artist or "", maxsplit=1)[0])
    title_full = _digest_normalize(title)
    title_base_raw = re.sub(r"\s*[\(\[].*?[\)\]]\s*$", "", title or "").strip()
    title_base = _digest_normalize(title_base_raw)
    keys: List[str] = []
    for a in (artist_full, artist_main):
        for t in (title_full, title_base):
            if a and t:
                key = f"{a}|{t}"
                if key not in keys:
                    keys.append(key)
    return keys


class AsyncDeezer:
    """Small public Deezer client for chart and genre metadata."""

    def __init__(self) -> None:
        self.base_url = str(getattr(config, "DEEZER_API_BASE_URL", "https://api.deezer.com") or "https://api.deezer.com").rstrip("/")
        self.timeout = float(getattr(config, "DEEZER_TIMEOUT", 10) or 10)
        self._semaphore = asyncio.Semaphore(int(getattr(config, "DEEZER_CONCURRENCY", 4) or 4))

    async def _get_json(self, session: aiohttp.ClientSession, path: str, params: Optional[Dict[str, Any]] = None) -> Optional[Dict[str, Any]]:
        if session is None or session.closed:
            return None
        url = f"{self.base_url}/{path.lstrip('/')}"
        try:
            async with self._semaphore:
                timeout = aiohttp.ClientTimeout(total=self.timeout)
                async with session.get(url, params=params or {}, timeout=timeout) as response:
                    if response.status == 429:
                        logger.warning("Deezer rate limit exceeded")
                        return None
                    if response.status != 200:
                        logger.warning("Deezer API error: HTTP %s", response.status)
                        return None
                    data = await response.json(content_type=None)
                    if isinstance(data, dict) and data.get("error"):
                        logger.warning("Deezer API returned error: %s", data.get("error"))
                        return None
                    return data if isinstance(data, dict) else None
        except asyncio.TimeoutError:
            logger.warning("Deezer API timeout")
        except Exception as exc:
            logger.warning("Deezer API request failed: %s", exc)
        return None

    async def get_chart_tracks(self, session: aiohttp.ClientSession, genre_id: int = 0, limit: int = 50) -> List[Dict[str, Any]]:
        limit = max(1, min(int(limit or 50), 100))
        data = await self._get_json(session, f"chart/{int(genre_id)}/tracks", {"limit": limit})
        rows = (data or {}).get("data") or []
        tracks: List[Dict[str, Any]] = []
        for item in rows:
            if not isinstance(item, dict):
                continue
            artist_obj = item.get("artist") or {}
            artist = str(artist_obj.get("name") or "").strip()
            title = str(item.get("title_short") or item.get("title") or "").strip()
            if not artist or not title:
                continue
            tracks.append({
                "artist": artist,
                "name": title,
                "rank": int(item.get("rank") or 0),
                "duration": int(item.get("duration") or 0),
                "deezer_id": item.get("id"),
                "link": item.get("link") or "",
                "type": "deezer_chart",
            })
        return tracks

# ==================== КЛАСС СЕССИЙ ПОИСКА ====================
class SearchSession:
    """Сессия поиска пользователя"""
    
    def __init__(self, session_id: str, user_id: int, query: str, results: List[Dict]):
        self.session_id = session_id
        self.user_id = user_id
        self.query = query
        self.results = results or []
        self.best_tracks = []
        self.original_message_id = None
        self.search_message_id = None
        self.chat_id = None
        self.message_thread_id = None
        self.created_at = time.time()
        self.last_accessed = time.time()

        # Stable track routing: uid -> track dict/TrackInfo
        self.tracks_by_uid: Dict[str, Any] = {}
        # Managed background enrich
        self.enrich_task: Optional[asyncio.Task] = None
        self.enrich_cancel_event: asyncio.Event = asyncio.Event()
        # Download-all UX
        self.download_all_cancel: bool = False
        self._download_all_last_update: float = 0.0

        self.reindex_tracks()

    def reindex_tracks(self) -> None:
        """Build uid index for results and best_tracks (best-effort)."""
        idx: Dict[str, Any] = {}
        try:
            for t in (self.results or []):
                uid = _track_uid_from_any(t)
                if uid:
                    idx[uid] = t
            for t in (self.best_tracks or []):
                uid = _track_uid_from_any(t)
                if uid and uid not in idx:
                    idx[uid] = t
        except Exception as e:
            logger.debug(f"reindex_tracks failed: {e}")
        self.tracks_by_uid = idx

    def get_track(self, ref: str, chat_type: Optional[str]) -> Optional[Any]:
        """Resolve a track either by stable uid or (legacy) by integer index.

        In groups we typically index into best_tracks; in private into results.
        """
        if not ref:
            return None
        # UID path
        if not str(ref).isdigit():
            t = self.tracks_by_uid.get(str(ref))
            if t is not None:
                return t
            # fallback: try compute for each track
            for t in (self.best_tracks or []) + (self.results or []):
                if _track_uid_from_any(t) == str(ref):
                    return t
            return None

        # legacy index path
        try:
            i = int(ref)
        except Exception:
            return None
        is_group = (chat_type in ("group", "supergroup"))
        preferred = (self.best_tracks if is_group else self.results) or []
        fallback = (self.results if is_group else self.best_tracks) or []
        source_list = preferred if preferred else fallback
        if 0 <= i < len(source_list):
            return source_list[i]
        return None
    
    def is_expired(self) -> bool:
        """Expire sessions after inactivity, not while the user is still paging."""
        timeout = int(getattr(config, "SEARCH_SESSION_TIMEOUT", 900) or 900)
        return time.time() - self.last_accessed > timeout
    
    def update_access(self):
        """Обновление времени доступа"""
        self.last_accessed = time.time()

# ==================== АСИНХРОННЫЙ МЕНЕДЖЕР СЕССИЙ ====================
class AsyncSessionManager:
    """Асинхронный менеджер сессий поиска (in-memory + optional Redis persistence)"""
    
    def __init__(self, cache_manager: Optional['AsyncCacheManager'] = None):
        self.cache_manager = cache_manager
        self.sessions: Dict[str, SearchSession] = {}
        self._cleanup_task = None
        self._background_tasks: Set[asyncio.Task] = set()
        self._session_access: Dict[str, float] = {}
        self._lock = asyncio.Lock()
    
    async def initialize(self):
        """Инициализация фоновой очистки"""
        await self._start_background_cleanup()

    def _spawn_task(self, coro, label: str) -> asyncio.Task:
        task = asyncio.create_task(coro, name=label)
        self._background_tasks.add(task)

        def _done(done_task: asyncio.Task) -> None:
            self._background_tasks.discard(done_task)
            if done_task.cancelled():
                return
            try:
                exc = done_task.exception()
            except Exception as task_exc:
                logger.warning("Session background task %s inspection failed: %s", label, task_exc)
                return
            if exc is not None:
                logger.error(
                    "Session background task %s failed",
                    label,
                    exc_info=(type(exc), exc, exc.__traceback__),
                )

        task.add_done_callback(_done)
        return task


    def _redis_enabled(self) -> bool:
        # Stateless callbacks: sessions must survive process restarts/sharding.
        # If Redis is available, always persist there.
        return getattr(self.cache_manager, "redis_client", None) is not None

    def _session_key(self, session_id: str) -> str:
        prefix = getattr(config, "SESSION_REDIS_PREFIX", "vlmb:session:")
        return f"{prefix}{session_id}"

    async def _persist_session(self, session: SearchSession):
        """Persist session to Redis (best-effort, non-fatal)."""
        if not self._redis_enabled():
            return
        try:
            payload = {
                "session_id": session.session_id,
                "user_id": session.user_id,
                "query": session.query,
                "results": session.results,
                "best_tracks": getattr(session, "best_tracks", []),
                "created_at": getattr(session, "created_at", None),
                "last_accessed": getattr(session, "last_accessed", None),
            }
            ttl = int(getattr(config, "SESSION_TTL_SECONDS", 21600))
            await self.cache_manager.redis_client.setex(self._session_key(session.session_id), ttl, json.dumps(payload, ensure_ascii=False, default=str))
        except Exception as e:
            logger.debug(f"Redis session persist failed for {session.session_id}: {e}")

    async def _load_session(self, session_id: str) -> Optional[SearchSession]:
        """Load session from Redis and rehydrate SearchSession."""
        if not self._redis_enabled():
            return None
        try:
            raw = await self.cache_manager.redis_client.get(self._session_key(session_id))
            if not raw:
                return None
            data = json.loads(raw)
            sess = SearchSession(data["session_id"], int(data["user_id"]), data.get("query") or "", data.get("results") or [])
            sess.best_tracks = data.get("best_tracks") or []
            # keep metadata if present
            try:
                if data.get("created_at") is not None:
                    sess.created_at = float(data.get("created_at"))
                if data.get("last_accessed") is not None:
                    sess.last_accessed = float(data.get("last_accessed"))
            except Exception:
                pass
            return sess
        except Exception as e:
            logger.debug(f"Redis session load failed for {session_id}: {e}")
            return None
    
    async def _start_background_cleanup(self):
        """Запуск фоновой очистки сессий"""
        async def cleanup_loop():
            while True:
                await asyncio.sleep(config.SESSION_CLEANUP_INTERVAL)
                await self.cleanup_expired()
        
        self._cleanup_task = self._spawn_task(cleanup_loop(), "session-cleanup")
    
    async def create_session(self, user_id: int, query: str, results: List[Dict]) -> str:
        """Create a session without holding the lock during Redis I/O."""
        async with self._lock:
            if len(self.sessions) >= int(getattr(config, "MAX_CACHED_SESSIONS", 2000) or 2000):
                self._force_cleanup_unlocked()

            # Session id must be unique even for back-to-back searches.
            session_id = f"{user_id}_{secrets.token_urlsafe(8)}"
            session = SearchSession(session_id, user_id, query, results)
            self.sessions[session_id] = session
            self._session_access[session_id] = time.time()

        await self._persist_session(session)
        return session_id

    async def get_session(self, session_id: str) -> Optional[SearchSession]:
        """Get a session without holding the global lock during Redis I/O."""
        async with self._lock:
            session = self.sessions.get(session_id)

        if session is None:
            loaded = await self._load_session(session_id)
            if loaded is None:
                return None
            async with self._lock:
                session = self.sessions.get(session_id) or loaded
                self.sessions[session_id] = session
                self._session_access[session_id] = time.time()

        async with self._lock:
            current = self.sessions.get(session_id)
            if current is None:
                return None
            if current.is_expired():
                self.sessions.pop(session_id, None)
                self._session_access.pop(session_id, None)
                return None
            current.update_access()
            self._session_access[session_id] = time.time()
            session = current

        self._spawn_task(self._persist_session(session), f"session-persist:{session_id}")
        return session

    async def cleanup_expired(self):
        """Очистка просроченных сессий"""
        async with self._lock:
            expired = [sid for sid, s in self.sessions.items() if s.is_expired()]
            for sid in expired:
                self.sessions.pop(sid, None)
                self._session_access.pop(sid, None)
        if expired:
            logger.info(f"Cleaned up {len(expired)} expired sessions")

    def _force_cleanup_unlocked(self) -> None:
        """Drop oldest sessions. Caller must hold ``self._lock``."""
        max_sessions = max(1, int(getattr(config, "MAX_CACHED_SESSIONS", 2000) or 2000))
        sorted_sessions = sorted(self._session_access.items(), key=lambda item: item[1])
        # Drop a proportional buffer to avoid cleaning on every new request.
        buffer_size = min(100, max(1, max_sessions // 10))
        target_size = max(0, max_sessions - buffer_size)
        drop_n = max(0, len(sorted_sessions) - target_size)
        for session_id, _ in sorted_sessions[:drop_n]:
            self.sessions.pop(session_id, None)
            self._session_access.pop(session_id, None)

    async def _force_cleanup(self) -> None:
        """Thread-safe public cleanup entry point."""
        async with self._lock:
            self._force_cleanup_unlocked()

    async def close(self) -> None:
        tasks = list(self._background_tasks)
        for task in tasks:
            task.cancel()
        if tasks:
            await asyncio.gather(*tasks, return_exceptions=True)
        self._background_tasks.clear()
        self._cleanup_task = None

    async def count_redis_sessions(self) -> int:
        """Считает количество сохранённых сессий в Redis.

        Используется только для диагностики/админки.
        Работает через cache_manager.redis_client и настройки SESSION_REDIS_PREFIX.
        """
        if not self._redis_enabled():
            return 0

        redis_cli = getattr(self.cache_manager, 'redis_client', None)
        if not redis_cli:
            return 0

        prefix = getattr(config, 'SESSION_REDIS_PREFIX', 'vlmb:session:')
        pattern = f"{prefix}*"

        try:
            if hasattr(redis_cli, 'scan_iter'):
                count = 0
                async for _ in redis_cli.scan_iter(match=pattern, count=500):
                    count += 1
                return count

            if hasattr(redis_cli, 'scan'):
                cursor = 0
                count = 0
                while True:
                    cursor, keys = await redis_cli.scan(cursor=cursor, match=pattern, count=500)
                    count += len(keys or [])
                    if cursor in (0, '0', None):
                        break
                return count

            if hasattr(redis_cli, 'keys'):
                keys = await redis_cli.keys(pattern)
                return len(keys or [])

        except Exception as e:
            logger.debug(f"Redis session count failed: {e}")

        return 0

# ==================== АСИНХРОННЫЙ МЕНЕДЖЕР ПОЛЬЗОВАТЕЛЕЙ ====================
class AsyncUserManager:
    """Асинхронный менеджер пользователей и лимитов"""
    
    def __init__(self):
        self.banned_users: Set[int] = set()
        self.requests: Dict[int, Dict[str, List[float]]] = {}
        self.limits = config.RATE_LIMITS
        self.user_cooldowns: Dict[int, float] = {}
        self._cleanup_task = None
        self._ban_refresh_lock = asyncio.Lock()
        self._last_ban_refresh = 0.0
    
    async def initialize(self):
        """Инициализация фоновых задач"""
        await self._start_background_cleanup()
    
    async def _start_background_cleanup(self):
        """Запуск фоновой очистки"""
        async def cleanup_loop():
            while True:
                await asyncio.sleep(config.SESSION_CLEANUP_INTERVAL)
                await self._cleanup_old_data()
        
        self._cleanup_task = asyncio.create_task(cleanup_loop(), name="user-manager-cleanup")

        def _cleanup_done(task: asyncio.Task) -> None:
            if task.cancelled():
                return
            try:
                exc = task.exception()
            except Exception:
                logger.exception("Failed to inspect user-manager cleanup task")
                return
            if exc is not None:
                logger.error(
                    "User-manager cleanup task failed",
                    exc_info=(type(exc), exc, exc.__traceback__),
                )

        self._cleanup_task.add_done_callback(_cleanup_done)
    
    async def _cleanup_old_data(self):
        """Очистка устаревших данных"""
        current_time = time.time()
        expired_users = []
        
        for user_id, user_requests in list(self.requests.items()):
            for action, timestamps in list(user_requests.items()):
                new_timestamps = [ts for ts in timestamps if current_time - ts < 3600]
                if new_timestamps:
                    user_requests[action] = new_timestamps
                else:
                    del user_requests[action]
            
            if not user_requests:
                expired_users.append(user_id)
        
        for user_id in expired_users:
            del self.requests[user_id]
        
        expired_cooldowns = [
            user_id for user_id, timestamp in self.user_cooldowns.items()
            if current_time - timestamp > 3600
        ]
        for user_id in expired_cooldowns:
            del self.user_cooldowns[user_id]
    
    async def validate_query(self, user_id: int, query: str) -> bool:
        """Minimal technical validation: a query must be non-empty and reasonably sized."""
        if await self.is_user_banned(user_id):
            return False
        q = (query or "").strip()
        if not q:
            return False
        return len(q) <= int(getattr(config, "MAX_QUERY_LENGTH", 200) or 200)
    
    async def refresh_bans(self, admin_db: Any = None, force: bool = False) -> None:
        """Refresh the in-memory ban set from the persistent SQLite table."""
        interval = max(1.0, float(getattr(config, "BAN_CACHE_REFRESH_SECONDS", 30) or 30))
        now = time.monotonic()
        if not force and now - self._last_ban_refresh < interval:
            return
        async with self._ban_refresh_lock:
            now = time.monotonic()
            if not force and now - self._last_ban_refresh < interval:
                return
            if admin_db is None:
                instance = globals().get("bot_instance")
                admin_db = getattr(instance, "admin_db", None) if instance else None
            if admin_db is None:
                return
            try:
                self.banned_users = set(await admin_db.get_banned_user_ids())
                self._last_ban_refresh = time.monotonic()
            except Exception:
                logger.exception("Failed to refresh persistent ban cache")

    async def is_user_banned(self, user_id: int) -> bool:
        """Check the persistent ban cache and refresh it periodically."""
        await self.refresh_bans(force=False)
        return int(user_id) in self.banned_users
    
    async def check_rate_limit(self, user_id: int, action: str) -> bool:
        """Проверка лимитов запросов"""
        if action not in self.limits:
            return True
            
        current_time = time.time()
        limit, window = self.limits[action]
        
        if user_id not in self.requests:
            self.requests[user_id] = {}
        if action not in self.requests[user_id]:
            self.requests[user_id][action] = []
        
        self.requests[user_id][action] = [
            ts for ts in self.requests[user_id][action] 
            if current_time - ts < window
        ]
        
        if len(self.requests[user_id][action]) >= limit:
            return False
        
        self.requests[user_id][action].append(current_time)
        return True
    
    async def check_cooldown(self, user_id: int) -> bool:
        """Проверка кулдауна"""
        now = time.time()
        if user_id in self.user_cooldowns and now - self.user_cooldowns[user_id] < config.COOLDOWN_SECONDS:
            return False
        self.user_cooldowns[user_id] = now
        return True

# ==================== АСИНХРОННЫЙ МЕНЕДЖЕР ФАЙЛОВ ====================
class AsyncFileManager:
    """Асинхронный менеджер для работы с файлами"""
    
    def __init__(self):
        self.temp_dir = tempfile.mkdtemp(prefix=config.TEMP_DIR_PREFIX)
        logger.info(f"Created temp directory: {self.temp_dir}")
    
    async def cleanup_old_files(self, max_age: int = None):
        """Асинхронная очистка старых файлов"""
        if max_age is None:
            max_age = config.TEMP_FILE_MAX_AGE
            
        current_time = time.time()
        deleted_count = 0
        
        try:
            files = await async_os.listdir(self.temp_dir)
            for filename in files:
                path = os.path.join(self.temp_dir, filename)
                try:
                    stat = await async_os.stat(path)
                    if current_time - stat.st_mtime > max_age:
                        await async_os.remove(path)
                        deleted_count += 1
                except Exception as e:
                    logger.error(f"Error cleaning up file {path}: {e}")
            
            if deleted_count > 0:
                logger.info(f"Cleaned up {deleted_count} temporary files")
        except Exception as e:
            logger.error(f"Error listing temp directory: {e}")

    async def close(self) -> None:
        """Remove the bot-owned temporary directory on shutdown."""
        temp_dir = self.temp_dir
        if not temp_dir or not os.path.isdir(temp_dir):
            return
        try:
            await asyncio.to_thread(shutil.rmtree, temp_dir, True)
            logger.info("Removed temp directory: %s", temp_dir)
        except Exception:
            logger.exception("Failed to remove temp directory %s", temp_dir)

# ==================== АСИНХРОННЫЙ МЕНЕДЖЕР ПАМЯТИ ====================

# ==================== ОСНОВНОЙ КЛАСС БОТА ====================

# ==================== МОНИТОРИНГ (безопасный no-op по умолчанию) ====================

class AsyncMusicBot:
    """Полностью асинхронный музыкальный бот с Яндекс.Музыкой"""
    
    def __init__(self):
        self.user_sessions: Dict[int, Dict] = {}
        # Compatibility mirror for legacy handlers; persistence boundary lives in storage.
        self.user_states: Dict[int, Dict] = {}
        self.state_store = InMemoryUserStateStore()
        self.session: Optional[aiohttp.ClientSession] = None
        self._initialized = False
        
        
        # Telegram API caches (reduce latency in groups)
        self._tg_me: Optional[Any] = None
        self._tg_me_ts: float = 0.0
        self._tg_chat_member_cache: Dict[int, Any] = {}  # chat_id -> {"ts": float, "member": obj}
        # Инициализация асинхронных менеджеров
        self.user_manager = AsyncUserManager()
        self.cache = AsyncCacheManager()
        self.file_manager = AsyncFileManager()
        self.lastfm = AsyncLastFM(config.LASTFM_API_KEY) if config.LASTFM_API_KEY else None
        self.deezer = AsyncDeezer()
        self.digest_store = None
        self.session_manager = AsyncSessionManager(self.cache)
        self.vk_token_manager = AsyncVKTokenManager()
        self.yandex_music = YandexMusicManager()
        self.youtube_music = YoutubeMusicManager()
        # VLMB 4.0 provider boundary. Existing managers remain the implementation;
        # application code can migrate to these adapters incrementally.
        self.providers = {
            "yandex": YandexProviderAdapter(self.yandex_music),
            "vk": VKProviderAdapter(self.search_vk_music),
            "youtube": YouTubeProviderAdapter(self.youtube_music),
        }
        self.provider_health = ProviderHealth()
        self.provider_router = ProviderRouter(
            cooldown_seconds=getattr(config, "PROVIDER_COOLDOWN_SECONDS", 30),
            failure_threshold=getattr(config, "PROVIDER_FAILURE_THRESHOLD", 2),
        )
        self.metrics = MetricsRegistry(getattr(config, "METRICS_HISTORY_SAMPLES", 5000))
        self.playlist_manager = PlaylistManager(2)
        self.download_queue = DownloadQueue(
            worker_count=getattr(config, "DOWNLOAD_QUEUE_WORKERS", 3),
            max_size=getattr(config, "DOWNLOAD_QUEUE_MAX_SIZE", 100),
        )
        self.global_download_semaphore = asyncio.Semaphore(max(1, int(getattr(config, "GLOBAL_DOWNLOAD_LIMIT", 12) or 12)))

        # User data (favorites/history/prefs)
        self.user_store: Optional[UserStore] = None
        
        configured_downloads = max(1, int(getattr(config, "MAX_CONCURRENT_DOWNLOADS", 5) or 5))
        max_file_size = max(1, int(getattr(config, "MAX_FILE_SIZE", 50 * 1024 * 1024) or 50 * 1024 * 1024))
        memory_budget = max(1, int(getattr(config, "AUDIO_MEMORY_BUDGET_MB", 512) or 512)) * 1024 * 1024
        # Download bytes are copied once into BytesIO before Telegram upload. Reserve
        # roughly 2x MAX_FILE_SIZE per slot to bound the worst-case memory spike.
        memory_slots = max(1, memory_budget // max(1, max_file_size * 2))
        self._download_concurrency = max(1, min(configured_downloads, memory_slots))
        if self._download_concurrency < configured_downloads:
            logger.warning(
                "Download concurrency reduced from %s to %s by AUDIO_MEMORY_BUDGET_MB=%s",
                configured_downloads, self._download_concurrency, memory_budget // (1024 * 1024),
            )
        self.download_semaphore = asyncio.Semaphore(self._download_concurrency)
        # Bound the complete in-memory audio lifecycle (download -> Telegram upload),
        # not only the network download itself. This prevents completed downloads
        # from piling up in RAM while Telegram uploads are still in progress.
        self.audio_memory_semaphore = asyncio.Semaphore(self._download_concurrency)
        # Per-chat download concurrency limits (private/group).
        self._chat_download_semaphores = {}  # chat_id -> asyncio.Semaphore
        self._chat_download_sem_last_used: Dict[int, float] = {}
        self._chat_download_sem_limits: Dict[int, int] = {}
        self._chat_download_limits_private = getattr(config, 'MAX_CONCURRENT_DOWNLOADS_PRIVATE', 4)
        self._chat_download_limits_group = getattr(config, 'MAX_CONCURRENT_DOWNLOADS_GROUP', 2)
        self.background_tasks: Set[asyncio.Task] = set()

        # --- Excel warmup (admin) state ---
        # Users in this set requested cancel for current warmup run.
        self.excel_warmup_cancel: Set[int] = set()
        # Users currently running warmup (to prevent duplicates).
        self.excel_warmup_processing: Set[int] = set()
        self._background_semaphore = asyncio.Semaphore(50)
        
        self._query_cache = {}
        self._user_cache = {}
        self._last_cache_cleanup = time.time()
        
        # Activity tracking
        self._last_user_activity = time.time()

        # --- TG file_id local LRU cache (fast-path, limits memory) ---
        self._tg_file_id_lru: "OrderedDict[str, str]" = OrderedDict()
        self._tg_file_id_lru_max: int = int(getattr(config, "TG_FILE_ID_LOCAL_CACHE_MAX", 50000) or 50000)
        self._tg_file_id_lock = asyncio.Lock()

        # --- SQLite reuse for tg_audio_cache (reduces connect/close overhead) ---
        self._tg_db_conn: Optional[aiosqlite.Connection] = None
        self._tg_db_lock = asyncio.Lock()

        # --- Excel warmup concurrency limiter ---
        self._excel_warmup_sem = asyncio.Semaphore(int(getattr(config, "EXCEL_WARMUP_CONCURRENCY", 2) or 2))

    

    def track_background_task(self, task: asyncio.Task, label: str = "background") -> asyncio.Task:
        """Register a task for shutdown and report otherwise-lost exceptions."""
        self.background_tasks.add(task)

        def _done(done_task: asyncio.Task) -> None:
            self.background_tasks.discard(done_task)
            if done_task.cancelled():
                return
            try:
                exc = done_task.exception()
            except Exception as task_exc:
                logger.warning("Background task %s inspection failed: %s", label, task_exc)
                return
            if exc is not None:
                logger.error(
                    "Background task %s failed",
                    label,
                    exc_info=(type(exc), exc, exc.__traceback__),
                )

        task.add_done_callback(_done)
        return task

    def create_background_task(self, coro, label: str = "background") -> asyncio.Task:
        return self.track_background_task(asyncio.create_task(coro, name=label), label)

    @staticmethod
    def _vk_info_is_expired(info: object) -> bool:
        """Normalize token 'is_expired' flag coming from DB/cache.

        В разных версиях БД/кода флаг мог оказаться bool/int/str.
        """
        try:
            if not isinstance(info, dict):
                return False
            v = info.get('is_expired', False)
            if v is None:
                return False
            if isinstance(v, bool):
                return v
            if isinstance(v, (int, float)):
                return bool(v)
            if isinstance(v, str):
                vv = v.strip().lower()
                return vv in ('1', 'true', 'yes', 'y', 'on')
        except Exception:
            return False
        return False

    def vk_enabled(self) -> bool:
        """VK доступен, если включён в конфиге и есть хотя бы один валидный токен.

        Источники токенов:
          1) vk_tokens.db (основной)
          2) config.VK_TOKEN (только фолбэк)
        """
        if not getattr(config, 'ENABLE_VK_MUSIC', True):
            return False

        # 1) DB tokens
        try:
            tm = getattr(self, 'vk_token_manager', None)
            tokens = getattr(tm, 'active_tokens', None) if tm else None
            if isinstance(tokens, dict) and tokens:
                for _tok, info in tokens.items():
                    if not self._vk_info_is_expired(info):
                        return True
        except Exception:
            pass

        # 2) Fallback token
        return bool(str(getattr(config, 'VK_TOKEN', '')).strip())






    def get_chat_download_semaphore(self, chat_id: int, chat_type: str = "private") -> asyncio.Semaphore:
        """Per-chat semaphore to limit concurrent downloads.

        Some codepaths expect this method on AsyncMusicBot. We keep the limiter here (not in session manager)
        to avoid attribute mismatches and keep button downloads stable.
        """
        # Lazy-init storage to stay backward compatible with older saved objects / hot reloads
        if not hasattr(self, "_chat_download_semaphores") or self._chat_download_semaphores is None:
            self._chat_download_semaphores: Dict[int, asyncio.Semaphore] = {}

        sem = self._chat_download_semaphores.get(chat_id)
        if sem is not None:
            self._chat_download_sem_last_used[chat_id] = time.monotonic()
            return sem

        # Limits (configurable)
        private_limit = getattr(config, "MAX_CONCURRENT_DOWNLOADS_PRIVATE", 4)
        group_limit = getattr(config, "MAX_CONCURRENT_DOWNLOADS_GROUP", 2)
        limit = private_limit if chat_type == "private" else group_limit
        try:
            limit_i = int(limit)
        except Exception:
            limit_i = 1
        if limit_i <= 0:
            limit_i = 1

        sem = asyncio.Semaphore(limit_i)
        self._chat_download_semaphores[chat_id] = sem
        self._chat_download_sem_last_used[chat_id] = time.monotonic()
        self._chat_download_sem_limits[chat_id] = limit_i
        return sem

    def _cleanup_chat_download_semaphores(self) -> int:
        """Remove idle per-chat semaphores without touching active/waiting users."""
        ttl = max(60.0, float(getattr(config, "CHAT_SEMAPHORE_TTL", 3600) or 3600))
        cutoff = time.monotonic() - ttl
        removed = 0
        for chat_id, sem in list(self._chat_download_semaphores.items()):
            if self._chat_download_sem_last_used.get(chat_id, 0.0) >= cutoff:
                continue
            limit = self._chat_download_sem_limits.get(chat_id, 1)
            waiters = getattr(sem, "_waiters", None)
            has_waiters = bool(waiters)
            is_active = int(getattr(sem, "_value", 0)) < int(limit)
            if has_waiters or is_active:
                continue
            self._chat_download_semaphores.pop(chat_id, None)
            self._chat_download_sem_last_used.pop(chat_id, None)
            self._chat_download_sem_limits.pop(chat_id, None)
            removed += 1
        return removed


    async def _get_bot_me_cached(self, bot: Any) -> Any:
        """Cache bot.get_me() result to avoid frequent Telegram API calls."""
        if self._tg_me is not None:
            return self._tg_me
        self._tg_me = await bot.get_me()
        self._tg_me_ts = time.time()
        return self._tg_me

    async def _get_bot_chat_member_cached(self, bot: Any, chat_id: int) -> Any:
        """Cache bot.get_chat_member(chat_id, me.id) with TTL (default 300s)."""
        me = await self._get_bot_me_cached(bot)
        ttl = getattr(config, "CHAT_MEMBER_CACHE_TTL", 300)
        now = time.time()
        rec = self._tg_chat_member_cache.get(chat_id)
        if rec and (now - float(rec.get("ts", 0.0)) < float(ttl)):
            return rec["member"]
        member = await bot.get_chat_member(chat_id, me.id)
        self._tg_chat_member_cache[chat_id] = {"ts": now, "member": member}
        return member

    async def initialize(self):
        """Полная инициализация бота"""
        await self.user_manager.initialize()
        await self.session_manager.initialize()
        await self.vk_token_manager.initialize()
        await self.yandex_music.initialize()
        await self.youtube_music.initialize()
        
        await self.init_session()
        # Reuse the shared HTTP session for Yandex.Music downloads.
        try:
            self.yandex_music.http_session = getattr(self, 'session', None)
        except Exception:
            pass
        await self.cache.init_redis(config.REDIS_URL)
        await self.download_queue.start(self._handle_queued_download)

        # Init user store (favorites/history/prefs) on the same DB file
        try:
            self.user_store = UserStore(getattr(config, "STATS_DB_PATH", "bot_stats.db"))
            await self.user_store.init()
        except Exception as e:
            logger.warning(f"UserStore init failed: {e}")
        
        await self._start_background_tasks()
        self._initialized = True
        logger.info("🔄 Асинхронный бот полностью инициализирован")
    
    async def _start_background_tasks(self):
        """Запуск фоновых задач (без warning 'was never awaited')."""
        # Собираем функции (не создаём корутины заранее, чтобы не было 'never awaited' при ошибках).
        task_factories = [
            self._periodic_cleanup,
            self._metrics_collector,
            self._cache_cleanup,
        ]
        # Keepalive может быть отключён в конфиге или отсутствовать в сборке.
        if hasattr(self, "_keepalive_loop") and callable(getattr(self, "_keepalive_loop")):
            task_factories.append(self._keepalive_loop)

        for fn in task_factories:
            try:
                self.create_background_task(fn(), getattr(fn, "__name__", "background"))
            except Exception as e:
                logger.warning(f"Failed to start background task {getattr(fn, '__name__', fn)}: {e}")

    async def init_session(self) -> None:
        """Инициализация HTTP сессии (ускорение + устойчивость).

        - Настраиваем отдельные таймауты connect/sock_read;
        - Пул соединений + DNS cache + keep-alive;
        - limit_per_host подстраивается под параллельные скачивания/поиск.
        """
        # Таймауты: total оставляем большим, но критично ограничить connect/sock_read,
        # чтобы подвисшие соединения не держали семафоры и не приводили к каскаду ошибок.
        total_to = float(getattr(config, "REQUEST_TIMEOUT", 25) or 25)
        connect_to = float(getattr(config, "HTTP_CONNECT_TIMEOUT", 8) or 8)
        sock_connect_to = float(getattr(config, "HTTP_SOCK_CONNECT_TIMEOUT", 8) or 8)
        sock_read_to = float(getattr(config, "HTTP_SOCK_READ_TIMEOUT", 25) or 25)

        timeout = aiohttp.ClientTimeout(
            total=total_to,
            connect=connect_to,
            sock_connect=sock_connect_to,
            sock_read=sock_read_to,
        )

        max_conn = int(getattr(config, "MAX_HTTP_CONNECTIONS", 140) or 140)

        # limit_per_host: если не задан — считаем от параллельности скачиваний/поиска.
        lph = int(getattr(config, "HTTP_LIMIT_PER_HOST", 0) or 0)
        if lph <= 0:
            try:
                dl = int(getattr(config, "MAX_CONCURRENT_DOWNLOADS", 25) or 25)
            except Exception:
                dl = 25
            # не раздуваем слишком сильно, но даём достаточно для пачек скачиваний
            lph = max(8, min(32, dl))

        if getattr(config, "ENABLE_CONNECTION_POOL", True):
            connector = aiohttp.TCPConnector(
                limit=max_conn,
                limit_per_host=lph,
                keepalive_timeout=float(getattr(config, 'HTTP_KEEPALIVE_TIMEOUT', 300) or 300),
                ttl_dns_cache=int(getattr(config, 'DNS_CACHE_TTL', 1800) or 1800),
                enable_cleanup_closed=True
            )
        else:
            connector = aiohttp.TCPConnector(limit=max_conn)

        self.session = aiohttp.ClientSession(
            timeout=timeout,
            connector=connector,
            headers={'User-Agent': getattr(config, "USER_AGENT", "Mozilla/5.0")}
        )

        # Reuse the same aiohttp session for Yandex.Music direct-link downloads
        try:
            if hasattr(self, 'yandex_music') and self.yandex_music:
                self.yandex_music.http_session = self.session
        except Exception:
            pass

        logger.info(f"HTTP сессия создана (limit={max_conn}, limit_per_host={lph})")
    
    async def _cleanup_caches(self):
        """Очистка кэшей"""
        current_time = time.time()
        
        expired_queries = [
            key for key, (_, timestamp) in self._query_cache.items()
            if current_time - timestamp > 300
        ]
        for key in expired_queries:
            del self._query_cache[key]
        
        expired_users = [
            user_id for user_id, timestamp in self._user_cache.items()
            if current_time - timestamp > 600
        ]
        for user_id in expired_users:
            del self._user_cache[user_id]
        
        if expired_queries or expired_users:
            logger.info(f"Очищены кэши: {len(expired_queries)} запросов, {len(expired_users)} пользователей")
    
    async def _periodic_cleanup(self):
        """Периодическая очистка"""
        while True:
            try:
                await asyncio.sleep(config.CLEANUP_INTERVAL)
                await self.file_manager.cleanup_old_files()
                await self.session_manager.cleanup_expired()
                removed_chat_sems = self._cleanup_chat_download_semaphores()
                removed_digest_locks = _cleanup_digest_send_locks()
                if removed_chat_sems or removed_digest_locks:
                    logger.info(
                        "Cleaned idle synchronization objects: chat_semaphores=%s digest_locks=%s",
                        removed_chat_sems, removed_digest_locks,
                    )
            except asyncio.CancelledError:
                raise
            except Exception:
                logger.exception("Ошибка в периодической очистке")
    
    async def _metrics_collector(self):
        """Сбор метрик"""
        while True:
            await asyncio.sleep(300)
            try:
                cache_stats = self.cache.get_cache_stats()
                cache_hits = cache_stats['hits']
                cache_misses = cache_stats['misses']
                total = cache_hits + cache_misses
                hit_ratio = cache_hits / total if total > 0 else 0
                
                logger.info("Performance metrics:")
                logger.info(f"  Cache: {cache_hits}/{total} ({hit_ratio:.1%}) hit ratio")
                logger.info(f"  Redis Cache: {cache_stats['redis_hits']}/{cache_stats['redis_hits'] + cache_stats['redis_misses']} ({cache_stats['redis_hit_rate']:.1%}) hit ratio")
                redis_sessions = await self.session_manager.count_redis_sessions()
                logger.info(f"  Active sessions (RAM): {len(self.session_manager.sessions)}")
                logger.info(f"  Active sessions (Redis): {redis_sessions}")
                memory_percent = _process_memory_percent()
                if memory_percent is not None:
                    logger.info(f"  Memory usage: {memory_percent:.1f}%")
                logger.info(f"  Background tasks: {len(self.background_tasks)}")
                provider_lines = self.provider_health.format_lines()
                if provider_lines:
                    logger.info("Provider health:")
                    for line in provider_lines:
                        logger.info(line)
                
            except Exception as e:
                logger.error(f"Metrics collection error: {e}")
    
    async def _cache_cleanup(self):
        """Очистка кэшей"""
        while True:
            await asyncio.sleep(600)
            try:
                current_time = time.time()
                if current_time - self._last_cache_cleanup > 600:
                    await self._cleanup_caches()
                    self._last_cache_cleanup = current_time
            except Exception as e:
                logger.error(f"Cache cleanup error: {e}")
    
#         async def _keepalive_loop(self):
# #         """Keepalive loop to avoid 'first request after idle' latency.
# 
#         It runs only when the bot was idle for a while and performs a tiny YM search
#         to keep DNS/TLS/HTTP sessions warm.
#         """
#         interval = int(getattr(config, "KEEPALIVE_INTERVAL_SECONDS", 300) or 300)
#         idle = int(getattr(config, "KEEPALIVE_IDLE_SECONDS", 120) or 120)
#         warmup_query = str(getattr(config, "KEEPALIVE_WARMUP_QUERY", "music") or "music")
# 
#         while True:
#             try:
#                 await asyncio.sleep(interval)
#                 # only warm up if no user activity for `idle` seconds
#                 if time.time() - getattr(self, "_last_user_activity", 0) < idle:
#                     continue
# 
#                 # ensure aiohttp session exists
#                 try:
#                     if not getattr(self, "session", None) or self.session.closed:
#                         await self.init_session()
#                 except Exception:
#                     pass
# 
#                 # warm up YM (VK keepalive omitted to avoid captcha loops)
#                 try:
#                     if getattr(config, "ENABLE_YANDEX_MUSIC", True) and getattr(self, "yandex_music", None) and self.yandex_music._initialized:
#                         await self.yandex_music.search_tracks(warmup_query, limit=1)
#                 except Exception as e:
#                     logger.debug(f"Keepalive YM failed: {e}")
#             except Exception as e:
#                 logger.debug(f"Keepalive loop error: {e}")
# 
# def mark_user_activity(self) -> None:
#         """Отметить пользовательскую активность"""
#         self._last_user_activity = time.time()
#     
    def mark_user_activity(self) -> None:
        """Отметить пользовательскую активность (для keepalive после простоя)."""
        self._last_user_activity = time.time()

    async def _keepalive_loop(self):
        """Keepalive loop to avoid first-request-after-idle latency.

        Запускается фоном. Если бот простаивал дольше KEEPALIVE_IDLE_SECONDS,
        выполняет лёгкий поиск в Яндекс.Музыке, чтобы прогреть DNS/TLS/сессию.
        """
        idle = float(getattr(config, "KEEPALIVE_IDLE_SECONDS", 120) or 120)
        interval = float(getattr(config, "KEEPALIVE_INTERVAL_SECONDS", 300) or 300)
        warmup_query = str(getattr(config, "KEEPALIVE_WARMUP_QUERY", "music") or "music")

        while True:
            try:
                await asyncio.sleep(interval)
                # only warm up if no user activity for `idle` seconds
                if time.time() - getattr(self, "_last_user_activity", 0) < idle:
                    continue

                # warm up YM only if enabled/initialized
                ym = getattr(self, "yandex_music", None) or getattr(self, "yandex_manager", None)
                if not ym:
                    continue

                try:
                    await ym.search_tracks(warmup_query, limit=1)
                except Exception as e:
                    logger.debug(f"Keepalive YM warmup failed: {e}")

            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.debug(f"Keepalive loop error: {e}")

    async def set_user_state(self, user_id: int, state: str, data: Dict = None):
        """Set state through the storage boundary while preserving legacy access."""
        payload = dict(data or {})
        await self.state_store.set(user_id, state, payload)
        self.user_states[user_id] = {
            'state': state, 'data': payload, 'timestamp': time.time()
        }

    async def get_user_state(self, user_id: int) -> Optional[Dict]:
        """Get state through storage; keep the legacy timestamp contract."""
        state_data = await self.state_store.get(user_id)
        if not state_data:
            return None
        legacy = self.user_states.get(user_id)
        if legacy and time.time() - legacy['timestamp'] < config.USER_STATE_TIMEOUT:
            return legacy
        if legacy:
            self.user_states.pop(user_id, None)
            await self.state_store.clear(user_id)
        return None

    async def clear_user_state(self, user_id: int):
        """Clear state through the storage boundary."""
        await self.state_store.clear(user_id)
        self.user_states.pop(user_id, None)
    
    @staticmethod
    def _artist_query(query: str) -> str:
        """Return the performer part of a query such as 'Ария - Штиль'."""
        q = _normalize_user_query(query)
        for separator in (" - ", " — ", " – "):
            if separator in q:
                left = q.split(separator, 1)[0].strip()
                if left:
                    return left
        return q

    @classmethod
    def rank_tracks_by_artist(cls, tracks: List[Dict[str, Any]], query: str) -> List[Dict[str, Any]]:
        """Production search ranking: relevance, artist/title match and deduplication."""
        return _rank_tracks(tracks, query)
    

    

    
    async def find_best_tracks(self, songs: List[Dict[str, Any]], query: str, count: int = 5) -> List[Dict[str, Any]]:
        """Compatibility wrapper around the single artist-first ranking rule."""
        ranked = self.rank_tracks_by_artist(songs, query)
        return ranked[:max(0, int(count or 0))]
    
    async def search_vk_music(self, query: str, count: int = 30) -> List[Dict[str, Any]]:
        """Search VK without duration, word, bitrate, or quality filters."""
        await self.ensure_session()
        retries = max(1, min(int(getattr(config, "VK_SEARCH_RETRIES", 2) or 2), 5))
        base_timeout = float(getattr(config, "VK_TIMEOUT", 4) or 4)
        backoff = float(getattr(config, "VK_RETRY_BACKOFF", 0.35) or 0.35)
        last_err = None

        for attempt in range(retries):
            vk_token = await self.vk_token_manager.get_next_valid_token()
            if not vk_token:
                logger.error("No valid VK tokens available")
                return []

            params = {
                "access_token": vk_token,
                "q": query,
                "count": min(int(count or 30), int(getattr(config, "MAX_SEARCH_RESULTS", 100) or 100)),
                "v": getattr(config, "VK_API_VERSION", "5.131"),
                "sort": getattr(config, "VK_SEARCH_SORT", 2),
                "auto_complete": getattr(config, "VK_SEARCH_AUTO_COMPLETE", 1),
            }
            timeout_s = base_timeout if attempt == 0 else min(base_timeout * 1.8, base_timeout + 5.0)

            try:
                async with self.session.get(
                    getattr(config, "VK_API_URL", "https://api.vk.com/method/audio.search"),
                    params=params,
                    timeout=timeout_s,
                ) as response:
                    if response.status != 200:
                        last_err = f"HTTP {response.status}"
                        if response.status in (429, 500, 502, 503, 504) and attempt + 1 < retries:
                            await asyncio.sleep(backoff * (1.8 ** attempt))
                            continue
                        return []

                    data = await response.json(content_type=None)
                    if isinstance(data, dict) and "error" in data:
                        err = data.get("error") or {}
                        code = int(err.get("error_code", 0) or 0)
                        last_err = f"VK error {code}: {err.get('error_msg', '')}"
                        if code in (5, 10, 14, 17, 18) and attempt + 1 < retries:
                            await asyncio.sleep(backoff * (1.8 ** attempt))
                            continue
                        return []

                    items = ((data.get("response") or {}).get("items") or []) if isinstance(data, dict) else []
                    results = []
                    for track in items:
                        url = track.get("url")
                        if not url:
                            continue
                        owner_id = track.get("owner_id")
                        audio_id = track.get("id")
                        vk_key = f"{owner_id}_{audio_id}" if owner_id is not None and audio_id is not None else None
                        results.append({
                            "url": url,
                            "artist": await self.clean_text(track.get("artist", "Неизвестный исполнитель")),
                            "title": await self.clean_text(track.get("title", "Неизвестный трек")),
                            "duration": track.get("duration", 0),
                            "source": "vk",
                            "vk_key": vk_key,
                            "vk_owner_id": owner_id,
                            "vk_audio_id": audio_id,
                            "vk_access_key": track.get("access_key"),
                        })
                    return results

            except (asyncio.TimeoutError, aiohttp.ClientError) as exc:
                last_err = exc
                if attempt + 1 < retries:
                    await asyncio.sleep(backoff * (1.8 ** attempt))
                    continue
                logger.warning(f"VK search failed for {query!r}: {exc}")
                return []
            except Exception as exc:
                last_err = exc
                if attempt + 1 < retries:
                    await asyncio.sleep(backoff * (1.8 ** attempt))
                    continue
                logger.error(f"VK search error for {query!r}: {exc}")
                return []

        if last_err:
            logger.debug(f"VK search exhausted retries for {query!r}: {last_err}")
        return []
    async def _provider_call(self, provider: str, operation: str, awaitable):
        """Execute provider operation with circuit-breaker routing and metrics."""
        started = time.perf_counter()
        try:
            result = await self.provider_router.call(provider, operation, lambda: awaitable)
            elapsed_ms = (time.perf_counter() - started) * 1000.0
            count = len(result) if isinstance(result, (list, tuple, dict, set)) else 0
            self.provider_health.record_success(provider, operation, elapsed_ms, count=count)
            await self.metrics.record(f"provider.{provider}.{operation}", ok=True, seconds=elapsed_ms / 1000.0)
            return result
        except Exception as exc:
            elapsed_ms = (time.perf_counter() - started) * 1000.0
            self.provider_health.record_failure(provider, operation, elapsed_ms, exc)
            await self.metrics.record(f"provider.{provider}.{operation}", ok=False, seconds=elapsed_ms / 1000.0)
            raise


    async def search_all_sources(self, query: str, limit: int = 30, preferred_source: Optional[str] = None) -> List[Dict[str, Any]]:
        """Search enabled sources in parallel, apply failover/circuit-breaker, then rank."""
        started = time.perf_counter()
        requested = max(1, int(limit or 30))
        multiplier = max(1, int(getattr(config, "MERGE_LIMIT_MULTIPLIER", 2) or 2))
        fetch_limit = requested * multiplier

        tasks: List[Tuple[str, asyncio.Task]] = []
        youtube_enabled = bool(
            getattr(config, "ENABLE_YOUTUBE_MUSIC", True) and self.youtube_music._initialized
        )
        direct_youtube_url = bool(
            youtube_enabled and self.youtube_music._YOUTUBE_URL_RE.match(str(query or "").strip())
        )
        if direct_youtube_url:
            # A pasted YouTube URL is an explicit provider choice; do not bury
            # the exact video behind unrelated VK/Yandex/YouTube search results.
            tasks.append(("yt", asyncio.create_task(self._provider_call("youtube", "search", self.youtube_music.search_tracks(query, 1)))))
        else:
            if self.vk_enabled():
                tasks.append(("vk", asyncio.create_task(self._provider_call("vk", "search", self.search_vk_music(query, fetch_limit)))))
            if getattr(config, "ENABLE_YANDEX_MUSIC", True) and self.yandex_music._initialized:
                tasks.append(("ym", asyncio.create_task(self._provider_call("yandex", "search", self._search_yandex_music_adapted(query, fetch_limit)))))
            if youtube_enabled:
                tasks.append(("yt", asyncio.create_task(self._provider_call("youtube", "search", self.youtube_music.search_tracks(query, fetch_limit)))))
        if not tasks:
            await self.metrics.record("search", ok=False, seconds=time.perf_counter() - started)
            return []

        gathered = await asyncio.gather(*(task for _, task in tasks), return_exceptions=True)
        buckets: Dict[str, List[Dict[str, Any]]] = {"vk": [], "ym": [], "yt": []}
        for (source, _), result in zip(tasks, gathered):
            if isinstance(result, Exception):
                logger.warning("%s search failed: %s", source.upper(), result)
                continue
            buckets[source] = result or []

        source_priority = str(preferred_source or getattr(config, "SOURCE_PRIORITY", "vk_first") or "vk_first").lower()
        if source_priority in ("yandex_first", "ym"):
            source_order = ("ym", "vk", "yt")
        elif source_priority in ("youtube_first", "youtube", "yt"):
            source_order = ("yt", "ym", "vk")
        else:
            # Preserve the former default order for fastest/vk_first/vk.
            source_order = ("vk", "ym", "yt")
        merged = [track for source in source_order for track in buckets[source]]

        # Remove only literal duplicates; this is not a relevance/quality criterion.
        unique: List[Dict[str, Any]] = []
        seen = set()
        for track in merged:
            identity = (
                str(track.get("source") or ""),
                str(track.get("track_id") or track.get("vk_key") or track.get("youtube_id") or ""),
                _norm_text(track.get("artist", "")),
                _norm_text(track.get("title", "")),
            )
            if identity in seen:
                continue
            seen.add(identity)
            unique.append(track)

        result = self.rank_tracks_by_artist(unique, query)[:requested]
        await self.metrics.record("search", ok=bool(result), seconds=time.perf_counter() - started)
        return result
    async def _search_yandex_music_adapted(self, query: str, limit: int) -> List[Dict[str, Any]]:
        """Convert ordinary Yandex.Music search results to the bot's common schema."""
        try:
            tracks = await self.yandex_music.search_tracks(query, max(1, int(limit or 10)))
        except Exception as exc:
            logger.warning(f"YM search failed for {query!r}: {exc}")
            tracks = []

        return [
            {
                "url": None,
                "artist": track.artist,
                "title": track.title,
                "duration": track.duration_sec,
                "source": "ym",
                "track_id": str(track.track_id),
                "ym_track": getattr(track, "ym_download_info", None) or getattr(track, "track_obj", None) or track,
                "vk_key": None,
            }
            for track in tracks
        ]


    @rate_limit("search")
    @validate_query
    async def safe_search_vk_music(self, query: str, user_id: int, count: int = 30) -> List[Dict[str, Any]]:
        """Cached multi-source search with one final artist-first ordering."""
        cache_key = _search_cache_key(query)
        cached = self._query_cache.get(cache_key)
        if cached and time.time() - cached[1] < 300:
            return self.rank_tracks_by_artist(cached[0], query)[:count]

        cached_results = await self.cache.get_cached_results(query)
        if cached_results:
            ranked = self.rank_tracks_by_artist(cached_results, query)[:count]
            self._query_cache[cache_key] = (ranked, time.time())
            return ranked

        preferred_source = None
        try:
            if self.user_store:
                prefs = await self.user_store.get_preferences(user_id)
                preferred_source = prefs.get("prefer_source") or None
                if preferred_source == "auto":
                    preferred_source = None
        except Exception:
            preferred_source = None
        results = await self.search_all_sources(query, count, preferred_source=preferred_source)
        if results:
            await self.cache.set_cached_results(query, results)
            self._query_cache[cache_key] = (results, time.time())

        return results

    async def _bg_enrich_session(self, session_id: str, query: str, user_id: int) -> None:
        """Load a larger result set in the background without applying extra filters."""
        if not getattr(config, "BG_ENRICH_SEARCH", True):
            return
        limit = int(getattr(config, "SESSION_ENRICH_LIMIT", 50) or 50)
        if not hasattr(self, "_enrich_semaphore"):
            self._enrich_semaphore = asyncio.Semaphore(int(getattr(config, "BG_ENRICH_CONCURRENCY", 2) or 2))

        try:
            async with self._enrich_semaphore:
                session = await self.session_manager.get_session(session_id)
                if not session or session.is_expired() or session.enrich_cancel_event.is_set():
                    return

                full_results = await self.search_all_sources(query, limit=limit)
                if len(full_results) <= len(session.results):
                    return

                session.results = self.rank_tracks_by_artist(full_results, query)
                session.best_tracks = session.results[:max(
                    int(getattr(config, "GROUP_CHAT_RESULTS_COUNT", 5) or 5),
                    int(getattr(config, "GROUP_CHAT_BEST_CACHE_COUNT", 10) or 10),
                )]
                session.reindex_tracks()
                await self.cache.set_cached_results(query, session.results)
                self._query_cache[_search_cache_key(query)] = (session.results, time.time())
                logger.info(f"Background search updated session {session_id}: {len(session.results)} results")
        except asyncio.CancelledError:
            return
        except Exception as exc:
            logger.debug(f"Background search failed for {session_id}: {exc}")




    def schedule_enrich_session(self, session_id: str, query: str, limit: int = 30) -> None:
        """Schedule managed background loading for an existing search session."""
        if not session_id or not query or not getattr(config, "BG_ENRICH_SEARCH", True):
            return

        async def runner():
            session = await self.session_manager.get_session(session_id)
            if not session:
                return
            if session.enrich_task and not session.enrich_task.done():
                session.enrich_cancel_event.set()
                session.enrich_task.cancel()
            session.enrich_cancel_event = asyncio.Event()
            await self._bg_enrich_session(session_id, query, session.user_id)

        task = self.create_background_task(runner(), f"search-enrich:{session_id}")

        async def store_task():
            session = await self.session_manager.get_session(session_id)
            if session:
                session.enrich_task = task
        self.create_background_task(store_task(), f"search-enrich-store:{session_id}")



    async def get_popular_tracks(self, limit: int = config.POPULAR_TRACKS_LIMIT) -> List[Dict[str, Any]]:
        """Получение популярных треков из чарта Deezer."""
        await self.ensure_session()
        return await self.deezer.get_chart_tracks(self.session, 0, limit)
    
    async def get_top_tracks_by_genre(self, genre: str, limit: int = config.POPULAR_TRACKS_LIMIT) -> List[Dict[str, Any]]:
        """Получение жанрового чарта Deezer."""
        await self.ensure_session()
        genre_id = DEEZER_GENRE_ALIASES.get(str(genre or "").casefold(), 0)
        return await self.deezer.get_chart_tracks(self.session, genre_id, limit)
    
    async def get_similar_artists(self, artist: str, limit: int = config.SIMILAR_ARTISTS_LIMIT) -> List[Dict[str, Any]]:
        """Получение похожих исполнителей"""
        if not self.lastfm:
            return []
        return await self.lastfm.get_similar_artists(self.session, artist, limit)
    
    @rate_limit("download")
    async def safe_download_audio(
        self,
        url: str,
        user_id: int,
        *,
        timeout_s: Optional[float] = None,
        retries: Optional[int] = None,
    ) -> bytes:
        """Безопасная загрузка аудио с проверкой лимитов.

        Для обычных пользовательских загрузок используются общие настройки.
        Рассылки могут передать короткий timeout/retries, чтобы одна протухшая
        ссылка не задерживала всю подборку на несколько минут.
        """
        async with self.download_semaphore:
            return await self.download_audio(url, timeout_s=timeout_s, retries=retries)

    async def download_audio(
        self,
        url: str,
        *,
        timeout_s: Optional[float] = None,
        retries: Optional[int] = None,
    ) -> bytes:
        """Загрузка аудиофайла (быстро + устойчиво).

        Улучшения:
        - экспоненциальный backoff + jitter вместо фиксированного sleep(1);
        - ретраи только на временные ошибки (timeout/5xx/429/conn reset);
        - отдельная обработка 403/404 (часто протухший VK-url) — быстро отдаём ошибку наверх,
          чтобы higher-level логика могла сделать fallback/refresh;
        - stream-chunks с fail-fast по MAX_FILE_SIZE.
        """
        await self.ensure_session()

        # Таймаут скачивания (поверх общего ClientSession timeout)
        try:
            dl_total = float(
                timeout_s if timeout_s is not None
                else (getattr(config, "DOWNLOAD_TIMEOUT", 20) or 20)
            )
        except Exception:
            dl_total = 20.0
        dl_total = max(3.0, min(dl_total, 120.0))
        req_timeout = aiohttp.ClientTimeout(total=dl_total)

        try:
            retry_count = int(
                retries if retries is not None
                else (getattr(config, "DOWNLOAD_RETRIES", 3) or 3)
            )
        except Exception:
            retry_count = 3
        retries = max(1, min(retry_count, 8))

        base_delay = float(getattr(config, "DOWNLOAD_RETRY_BASE_DELAY", 0.7) or 0.7)
        max_delay = float(getattr(config, "DOWNLOAD_RETRY_MAX_DELAY", 6.0) or 6.0)

        max_size = int(getattr(config, "MAX_FILE_SIZE", 50 * 1024 * 1024) or (50 * 1024 * 1024))

        last_exc: Exception = None  # type: ignore

        for attempt in range(retries):
            try:
                async with self.session.get(url, timeout=req_timeout) as response:
                    st = int(response.status)

                    # Часто означает протухшую ссылку VK: не тратим ретраи
                    if st in (403, 404):
                        raise Exception(f"HTTP {st} (link expired)")

                    # Временные статусы — можно ретраить
                    if st in (429, 500, 502, 503, 504):
                        raise Exception(f"HTTP {st}")

                    if st != 200:
                        raise Exception(f"HTTP {st}")

                    buf = BytesIO()
                    total = 0

                    async for chunk in response.content.iter_chunked(256 * 1024):
                        if not chunk:
                            continue
                        total += len(chunk)
                        if total > max_size:
                            raise Exception(f"Файл слишком большой: {total} байт")
                        buf.write(chunk)

                    content = buf.getvalue()
                    if not content:
                        raise Exception("Empty file")

                    return content

            except asyncio.TimeoutError as e:
                last_exc = e
            except aiohttp.ClientError as e:
                last_exc = e
            except Exception as e:
                last_exc = e
                msg = str(e).lower()
                if "link expired" in msg or "http 403" in msg or "http 404" in msg:
                    raise

            if attempt >= retries - 1:
                raise Exception("Все попытки загрузки не удались") from last_exc

            delay = min(max_delay, base_delay * (1.8 ** attempt))
            delay = delay * (1.0 + (random.random() * 0.35))
            await asyncio.sleep(delay)

        raise Exception("Все попытки загрузки не удались")

    async def download_audio_any_source(
        self,
        track_info: Dict[str, Any],
        user_id: int,
        *,
        fast: bool = False,
    ) -> bytes:
        """Download audio from VK, Yandex Music or YouTube with bounded fallbacks."""
        source = _source_code(track_info.get('source', 'vk'))
        fast_timeout = float(getattr(config, "DIGEST_DOWNLOAD_TIMEOUT", 10.0) or 10.0) if fast else None
        fast_retries = int(getattr(config, "DIGEST_DOWNLOAD_RETRIES", 1) or 1) if fast else None
        ym_fast_timeout = float(getattr(config, "DIGEST_YM_DOWNLOAD_TIMEOUT", 16.0) or 16.0) if fast else None
        yt_fast_timeout = float(getattr(config, "DIGEST_YOUTUBE_DOWNLOAD_TIMEOUT", 45.0) or 45.0) if fast else None

        async def _download_vk(url: str) -> bytes:
            return await self._provider_call(
                "vk",
                "download",
                self.safe_download_audio(
                    url,
                    user_id,
                    timeout_s=fast_timeout,
                    retries=fast_retries,
                ),
            )

        async def _download_yt(track: Dict[str, Any]) -> bytes:
            if not getattr(config, "ENABLE_YOUTUBE_MUSIC", True) or not self.youtube_music._initialized:
                raise RuntimeError("YouTube источник недоступен")
            async with self.download_semaphore:
                coro = self.youtube_music.download_track_bytes(track)
                operation = asyncio.wait_for(coro, timeout=yt_fast_timeout) if fast and yt_fast_timeout is not None else coro
                data, extension = await self._provider_call("youtube", "download", operation)
            track_info['audio_ext'] = extension
            return data

        async def _search_download_yt(query: str) -> Optional[bytes]:
            if not getattr(config, "ENABLE_YOUTUBE_MUSIC", True) or not self.youtube_music._initialized:
                return None
            tracks = await self.youtube_music.search_tracks(query, 1)
            if not tracks:
                return None
            data = await _download_yt(tracks[0])
            track_info['audio_ext'] = tracks[0].get('audio_ext') or track_info.get('audio_ext')
            return data

        async def _download_ym(track: Dict[str, Any]) -> Optional[bytes]:
            if not self.yandex_music._initialized:
                return None
            ym_track = TrackInfo(
                idx=0,
                track_id=track.get('track_id'),
                title=str(track.get('title') or ''),
                artist=str(track.get('artist') or ''),
                album=str(track.get('album') or ''),
                duration_sec=int(track.get('duration') or 0),
                source='ym',
            )
            async with self.download_semaphore:
                if fast and ym_fast_timeout is not None:
                    audio_data = await self._provider_call(
                        "yandex",
                        "download",
                        asyncio.wait_for(self.yandex_music.download_track_bytes(ym_track), timeout=ym_fast_timeout),
                    )
                else:
                    audio_data = await self._provider_call(
                        "yandex", "download", self.yandex_music.download_track_bytes(ym_track)
                    )
            if audio_data:
                track_info['audio_ext'] = 'mp3'
                return audio_data
            if fast:
                return None

            fd, temp_file = tempfile.mkstemp(suffix='.mp3')
            os.close(fd)
            try:
                async with self.download_semaphore:
                    success = await self._provider_call(
                        "yandex", "download", self.yandex_music.download_track(ym_track, temp_file)
                    )
                if not success:
                    return None
                async with aiofiles.open(temp_file, 'rb') as file_obj:
                    data = await file_obj.read()
                track_info['audio_ext'] = 'mp3'
                return data or None
            finally:
                try:
                    if os.path.exists(temp_file):
                        os.unlink(temp_file)
                except OSError:
                    logger.debug("Failed to remove temporary YM file: %s", temp_file)

        query = f"{track_info.get('artist', '')} {track_info.get('title', '')}".strip()

        if source == 'yt':
            try:
                return await _download_yt(track_info)
            except Exception as exc:
                logger.warning("YouTube download failed for %s: %s", query, exc)
                if self.yandex_music._initialized:
                    ym_tracks = await self.yandex_music.search_tracks(query, 1)
                    if ym_tracks:
                        fallback = await _download_ym({
                            'artist': ym_tracks[0].artist,
                            'title': ym_tracks[0].title,
                            'duration': ym_tracks[0].duration_sec,
                            'source': 'ym',
                            'track_id': ym_tracks[0].track_id,
                        })
                        if fallback:
                            return fallback
                if self.vk_enabled():
                    vk_tracks = await self.search_vk_music(query, 1)
                    if vk_tracks and vk_tracks[0].get('url'):
                        return await _download_vk(vk_tracks[0]['url'])
                raise RuntimeError(f"Ошибка скачивания YouTube: {exc}") from exc

        if source == 'ym':
            try:
                data = await _download_ym(track_info)
                if data:
                    return data
                raise RuntimeError("Яндекс.Музыка не вернула аудио")
            except Exception as exc:
                logger.warning("Yandex.Music download failed for %s: %s", query, exc)
                if self.vk_enabled():
                    try:
                        vk_tracks = await self.search_vk_music(query, 1)
                        if vk_tracks and vk_tracks[0].get('url'):
                            return await _download_vk(vk_tracks[0]['url'])
                    except Exception as vk_exc:
                        logger.warning("VK fallback failed for %s: %s", query, vk_exc)
                try:
                    yt_data = await _search_download_yt(query)
                    if yt_data:
                        return yt_data
                except Exception as yt_exc:
                    logger.warning("YouTube fallback failed for %s: %s", query, yt_exc)
                raise RuntimeError(f"Ошибка скачивания: {exc}") from exc

        url = str(track_info.get('url') or '').strip()
        if not url:
            if self.yandex_music._initialized:
                ym_tracks = await self.yandex_music.search_tracks(query, 1)
                if ym_tracks:
                    data = await _download_ym({
                        'artist': ym_tracks[0].artist,
                        'title': ym_tracks[0].title,
                        'duration': ym_tracks[0].duration_sec,
                        'source': 'ym',
                        'track_id': ym_tracks[0].track_id,
                    })
                    if data:
                        return data
            yt_data = await _search_download_yt(query)
            if yt_data:
                return yt_data
            raise RuntimeError("Ссылка для скачивания недоступна")

        try:
            track_info['audio_ext'] = track_info.get('audio_ext') or 'mp3'
            return await _download_vk(url)
        except Exception as exc:
            logger.warning("Primary source download failed (%s): %s", source, exc)

            if source == 'vk' and self.vk_enabled():
                message = str(exc).lower()
                expired_hint = any(token in message for token in ("http 403", "http 404", "expired", "link"))
                if expired_hint:
                    try:
                        vk_tracks = await self.search_vk_music(query, 1)
                        if vk_tracks and vk_tracks[0].get('url'):
                            return await _download_vk(vk_tracks[0]['url'])
                    except Exception as refresh_exc:
                        logger.debug("VK refresh-after-fail failed: %s", refresh_exc)

            if self.yandex_music._initialized:
                try:
                    ym_tracks = await self.yandex_music.search_tracks(query, 1)
                    if ym_tracks:
                        data = await _download_ym({
                            'artist': ym_tracks[0].artist,
                            'title': ym_tracks[0].title,
                            'duration': ym_tracks[0].duration_sec,
                            'source': 'ym',
                            'track_id': ym_tracks[0].track_id,
                        })
                        if data:
                            return data
                except Exception as ym_exc:
                    logger.warning("Yandex fallback failed for %s: %s", query, ym_exc)

            try:
                yt_data = await _search_download_yt(query)
                if yt_data:
                    return yt_data
            except Exception as yt_exc:
                logger.warning("YouTube fallback failed for %s: %s", query, yt_exc)
            raise

    async def _handle_queued_download(self, job):
        """Worker callback used by playlist/batch jobs."""
        payload = dict(job.payload or {})
        update = payload.get("update")
        context = payload.get("context")
        track = payload.get("track")
        if update is None or context is None or not isinstance(track, dict):
            raise RuntimeError("invalid queued download payload")
        async with self.global_download_semaphore:
            return await download_and_send_audio(update, context, update.effective_message, track, update.effective_message.message_id)

    async def clean_text(self, text: str) -> str:
        """Асинхронная очистка текста"""
        if not text:
            return ""
        
        loop = asyncio.get_running_loop()
        try:
            cleaned = await loop.run_in_executor(
                None,
                lambda: ' '.join(text.split())
            )
            cleaned = await loop.run_in_executor(
                None,
                lambda: re.sub(r'<[^>]+>', '', cleaned)
            )
            return cleaned.strip()
        except Exception:
            return text.strip()
    
    def create_safe_filename(self, artist: str, title: str, extension: str = "mp3") -> str:
        """Создание безопасного имени файла с фактическим расширением аудио."""
        safe_artist = re.sub(r'[<>:"/\\|?*]', '', artist).strip() or "Unknown Artist"
        safe_title = re.sub(r'[<>:"/\\|?*]', '', title).strip() or "Unknown Track"
        safe_extension = re.sub(r'[^a-zA-Z0-9]', '', str(extension or "mp3")).lower() or "mp3"
        suffix = f".{safe_extension}"
        filename = f"{safe_artist} - {safe_title}{suffix}"

        if len(filename) > config.MAX_FILENAME_LENGTH:
            max_len = max(1, (config.MAX_FILENAME_LENGTH - len(suffix) - 3) // 2)
            safe_artist = safe_artist[:max_len].strip()
            safe_title = safe_title[:max_len].strip()
            filename = f"{safe_artist} - {safe_title}{suffix}"

            if len(filename) > config.MAX_FILENAME_LENGTH:
                stem_limit = max(1, config.MAX_FILENAME_LENGTH - len(suffix))
                filename = filename[:stem_limit].rstrip() + suffix

        return filename
    
    def _tg_query_key(self, artist: str, title: str, vk_key: str = None) -> str:
        """Ключ для кэша Telegram file_id"""
        if vk_key:
            return f"vk:{str(vk_key).strip()}"
        s = f"{artist} - {title}".strip().lower()
        s = re.sub(r"\s+", " ", s)
        return f"at:{s}"
    
    def _tg_lru_get(self, key: str) -> Optional[str]:
        if not key:
            return None
        try:
            if key in self._tg_file_id_lru:
                val = self._tg_file_id_lru.pop(key)
                # mark as recently used
                self._tg_file_id_lru[key] = val
                return val
        except Exception:
            return None
        return None

    def _tg_lru_set(self, key: str, file_id: str) -> None:
        if not key or not file_id:
            return
        try:
            if key in self._tg_file_id_lru:
                self._tg_file_id_lru.pop(key, None)
            self._tg_file_id_lru[key] = file_id
            # trim
            max_n = self._tg_file_id_lru_max if self._tg_file_id_lru_max and self._tg_file_id_lru_max > 0 else 50000
            while len(self._tg_file_id_lru) > max_n:
                self._tg_file_id_lru.popitem(last=False)
        except Exception:
            pass

    async def peek_tg_file_id_local_bulk(self, items: List[Tuple[str, str, Optional[str]]]) -> Set[str]:
        """Fast local-only check whether we already have Telegram file_id in LRU.

        Returns a set of tg_query_key values that hit in local memory. No Redis/SQLite IO.
        """
        keys: Set[str] = set()
        if not items:
            return keys
        try:
            async with self._tg_file_id_lock:
                for artist, title, vk_key in items:
                    k = self._tg_query_key(artist, title, vk_key)
                    if k in self._tg_file_id_lru:
                        # do not reorder here (pure peek)
                        keys.add(k)
        except Exception:
            return keys
        return keys


    async def _ensure_tg_db_conn(self) -> aiosqlite.Connection:
        """Keep a single aiosqlite connection for tg_audio_cache to avoid per-call connect overhead."""
        if getattr(self, "_tg_db_conn", None) is not None:
            return self._tg_db_conn

        async with self._tg_db_lock:
            if getattr(self, "_tg_db_conn", None) is not None:
                return self._tg_db_conn

            db_path = await self._tg_db_path()
            conn = await _sqlite_connect(
                db_path,
                timeout=float(getattr(config, "DATABASE_TIMEOUT", 20) or 20),
            )

            # Pragmas tuned for many small reads/writes (safe enough for cache DB)
            try:
                if getattr(config, "DB_WAL_MODE", True):
                    await conn.execute("PRAGMA journal_mode=WAL;")
                await conn.execute("PRAGMA synchronous=NORMAL;")
                await conn.execute("PRAGMA temp_store=MEMORY;")
                # ~64MB page cache (negative means KB)
                await conn.execute("PRAGMA cache_size=-65536;")
            except Exception as exc:
                logger.warning("TG cache SQLite pragma setup failed: %s", exc)

            await self._ensure_tg_audio_cache_table(conn)
            self._tg_db_conn = conn
            return conn


    async def _tg_db_path(self) -> str:
        return getattr(config, "TG_FILE_CACHE_DB", "bot_stats.db")

    async def _ensure_tg_audio_cache_table(self, conn: aiosqlite.Connection) -> None:
        await conn.execute(
            """CREATE TABLE IF NOT EXISTS tg_audio_cache(
                query_key TEXT PRIMARY KEY,
                file_id TEXT,
                artist TEXT,
                title TEXT,
                duration INTEGER,
                unique_id TEXT,
                created_at INTEGER
            )"""
        )
        await conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_tg_audio_cache_created_at ON tg_audio_cache(created_at)"
        )
        await conn.commit()

    async def get_tg_file_id(self, artist: str, title: str, vk_key: str = None) -> Optional[str]:
        """Получить сохранённый Telegram file_id для трека (по artist+title).

        Используется для мгновенной отдачи аудио из Telegram без повторного скачивания.
        """
        key = self._tg_query_key(artist, title, vk_key)

        # 0) ultra-fast in-memory LRU
        try:
            async with self._tg_file_id_lock:
                lru_hit = self._tg_lru_get(key)
            if lru_hit:
                return lru_hit
        except Exception:
            pass

        cache_key = f"tg_file_id:{key}"

        # 1) in-process cache / Redis
        try:
            cached = await self.cache.get(cache_key)
            if isinstance(cached, dict) and cached.get("file_id"):
                file_id = cached["file_id"]
                try:
                    async with self._tg_file_id_lock:
                        self._tg_lru_set(key, file_id)
                except Exception:
                    pass
                return file_id
        except Exception:
            pass

        # 2) SQLite (single reused connection)
        try:
            conn = await self._ensure_tg_db_conn()
            async with self._tg_db_lock:
                cur = await conn.execute(
                    "SELECT file_id FROM tg_audio_cache WHERE query_key = ?",
                    (key,)
                )
                row = await cur.fetchone()
                await cur.close()

            if row and row[0]:
                file_id = row[0]
                # warm both caches
                try:
                    await self.cache.set(cache_key, {"file_id": file_id}, ttl=60 * 60 * 24 * 30)
                except Exception:
                    pass
                try:
                    async with self._tg_file_id_lock:
                        self._tg_lru_set(key, file_id)
                except Exception:
                    pass
                return file_id
        except Exception:
            pass

        return None

    async def set_tg_file_id(
        self,
        artist: str,
        title: str,
        file_id: str,
        unique_id: Optional[str] = None,
        duration: Optional[int] = None,
        vk_key: Optional[str] = None,
    ) -> None:
        """Сохранить Telegram file_id для трека (по artist+title)."""
        if not file_id:
            return

        key = self._tg_query_key(artist, title, vk_key)
        cache_key = f"tg_file_id:{key}"

        # update in-memory LRU first
        try:
            async with self._tg_file_id_lock:
                self._tg_lru_set(key, file_id)
        except Exception:
            pass

        # write to cache / Redis
        try:
            await self.cache.set(cache_key, {"file_id": file_id}, ttl=60 * 60 * 24 * 30)
        except Exception:
            pass

        # write to SQLite
        try:
            conn = await self._ensure_tg_db_conn()
            async with self._tg_db_lock:
                await conn.execute(
                    "INSERT OR REPLACE INTO tg_audio_cache(query_key, file_id, artist, title, duration, unique_id, created_at) "
                    "VALUES(?, ?, ?, ?, ?, ?, ?)",
                    (key, file_id, artist, title, duration, unique_id, int(time.time()))
                )
                await conn.commit()
        except Exception as e:
            logger.warning(f"TG file_id db write failed: {e}")

    async def delete_tg_file_id(self, artist: str, title: str, vk_key: str = None) -> None:
        """Удалить Telegram file_id (если стал невалидным)."""
        key = self._tg_query_key(artist, title, vk_key)
        cache_key = f"tg_file_id:{key}"

        try:
            await self.cache.delete(cache_key)
        except Exception:
            pass

        try:
            conn = await self._ensure_tg_db_conn()
            async with self._tg_db_lock:
                await conn.execute("DELETE FROM tg_audio_cache WHERE query_key = ?", (key,))
                await conn.commit()
        except Exception:
            pass


    async def ensure_session(self) -> None:
        """Проверка инициализации сессии"""
        if not self._initialized:
            await self.initialize()
    
    async def close_session(self) -> None:
        """Close network clients, SQLite handles and background tasks."""
        tasks = list(getattr(self, "background_tasks", set()) or [])
        user_cleanup_task = getattr(getattr(self, "user_manager", None), "_cleanup_task", None)
        if user_cleanup_task is not None:
            tasks.append(user_cleanup_task)

        for task in tasks:
            try:
                task.cancel()
            except Exception:
                pass
        if tasks:
            await asyncio.gather(*tasks, return_exceptions=True)
        self.background_tasks.clear()

        session_manager = getattr(self, "session_manager", None)
        if session_manager is not None and hasattr(session_manager, "close"):
            await session_manager.close()

        if self.session and not self.session.closed:
            await self.session.close()
            logger.info("HTTP session closed")
        self._initialized = False

        try:
            if self._tg_db_conn is not None:
                await self._tg_db_conn.close()
                self._tg_db_conn = None
        except Exception as exc:
            logger.debug(f"TG cache DB close failed: {exc}")

        redis_client = getattr(getattr(self, "cache", None), "redis_client", None)
        if redis_client is not None:
            try:
                close = getattr(redis_client, "aclose", None) or getattr(redis_client, "close", None)
                if close is not None:
                    result = close()
                    if asyncio.iscoroutine(result):
                        await result
            except Exception as exc:
                logger.debug(f"Redis close failed: {exc}")

        youtube_manager = getattr(self, "youtube_music", None)
        if youtube_manager is not None and hasattr(youtube_manager, "close"):
            try:
                await youtube_manager.close()
            except Exception as exc:
                logger.debug("YouTube manager close failed: %s", exc)

        file_manager = getattr(self, "file_manager", None)
        if file_manager is not None and hasattr(file_manager, "close"):
            await file_manager.close()

        executor = getattr(getattr(self, "yandex_music", None), "executor", None)
        if executor is not None:
            try:
                executor.shutdown(wait=False, cancel_futures=True)
            except Exception:
                pass

# ==================== ГЛОБАЛЬНЫЕ ПЕРЕМЕННЫЕ ====================
bot_instance: Optional[AsyncMusicBot] = None


async def get_bot_username(context: ContextTypes.DEFAULT_TYPE) -> str:
    """Get bot username with caching to avoid frequent get_me() calls."""
    try:
        cached = getattr(bot_instance, "bot_username", None) if bot_instance else None
        if cached:
            return cached
    except Exception:
        pass

    me = await context.bot.get_me()
    username = me.username or ""
    if bot_instance:
        try:
            bot_instance.bot_username = username
        except Exception:
            pass
    return username

# ==================== ОБРАБОТЧИКИ ====================
def _bot_identity_for_usage(context: CallbackContext) -> Tuple[int, str]:
    bot = getattr(context, "bot", None)
    try:
        bot_id = int(getattr(bot, "id", 0) or 0)
    except Exception:
        bot_id = 0
    try:
        bot_username = getattr(bot, "username", "") or ""
    except Exception:
        bot_username = ""
    username = str(
        (getattr(bot_instance, "bot_username", "") if bot_instance else "")
        or bot_username
        or ""
    ).lstrip("@").lower()
    return bot_id, username


def _group_message_uses_bot(message: Any, bot_id: int, bot_username: str) -> bool:
    """Recognize only explicit bot usage; ordinary group chatter is ignored."""
    if not message:
        return False
    replied = getattr(message, "reply_to_message", None)
    replied_user = getattr(replied, "from_user", None) if replied else None
    if bot_id > 0 and replied_user and int(getattr(replied_user, "id", 0) or 0) == bot_id:
        return True

    text = str(getattr(message, "text", None) or getattr(message, "caption", None) or "").strip()
    if not text:
        return False

    first_token = text.split(None, 1)[0]
    command_match = re.fullmatch(r"/[A-Za-z0-9_]+(?:@([A-Za-z0-9_]+))?", first_token)
    if command_match:
        target = (command_match.group(1) or "").lower()
        if not target or (bot_username and target == bot_username):
            return True

    if bot_username and re.search(rf"(?<![A-Za-z0-9_])@{re.escape(bot_username)}\b", text, re.IGNORECASE):
        return True
    return bool(re.match(r"^найти(?:\s+|$)", text, re.IGNORECASE))


def _unique_usage_context(update: Update, context: CallbackContext) -> Optional[str]:
    user = getattr(update, "effective_user", None)
    if not user or bool(getattr(user, "is_bot", False)):
        return None
    try:
        user_id = int(getattr(user, "id", 0) or 0)
    except (TypeError, ValueError):
        return None
    if user_id <= 0 or user_id in AdminDB._TELEGRAM_SERVICE_USER_IDS:
        return None

    chat = getattr(update, "effective_chat", None)
    if not chat:
        query = getattr(update, "callback_query", None)
        callback_message = getattr(query, "message", None) if query else None
        chat = getattr(callback_message, "chat", None)
    chat_type = str(getattr(chat, "type", "") or "").lower()
    if chat_type == "private":
        return "private" if (getattr(update, "effective_message", None) or getattr(update, "callback_query", None)) else None
    if chat_type not in ("group", "supergroup"):
        return None
    if getattr(update, "callback_query", None) is not None:
        return "group"
    bot_id, bot_username = _bot_identity_for_usage(context)
    return "group" if _group_message_uses_bot(getattr(update, "effective_message", None), bot_id, bot_username) else None


async def track_unique_user_usage(update: Update, context: CallbackContext) -> None:
    """Early, non-blocking accounting handler registered in its own group."""
    try:
        admin_db = getattr(bot_instance, "admin_db", None) if bot_instance else None
        if admin_db is None:
            return
        context_name = _unique_usage_context(update, context)
        if context_name:
            await admin_db.record_user_usage(update.effective_user, context_name)
    except Exception:
        logger.exception("Unique user usage handler failed")


async def observe_bot_chat_membership(update: Update, context: CallbackContext) -> None:
    """Track when the bot is added to, promoted in, restricted in, or removed from a chat."""
    try:
        admin_db = getattr(bot_instance, "admin_db", None) if bot_instance else None
        if admin_db is not None:
            await admin_db.track_chat_membership(update)
    except Exception:
        logger.exception("Bot chat membership handler failed")


async def user_check(update: Update, enforce_cooldown: bool = True) -> bool:
    """Асинхронная проверка пользователя.

    Navigation callbacks must not be blocked by the short anti-spam cooldown;
    expensive actions have their own rate limits.
    """
    user = None
    chat = None
    
    if update.message:
        user = update.effective_user
        chat = update.effective_chat
    elif update.callback_query:
        user = update.callback_query.from_user
        chat = update.callback_query.message.chat
    else:
        return False
    
    if not user:
        return False

    # mark activity for keepalive logic
    try:
        bot_instance.mark_user_activity()
    except Exception:
        pass
    
    if await bot_instance.user_manager.is_user_banned(user.id):
        logger.warning(f"Banned user {user.id} tried to access bot")
        try:
            if update.callback_query:
                await safe_answer_callback(getattr(update, 'callback_query', None), "❌ Ваш доступ к боту ограничен.", show_alert=True)
            else:
                await update.message.reply_text("❌ Ваш доступ к боту ограничен.")
        except Exception as e:
            logger.error(f"Error sending ban message: {e}")
        return False

    if enforce_cooldown and not await bot_instance.user_manager.check_cooldown(user.id):
        logger.debug(f"Rate limit exceeded for user {user.id} in chat {chat.id}")
        try:
            if update.callback_query:
                await safe_answer_callback(getattr(update, 'callback_query', None), "⏳ Слишком частые запросы. Подождите немного.", show_alert=True)
            else:
                await update.message.reply_text("⏳ Слишком частые запросы. Подождите немного.")
        except Exception as e:
            logger.error(f"Error sending rate limit message: {e}")
        return False

    return True

def _favorite_deep_link_uid(track: Dict[str, Any]) -> Optional[str]:
    """Return the exact stable track UID for Telegram callback_data.

    This is no longer a deep-link payload. Callback data may contain ':' and
    therefore can preserve provider-prefixed UIDs such as vk:..., yt:... and
    ym:... exactly as they are stored in user_history.
    """
    try:
        uid = str(_track_uid_from_any(track) or "").strip()
        if not uid:
            return None
        uid = re.sub(r"[^A-Za-z0-9_:-]", "", uid)
        return uid or None
    except Exception:
        return None


async def _favorite_reply_markup(context: CallbackContext, track: Dict[str, Any]) -> Optional[InlineKeyboardMarkup]:
    """Build a native Telegram callback button for a downloaded track.

    Do not use a t.me deep-link here: Telegram Desktop can open the bot with
    the deep-link text merely prefilled in the compose box instead of sending
    /start to the bot. A callback query is delivered directly to this bot and
    can therefore add the exact downloaded track immediately.
    """
    try:
        uid = _favorite_deep_link_uid(track)
        if not uid:
            return None
        # uid is a short, Telegram-safe stable track identifier. Keep the
        # callback payload well below Telegram's 64-byte callback_data limit.
        callback_data = f"fav_audio:{uid}"
        return InlineKeyboardMarkup([[
            InlineKeyboardButton("❤️ Добавить в избранное", callback_data=callback_data),
        ]])
    except Exception:
        logger.debug("Failed to build favorite callback", exc_info=True)
        return None


async def download_and_send_audio(update: Update, context: CallbackContext, message, track: Dict, original_message_id: int = None) -> bool:
    """Download and send a track; return whether Telegram delivery succeeded."""
    user_id = update.effective_user.id if update and update.effective_user else message.from_user.id
    loading_message = None
    audio_file = None
    memory_slot_acquired = False

    logger.info(f"Starting download from {track.get('source', 'vk')}: {track['artist']} - {track['title']}")
    
    try:
        filename = bot_instance.create_safe_filename(track['artist'], track['title'])

        cached_file_id = None
        try:
            cached_file_id = await bot_instance.get_tg_file_id(track['artist'], track['title'], track.get('vk_key'))
        except Exception:
            cached_file_id = None

        if cached_file_id:
            send_kwargs_fast = {
                'chat_id': message.chat.id,
                'message_thread_id': getattr(message, 'message_thread_id', None),
                'audio': cached_file_id,
                'title': track['title'][:64],
                'performer': track['artist'][:64],
                'caption': getattr(config, 'TRACK_CREDIT_CAPTION', None),
                'parse_mode': 'HTML'
            }
            favorite_markup = await _favorite_reply_markup(context, track)
            if favorite_markup is not None:
                send_kwargs_fast['reply_markup'] = favorite_markup

            if message.chat.type in ["group", "supergroup"] and original_message_id:
                send_kwargs_fast['reply_to_message_id'] = original_message_id

            try:
                _t0 = time.monotonic()
                logger.info(
                    "TG upload start (cached file_id) chat_id=%s thread_id=%s reply_to=%s file_id=%s title=%r performer=%r",
                    message.chat.id,
                    getattr(message, 'message_thread_id', None),
                    send_kwargs_fast.get('reply_to_message_id'),
                    _short_id(cached_file_id, 18),
                    track.get('title','')[:64],
                    track.get('artist','')[:64],
                )
                sent_msg = await safe_send_audio(context.bot, **send_kwargs_fast)
                logger.info(
                    "TG upload OK (cached file_id) dt=%.2fs message_id=%s",
                    time.monotonic() - _t0,
                    getattr(sent_msg, 'message_id', None),
                )

                try:
                    if hasattr(bot_instance,'admin_db') and bot_instance.admin_db:
                        await bot_instance.admin_db.log_download(user_id, message.chat.id, track.get('source',''))
                except Exception:
                    logger.debug("log_download failed", exc_info=True)
                # history / last
                try:
                    if getattr(bot_instance, 'user_store', None):
                        uid = _track_uid_from_any(track)
                        await bot_instance.user_store.add_history(user_id, uid, track)
                        await bot_instance.user_store.set_last(user_id, uid, track)
                        await bot_instance.user_store.set_pending_favorite(user_id, uid, track)
                except Exception:
                    logger.debug("history update failed", exc_info=True)
                return True
            except BadRequest as e:
                msg = str(e).lower()

                # If the message we reply to was deleted (e.g. the search list was closed),
                # Telegram returns "Message to be replied not found". In this case retry without reply_to.
                if ("message to be replied not found" in msg) or ("reply message not found" in msg):
                    try:
                        send_kwargs_fast.pop('reply_to_message_id', None)
                        _t0 = time.monotonic()
                        logger.info(
                            "TG upload start (cached file_id, no reply_to) chat_id=%s thread_id=%s file_id=%s",
                            message.chat.id,
                            getattr(message, 'message_thread_id', None),
                            _short_id(cached_file_id, 18),
                        )
                        sent_msg = await safe_send_audio(context.bot, **send_kwargs_fast)
                        logger.info(
                            "TG upload OK (cached file_id, no reply_to) dt=%.2fs message_id=%s",
                            time.monotonic() - _t0,
                            getattr(sent_msg, 'message_id', None),
                        )

                        try:
                            if hasattr(bot_instance,'admin_db') and bot_instance.admin_db:
                                await bot_instance.admin_db.log_download(user_id, message.chat.id, track.get('source',''))
                        except Exception:
                            logger.debug("log_download failed", exc_info=True)
                        try:
                            if getattr(bot_instance, 'user_store', None):
                                uid = _track_uid_from_any(track)
                                await bot_instance.user_store.add_history(user_id, uid, track)
                                await bot_instance.user_store.set_last(user_id, uid, track)
                                await bot_instance.user_store.set_pending_favorite(user_id, uid, track)
                        except Exception:
                            logger.debug("history update failed", exc_info=True)
                        return True
                    except TelegramError:
                        pass

                if ("wrong file identifier" in msg) or ("file_id_invalid" in msg) or ("file id invalid" in msg):
                    try:
                        await bot_instance.delete_tg_file_id(track['artist'], track['title'], track.get('vk_key'))
                    except Exception:
                        pass
                raise
            except TelegramError:
                pass
        
        await bot_instance.audio_memory_semaphore.acquire()
        memory_slot_acquired = True

        await safe_answer_callback(getattr(update, 'callback_query', None), 
                                 f"{_track_duration_text(track)} ⬇️ Загружаю из {_source_name(track.get('source'), genitive=True)}: {track['artist']} - {track['title']}",
                                 show_alert=False)
        
        safe_artist = html.escape(str(track.get('artist', '')))
        safe_title = html.escape(str(track.get('title', '')))
        src_label = _source_name(track.get('source'), genitive=True)
        loading_text = f"⬇️ <b>Загружаю из {src_label}</b>\n{_track_duration_text(track)} {safe_artist} - {safe_title}\n⏳ Пожалуйста, подождите..."
        
        loading_kwargs = {
            'chat_id': message.chat.id,
            'message_thread_id': getattr(message, 'message_thread_id', None),
            'text': loading_text,
            'parse_mode': 'HTML'
        }
        
        if message.chat.type in ["group", "supergroup"] and original_message_id:
            loading_kwargs['reply_to_message_id'] = original_message_id
        
        try:
            loading_message = await safe_send_message(context.bot, **loading_kwargs)
        except BadRequest as e:
            msg = str(e).lower()
            if (("message to be replied not found" in msg) or ("reply message not found" in msg)) and ('reply_to_message_id' in loading_kwargs):
                loading_kwargs.pop('reply_to_message_id', None)
                loading_message = await safe_send_message(context.bot, **loading_kwargs)
            else:
                raise
        

        _dl_t0 = time.monotonic()
        audio_data = await bot_instance.download_audio_any_source(track, user_id)
        _dl_dt = time.monotonic() - _dl_t0
        try:
            logger.info(
                "Downloaded audio bytes=%s dt=%.2fs src=%s %s - %s",
                len(audio_data) if audio_data is not None else None,
                _dl_dt,
                track.get('source',''),
                track.get('artist',''),
                track.get('title',''),
            )
        except Exception:
            pass

        
        audio_size = len(audio_data)
        filename = bot_instance.create_safe_filename(
            track['artist'], track['title'], track.get('audio_ext') or 'mp3'
        )
        audio_file = BytesIO(audio_data)
        del audio_data
        audio_file.name = filename
        
        send_kwargs = {
            'chat_id': message.chat.id,
            'message_thread_id': getattr(message, 'message_thread_id', None),
            'audio': audio_file,
            'title': track['title'][:64],
            'performer': track['artist'][:64],
            'caption': getattr(config, 'TRACK_CREDIT_CAPTION', None),
            'parse_mode': 'HTML'
        }
        favorite_markup = await _favorite_reply_markup(context, track)
        if favorite_markup is not None:
            send_kwargs['reply_markup'] = favorite_markup
        
        if message.chat.type in ["group", "supergroup"] and original_message_id:
            send_kwargs['reply_to_message_id'] = original_message_id
        

        try:
            _t0 = time.monotonic()
            size_b = audio_size
            logger.info(
                "TG upload start (BytesIO) chat_id=%s thread_id=%s reply_to=%s size_bytes=%s title=%r performer=%r src=%s",
                message.chat.id,
                getattr(message, 'message_thread_id', None),
                send_kwargs.get('reply_to_message_id'),
                size_b,
                track.get('title','')[:64],
                track.get('artist','')[:64],
                track.get('source',''),
            )
            sent_msg = await safe_send_audio(context.bot, **send_kwargs)
            logger.info(
                "TG upload OK (BytesIO) dt=%.2fs message_id=%s file_id=%s unique_id=%s",
                time.monotonic() - _t0,
                getattr(sent_msg, 'message_id', None),
                _short_id(getattr(getattr(sent_msg,'audio',None),'file_id',None), 18),
                _short_id(getattr(getattr(sent_msg,'audio',None),'file_unique_id',None), 18),
            )

        except BadRequest as e:
            msg = str(e).lower()
            if (("message to be replied not found" in msg) or ("reply message not found" in msg)) and ('reply_to_message_id' in send_kwargs):
                send_kwargs.pop('reply_to_message_id', None)
                sent_msg = await safe_send_audio(context.bot, **send_kwargs)
            else:
                raise
        try:
            if hasattr(bot_instance,'admin_db') and bot_instance.admin_db:
                await bot_instance.admin_db.log_download(user_id, message.chat.id, track.get('source',''))
        except Exception:
            logger.debug("log_download failed", exc_info=True)

        # history / last
        try:
            if getattr(bot_instance, 'user_store', None):
                uid = _track_uid_from_any(track)
                await bot_instance.user_store.add_history(user_id, uid, track)
                await bot_instance.user_store.set_last(user_id, uid, track)
                await bot_instance.user_store.set_pending_favorite(user_id, uid, track)
        except Exception:
            logger.debug("history update failed", exc_info=True)

        # Mandatory file_id cache: persist after successful send (don't silently skip)
        if sent_msg and getattr(sent_msg, 'audio', None):
            try:
                await bot_instance.set_tg_file_id(
                    track['artist'], track['title'],
                    sent_msg.audio.file_id,
                    getattr(sent_msg.audio, 'file_unique_id', None),
                    getattr(sent_msg.audio, 'duration', None),
                    track.get('vk_key')
                )
            except Exception as e:
                # Keep UX working, but log loudly and at least warm local LRU
                logger.exception(f"Failed to persist Telegram file_id cache for {track.get('artist')} - {track.get('title')}: {e}")
                try:
                    k = bot_instance._tg_query_key(track.get('artist',''), track.get('title',''), track.get('vk_key'))
                    async with bot_instance._tg_file_id_lock:
                        bot_instance._tg_lru_set(k, getattr(sent_msg.audio, 'file_id', None))
                except Exception:
                    pass
        return True

    except RateLimitExceeded as e:
        error_msg = f"❌ <b>Слишком частые загрузки</b>\n{str(e)}"
        await safe_answer_callback(getattr(update, 'callback_query', None), error_msg, show_alert=True)
        return False
    except Exception as e:
        logger.exception("Ошибка при скачивании")
        error_msg = "❌ <b>Ошибка при скачивании</b>\nПопробуйте другой трек или повторите позже."
        if "preview" in str(e).lower() or "30" in str(e):
            error_msg = "❌ Только preview (30 сек). Для скачивания полной версии нужна подписка Яндекс.Плюс."
        await safe_answer_callback(getattr(update, 'callback_query', None), error_msg, show_alert=True)
        return False
    finally:
        if audio_file is not None:
            try:
                audio_file.close()
            except Exception:
                pass
        if memory_slot_acquired:
            try:
                bot_instance.audio_memory_semaphore.release()
            except Exception:
                logger.debug("Audio memory semaphore release failed", exc_info=True)
        if loading_message is not None:
            try:
                await safe_delete_message(
                    context.bot,
                    loading_message.chat_id,
                    loading_message.message_id,
                )
            except Exception as exc:
                logger.debug("Loading message cleanup failed: %s", exc)


def _normalize_user_query(query: str) -> str:
    """Keep the user's query intact; only trim and collapse whitespace."""
    return re.sub(r"\s+", " ", (query or "").strip())


def _norm_text(value: str) -> str:
    """Normalization used only for case-insensitive performer comparison."""
    value = (value or "").casefold().replace("ё", "е")
    value = re.sub(r"[^\w\s]", " ", value, flags=re.UNICODE)
    return re.sub(r"\s+", " ", value, flags=re.UNICODE).strip()


async def process_search(update: Update, context: CallbackContext, query: str) -> None:
    """Run a search and render it in the same private chat or forum topic."""
    user_id = update.effective_user.id
    chat = update.effective_chat
    chat_type = chat.type
    query = _normalize_user_query(query)
    logger.info(f"Search process started - User: {user_id}, Chat: {chat_type}, Query: {query!r}")

    try:
        sources = []
        if bot_instance.yandex_music._initialized:
            sources.append("🎵 Яндекс.Музыка")
        if bot_instance.vk_enabled():
            sources.append("🎶 VK")
        if getattr(config, "ENABLE_YOUTUBE_MUSIC", True) and bot_instance.youtube_music._initialized:
            sources.append("▶️ YouTube")
        sources_text = " + ".join(sources) if sources else "неизвестный источник"

        effective_message = update.effective_message
        send_kwargs = {
            "chat_id": chat.id,
            "text": f"🔍 <b>Ищем в {sources_text}:</b> <i>{_esc(query)}</i>\n\n⏳ Обрабатываю запрос...",
            "parse_mode": "HTML",
        }
        if effective_message and chat_type in ("group", "supergroup"):
            send_kwargs["reply_to_message_id"] = effective_message.message_id
            thread_id = getattr(effective_message, "message_thread_id", None)
            if thread_id is not None:
                send_kwargs["message_thread_id"] = thread_id

        search_message = await safe_send_message(context.bot, **send_kwargs)
        songs = bot_instance.rank_tracks_by_artist(
            await bot_instance.safe_search_vk_music(query, user_id), query
        )

        ym_count = sum(1 for song in songs if song.get("source") == "ym")
        vk_count = sum(1 for song in songs if song.get("source") == "vk")
        yt_count = sum(1 for song in songs if song.get("source") == "yt")
        logger.info(f"Search completed - Found {len(songs)} tracks for query: {query!r}")

        try:
            if getattr(bot_instance, "admin_db", None):
                await bot_instance.admin_db.log_search(user_id, chat.id, query, ym_count, vk_count, yt_count, len(songs))
                await bot_instance.admin_db.log_user_action(
                    user_id, chat.id, "search",
                    {"query": query, "ym": ym_count, "vk": vk_count, "yt": yt_count, "total": len(songs)},
                )
        except Exception:
            logger.debug("Search statistics write failed", exc_info=True)

        if not songs:
            await safe_edit_text(
                search_message,
                f"❌ <b>По запросу ничего не найдено</b>\n\n<i>Запрос:</i> {_esc(query)}",
                parse_mode="HTML",
            )
            return

        session_id = await bot_instance.session_manager.create_session(user_id, query, songs)
        session = await bot_instance.session_manager.get_session(session_id)
        if session:
            session.chat_id = chat.id
            session.search_message_id = search_message.message_id
            if effective_message:
                session.original_message_id = effective_message.message_id
                session.message_thread_id = getattr(effective_message, "message_thread_id", None)

        bot_instance.schedule_enrich_session(
            session_id, query, int(getattr(config, "SESSION_ENRICH_LIMIT", 30) or 30)
        )

        if chat_type in ("group", "supergroup"):
            top_n = int(getattr(config, "GROUP_CHAT_RESULTS_COUNT", 5) or 5)
            best_count = max(top_n, int(getattr(config, "GROUP_CHAT_BEST_CACHE_COUNT", 10) or 10))
            best_tracks = songs[:best_count]
            top_tracks = best_tracks[:top_n]
            if session:
                session.best_tracks = best_tracks
                session.reindex_tracks()

            rows = []
            max_len = int(getattr(config, "BUTTON_TEXT_MAX_LENGTH", 48) or 48)
            for index, track in enumerate(top_tracks):
                icon = _source_icon(track.get("source"))
                label = f"{_track_duration_text(track)} {icon} {track.get('artist') or 'Неизвестный исполнитель'} — {track.get('title') or 'Неизвестный трек'}"
                if len(label) > max_len:
                    label = label[:max_len - 1].rstrip() + "…"
                uid = _track_uid_from_any(track) or str(index)
                rows.append([InlineKeyboardButton(label, callback_data=f"dl:{session_id}:{uid}")])

            rows.append([InlineKeyboardButton("👥 Похожие", callback_data=f"similar_search:{session_id}")])
            if len(best_tracks) > len(top_tracks):
                rows.append([InlineKeyboardButton("➕ Показать ещё", callback_data=f"more:{session_id}")])
            rows.append([InlineKeyboardButton("❌ Закрыть список", callback_data="close_search")])

            source_parts = []
            if ym_count:
                source_parts.append(f"🎵 Яндекс: {ym_count}")
            if vk_count:
                source_parts.append(f"🎶 VK: {vk_count}")
            if yt_count:
                source_parts.append(f"▶️ YouTube: {yt_count}")
            source_info = ("\n" + " | ".join(source_parts)) if source_parts else ""

            await safe_edit_text(
                search_message,
                f"✅ <b>Топ-{len(top_tracks)} по запросу:</b> <i>{_esc(query)}</i>{source_info}\n\nВыберите трек кнопкой ниже:",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup(rows),
                disable_web_page_preview=True,
            )
            return

        await show_songs_page(update, context, search_message, user_id, 0, session_id)
    except Exception as exc:
        logger.error(f"Search processing error: {exc}", exc_info=True)
        target = update.effective_message
        if target:
            await target.reply_text("❌ <b>Произошла ошибка при поиске</b>\nПопробуйте позже.", parse_mode="HTML")

async def show_songs_page(update: Update, context: CallbackContext, message, user_id: int, page: int, session_id: str):
    """Показ результатов поиска (постранично). В сообщении — кратко, названия треков — в кнопках."""
    session = await bot_instance.session_manager.get_session(session_id)
    if not session:
        await safe_edit_text(message, "❌ Сессия устарела. Начните поиск заново.")
        return

    all_songs = session.results or []
    query = getattr(session, "query", "") or ""

    # Сколько треков на страницу (в личке — много; в группах этот экран почти не используется)
    songs_per_page = int(getattr(config, "SONGS_PER_PAGE", 10) or 10)
    total_pages = max(1, math.ceil(len(all_songs) / songs_per_page))
    page = max(0, min(int(page or 0), total_pages - 1))

    start_idx = page * songs_per_page
    end_idx = start_idx + songs_per_page
    current_songs = all_songs[start_idx:end_idx]

    ym_count = sum(1 for s in all_songs if s.get('source') == 'ym')
    vk_count = sum(1 for s in all_songs if s.get('source') == 'vk')
    yt_count = sum(1 for s in all_songs if s.get('source') == 'yt')

    source_parts = []
    if ym_count:
        source_parts.append(f"🎵 Яндекс: {ym_count}")
    if vk_count:
        source_parts.append(f"🎶 VK: {vk_count}")
    if yt_count:
        source_parts.append(f"▶️ YouTube: {yt_count}")
    sources_info = ("\n" + " | ".join(source_parts)) if source_parts else ""

    results_text = (
        f"🎵 <b>{_esc(query)}</b>\n"
        f"Найдено <b>{len(all_songs)}</b> · страница <b>{page + 1}/{total_pages}</b>"
        f"{sources_info}\n\n"
        "Нажмите на трек для скачивания. После скачивания появится кнопка «Добавить в избранное»."
    )

    def _truncate_btn(text: str, max_len: int) -> str:
        text = (text or "").strip()
        if max_len and len(text) > max_len:
            return text[: max_len - 1].rstrip() + "…"
        return text

    btn_max = int(getattr(config, "BUTTON_TEXT_MAX_LENGTH", 48) or 48)

    keyboard_rows = []
    # Fast local-only check for Telegram file_id cache (to show ⚡ and enable fileID-first UX)
    local_fileid_keys: Set[str] = set()
    try:
        peek_items = [((s.get('artist') or ''), (s.get('title') or ''), s.get('vk_key')) for s in (current_songs or [])]
        local_fileid_keys = await bot_instance.peek_tg_file_id_local_bulk(peek_items)
    except Exception:
        local_fileid_keys = set()

    for i, song in enumerate(current_songs, start=start_idx + 1):
        global_idx = i - 1
        uid = _track_uid_from_any(song) or str(global_idx)
        source_icon = _source_icon(song.get('source'))
        artist_raw = (song.get("artist") or "Неизвестный исполнитель").strip()
        title_raw = (song.get("title") or "Неизвестный трек").strip()
        src = (song.get('source') or 'vk')
        badge = _source_badge(src)
        k = bot_instance._tg_query_key(artist_raw, title_raw, song.get('vk_key'))
        speed = '⚡' if k in local_fileid_keys else ''
        prefix = f"{i:02d} · {_track_duration_text(song)}"
        badges = f" {speed}" if speed else ""
        label = _truncate_btn(f"{prefix}{badges} · {source_icon} {artist_raw} — {title_raw}", btn_max)
        dl_btn = InlineKeyboardButton(label, callback_data=f"dl:{session_id}:{uid}")
        keyboard_rows.append([dl_btn])

    pagination_buttons = []
    if page > 0:
        pagination_buttons.append(InlineKeyboardButton("⬅️ Назад", callback_data=f"page:{session_id}:{page-1}"))
    pagination_buttons.append(InlineKeyboardButton(f"{page+1}/{total_pages}", callback_data="noop"))
    if page < total_pages - 1:
        pagination_buttons.append(InlineKeyboardButton("Вперед ➡️", callback_data=f"page:{session_id}:{page+1}"))
    if pagination_buttons:
        keyboard_rows.append(pagination_buttons)

    # Primary actions: keep the most useful actions together instead of stacking wide buttons.
    if len(all_songs or []) > 0:
        actions = [InlineKeyboardButton("⬇️ Первый", callback_data=f"dlbest:{session_id}")]
        if getattr(message, "chat", None) and getattr(message.chat, "type", None) == "private":
            actions.append(InlineKeyboardButton(f"⬇️ Все · {len(all_songs)}", callback_data=f"dlall:{session_id}"))
        keyboard_rows.append(actions)

    keyboard_rows.append([
        InlineKeyboardButton("👥 Похожие", callback_data=f"similar_search:{session_id}"),
        InlineKeyboardButton("✕ Закрыть", callback_data="close_search"),
    ])

    await safe_edit_text(
        message,
        results_text,
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(keyboard_rows),
        disable_web_page_preview=True
    )

    if user_id in bot_instance.user_sessions:
        bot_instance.user_sessions[user_id]["current_page"] = page


async def handle_download_callback(update: Update, context: CallbackContext, query, session_id: str, track_ref: str):
    """Обработка callback загрузки.

    track_ref can be either a stable uid (preferred) or a legacy integer index.
    """
    logger.info(f"Processing download callback: session_id={session_id}, track_ref={track_ref}")
    try:
        if hasattr(bot_instance, "admin_db") and bot_instance.admin_db and getattr(query, "from_user", None):
            await bot_instance.admin_db.log_user_action(
                query.from_user.id,
                query.message.chat.id if getattr(query, "message", None) and getattr(query.message, "chat", None) else 0,
                "download_click",
                {"session_id": session_id, "track_ref": track_ref}
            )
    except Exception:
        pass


    session_data = await bot_instance.session_manager.get_session(session_id)
    if not session_data:
        logger.warning(f"Session not found: {session_id}")
        await safe_answer_callback(query, "❌ Сессия устарела. Выполните поиск заново.", show_alert=True)
        return

    chat_type = query.message.chat.type if query and query.message and getattr(query.message, 'chat', None) else None
    original_message_id = query.message.message_id if chat_type in ("group", "supergroup") else session_data.original_message_id

    track = session_data.get_track(str(track_ref), chat_type)
    if not track:
        await safe_answer_callback(query, "❌ Трек недоступен. Выполните поиск заново.", show_alert=True)
        return

    # Per-chat download limiter
    chat_id = int(update.effective_chat.id) if update and update.effective_chat else int(query.message.chat_id)
    sem = bot_instance.get_chat_download_semaphore(chat_id, chat_type or 'private')
    async with sem_guard(sem):
        await download_and_send_audio(update, context, query.message, track, original_message_id)



async def handle_direct_user_track_download(update: Update, context: CallbackContext, query, track: Dict[str, Any]) -> None:
    """Download a track dict from favorites/history (no session)."""
    try:
        await safe_answer_callback(query, "⬇️ Загружаю...", show_alert=False)
    except Exception:
        pass
    chat = getattr(query, 'message', None)
    if not chat:
        return
    chat_type = getattr(chat.chat, 'type', 'private') if getattr(chat, 'chat', None) else 'private'
    chat_id = int(chat.chat.id) if getattr(chat, 'chat', None) else int(update.effective_chat.id)
    sem = bot_instance.get_chat_download_semaphore(chat_id, chat_type)
    original_message_id = chat.message_id if chat_type in ('group','supergroup') else None
    async with sem_guard(sem):
        await download_and_send_audio(update, context, chat, track, original_message_id)


async def handle_download_best_callback(update: Update, context: CallbackContext, query, session_id: str):
    """Скачать первый результат текущей выдачи."""
    session = await bot_instance.session_manager.get_session(session_id)
    if not session:
        await safe_answer_callback(query, "❌ Сессия устарела. Выполните поиск заново.", show_alert=True)
        return

    chat_type = query.message.chat.type if query and query.message else "private"
    tracks = session.best_tracks if chat_type in ("group", "supergroup") else session.results
    if not tracks:
        await safe_answer_callback(query, "❌ В сессии нет треков.", show_alert=True)
        return

    await safe_answer_callback(query, "⬇️ Загружаю первый результат…", show_alert=False)
    original_message_id = query.message.message_id if chat_type in ("group", "supergroup") else session.original_message_id
    sem = bot_instance.get_chat_download_semaphore(query.message.chat.id, chat_type)
    async with sem_guard(sem):
        await download_and_send_audio(update, context, query.message, tracks[0], original_message_id)


async def handle_download_all_callback(update: Update, context: CallbackContext, query, session_id: str):
    """Скачать все найденные треки (только в личке)"""
    logger.info(f"Processing download-all callback: session_id={session_id}")
    try:
        if hasattr(bot_instance, "admin_db") and bot_instance.admin_db and getattr(query, "from_user", None):
            await bot_instance.admin_db.log_user_action(
                query.from_user.id,
                query.message.chat.id if getattr(query, "message", None) and getattr(query.message, "chat", None) else 0,
                "download_all",
                {"session_id": session_id}
            )
    except Exception:
        pass


    session_data = await bot_instance.session_manager.get_session(session_id)
    if not session_data:
        await safe_answer_callback(query, "❌ Сессия устарела. Выполните поиск заново.", show_alert=True)
        return

    # Разрешаем только в личке
    if not (query and query.message and getattr(query.message, 'chat', None) and getattr(query.message.chat, 'type', None) == "private"):
        await safe_answer_callback(query, "❌ Эта кнопка работает только в личных сообщениях.", show_alert=True)
        return

    # Берём полный список результатов (в личке всегда results)
    tracks = (getattr(session_data, "results", None) or getattr(session_data, "best_tracks", None) or [])
    if not tracks:
        await safe_answer_callback(query, "❌ В этой сессии нет треков. Выполните поиск заново.", show_alert=True)
        return

    await safe_answer_callback(query, "⬇️ Начинаю загрузку…", show_alert=False)

    # Ограничение на массовую загрузку (чтобы не улететь в flood/ограничения Telegram)
    limit = int(getattr(config, "DOWNLOAD_ALL_LIMIT", 50) or 50)
    total = len(tracks)
    if total > limit:
        tracks_to_send = tracks[:limit]
        note = f"\n\n⚠️ <i>Показано {limit} из {total}. Лимит можно изменить параметром DOWNLOAD_ALL_LIMIT в config.py</i>"
    else:
        tracks_to_send = tracks
        note = ""

    # FileID-first: if we already have Telegram file_id locally, send those first (feels instant)
    try:
        items = [((t.get('artist') or ''), (t.get('title') or ''), t.get('vk_key')) for t in (tracks_to_send or [])[:min(30, len(tracks_to_send or []))]]
        local_keys = await bot_instance.peek_tg_file_id_local_bulk(items)
        if local_keys:
            def _has_local(t):
                k = bot_instance._tg_query_key(t.get('artist',''), t.get('title',''), t.get('vk_key'))
                return 1 if k in local_keys else 0
            tracks_to_send = sorted(tracks_to_send, key=_has_local, reverse=True)
    except Exception:
        pass


    # Сообщение прогресса + кнопка отмены
    progress_text = f"⬇️ <b>Скачиваю все треки</b>\nВсего: {len(tracks_to_send)}{note}\n\n⏳ 0/{len(tracks_to_send)}"
    try:
        # reset cancel flag
        try:
            session_data.download_all_cancel = False
        except Exception:
            pass
        progress_msg = await safe_send_message(context.bot, 
            chat_id=query.message.chat.id,
            text=progress_text,
            parse_mode="HTML",
            message_thread_id=getattr(query.message, "message_thread_id", None),
            disable_web_page_preview=True,
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⏹️ Остановить", callback_data=f"dlall_cancel:{session_id}")]])
        )
    except Exception:
        progress_msg = None

    

    # Пер-чат лимитер (в личке) + параллельная отправка пачки
    chat_id = int(query.message.chat.id)
    sem = bot_instance.get_chat_download_semaphore(chat_id, "private")

    # Сколько воркеров запускать для download-all.
    # Реальный параллелизм всё равно ограничится семафором `sem` и глобальным `download_semaphore`.
    try:
        workers = int(getattr(config, "DOWNLOAD_ALL_WORKERS", 0) or 0)
    except Exception:
        workers = 0
    if workers <= 0:
        # По умолчанию — как лимит скачиваний для лички (быстро, но без перегруза)
        workers = int(getattr(config, "MAX_CONCURRENT_DOWNLOADS_PRIVATE", 6) or 6)
    workers = max(1, min(workers, 12))

    ok = 0
    fail = 0
    done = 0
    total_to_send = len(tracks_to_send)

    counter_lock = asyncio.Lock()
    q: asyncio.Queue = asyncio.Queue()

    for t in tracks_to_send:
        q.put_nowait(t)

    async def _maybe_update_progress(force: bool = False):
        if not progress_msg:
            return
        now = time.time()
        last = float(getattr(session_data, '_download_all_last_update', 0.0) or 0.0)
        # не флудим правками: раз в 1.2с или по force
        if (not force) and (now - last) < 1.2:
            return
        try:
            session_data._download_all_last_update = now
        except Exception:
            pass
        try:
            await safe_edit_text(
                progress_msg,
                f"⬇️ <b>Скачиваю все треки</b>\nВсего: {total_to_send}{note}\n\n⏳ {done}/{total_to_send}\n✅ Успешно: {ok}\n❌ Ошибок: {fail}"
                + ("\n\n⏹️ <i>Остановлено пользователем</i>" if getattr(session_data, 'download_all_cancel', False) else ""),
                parse_mode="HTML",
                disable_web_page_preview=True,
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⏹️ Остановить", callback_data=f"dlall_cancel:{session_id}")]])
                if not getattr(session_data, 'download_all_cancel', False) else None
            )
        except Exception as e:
            logger.debug(f"download-all progress update failed: {e}")

    async def _worker(worker_idx: int):
        nonlocal ok, fail, done
        while True:
            try:
                if getattr(session_data, 'download_all_cancel', False):
                    break
            except Exception:
                pass

            try:
                track = q.get_nowait()
            except asyncio.QueueEmpty:
                break

            try:
                async with sem_guard(sem):
                    success = await download_and_send_audio(
                        update, context, query.message, track, original_message_id=None
                    )
            except Exception as e:
                logger.warning(f"download-all worker#{worker_idx} failed: {e}")
                success = False

            async with counter_lock:
                done += 1
                if success:
                    ok += 1
                else:
                    fail += 1

            await _maybe_update_progress(force=(done == 1 or done == total_to_send or done % 3 == 0))

            try:
                q.task_done()
            except Exception:
                pass

    worker_tasks = [asyncio.create_task(_worker(i+1)) for i in range(workers)]
    try:
        await asyncio.gather(*worker_tasks, return_exceptions=True)
    finally:
        for t in worker_tasks:
            try:
                if not t.done():
                    t.cancel()
            except Exception:
                pass
        await _maybe_update_progress(force=True)

    # Финал
    if progress_msg:
        stopped = bool(getattr(session_data, 'download_all_cancel', False))
        text = ("⏹️ <b>Остановлено</b>" if stopped else "✅ <b>Готово!</b>")
        text += f"\nВсего отправлено: {ok}\nОшибок: {fail}"
        try:
            await safe_edit_text(progress_msg, text, parse_mode="HTML", disable_web_page_preview=True)
        except Exception as e:
            logger.debug(f"download-all final update failed: {e}")


async def handle_page_callback(update: Update, context: CallbackContext, query, session_id: str, page: int):
    """Обработка пагинации"""
    logger.info(f"Processing page callback: session_id={session_id}, page={page}")
    await safe_answer_callback(query)
    await show_songs_page(update, context, query.message, query.from_user.id, page, session_id)

async def handle_more_callback(update: Update, context: CallbackContext, query, session_id: str):
    """Показать расширенный список в группе или теме форума."""
    session = await bot_instance.session_manager.get_session(session_id)
    if not session or not getattr(session, "best_tracks", None):
        await safe_answer_callback(query, "❌ Сессия устарела. Повторите поиск.", show_alert=True)
        return

    await safe_answer_callback(query)
    limit = int(getattr(config, "GROUP_CHAT_BEST_CACHE_COUNT", 10) or 10)
    tracks = session.best_tracks[:max(1, limit)]
    rows = []
    max_len = int(getattr(config, "BUTTON_TEXT_MAX_LENGTH", 48) or 48)
    for index, track in enumerate(tracks):
        icon = _source_icon(track.get("source"))
        label = f"{_track_duration_text(track)} {index + 1}. {icon} {track.get('artist', '')} — {track.get('title', '')}"
        if len(label) > max_len:
            label = label[:max_len - 1].rstrip() + "…"
        rows.append([InlineKeyboardButton(
            label,
            callback_data=f"dl:{session_id}:{_track_uid_from_any(track) or index}",
        )])
    rows.append([InlineKeyboardButton("❌ Закрыть список", callback_data="close_search")])

    await safe_edit_message_text(
        query,
        f"✅ <b>Найденные треки по запросу:</b>\n\n<i>{_esc(session.query)}</i>\n\nВыберите трек:",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(rows),
    )

# ==================== НОВЫЕ ФУНКЦИИ ДЛЯ КНОПОК ====================
async def show_charts_menu(update: Update, context: CallbackContext):
    """Показать меню чартов"""
    try:
        q = update.callback_query
        if q and hasattr(bot_instance, "admin_db") and bot_instance.admin_db:
            await bot_instance.admin_db.log_user_action(q.from_user.id, q.message.chat.id, "menu_charts", {})
    except Exception:
        pass

    query = update.callback_query
    
    keyboard = [
        [InlineKeyboardButton("🌍 Глобальные хиты", callback_data="charts:global")],
        [InlineKeyboardButton("📅 Недельные хиты", callback_data="charts:weekly")],
        [InlineKeyboardButton("🎸 Рок", callback_data="charts:rock")],
        [InlineKeyboardButton("🎵 Поп", callback_data="charts:pop")],
        [InlineKeyboardButton("🎤 Хип-хоп", callback_data="charts:hiphop")],
        [InlineKeyboardButton("🎧 Электроника", callback_data="charts:electronic")],
        [InlineKeyboardButton("🔙 Назад", callback_data="main_menu")]
    ]
    
    await query.edit_message_text(
        text="📊 <b>Топ-чарты Last.fm</b>\n\nВыберите категорию:",
        parse_mode='HTML',
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def show_similar_artists_menu(update: Update, context: CallbackContext):
    """Показать меню поиска похожих исполнителей"""
    try:
        q = update.callback_query
        if q and hasattr(bot_instance, "admin_db") and bot_instance.admin_db:
            await bot_instance.admin_db.log_user_action(q.from_user.id, q.message.chat.id, "menu_similar_artists", {})
    except Exception:
        pass

    query = update.callback_query
    
    await query.edit_message_text(
        text="👥 <b>Похожие исполнители</b>\n\nВведите имя исполнителя для поиска похожих:",
        parse_mode='HTML'
    )
    await bot_instance.set_user_state(query.from_user.id, "similar_artists")

async def handle_charts_selection(update: Update, context: CallbackContext, chart_type: str):
    """Обработка выбора чарта"""
    try:
        q = update.callback_query
        if q and hasattr(bot_instance, "admin_db") and bot_instance.admin_db:
            await bot_instance.admin_db.log_user_action(
                q.from_user.id, q.message.chat.id, "charts_select", {"chart_type": chart_type}
            )
    except Exception:
        pass

    query = update.callback_query
    user_id = query.from_user.id
    
    await safe_answer_callback(query, f"Загружаем {chart_type} чарт...")
    
    try:
        tracks = []
        genre_name = ""
        
        if chart_type == "global":
            tracks = await bot_instance.get_popular_tracks()
            genre_name = "Глобальные хиты"
        elif chart_type == "weekly":
            tracks = await bot_instance.get_popular_tracks()
            genre_name = "Недельные хиты"
        else:
            genre_map = {
                "rock": "rock",
                "pop": "pop",
                "hiphop": "hip-hop",
                "electronic": "electronic",
                "metal": "metal",
                "jazz": "jazz"
            }
            genre_key = genre_map.get(chart_type, chart_type)
            tracks = await bot_instance.get_top_tracks_by_genre(genre_key)
            genre_name = chart_type.capitalize()
        
        if not tracks:
            await query.edit_message_text(
                text=f"❌ <b>Не удалось загрузить {genre_name}</b>\n\nПопробуйте позже.",
                parse_mode='HTML'
            )
            return
        
        # Создаем сессию для чартов
        session_id = await bot_instance.session_manager.create_session(
            user_id, 
            f"charts:{chart_type}", 
            [{"artist": t["artist"], "title": t["name"], "duration": int(t.get("duration") or 0), "source": "chart"} for t in tracks[:10]]
        )
        
        # Показываем треки
        message_text = f"📊 <b>{genre_name} (топ-10)</b>\n\n"
        for i, track in enumerate(tracks[:10], 1):
            message_text += f"{_track_duration_text(track)} {i}. <b>{html.escape(str(track.get('artist','')))}</b> — {html.escape(str(track.get('name','')))}\n"
        
        message_text += "\nВыберите номер трека для поиска:"
        
        keyboard = []
        row = []
        # В callback кладём индекс (0..9). Раньше здесь использовалась переменная uid,
        # которой в этом скоупе нет → NameError и пользователь видел "не удалось".
        for i in range(1, 11):
            row.append(InlineKeyboardButton(str(i), callback_data=f"chart_dl:{session_id}:{i-1}"))
            if len(row) == 5:
                keyboard.append(row)
                row = []
        if row:
            keyboard.append(row)
        
        keyboard.append([InlineKeyboardButton("🔙 Назад к чартам", callback_data="charts_menu")])
        
        await query.edit_message_text(
            text=message_text,
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup(keyboard),
            disable_web_page_preview=True
        )
        
    except Exception as e:
        logger.error(f"Error loading charts: {e}")
        await query.edit_message_text(
            text="❌ <b>Ошибка при загрузке чартов</b>\n\nПопробуйте позже.",
            parse_mode='HTML'
        )

async def handle_chart_download(update: Update, context: CallbackContext, session_id: str, track_ref: str):
    """Обработка скачивания трека из чарта"""
    query = update.callback_query
    session = await bot_instance.session_manager.get_session(session_id)
    
    if not session:
        await safe_answer_callback(query, "❌ Трек недоступен", show_alert=True)
        return

    track_info = session.get_track(str(track_ref), chat_type=None) if hasattr(session, 'get_track') else None
    if not track_info:
        await safe_answer_callback(query, "❌ Трек недоступен", show_alert=True)
        return

    # Ищем трек по названию
    search_query = f"{track_info['artist']} - {track_info['title']}"
    await safe_answer_callback(query, f"Ищем: {search_query}")
    
    # Вызываем поиск
    await process_search(update, context, search_query)


def _guess_artist_for_similar_from_session(session) -> str:
    """Пытаемся определить исполнителя для 'похожих' из поисковой сессии."""
    try:
        q = (getattr(session, 'query', '') or '').strip()
        if ' - ' in q:
            left = q.split(' - ', 1)[0].strip()
            if left:
                return left
        # fallback: первый найденный трек
        tracks = (getattr(session, 'best_tracks', None) or getattr(session, 'results', None) or [])
        if tracks:
            a = (tracks[0].get('artist') or '').strip()
            if a:
                return a
        return q
    except Exception:
        return (getattr(session, 'query', '') or '').strip()


async def _render_group_search_from_session(query, session_id: str) -> None:
    """Вернуть пользователя к экрану выдачи (группа/супергруппа) без повторного поиска."""
    session = await bot_instance.session_manager.get_session(session_id)
    if not session:
        await safe_answer_callback(query, "❌ Сессия устарела", show_alert=True)
        return

    top_n = int(getattr(config, "GROUP_CHAT_RESULTS_COUNT", 5) or 5)
    best_tracks = (getattr(session, 'best_tracks', None) or getattr(session, 'results', None) or [])
    top_tracks = (best_tracks or [])[:top_n]
    if not top_tracks:
        await safe_edit_message_text(query, "❌ Сессия устарела. Начните поиск заново.")
        return

    qtxt = _esc(getattr(session, 'query', '') or '')

    def _truncate_btn(text: str, max_len: int) -> str:
        try:
            text = (text or '').strip()
            if max_len and len(text) > max_len:
                return text[: max_len - 1].rstrip() + '…'
            return text
        except Exception:
            return text

    btn_max = int(getattr(config, "BUTTON_TEXT_MAX_LENGTH", 48) or 48)
    keyboard_rows = []
    for i, track in enumerate(top_tracks):
        source_icon = _source_icon(track.get('source'))
        artist_raw = (track.get('artist') or 'Неизвестный исполнитель')
        title_raw = (track.get('title') or 'Неизвестный трек')
        label = _truncate_btn(f"{_track_duration_text(track)} {source_icon} {artist_raw} — {title_raw}", btn_max)
        uid = _track_uid_from_any(track) or str(i)
        keyboard_rows.append([InlineKeyboardButton(label, callback_data=f"dl:{session_id}:{uid}")])

    # То, что пользователь просил: выдача + отдельная кнопка
    keyboard_rows.append([InlineKeyboardButton("👥 Найти похожих исполнителей", callback_data=f"similar_search:{session_id}")])

    if len(best_tracks) > len(top_tracks):
        keyboard_rows.append([InlineKeyboardButton("➕ Показать ещё", callback_data=f"more:{session_id}")])

    keyboard_rows.append([InlineKeyboardButton("❌ Закрыть список", callback_data="close_search")])

    message_text = (
        f"✅ <b>Топ-{len(top_tracks)} по запросу:</b> <i>{qtxt}</i>\n\n"
        "Выберите трек кнопкой ниже:"
    )
    await safe_edit_message_text(query, message_text, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard_rows), disable_web_page_preview=True)


async def handle_similar_from_search_callback(update: Update, context: CallbackContext, session_id: str):
    """Показ похожих исполнителей прямо из выдачи поиска (кнопка под результатами)."""
    query = update.callback_query
    session = await bot_instance.session_manager.get_session(session_id)
    if not session:
        await safe_answer_callback(query, "❌ Сессия устарела", show_alert=True)
        return

    artist_name = _guess_artist_for_similar_from_session(session)
    if not artist_name or len(artist_name) < 2:
        await safe_answer_callback(query, "❌ Не удалось определить исполнителя", show_alert=True)
        return

    await safe_answer_callback(query, "👥 Ищу похожих…")
    try:
        similar_artists = await bot_instance.get_similar_artists(artist_name)
    except Exception as e:
        logger.error(f"Error getting similar artists from search ({artist_name}): {e}")
        similar_artists = []

    if not similar_artists:
        await safe_edit_message_text(
            query,
            f"❌ <b>Не удалось найти похожих исполнителей для:</b>\n<i>{_esc(artist_name)}</i>",
            parse_mode='HTML'
        )
        return

    message_text = f"👥 <b>Похожие на { _esc(artist_name) }:</b>\n\n"
    for i, artist in enumerate(similar_artists[:10], 1):
        name = artist.get('name') if isinstance(artist, dict) else str(artist)
        match = artist.get('match') if isinstance(artist, dict) else None
        try:
            match_percent = int(float(match) * 100) if match is not None else None
        except Exception:
            match_percent = None
        if match_percent is not None:
            message_text += f"{i}. <b>{_esc(str(name))}</b> ({match_percent}%)\n"
        else:
            message_text += f"{i}. <b>{_esc(str(name))}</b>\n"

    keyboard = [
        [InlineKeyboardButton(f"🎵 Искать треки {a.get('name','')}", callback_data=f"search:{a.get('name','')}")]
        for a in (similar_artists[:3] if isinstance(similar_artists[0], dict) else [])
    ]

    # Если формат API вернул список строк/объектов не dict — сделаем кнопки вручную
    if not keyboard:
        keyboard = []
        for a in similar_artists[:3]:
            name = a.get('name') if isinstance(a, dict) else str(a)
            keyboard.append([InlineKeyboardButton(f"🎵 Искать треки {name}", callback_data=f"search:{name}")])

    keyboard.append([InlineKeyboardButton("🔙 Назад к результатам", callback_data=f"back_search:{session_id}")])

    await safe_edit_message_text(
        query,
        message_text,
        parse_mode='HTML',
        reply_markup=InlineKeyboardMarkup(keyboard),
        disable_web_page_preview=True
    )


async def handle_back_search_callback(update: Update, context: CallbackContext, session_id: str) -> None:
    """Вернуться к результатам поиска после показа похожих исполнителей."""
    query = update.callback_query
    try:
        chat_type = getattr(getattr(query, 'message', None), 'chat', None)
        chat_type = getattr(chat_type, 'type', None)
    except Exception:
        chat_type = None

    if chat_type in ("group", "supergroup"):
        await safe_answer_callback(query)
        await _render_group_search_from_session(query, session_id)
        return

    # private / other: просто перерисуем страницу 1
    session = await bot_instance.session_manager.get_session(session_id)
    if not session:
        await safe_answer_callback(query, "❌ Сессия устарела", show_alert=True)
        return
    user_id = getattr(getattr(query, 'from_user', None), 'id', 0) or 0
    await safe_answer_callback(query)
    await show_songs_page(update, context, query.message, user_id, 0, session_id)

async def handle_similar_artists_input(update: Update, context: CallbackContext):
    """Обработка ввода исполнителя для поиска похожих"""
    artist_name = update.message.text.strip()
    
    if len(artist_name) < 2:
        await update.message.reply_text("❌ <b>Слишком короткое имя исполнителя</b>\nМинимум 2 символа.", parse_mode='HTML')
        return
    
    try:
        similar_artists = await bot_instance.get_similar_artists(artist_name)
        
        if not similar_artists:
            await bot_instance.clear_user_state(update.effective_user.id)
            await update.message.reply_text(
                f"❌ <b>Не удалось найти похожих исполнителей для:</b>\n<i>{_esc(artist_name)}</i>\n\nМожно снова выбрать действие в меню ниже.",
                parse_mode='HTML',
                reply_markup=private_main_keyboard(),
            )
            return
        
        message_text = f"👥 <b>Похожие на {artist_name}:</b>\n\n"
        for i, artist in enumerate(similar_artists[:8], 1):
            match_percent = int(artist['match'] * 100)
            message_text += f"{i}. <b>{artist['name']}</b> ({match_percent}%)\n"
        
        keyboard = [
            [InlineKeyboardButton(f"🎵 Искать треки {artist['name']}", callback_data=f"search:{artist['name']}")]
            for artist in similar_artists[:3]
        ]
        keyboard.append([InlineKeyboardButton("🔙 Назад", callback_data="similar_menu")])
        
        await bot_instance.clear_user_state(update.effective_user.id)
        await update.message.reply_text(
            text=message_text,
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup(keyboard),
            disable_web_page_preview=True
        )
        
    except Exception as e:
        logger.error(f"Error getting similar artists: {e}")
        await bot_instance.clear_user_state(update.effective_user.id)
        await update.message.reply_text(
            "❌ <b>Ошибка при поиске похожих исполнителей</b>\n\nПопробуйте позже.",
            parse_mode='HTML',
            reply_markup=private_main_keyboard(),
        )

async def show_help_menu(update: Update, context: CallbackContext):
    """Показать меню помощи"""
    try:
        q = update.callback_query
        if q and hasattr(bot_instance, "admin_db") and bot_instance.admin_db:
            await bot_instance.admin_db.log_user_action(q.from_user.id, q.message.chat.id, "menu_help", {})
    except Exception:
        pass

    query = update.callback_query
    
    help_text = """
🆘 <b>Помощь по использованию бота</b>

<b>Основные команды:</b>
• /start - Запустить бота
• /search [запрос] - Поиск музыки
• /help - Показать это меню

<b>Как искать:</b>
• Введите название трека или исполнителя
• Формат: "Исполнитель - Название трека"
• Используйте @бот [запрос] в группах

<b>Источники музыки:</b>
• 🎵 Яндекс.Музыка - высокое качество
• 🎶 VK - большой каталог

<b>Дополнительные функции:</b>
• 📊 Чарты - популярные треки
• 👥 Похожие исполнители
"""
    
    keyboard = [
        [InlineKeyboardButton("🔍 Начать поиск", callback_data="new_search")],
        [InlineKeyboardButton("📊 Чарты", callback_data="charts_menu")],
        [InlineKeyboardButton("👥 Похожие исполнители", callback_data="similar_menu")],
        [InlineKeyboardButton("🔙 Главное меню", callback_data="main_menu")]
    ]
    
    await query.edit_message_text(
        text=help_text,
        parse_mode='HTML',
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def show_main_menu(query):
    """Показать главное меню"""
    try:
        if query and getattr(query, "from_user", None) and hasattr(bot_instance, "admin_db") and bot_instance.admin_db:
            await bot_instance.admin_db.log_user_action(
                query.from_user.id,
                query.message.chat.id if getattr(query, "message", None) and getattr(query.message, "chat", None) else 0,
                "menu_main",
                {}
            )
    except Exception:
        pass

    keyboard = [
        [InlineKeyboardButton("🔍 Поиск музыки", callback_data="new_search")],
        [InlineKeyboardButton("🎲 Подборка по жанру", callback_data="mix:new")],
        [InlineKeyboardButton("📊 Чарты", callback_data="charts_menu")],
        [InlineKeyboardButton("👥 Похожие исполнители", callback_data="similar_menu")],
        [InlineKeyboardButton("🆘 Помощь", callback_data="help_menu")]
    ]

    # Админка показывается только админам
    try:
        if query and getattr(query, "from_user", None) and _is_admin(query.from_user.id):
            keyboard.append([InlineKeyboardButton("🛠 Админка", callback_data="admin_menu")])
    except Exception:
        pass

    
    await safe_edit_message_text(query, 
        "🏠 <b>Главное меню</b>\n\nВыберите действие из меню ниже:",
        parse_mode='HTML',
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def button_handler(update: Update, context: CallbackContext) -> None:
    """Route callback buttons exactly once and preserve informative alerts."""
    query = update.callback_query
    if query is None:
        return

    user = update.effective_user
    if not user:
        await safe_answer_callback(query, "❌ Ошибка пользователя", show_alert=True)
        return
    
    if not await user_check(update, enforce_cooldown=False):
        return

    # Регистрируем чат/пользователя (для статистики/экспорта/рассылки)
    try:
        if hasattr(bot_instance, "admin_db") and bot_instance.admin_db:
            await bot_instance.admin_db.track_chat(update, bot=context.bot)
            await bot_instance.admin_db.track_user(user)
    except Exception:
        pass

    try:
        callback_message = getattr(query, "message", None)
        callback_chat = getattr(callback_message, "chat", None)
        logger.info(
            "Processing callback in chat %s (%s): %s",
            getattr(callback_chat, "id", "inline"),
            getattr(callback_chat, "type", "inline"),
            query.data,
        )

        callback_data = query.data or ""

        if callback_data.startswith("mix:"):
            await handle_user_mix_callback(update, context)
            return

        if callback_data.startswith("digest:"):
            await handle_digest_callback(update, context)
            return

        # --- Admin access gate (protect ALL admin callbacks) ---
        if (
            callback_data.startswith("admin_")
            or callback_data in {"tokens_menu", "add_token", "refresh_tokens", "remove_token_menu"}
            or callback_data.startswith("del_token:")
        ):
            if not _is_admin(query.from_user.id):
                await safe_answer_callback(query, "⛔️ Доступ только для админов", show_alert=True)
                return

        if callback_data.startswith("user_settings:"):
            if not await user_check(update):
                return
            parts = callback_data.split(":")
            try:
                if parts[1] == "source":
                    await bot_instance.user_store.set_preferences(query.from_user.id, prefer_source=parts[2])
                elif parts[1] == "bitrate":
                    await bot_instance.user_store.set_preferences(query.from_user.id, prefer_bitrate_kbps=int(parts[2]))
                await safe_answer_callback(query, "Настройки сохранены")
                prefs = await bot_instance.user_store.get_preferences(query.from_user.id)
                await safe_edit_message_text(
                    query,
                    "⚙️ <b>Настройки сохранены</b>\n\n"
                    f"Источник: <b>{_esc(str(prefs.get('prefer_source') or 'auto'))}</b>\n"
                    f"Качество: <b>{int(prefs.get('prefer_bitrate_kbps') or 192)} kbps</b>",
                    parse_mode="HTML",
                )
            except Exception as exc:
                await safe_answer_callback(query, "Не удалось сохранить", show_alert=True)
                logger.warning("User settings update failed: %s", exc)
            return

        # --- Admin panel callbacks (admins only) ---
        if callback_data == "admin_menu":
            if _is_admin(query.from_user.id):
                await safe_answer_callback(query)
                await bot_instance.clear_user_state(query.from_user.id)
                await safe_edit_message_text(
                    query,
                    "🛠️ <b>VLMB Admin</b>\n\nВыберите раздел:",
                    parse_mode='HTML',
                    reply_markup=_admin_keyboard()
                )
            return

        if callback_data == "admin_close":
            await safe_answer_callback(query)
            await bot_instance.clear_user_state(query.from_user.id)
            try:
                await safe_edit_message_text(query, "✅ Админка закрыта.", parse_mode="HTML")
            except Exception:
                pass
            return

        if callback_data == "admin_fileid":
            if not _is_admin(query.from_user.id):
                await safe_answer_callback(query, "⛔️ Доступ только для админов", show_alert=True)
                return
            await safe_answer_callback(query)
            user_id = query.from_user.id
            await bot_instance.set_user_state(user_id, "admin_collect_fileid", {"ts": time.time()})
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton("🛑 Выйти из режима", callback_data="admin_fileid_stop")],
                [InlineKeyboardButton("🔙 Назад", callback_data="admin_menu")]
            ])
            await safe_edit_message_text(
                query,
                "📎 <b>Режим приёма file_id</b>\n\n"
                "Просто <b>перешли сюда любой файл</b> (документ/аудио/видео/фото/голосовое и т.д.).\n"
                "Бот поймает файл и запишет его <code>file_id</code> в базу.\n\n"
                "Чтобы выйти — нажми «🛑 Выйти из режима».",
                parse_mode="HTML",
                reply_markup=kb,
                disable_web_page_preview=True
            )
            return

        if callback_data == "admin_fileid_stop":
            if _is_admin(query.from_user.id):
                await safe_answer_callback(query)
                await bot_instance.clear_user_state(query.from_user.id)
                await safe_edit_message_text(
                    query,
                    "✅ Режим приёма file_id выключен.\n\nВыберите раздел:",
                    parse_mode="HTML",
                    reply_markup=_admin_keyboard()
                )
            return

        if callback_data.startswith("admin_stats"):
            days = 7
            if ":" in callback_data:
                try:
                    days = int(callback_data.split(":", 1)[1])
                except Exception:
                    days = 7
            await _show_admin_stats(query, days=days)
            return

        if callback_data.startswith("admin_export:"):
            kind = callback_data.split(":", 1)[1] if ":" in callback_data else ""
            await _handle_admin_export(update, context, kind)
            return

        if callback_data in {"admin_monitor", "admin_monitor:refresh"}:
            await _show_admin_monitor(query)
            return

        if callback_data.startswith("admin_cache"):
            if ":" in callback_data:
                action = callback_data.split(":", 1)[1]
                await _handle_admin_cache_action(query, action)
            else:
                await _show_admin_cache_menu(query)
            return


        if callback_data.startswith("admin_settings"):
            # ВАЖНО: _show_admin_settings/_handle_admin_settings_action принимают (update, context, ...)
            if ":" in callback_data:
                await _handle_admin_settings_action(update, context, callback_data)
            else:
                await _show_admin_settings(update, context)
            return

        if callback_data.startswith("admin_moderation"):
            if ":" in callback_data:
                action = callback_data.split(":", 1)[1]
                await _handle_admin_moderation_action(query, action)
            else:
                await _show_admin_moderation(query)
            return

        if callback_data.startswith("admin_groups"):
            page = 0
            if ":" in callback_data:
                try:
                    page = int(callback_data.split(":", 1)[1])
                except (TypeError, ValueError):
                    page = 0
            await _show_admin_groups(query, page=page)
            return

        if callback_data.startswith("admin_broadcast"):
            if ":" in callback_data:
                action = callback_data.split(":", 1)[1]
                await _handle_admin_broadcast_action(query, action, context=context)
            else:
                await _show_admin_broadcast(query, context=context)
            return


        if callback_data == "tokens_menu":
            await show_tokens_menu(update, context)
            return
        if callback_data == "add_token":
            await handle_add_token(update, context)
            return
        if callback_data == "refresh_tokens":
            if _is_admin(query.from_user.id):
                await safe_answer_callback(query)
                text, kb = await _render_tokens_menu()
                await safe_edit_message_text(query, text, parse_mode='HTML', reply_markup=kb, disable_web_page_preview=True)
            return
        if callback_data == "remove_token_menu":
            await show_remove_token_menu(update, context)
            return
        if callback_data.startswith("del_token:"):
            try:
                idx = int(callback_data.split(":", 1)[1])
            except Exception:
                idx = -1
            await handle_delete_token_callback(update, context, idx)
            return
        
        if callback_data.startswith("pldl:") or callback_data.startswith("plall:"):
            parts = callback_data.split(":")
            sid = parts[1] if len(parts) > 1 else ""
            sess = await bot_instance.session_manager.get_session(sid)
            if not sess or sess.user_id != query.from_user.id:
                await safe_answer_callback(query, "Список устарел", show_alert=True)
                return
            if callback_data.startswith("pldl:") and len(parts) >= 3:
                tr = sess.get_track(parts[2], query.message.chat.type if query.message else None)
                if not tr:
                    await safe_answer_callback(query, "Трек не найден", show_alert=True)
                    return
                await safe_answer_callback(query, "Добавлено в очередь")
                try:
                    job = await bot_instance.download_queue.submit(query.from_user.id, {"update": update, "context": context, "track": tr}, priority=50)
                    await safe_edit_message_text(query, f"⬇️ Трек добавлен в очередь. ID: <code>{job.job_id[:8]}</code>", parse_mode="HTML")
                except Exception as exc:
                    await safe_answer_callback(query, "Очередь переполнена", show_alert=True)
                    logger.warning("Playlist queue submit failed: %s", exc)
                return
            if callback_data.startswith("plall:"):
                await safe_answer_callback(query, "Playlist добавлен в очередь")
                submitted = 0
                for tr in sess.results[:int(getattr(config, "MAX_PLAYLIST_TRACKS", 100))]:
                    try:
                        await bot_instance.download_queue.submit(query.from_user.id, {"update": update, "context": context, "track": tr}, priority=80)
                        submitted += 1
                    except Exception:
                        break
                await safe_edit_message_text(query, f"📥 В очередь добавлено: <b>{submitted}</b> треков", parse_mode="HTML")
                return

        # No-op кнопки (индикатор страницы и т.п.)
        if callback_data in ("current_page", "noop"):
            await safe_answer_callback(query)
            return

        if callback_data.startswith('dlbest:'):
            parts = callback_data.split(':', 1)
            if len(parts) == 2 and parts[1]:
                session_id = parts[1]
                await handle_download_best_callback(update, context, query, session_id)
            else:
                await safe_answer_callback(query, "❌ Ошибка в данных сессии", show_alert=True)
            return

        if callback_data.startswith('dlall:'):
            parts = callback_data.split(':', 1)
            if len(parts) == 2 and parts[1]:
                session_id = parts[1]
                await handle_download_all_callback(update, context, query, session_id)
            else:
                await safe_answer_callback(query, "❌ Ошибка в данных сессии", show_alert=True)

        elif callback_data.startswith('dlall_cancel:'):
            parts = callback_data.split(':', 1)
            session_id = parts[1] if len(parts) == 2 else ""
            sess = await bot_instance.session_manager.get_session(session_id) if session_id else None
            if sess:
                try:
                    sess.download_all_cancel = True
                except Exception:
                    pass
            await safe_answer_callback(query, "⏹️ Останавливаю...", show_alert=False)

        elif callback_data.startswith('fav_audio:') and hasattr(bot_instance, 'user_store'):
            uid = callback_data.split(':', 1)[1].strip() if ':' in callback_data else ""
            if not uid:
                await safe_answer_callback(query, "❌ Не удалось определить трек", show_alert=True)
                return
            track = await bot_instance.user_store.get_history_track(query.from_user.id, uid)
            if not track:
                await safe_answer_callback(query, "❌ Трек не найден в истории", show_alert=True)
                return
            if await bot_instance.user_store.is_favorite(query.from_user.id, uid):
                await safe_answer_callback(query, "❤️ Трек уже в избранном", show_alert=False)
                try:
                    await query.message.edit_reply_markup(
                        reply_markup=InlineKeyboardMarkup([[
                            InlineKeyboardButton("💔 Убрать из избранного", callback_data=f"fav_audio_remove:{uid}"),
                        ]])
                    )
                except Exception:
                    pass
                return
            await bot_instance.user_store.add_favorite(query.from_user.id, uid, track)
            await safe_answer_callback(query, "❤️ Добавлено в избранное", show_alert=False)
            try:
                await query.message.edit_reply_markup(
                    reply_markup=InlineKeyboardMarkup([[
                        InlineKeyboardButton("💔 Убрать из избранного", callback_data=f"fav_audio_remove:{uid}"),
                    ]])
                )
            except Exception:
                pass

        elif callback_data.startswith('fav_audio_remove:') and hasattr(bot_instance, 'user_store'):
            uid = callback_data.split(':', 1)[1].strip() if ':' in callback_data else ""
            if not uid:
                await safe_answer_callback(query, "❌ Не удалось определить трек", show_alert=True)
                return
            await bot_instance.user_store.remove_favorite(query.from_user.id, uid)
            await safe_answer_callback(query, "💔 Убрано из избранного", show_alert=False)
            try:
                await query.message.edit_reply_markup(
                    reply_markup=InlineKeyboardMarkup([[
                        InlineKeyboardButton("❤️ Добавить в избранное", callback_data=f"fav_audio:{uid}"),
                    ]])
                )
            except Exception:
                pass

        elif callback_data.startswith('favtoggle:') and hasattr(bot_instance, 'user_store'):
            parts = callback_data.split(':', 2)
            if len(parts) != 3:
                await safe_answer_callback(query, "❌ Ошибка в данных избранного", show_alert=True)
                return
            session_id, track_ref = parts[1], parts[2]
            session = await bot_instance.session_manager.get_session(session_id)
            if not session:
                await safe_answer_callback(query, "❌ Результаты поиска устарели. Выполните поиск заново.", show_alert=True)
                return
            chat_type = getattr(getattr(query, 'message', None), 'chat', None)
            chat_type = getattr(chat_type, 'type', None)
            track = session.get_track(str(track_ref), chat_type=chat_type) if hasattr(session, 'get_track') else None
            if not track:
                await safe_answer_callback(query, "❌ Трек недоступен. Выполните поиск заново.", show_alert=True)
                return
            uid = _track_uid_from_any(track)
            if not uid:
                await safe_answer_callback(query, "❌ Не удалось определить трек", show_alert=True)
                return
            if await bot_instance.user_store.is_favorite(query.from_user.id, uid):
                await bot_instance.user_store.remove_favorite(query.from_user.id, uid)
                await safe_answer_callback(query, "💔 Убрано из избранного", show_alert=False)
            else:
                await bot_instance.user_store.add_favorite(query.from_user.id, uid, track)
                await safe_answer_callback(query, "❤️ Добавлено в избранное", show_alert=False)

        elif callback_data.startswith('favdl:') and hasattr(bot_instance, 'user_store'):
            uid = callback_data.split(':', 1)[1] if ':' in callback_data else ""
            tr = await bot_instance.user_store.get_favorite(query.from_user.id, uid)
            if not tr:
                await safe_answer_callback(query, "❌ Трек не найден в избранном", show_alert=True)
            else:
                await handle_direct_user_track_download(update, context, query, tr)

        elif callback_data.startswith('favrm:') and hasattr(bot_instance, 'user_store'):
            uid = callback_data.split(':', 1)[1] if ':' in callback_data else ""
            await bot_instance.user_store.remove_favorite(query.from_user.id, uid)
            await safe_answer_callback(query, "💔 Убрано из избранного", show_alert=True)
            try:
                await query.message.edit_reply_markup(reply_markup=None)
            except Exception:
                pass

        elif callback_data.startswith('histdl:') and hasattr(bot_instance, 'user_store'):
            uid = callback_data.split(':', 1)[1] if ':' in callback_data else ""
            tr = await bot_instance.user_store.get_history_track(query.from_user.id, uid)
            if not tr:
                await safe_answer_callback(query, "❌ Трек не найден", show_alert=True)
            else:
                await handle_direct_user_track_download(update, context, query, tr)

        elif callback_data.startswith('dl:'):
            parts = callback_data.split(':', 2)
            if len(parts) == 3:
                session_id = parts[1]
                track_ref = parts[2]
                await handle_download_callback(update, context, query, session_id, track_ref)
            else:
                await safe_answer_callback(query, "❌ Ошибка в данных трека", show_alert=True)
        
        elif callback_data.startswith('page:'):
            parts = callback_data.split(':')
            if len(parts) >= 3:
                session_id = parts[1]
                try:
                    page = max(0, int(parts[2]))
                except (TypeError, ValueError):
                    await safe_answer_callback(query, "❌ Ошибка в данных пагинации", show_alert=True)
                    return
                await handle_page_callback(update, context, query, session_id, page)
            else:
                await safe_answer_callback(query, "❌ Ошибка в данных пагинации", show_alert=True)
        
        elif callback_data.startswith('more:'):
            parts = callback_data.split(':', 1)
            if len(parts) == 2:
                session_id = parts[1]
                await handle_more_callback(update, context, query, session_id)
            else:
                await safe_answer_callback(query, "❌ Ошибка в данных", show_alert=True)
        
        elif callback_data.startswith('chart_dl:'):
            parts = callback_data.split(':')
            if len(parts) >= 3:
                session_id = parts[1]
                track_ref = parts[2]
                await handle_chart_download(update, context, session_id, track_ref)
        
        elif callback_data.startswith('charts:'):
            chart_type = callback_data.split(':')[1]
            await handle_charts_selection(update, context, chart_type)

        elif callback_data.startswith('similar_search:'):
            parts = callback_data.split(':', 1)
            if len(parts) == 2 and parts[1]:
                await handle_similar_from_search_callback(update, context, parts[1])
            else:
                await safe_answer_callback(query, "❌ Ошибка в данных", show_alert=True)

        elif callback_data.startswith('back_search:'):
            parts = callback_data.split(':', 1)
            if len(parts) == 2 and parts[1]:
                await handle_back_search_callback(update, context, parts[1])
            else:
                await safe_answer_callback(query, "❌ Ошибка в данных", show_alert=True)
        
        elif callback_data.startswith('search:'):
            artist_name = callback_data.split(':', 1)[1]
            await safe_answer_callback(query, f"🔍 Ищу: {artist_name}", show_alert=False)
            await process_search(update, context, artist_name)
        
        elif callback_data == "main_menu":
            await safe_answer_callback(query)
            await bot_instance.clear_user_state(user.id)
            await show_main_menu(query)
        
        elif callback_data == "charts_menu":
            await safe_answer_callback(query)
            await show_charts_menu(update, context)
        
        elif callback_data == "similar_menu":
            await safe_answer_callback(query)
            await show_similar_artists_menu(update, context)
        
        elif callback_data == "help_menu":
            await safe_answer_callback(query)
            await show_help_menu(update, context)
        
        elif callback_data == "new_search":
            await safe_answer_callback(query)
            await bot_instance.clear_user_state(user.id)
            await safe_edit_message_text(query, "🔍 <b>Введите название трека или исполнителя:</b>", parse_mode='HTML')
        
        elif callback_data == "close_search":
            # Close search list. In groups we delete the list message полностью (без служебного текста),
            # в личке — показываем подсказку.
            if user.id in bot_instance.user_sessions:
                del bot_instance.user_sessions[user.id]

            chat_type = getattr(query.message.chat, "type", None) if query.message else None
            if chat_type in ("group", "supergroup"):
                # Удаляем сообщение со списком
                try:
                    await safe_delete_message(context.bot, chat_id=query.message.chat_id, message_id=query.message.message_id)
                except Exception:
                    # Если удалить не удалось — просто убираем клавиатуру и текст
                    await safe_edit_message_text(query, " ", parse_mode=None)
                # Коротко отвечаем на callback без текста в чат
                await safe_answer_callback(query)
            else:
                await safe_answer_callback(query)
                await safe_edit_message_text(query, "✅ Список поиска закрыт. Используйте /search для нового поиска.", parse_mode='HTML')
        
        else:
            await safe_answer_callback(query, "❌ Неизвестная команда", show_alert=True)
                    
    except Exception as e:
        logger.error(f"Error in button handler: {e}", exc_info=True)
        await safe_answer_callback(query, "❌ Произошла ошибка при обработке запроса", show_alert=True)

async def handle_add_token(update: Update, context: CallbackContext):
    """Добавление нового токена"""
    query = update.callback_query
    await safe_answer_callback(query)
    
    await bot_instance.set_user_state(query.from_user.id, "add_token")
    await query.edit_message_text(
        text="➕ <b>Добавление VK токена</b>\n\nОтправьте токен в формате:\n\n<code>vk1.a.token_here</code>\n\nПосле токена можно добавить описание через пробел.",
        parse_mode='HTML'
    )

async def handle_token_input(update: Update, context: CallbackContext):
    """Обработка ввода токена"""
    user_id = update.effective_user.id
    text = update.message.text.strip()
    
    if user_id not in config.ADMIN_IDS:
        await update.message.reply_text("❌ У вас нет прав для добавления токенов.")
        return
    
    state_data = await bot_instance.get_user_state(user_id)
    if not state_data or state_data['state'] != 'add_token':
        return
    
    parts = text.split(maxsplit=1)
    token = parts[0]
    description = parts[1] if len(parts) > 1 else None
    
    if not token.startswith('vk1.a.'):
        await update.message.reply_text("❌ <b>Неверный формат токена</b>\nТокен должен начинаться с 'vk1.a.'", parse_mode='HTML')
        return
    
    success = await bot_instance.vk_token_manager.add_token(token, user_id, description)
    
    if success:
        await update.message.reply_text(
            f"✅ <b>Токен успешно добавлен</b>\n\nТокен: <code>{token[:20]}...</code>\nОписание: {description or 'Нет'}",
            parse_mode='HTML'
        )
    else:
        await update.message.reply_text("❌ <b>Ошибка при добавлении токена</b>\nВозможно, токен уже существует.", parse_mode='HTML')
    
    await bot_instance.clear_user_state(user_id)



# ==================== ADMIN DATASTORE & FEATURES ====================
# ==================== ADMIN DATASTORE & FEATURES ====================
class AdminDB:
    """Хранилище и сервисы админки (SQLite через aiosqlite).

    Важно:
      - Телефон Telegram бот НЕ может получить сам. Он появится только если пользователь
        сам отправит контакт (кнопка "Поделиться контактом") или перешлёт его.
      - Чаты/каналы, где бот админ, бот узнаёт только после того, как в этот чат/канал
        прилетит апдейт (сообщение/пост/коллбек и т.п.). Поэтому для "всех админ-чатов"
        нужен хотя бы один апдейт из каждого такого места.
    """

    _USER_USAGE_CACHE_MAX = 50_000
    _USER_USAGE_WRITE_INTERVAL = 60.0
    _TELEGRAM_SERVICE_USER_IDS = frozenset({777000})

    def __init__(self, db_path: str):
        self.db_path = db_path
        self._write_lock = asyncio.Lock()
        # Bounded write-throttle cache. SQLite remains the source of truth.
        self._user_usage_cache: "OrderedDict[Tuple[int, str], float]" = OrderedDict()

    @staticmethod
    def _quote_identifier(name: str) -> str:
        return '"' + str(name).replace('"', '""') + '"'

    async def _table_columns(self, conn: Any, table: str) -> Set[str]:
        """Return columns only after verifying the table through sqlite_master."""
        cur = await conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
            (str(table),),
        )
        exists = await cur.fetchone()
        await cur.close()
        if not exists:
            return set()
        quoted = self._quote_identifier(table)
        cur = await conn.execute(f"PRAGMA table_info({quoted})")
        rows = await cur.fetchall()
        await cur.close()
        return {str(row[1]) for row in rows if len(row) > 1 and row[1]}

    @staticmethod
    def _usage_timestamp_columns(columns: Set[str]) -> Tuple[Optional[str], Optional[str]]:
        first = next((name for name in ("first_seen", "created_at", "ts", "last_seen") if name in columns), None)
        last = next((name for name in ("last_seen", "ts", "created_at", "first_seen") if name in columns), None)
        return first, last

    async def _recover_unique_user_stats(self, conn: Any) -> int:
        """Best-effort, schema-aware and idempotent recovery from existing tables."""
        cur = await conn.execute(
            "SELECT name FROM sqlite_master "
            "WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
        )
        tables = [str(row[0]) for row in await cur.fetchall() if row and row[0]]
        await cur.close()
        table_columns: Dict[str, Set[str]] = {}
        for table in tables:
            table_columns[table] = await self._table_columns(conn, table)

        recovered: Dict[Tuple[int, str], Tuple[str, str]] = {}
        fallback_ts = _utc_now().isoformat()

        def merge_rows(rows: List[Any]) -> None:
            for row in rows:
                if not row or len(row) < 4:
                    continue
                try:
                    user_id = int(row[0] or 0)
                except (TypeError, ValueError):
                    continue
                context_name = str(row[1] or "").lower()
                context_name = "group" if context_name in ("group", "supergroup") else context_name
                if (
                    user_id <= 0
                    or user_id in self._TELEGRAM_SERVICE_USER_IDS
                    or context_name not in ("private", "group")
                ):
                    continue
                first_seen = str(row[2] or fallback_ts)
                last_seen = str(row[3] or first_seen)
                key = (user_id, context_name)
                previous = recovered.get(key)
                if previous is None:
                    recovered[key] = (first_seen, last_seen)
                else:
                    recovered[key] = (min(previous[0], first_seen), max(previous[1], last_seen))

        chats_columns = table_columns.get("chats", set())
        can_join_chats = {"chat_id", "chat_type"}.issubset(chats_columns)

        for table in tables:
            if table == "user_usage_context":
                continue
            columns = table_columns.get(table, set())
            if "user_id" not in columns:
                continue

            quoted_table = self._quote_identifier(table)
            first_col, last_col = self._usage_timestamp_columns(columns)
            first_expr = f"CAST(src.{self._quote_identifier(first_col)} AS TEXT)" if first_col else "?"
            last_expr = f"CAST(src.{self._quote_identifier(last_col)} AS TEXT)" if last_col else first_expr
            bot_filter = ""
            if "is_bot" in columns:
                bot_filter = f" AND COALESCE(CAST(src.{self._quote_identifier('is_bot')} AS INTEGER), 0)=0"

            direct_context_col = next((name for name in ("context", "chat_type") if name in columns), None)
            if direct_context_col:
                context_expr = (
                    f"CASE LOWER(CAST(src.{self._quote_identifier(direct_context_col)} AS TEXT)) "
                    "WHEN 'private' THEN 'private' "
                    "WHEN 'group' THEN 'group' "
                    "WHEN 'supergroup' THEN 'group' END"
                )
                sql = (
                    f"SELECT CAST(src.{self._quote_identifier('user_id')} AS INTEGER), {context_expr}, "
                    f"MIN(COALESCE({first_expr}, ?)), MAX(COALESCE({last_expr}, ?)) "
                    f"FROM {quoted_table} src "
                    f"WHERE CAST(src.{self._quote_identifier('user_id')} AS INTEGER)>0 "
                    f"AND {context_expr} IS NOT NULL{bot_filter} "
                    f"GROUP BY CAST(src.{self._quote_identifier('user_id')} AS INTEGER), {context_expr}"
                )
                # The final two parameters are COALESCE fallbacks. Placeholder
                # expressions (when no timestamp column exists) need their own values.
                placeholder_count = sql.count("?")
                cur = await conn.execute(sql, tuple([fallback_ts] * placeholder_count))
                merge_rows(await cur.fetchall())
                await cur.close()
                continue

            if "chat_id" in columns and can_join_chats and table != "chats":
                context_expr = (
                    "CASE LOWER(CAST(c.chat_type AS TEXT)) "
                    "WHEN 'private' THEN 'private' "
                    "WHEN 'group' THEN 'group' "
                    "WHEN 'supergroup' THEN 'group' END"
                )
                sql = (
                    f"SELECT CAST(src.{self._quote_identifier('user_id')} AS INTEGER), {context_expr}, "
                    f"MIN(COALESCE({first_expr}, ?)), MAX(COALESCE({last_expr}, ?)) "
                    f"FROM {quoted_table} src "
                    "JOIN chats c ON c.chat_id=src.chat_id "
                    f"WHERE CAST(src.{self._quote_identifier('user_id')} AS INTEGER)>0 "
                    f"AND {context_expr} IS NOT NULL{bot_filter} "
                    f"GROUP BY CAST(src.{self._quote_identifier('user_id')} AS INTEGER), {context_expr}"
                )
                placeholder_count = sql.count("?")
                cur = await conn.execute(sql, tuple([fallback_ts] * placeholder_count))
                merge_rows(await cur.fetchall())
                await cur.close()

        if recovered:
            await conn.executemany(
                "INSERT INTO user_usage_context(user_id,context,first_seen,last_seen) VALUES(?,?,?,?) "
                "ON CONFLICT(user_id,context) DO UPDATE SET "
                "first_seen=MIN(user_usage_context.first_seen, excluded.first_seen), "
                "last_seen=MAX(user_usage_context.last_seen, excluded.last_seen)",
                [
                    (user_id, context_name, timestamps[0], timestamps[1])
                    for (user_id, context_name), timestamps in recovered.items()
                ],
            )
        return len(recovered)

    async def _initialize_unique_user_stats(self, conn: Any) -> None:
        await conn.execute("""CREATE TABLE IF NOT EXISTS user_usage_context (
            user_id INTEGER NOT NULL,
            context TEXT NOT NULL CHECK(context IN ('private', 'group')),
            first_seen TEXT NOT NULL,
            last_seen TEXT NOT NULL,
            PRIMARY KEY (user_id, context)
        )""")
        await self._recover_unique_user_stats(conn)
        await conn.commit()

        cur = await conn.execute(
            "SELECT user_id, context FROM user_usage_context "
            "WHERE user_id>0 ORDER BY last_seen DESC LIMIT ?",
            (self._USER_USAGE_CACHE_MAX,),
        )
        rows = await cur.fetchall()
        await cur.close()
        self._user_usage_cache.clear()
        for user_id, context_name in reversed(rows):
            self._user_usage_cache[(int(user_id), str(context_name))] = 0.0

        private_count, group_count, total_count = await self.get_unique_user_counts(conn=conn)
        logger.info("Unique user stats migration completed")
        logger.info(
            "Unique user stats ready: private=%s groups=%s total=%s",
            private_count,
            group_count,
            total_count,
        )

    async def init(self):
        async with _sqlite_connection(self.db_path) as conn:
            try:
                if getattr(config, "DB_WAL_MODE", True):
                    await conn.execute("PRAGMA journal_mode=WAL")
                await conn.execute("PRAGMA synchronous=NORMAL")
            except Exception as exc:
                logger.warning("AdminDB SQLite pragma setup failed: %s", exc)
            await conn.execute("""CREATE TABLE IF NOT EXISTS admin_settings(
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )""")
            await conn.execute("""CREATE TABLE IF NOT EXISTS bans(
                user_id INTEGER PRIMARY KEY,
                reason TEXT,
                banned_by INTEGER,
                banned_at TEXT NOT NULL
            )""")
            await conn.execute("""CREATE TABLE IF NOT EXISTS chats(
                chat_id INTEGER PRIMARY KEY,
                chat_type TEXT,
                title TEXT,
                username TEXT,
                last_seen TEXT NOT NULL,
                bot_is_admin INTEGER NOT NULL DEFAULT 0,
                bot_status TEXT,
                can_post_messages INTEGER NOT NULL DEFAULT 0,
                can_send_messages INTEGER NOT NULL DEFAULT 0
            )""")
            await conn.execute("""CREATE TABLE IF NOT EXISTS users(
                user_id INTEGER PRIMARY KEY,
                first_name TEXT,
                last_name TEXT,
                username TEXT,
                phone TEXT,
                created_at TEXT NOT NULL,
                last_seen TEXT NOT NULL
            )""")
            await conn.execute("""CREATE TABLE IF NOT EXISTS stats_daily(
                day TEXT PRIMARY KEY,
                searches INTEGER NOT NULL DEFAULT 0,
                downloads INTEGER NOT NULL DEFAULT 0,
                ym_searches INTEGER NOT NULL DEFAULT 0,
                vk_searches INTEGER NOT NULL DEFAULT 0,
                yt_searches INTEGER NOT NULL DEFAULT 0,
                errors INTEGER NOT NULL DEFAULT 0
            )""")
            await conn.execute("""CREATE TABLE IF NOT EXISTS stats_events(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT NOT NULL,
                event TEXT NOT NULL,
                user_id INTEGER,
                chat_id INTEGER,
                source TEXT,
                extra TEXT
            )""")
            await conn.execute("""CREATE TABLE IF NOT EXISTS user_actions(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ts TEXT NOT NULL,
                user_id INTEGER,
                chat_id INTEGER,
                action TEXT NOT NULL,
                payload TEXT
            )""")
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_user_actions_ts ON user_actions(ts)")
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_user_actions_user ON user_actions(user_id)")
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_stats_events_ts ON stats_events(ts)")
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_stats_events_user ON stats_events(user_id)")
            await conn.execute("""CREATE TABLE IF NOT EXISTS forwarded_file_ids(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_id TEXT NOT NULL UNIQUE,
                file_unique_id TEXT,
                file_type TEXT,
                file_name TEXT,
                mime_type TEXT,
                file_size INTEGER,
                added_by INTEGER,
                added_at TEXT NOT NULL
            )""")
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_forwarded_file_ids_type ON forwarded_file_ids(file_type)")
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_forwarded_file_ids_added_at ON forwarded_file_ids(added_at)")
            await conn.commit()

            # мягкие миграции под старые БД
            async def _ensure_column(table: str, col: str, col_def: str):
                try:
                    cur = await conn.execute(f"PRAGMA table_info({table})")
                    cols = [r[1] for r in await cur.fetchall()]
                    await cur.close()
                    if col not in cols:
                        await conn.execute(f"ALTER TABLE {table} ADD COLUMN {col_def}")
                except Exception:
                    logger.exception("SQLite migration failed: table=%s column=%s", table, col)

            await _ensure_column("chats", "bot_is_admin", "bot_is_admin INTEGER NOT NULL DEFAULT 0")
            await _ensure_column("chats", "bot_status", "bot_status TEXT")
            await _ensure_column("chats", "can_post_messages", "can_post_messages INTEGER NOT NULL DEFAULT 0")
            await _ensure_column("chats", "can_send_messages", "can_send_messages INTEGER NOT NULL DEFAULT 0")
            await _ensure_column("users", "phone", "phone TEXT")
            await _ensure_column("stats_daily", "yt_searches", "yt_searches INTEGER NOT NULL DEFAULT 0")
            await conn.commit()

            try:
                await self._initialize_unique_user_stats(conn)
            except Exception:
                logger.exception("Unique user stats initialization failed")

    # ---------- unique user usage ----------
    async def record_user_usage(self, user: Any, context_name: str) -> bool:
        """Persist a human user's private/group usage without affecting bot handling."""
        try:
            if not user or bool(getattr(user, "is_bot", False)):
                return False
            user_id = int(getattr(user, "id", 0) or 0)
            context_name = str(context_name or "").lower()
            if (
                user_id <= 0
                or user_id in self._TELEGRAM_SERVICE_USER_IDS
                or context_name not in ("private", "group")
            ):
                return False

            key = (user_id, context_name)
            now_mono = time.monotonic()
            cached_at = self._user_usage_cache.get(key)
            if cached_at is not None and now_mono - cached_at < self._USER_USAGE_WRITE_INTERVAL:
                self._user_usage_cache.move_to_end(key)
                return True

            now = _utc_now().isoformat()
            async with self._write_lock:
                # Re-check after awaiting the lock so concurrent updates for the
                # same pair do not produce duplicate writes.
                cached_at = self._user_usage_cache.get(key)
                if cached_at is not None and now_mono - cached_at < self._USER_USAGE_WRITE_INTERVAL:
                    self._user_usage_cache.move_to_end(key)
                    return True
                async with _sqlite_connection(self.db_path) as conn:
                    await conn.execute(
                        "INSERT INTO user_usage_context(user_id,context,first_seen,last_seen) VALUES(?,?,?,?) "
                        "ON CONFLICT(user_id,context) DO UPDATE SET last_seen=excluded.last_seen",
                        (user_id, context_name, now, now),
                    )
                    await conn.commit()
                self._user_usage_cache[key] = now_mono
                self._user_usage_cache.move_to_end(key)
                while len(self._user_usage_cache) > self._USER_USAGE_CACHE_MAX:
                    self._user_usage_cache.popitem(last=False)
            return True
        except Exception:
            logger.exception("Failed to record unique user usage")
            return False

    async def get_unique_user_counts(self, conn: Any = None) -> Tuple[int, int, int]:
        """Return all-time private, group and de-duplicated total user counts."""
        owns_connection = conn is None
        db = conn
        try:
            if db is None:
                db = await _sqlite_connect(self.db_path)
            cur = await db.execute("""
                SELECT
                    COUNT(DISTINCT CASE WHEN context = 'private' THEN user_id END),
                    COUNT(DISTINCT CASE WHEN context = 'group' THEN user_id END),
                    COUNT(DISTINCT user_id)
                FROM user_usage_context
                WHERE user_id > 0
            """)
            row = await cur.fetchone()
            await cur.close()
            return (
                int(row[0] or 0) if row else 0,
                int(row[1] or 0) if row else 0,
                int(row[2] or 0) if row else 0,
            )
        except Exception:
            logger.exception("Failed to read unique user stats")
            return 0, 0, 0
        finally:
            if owns_connection and db is not None:
                await db.close()

    # ---------- settings ----------
    async def set_setting(self, key: str, value: Any):
        v = json.dumps(value, ensure_ascii=False)
        now = _utc_now().isoformat()
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute(
                "INSERT INTO admin_settings(key,value,updated_at) VALUES(?,?,?) "
                "ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at",
                (key, v, now)
            )
            await conn.commit()

    async def get_setting(self, key: str, default: Any = None) -> Any:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute("SELECT value FROM admin_settings WHERE key=?", (key,))
            row = await cur.fetchone()
            await cur.close()
        if not row:
            return default
        try:
            return json.loads(row[0])
        except Exception:
            return default

    async def apply_runtime_settings(self):
        """Подхватывает настройки из БД и применяет к config на лету."""
        try:
            enable_ym = await self.get_setting("enable_ym", getattr(config, "ENABLE_YANDEX_MUSIC", True))
            enable_vk = await self.get_setting("enable_vk", getattr(config, "ENABLE_VK_MUSIC", True))
            enable_yt = await self.get_setting("enable_yt", getattr(config, "ENABLE_YOUTUBE_MUSIC", True))
            priority = await self.get_setting(
                "source_priority",
                getattr(config, "SOURCE_PRIORITY", "vk_first"),
            )  # fastest|ym|vk|yt|yandex_first|vk_first|youtube_first
            setattr(config, "ENABLE_YANDEX_MUSIC", bool(enable_ym))
            setattr(config, "ENABLE_VK_MUSIC", bool(enable_vk))
            setattr(config, "ENABLE_YOUTUBE_MUSIC", bool(enable_yt))
            setattr(config, "SOURCE_PRIORITY", priority)
        except Exception as e:
            logger.warning(f"AdminDB apply settings failed: {e}")

    # ---------- chat registry ----------
    async def track_chat(self, update: Update, bot: Any = None):
        """Регистрирует чат в БД + (опционально) фиксирует статус бота в чате."""
        try:
            chat = update.effective_chat
            if not chat:
                return
            title = getattr(chat, "title", None) or ""
            username = getattr(chat, "username", None) or ""
            now = _utc_now().isoformat()

            bot_is_admin = 0
            bot_status = None
            can_post_messages = 0
            can_send_messages = 0
            chat_type = str(chat.type)
            membership_checked = chat_type not in ("group", "supergroup", "channel")

            try:
                if bot is not None and chat_type in ("group", "supergroup", "channel"):
                    if bot_instance is not None:
                        member = await bot_instance._get_bot_chat_member_cached(bot, chat.id)
                    else:
                        me = await bot.get_me()
                        member = await bot.get_chat_member(chat.id, me.id)
                    membership_checked = True
                    bot_status = getattr(member, "status", None)
                    bot_is_admin = 1 if bot_status in ("administrator", "creator") else 0
                    if getattr(member, "can_post_messages", None) is True:
                        can_post_messages = 1
                    if getattr(member, "can_send_messages", None) is True:
                        can_send_messages = 1
            except Exception as exc:
                logger.debug("Unable to determine bot membership in chat %s: %s", chat.id, exc)

            async with _sqlite_connection(self.db_path) as conn:
                if membership_checked:
                    await conn.execute(
                        "INSERT INTO chats(chat_id,chat_type,title,username,last_seen,bot_is_admin,bot_status,can_post_messages,can_send_messages) "
                        "VALUES(?,?,?,?,?,?,?,?,?) "
                        "ON CONFLICT(chat_id) DO UPDATE SET "
                        "chat_type=excluded.chat_type, title=excluded.title, username=excluded.username, last_seen=excluded.last_seen, "
                        "bot_is_admin=excluded.bot_is_admin, bot_status=excluded.bot_status, "
                        "can_post_messages=excluded.can_post_messages, can_send_messages=excluded.can_send_messages",
                        (int(chat.id), chat_type, title, username, now, int(bot_is_admin), bot_status, int(can_post_messages), int(can_send_messages)),
                    )
                else:
                    # A transient Telegram error must not erase the last known
                    # administrator status for an already registered chat.
                    await conn.execute(
                        "INSERT INTO chats(chat_id,chat_type,title,username,last_seen,bot_is_admin,bot_status,can_post_messages,can_send_messages) "
                        "VALUES(?,?,?,?,?,0,NULL,0,0) "
                        "ON CONFLICT(chat_id) DO UPDATE SET "
                        "chat_type=excluded.chat_type, title=excluded.title, username=excluded.username, last_seen=excluded.last_seen",
                        (int(chat.id), chat_type, title, username, now),
                    )
                await conn.commit()
        except Exception:
            logger.exception("Failed to track chat metadata")

    async def track_chat_membership(self, update: Update) -> None:
        """Persist authoritative bot membership changes from MY_CHAT_MEMBER updates."""
        try:
            membership = getattr(update, "my_chat_member", None)
            chat = getattr(membership, "chat", None) or getattr(update, "effective_chat", None)
            new_member = getattr(membership, "new_chat_member", None)
            if not membership or not chat or not new_member:
                return

            chat_type = str(getattr(chat, "type", "") or "")
            if chat_type not in ("group", "supergroup", "channel"):
                return

            status = str(getattr(new_member, "status", "") or "")
            title = str(getattr(chat, "title", "") or "")
            username = str(getattr(chat, "username", "") or "")
            now = _utc_now().isoformat()
            bot_is_admin = 1 if status in ("administrator", "creator") else 0
            can_post_messages = 1 if getattr(new_member, "can_post_messages", None) is True else 0

            if status in ("member", "administrator", "creator"):
                can_send_messages = 1
            elif status == "restricted":
                can_send_messages = 1 if getattr(new_member, "can_send_messages", None) is True else 0
            else:
                can_send_messages = 0

            async with _sqlite_connection(self.db_path) as conn:
                await conn.execute(
                    "INSERT INTO chats(chat_id,chat_type,title,username,last_seen,bot_is_admin,bot_status,can_post_messages,can_send_messages) "
                    "VALUES(?,?,?,?,?,?,?,?,?) "
                    "ON CONFLICT(chat_id) DO UPDATE SET "
                    "chat_type=excluded.chat_type,title=excluded.title,username=excluded.username,last_seen=excluded.last_seen," 
                    "bot_is_admin=excluded.bot_is_admin,bot_status=excluded.bot_status," 
                    "can_post_messages=excluded.can_post_messages,can_send_messages=excluded.can_send_messages",
                    (
                        int(chat.id), chat_type, title, username, now, bot_is_admin, status,
                        can_post_messages, can_send_messages,
                    ),
                )
                await conn.commit()
        except Exception:
            logger.exception("Failed to track bot chat membership")

    async def count_group_chats(self, active_only: bool = True) -> int:
        where = "WHERE chat_type IN ('group','supergroup')"
        if active_only:
            where += " AND COALESCE(bot_status,'') NOT IN ('left','kicked')"
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(f"SELECT COUNT(1) FROM chats {where}")
            row = await cur.fetchone()
            await cur.close()
        return int(row[0] if row and row[0] is not None else 0)

    async def count_admin_group_chats(self, active_only: bool = True) -> int:
        where = "WHERE chat_type IN ('group','supergroup') AND bot_is_admin=1"
        if active_only:
            where += " AND COALESCE(bot_status,'') NOT IN ('left','kicked')"
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(f"SELECT COUNT(1) FROM chats {where}")
            row = await cur.fetchone()
            await cur.close()
        return int(row[0] if row and row[0] is not None else 0)

    async def get_group_chat_ids(self, active_only: bool = True) -> List[int]:
        where = "WHERE chat_type IN ('group','supergroup')"
        if active_only:
            where += " AND COALESCE(bot_status,'') NOT IN ('left','kicked')"
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(f"SELECT chat_id FROM chats {where} ORDER BY last_seen DESC")
            rows = await cur.fetchall()
            await cur.close()
        return [int(row[0]) for row in rows]

    async def list_group_chats(
        self,
        active_only: bool = True,
        limit: int = 10,
        offset: int = 0,
    ) -> List[Dict[str, Any]]:
        where = "WHERE chat_type IN ('group','supergroup')"
        if active_only:
            where += " AND COALESCE(bot_status,'') NOT IN ('left','kicked')"
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT chat_id,chat_type,title,username,bot_is_admin,bot_status,can_send_messages,last_seen "
                f"FROM chats {where} ORDER BY last_seen DESC LIMIT ? OFFSET ?",
                (max(1, int(limit)), max(0, int(offset))),
            )
            rows = await cur.fetchall()
            await cur.close()
        return [
            {
                "chat_id": int(row[0]),
                "chat_type": row[1] or "",
                "title": row[2] or "",
                "username": row[3] or "",
                "bot_is_admin": int(row[4] or 0),
                "bot_status": row[5] or "",
                "can_send_messages": int(row[6] or 0),
                "last_seen": row[7] or "",
            }
            for row in rows
        ]

    async def get_group_chat(self, chat_id: int) -> Optional[Dict[str, Any]]:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT chat_id,chat_type,title,username,bot_is_admin,bot_status,can_send_messages,last_seen "
                "FROM chats WHERE chat_id=? AND chat_type IN ('group','supergroup')",
                (int(chat_id),),
            )
            row = await cur.fetchone()
            await cur.close()
        if not row:
            return None
        return {
            "chat_id": int(row[0]),
            "chat_type": row[1] or "",
            "title": row[2] or "",
            "username": row[3] or "",
            "bot_is_admin": int(row[4] or 0),
            "bot_status": row[5] or "",
            "can_send_messages": int(row[6] or 0),
            "last_seen": row[7] or "",
        }

    async def get_chat_ids(self, target: str = "all") -> List[int]:
        """target: all|private|groups"""
        where = ""
        if target == "private":
            where = "WHERE chat_type='private'"
        elif target == "groups":
            where = "WHERE chat_type IN ('group','supergroup')"
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(f"SELECT chat_id FROM chats {where}")
            rows = await cur.fetchall()
            await cur.close()
        return [int(r[0]) for r in rows]

    async def get_admin_chat_ids(self) -> List[int]:
        """Чаты/каналы, где бот зафиксирован админом."""
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT chat_id FROM chats WHERE chat_type IN ('group','supergroup','channel') AND bot_is_admin=1"
            )
            rows = await cur.fetchall()
            await cur.close()
        return [int(r[0]) for r in rows]

    async def export_admin_chats_rows(self) -> List[Dict[str, Any]]:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT chat_id, chat_type, title, username, bot_is_admin, bot_status, can_post_messages, can_send_messages, last_seen "
                "FROM chats WHERE chat_type IN ('group','supergroup','channel') ORDER BY last_seen DESC"
            )
            rows = await cur.fetchall()
            await cur.close()
        out = []
        for r in rows:
            out.append({
                "chat_id": r[0],
                "chat_type": r[1] or "",
                "title": r[2] or "",
                "username": r[3] or "",
                "bot_is_admin": int(r[4] or 0),
                "bot_status": r[5] or "",
                "can_post_messages": int(r[6] or 0),
                "can_send_messages": int(r[7] or 0),
                "last_seen": r[8] or "",
            })
        return out

    # ---------- user registry ----------
    async def track_user(self, user: Any):
        """Регистрирует пользователя (для статистики/экспорта/рассылки)."""
        try:
            if not user:
                return
            now = _utc_now().isoformat()
            async with _sqlite_connection(self.db_path) as conn:
                await conn.execute(
                    "INSERT INTO users(user_id,first_name,last_name,username,phone,created_at,last_seen) "
                    "VALUES(?,?,?,?,?,?,?) "
                    "ON CONFLICT(user_id) DO UPDATE SET "
                    "first_name=excluded.first_name, last_name=excluded.last_name, username=excluded.username, "
                    "last_seen=excluded.last_seen",
                    (
                        int(user.id),
                        getattr(user, "first_name", None),
                        getattr(user, "last_name", None),
                        getattr(user, "username", None),
                        None,
                        now,
                        now,
                    )
                )
                await conn.commit()
        except Exception:
            logger.exception("Failed to track user metadata")

    async def set_user_phone(self, user_id: int, phone: str):
        try:
            if not user_id or not phone:
                return
            phone = str(phone).strip()
            if not phone:
                return
            async with _sqlite_connection(self.db_path) as conn:
                await conn.execute("UPDATE users SET phone=? WHERE user_id=?", (phone, int(user_id)))
                await conn.commit()
        except Exception:
            logger.exception("Failed to save user phone")

    async def get_user_ids(self) -> List[int]:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute("SELECT user_id FROM users ORDER BY last_seen DESC")
            rows = await cur.fetchall()
            await cur.close()
        return [int(r[0]) for r in rows]

    async def get_user(self, user_id: int) -> Optional[Dict[str, Any]]:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT user_id,first_name,last_name,username,last_seen FROM users WHERE user_id=?",
                (int(user_id),),
            )
            row = await cur.fetchone()
            await cur.close()
        if not row:
            return None
        return {
            "user_id": int(row[0]),
            "first_name": row[1] or "",
            "last_name": row[2] or "",
            "username": row[3] or "",
            "last_seen": row[4] or "",
        }

    async def count_users(self) -> int:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute("SELECT COUNT(1) FROM users")
            row = await cur.fetchone()
            await cur.close()
        return int(row[0] if row and row[0] is not None else 0)

    async def export_users_rows(self) -> List[Dict[str, Any]]:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT user_id, first_name, last_name, username, phone, created_at, last_seen "
                "FROM users ORDER BY last_seen DESC"
            )
            rows = await cur.fetchall()
            await cur.close()
        out = []
        for r in rows:
            out.append({
                "user_id": r[0],
                "first_name": r[1] or "",
                "last_name": r[2] or "",
                "username": r[3] or "",
                "phone": r[4] or "",
                "created_at": r[5] or "",
                "last_seen": r[6] or "",
            })
        return out

    # ---------- bans ----------
    async def is_banned(self, user_id: int) -> bool:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute("SELECT 1 FROM bans WHERE user_id=?", (int(user_id),))
            row = await cur.fetchone()
            await cur.close()
        return bool(row)

    async def ban_user(self, user_id: int, banned_by: int, reason: str = "") -> bool:
        now = _utc_now().isoformat()
        async with self._write_lock:
            async with _sqlite_connection(self.db_path) as conn:
                await conn.execute(
                    "INSERT INTO bans(user_id,reason,banned_by,banned_at) VALUES(?,?,?,?) "
                    "ON CONFLICT(user_id) DO UPDATE SET reason=excluded.reason, banned_by=excluded.banned_by, banned_at=excluded.banned_at",
                    (int(user_id), reason, int(banned_by), now)
                )
                await conn.commit()
        return True

    async def unban_user(self, user_id: int) -> bool:
        async with self._write_lock:
            async with _sqlite_connection(self.db_path) as conn:
                cur = await conn.execute("DELETE FROM bans WHERE user_id=?", (int(user_id),))
                changed = int(cur.rowcount or 0) > 0
                await cur.close()
                await conn.commit()
        return changed

    async def get_banned_user_ids(self) -> Set[int]:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute("SELECT user_id FROM bans")
            rows = await cur.fetchall()
            await cur.close()
        return {int(row[0]) for row in rows}

    async def list_bans(self, limit: int = 50) -> List[Dict[str, Any]]:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT user_id, reason, banned_by, banned_at FROM bans ORDER BY banned_at DESC LIMIT ?",
                (int(limit),)
            )
            rows = await cur.fetchall()
            await cur.close()
        out = []
        for user_id, reason, banned_by, banned_at in rows:
            out.append({"user_id": user_id, "reason": reason or "", "banned_by": banned_by, "banned_at": banned_at})
        return out

    # ---------- forwarded file_id pool ----------
    async def add_forwarded_file_id(
        self,
        file_id: str,
        file_unique_id: str = None,
        file_type: str = None,
        file_name: str = None,
        mime_type: str = None,
        file_size: int = None,
        added_by: int = None,
    ) -> bool:
        if not file_id:
            return False
        now = _utc_now().isoformat()
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute(
                """
                INSERT OR IGNORE INTO forwarded_file_ids(
                    file_id, file_unique_id, file_type, file_name, mime_type, file_size, added_by, added_at
                ) VALUES(?,?,?,?,?,?,?,?)
                """,
                (file_id, file_unique_id, file_type, file_name, mime_type, file_size, added_by, now)
            )
            await conn.commit()
        return True

    async def count_forwarded_file_ids(self) -> int:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute("SELECT COUNT(1) FROM forwarded_file_ids")
            row = await cur.fetchone()
            await cur.close()
        return int(row[0] if row and row[0] is not None else 0)

    async def count_cached_audio_file_ids(self) -> int:
        """Количество file_id в кэше tg_audio_cache (прогрев/скачивания)."""
        async with _sqlite_connection(self.db_path) as conn:
            try:
                cur = await conn.execute("SELECT COUNT(1) FROM tg_audio_cache")
                row = await cur.fetchone()
                await cur.close()
                return int(row[0] if row and row[0] is not None else 0)
            except Exception:
                return 0

    # ---------- stats ----------
    async def _bump_daily(self, day: str, field: str, inc: int = 1):
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute(
                f"INSERT INTO stats_daily(day,{field}) VALUES(?,?) "
                f"ON CONFLICT(day) DO UPDATE SET {field}={field}+excluded.{field}",
                (day, int(inc))
            )
            await conn.commit()

    async def log_event(self, event: str, user_id: int = None, chat_id: int = None, source: str = None, extra: Any = None):
        ts = _utc_now().isoformat()
        try:
            extra_s = json.dumps(extra, ensure_ascii=False) if extra is not None else None
        except Exception:
            extra_s = None
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute(
                "INSERT INTO stats_events(ts,event,user_id,chat_id,source,extra) VALUES(?,?,?,?,?,?)",
                (ts, event, user_id, chat_id, source, extra_s)
            )
            await conn.commit()

    async def log_search(
        self,
        user_id: int,
        chat_id: int,
        query: str,
        ym_count: int,
        vk_count: int,
        yt_count: int,
        total: int,
    ):
        day = _utc_now().date().isoformat()
        ts = _utc_now().isoformat()
        extra_s = json.dumps(
            {"q": query[:200], "ym": ym_count, "vk": vk_count, "yt": yt_count, "total": total},
            ensure_ascii=False,
        )
        async with self._write_lock:
            async with _sqlite_connection(self.db_path) as conn:
                await conn.execute(
                    "INSERT INTO stats_daily(day,searches,ym_searches,vk_searches,yt_searches) VALUES(?,?,?,?,?) "
                    "ON CONFLICT(day) DO UPDATE SET "
                    "searches=searches+excluded.searches, "
                    "ym_searches=ym_searches+excluded.ym_searches, "
                    "vk_searches=vk_searches+excluded.vk_searches, "
                    "yt_searches=yt_searches+excluded.yt_searches",
                    (day, 1, 1 if ym_count else 0, 1 if vk_count else 0, 1 if yt_count else 0),
                )
                await conn.execute(
                    "INSERT INTO stats_events(ts,event,user_id,chat_id,source,extra) VALUES(?,?,?,?,?,?)",
                    (ts, "search", user_id, chat_id, None, extra_s),
                )
                await conn.commit()

    async def log_download(self, user_id: int, chat_id: int, source: str):
        day = _utc_now().date().isoformat()
        ts = _utc_now().isoformat()
        async with self._write_lock:
            async with _sqlite_connection(self.db_path) as conn:
                await conn.execute(
                    "INSERT INTO stats_daily(day,downloads) VALUES(?,1) "
                    "ON CONFLICT(day) DO UPDATE SET downloads=downloads+1",
                    (day,),
                )
                await conn.execute(
                    "INSERT INTO stats_events(ts,event,user_id,chat_id,source,extra) VALUES(?,?,?,?,?,?)",
                    (ts, "download", user_id, chat_id, source, None),
                )
                await conn.commit()

    async def log_error(self, user_id: int = None, chat_id: int = None, where: str = "", err: str = ""):
        day = _utc_now().date().isoformat()
        ts = _utc_now().isoformat()
        extra_s = json.dumps({"where": where[:120], "err": (err or "")[:300]}, ensure_ascii=False)
        async with self._write_lock:
            async with _sqlite_connection(self.db_path) as conn:
                await conn.execute(
                    "INSERT INTO stats_daily(day,errors) VALUES(?,1) "
                    "ON CONFLICT(day) DO UPDATE SET errors=errors+1",
                    (day,),
                )
                await conn.execute(
                    "INSERT INTO stats_events(ts,event,user_id,chat_id,source,extra) VALUES(?,?,?,?,?,?)",
                    (ts, "error", user_id, chat_id, None, extra_s),
                )
                await conn.commit()

    async def log_user_action(self, user_id: int, chat_id: int, action: str, payload: Any = None):
        """Лог конкретных действий пользователей (для админки/Excel)."""
        try:
            ts = _utc_now().isoformat()
            payload_s = None
            if payload is not None:
                try:
                    payload_s = json.dumps(payload, ensure_ascii=False)
                except Exception:
                    payload_s = str(payload)
            async with _sqlite_connection(self.db_path) as conn:
                await conn.execute(
                    "INSERT INTO user_actions(ts,user_id,chat_id,action,payload) VALUES(?,?,?,?,?)",
                    (ts, int(user_id or 0), int(chat_id or 0), str(action), payload_s)
                )
                await conn.commit()
        except Exception:
            logger.exception("Failed to log user action %s for user %s", action, user_id)

    async def export_user_actions_rows(self, days: int = 7, limit: int = 200000) -> List[Dict[str, Any]]:
        """Экспорт действий за последние N дней (для Excel)."""
        try:
            since = (_utc_now() - timedelta(days=max(1, int(days or 1)))).isoformat()
        except Exception:
            since = (_utc_now() - timedelta(days=7)).isoformat()

        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                """
                SELECT a.ts, a.user_id, u.username, u.first_name, u.last_name, a.chat_id, a.action, a.payload
                FROM user_actions a
                LEFT JOIN users u ON u.user_id = a.user_id
                WHERE a.ts >= ?
                ORDER BY a.ts DESC
                LIMIT ?
                """,
                (since, int(limit))
            )
            rows = await cur.fetchall()
            await cur.close()

        out: List[Dict[str, Any]] = []
        for r in rows:
            out.append({
                "ts": r[0] or "",
                "user_id": r[1] or 0,
                "username": r[2] or "",
                "first_name": r[3] or "",
                "last_name": r[4] or "",
                "chat_id": r[5] or 0,
                "action": r[6] or "",
                "payload": r[7] or "",
            })
        return out


    async def get_stats(self, days: int = 7) -> List[Dict[str, Any]]:
        start_day = (_utc_now().date() - timedelta(days=max(days, 1)-1)).isoformat()
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT day, searches, downloads, ym_searches, vk_searches, yt_searches, errors FROM stats_daily WHERE day>=? ORDER BY day ASC",
                (start_day,)
            )
            rows = await cur.fetchall()
            await cur.close()
        out = []
        for day, searches, downloads, ym_s, vk_s, yt_s, errors in rows:
            out.append({
                "day": day, "searches": searches, "downloads": downloads,
                "ym": ym_s, "vk": vk_s, "yt": yt_s, "errors": errors,
            })
        return out

    async def get_broadcast_stats(self, days: int = 7) -> Dict[str, int]:
        since = (_utc_now() - timedelta(days=max(1, int(days or 1)))).isoformat()
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT extra FROM stats_events WHERE event='broadcast' AND ts>=? ORDER BY ts DESC",
                (since,),
            )
            rows = await cur.fetchall()
            await cur.close()
        result = {"runs": len(rows), "recipients": 0, "ok": 0, "fail": 0}
        for row in rows:
            try:
                payload = json.loads(row[0] or "{}")
            except Exception:
                continue
            for key in ("recipients", "ok", "fail"):
                try:
                    result[key] += max(0, int(payload.get(key) or 0))
                except (TypeError, ValueError):
                    continue
        return result


class GroupDigestStore:
    """SQLite settings for group-admin music digests and observed forum topics."""

    def __init__(self, db_path: str):
        self.db_path = db_path
        self._audio_index: Dict[str, Dict[str, Any]] = {}
        self._audio_index_ts = 0.0
        self._audio_index_lock = asyncio.Lock()

    async def init(self) -> None:
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS group_topics(
                    chat_id INTEGER NOT NULL,
                    thread_id INTEGER NOT NULL,
                    title TEXT,
                    last_seen REAL NOT NULL,
                    PRIMARY KEY(chat_id, thread_id)
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS group_digest_subscriptions(
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    chat_id INTEGER NOT NULL,
                    thread_id INTEGER NOT NULL DEFAULT 0,
                    genre_id INTEGER NOT NULL,
                    genre_name TEXT NOT NULL,
                    schedule_key TEXT NOT NULL,
                    track_count INTEGER NOT NULL DEFAULT 5,
                    enabled INTEGER NOT NULL DEFAULT 1,
                    created_by INTEGER NOT NULL,
                    next_run_at REAL NOT NULL,
                    last_run_at REAL,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS group_digest_sent_tracks(
                    subscription_id INTEGER NOT NULL,
                    track_key TEXT NOT NULL,
                    sent_at REAL NOT NULL,
                    PRIMARY KEY(subscription_id, track_key)
                )
            """)
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_digest_due ON group_digest_subscriptions(enabled, next_run_at)")
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_digest_chat ON group_digest_subscriptions(chat_id)")
            await conn.execute("CREATE INDEX IF NOT EXISTS idx_digest_sent_at ON group_digest_sent_tracks(sent_at)")
            await conn.commit()

    async def remember_topic(self, chat_id: int, thread_id: Optional[int], title: Optional[str] = None) -> None:
        if not thread_id:
            return
        supplied_title = bool((title or "").strip())
        topic_title = ((title or f"Раздел #{int(thread_id)}").strip())[:128]
        async with _sqlite_connection(self.db_path) as conn:
            if supplied_title:
                await conn.execute(
                    "INSERT INTO group_topics(chat_id,thread_id,title,last_seen) VALUES(?,?,?,?) "
                    "ON CONFLICT(chat_id,thread_id) DO UPDATE SET title=excluded.title,last_seen=excluded.last_seen",
                    (int(chat_id), int(thread_id), topic_title, time.time()),
                )
            else:
                await conn.execute(
                    "INSERT INTO group_topics(chat_id,thread_id,title,last_seen) VALUES(?,?,?,?) "
                    "ON CONFLICT(chat_id,thread_id) DO UPDATE SET last_seen=excluded.last_seen",
                    (int(chat_id), int(thread_id), topic_title, time.time()),
                )
            await conn.commit()

    async def list_topics(self, chat_id: int, limit: int = 20) -> List[Dict[str, Any]]:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT thread_id,title,last_seen FROM group_topics WHERE chat_id=? ORDER BY last_seen DESC LIMIT ?",
                (int(chat_id), int(limit)),
            )
            rows = await cur.fetchall()
            await cur.close()
        return [{"thread_id": int(r[0]), "title": r[1] or f"Раздел #{r[0]}", "last_seen": float(r[2] or 0)} for r in rows]

    async def create_subscription(
        self,
        chat_id: int,
        thread_id: int,
        genre_id: int,
        genre_name: str,
        schedule_key: str,
        track_count: int,
        created_by: int,
    ) -> int:
        now = _utc_now().isoformat()
        next_run = _digest_next_run(schedule_key)
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "INSERT INTO group_digest_subscriptions("
                "chat_id,thread_id,genre_id,genre_name,schedule_key,track_count,enabled,created_by,next_run_at,created_at,updated_at"
                ") VALUES(?,?,?,?,?,?,1,?,?,?,?)",
                (
                    int(chat_id), int(thread_id or 0), int(genre_id), str(genre_name), str(schedule_key),
                    max(1, min(int(track_count), 10)), int(created_by), float(next_run), now, now,
                ),
            )
            await conn.commit()
            return int(cur.lastrowid)

    @staticmethod
    def _row_to_subscription(row: Any) -> Dict[str, Any]:
        return {
            "id": int(row[0]), "chat_id": int(row[1]), "thread_id": int(row[2] or 0),
            "genre_id": int(row[3]), "genre_name": row[4] or "Все жанры",
            "schedule_key": row[5] or "d18", "track_count": int(row[6] or 5),
            "enabled": bool(row[7]), "created_by": int(row[8] or 0),
            "next_run_at": float(row[9] or 0), "last_run_at": float(row[10] or 0),
        }

    async def list_subscriptions(self, chat_id: int) -> List[Dict[str, Any]]:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT id,chat_id,thread_id,genre_id,genre_name,schedule_key,track_count,enabled,created_by,next_run_at,last_run_at "
                "FROM group_digest_subscriptions WHERE chat_id=? ORDER BY id DESC",
                (int(chat_id),),
            )
            rows = await cur.fetchall()
            await cur.close()
        return [self._row_to_subscription(r) for r in rows]

    async def get_subscription(self, subscription_id: int, chat_id: Optional[int] = None) -> Optional[Dict[str, Any]]:
        sql = (
            "SELECT id,chat_id,thread_id,genre_id,genre_name,schedule_key,track_count,enabled,created_by,next_run_at,last_run_at "
            "FROM group_digest_subscriptions WHERE id=?"
        )
        args: List[Any] = [int(subscription_id)]
        if chat_id is not None:
            sql += " AND chat_id=?"
            args.append(int(chat_id))
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(sql, tuple(args))
            row = await cur.fetchone()
            await cur.close()
        return self._row_to_subscription(row) if row else None

    async def set_enabled(self, subscription_id: int, chat_id: int, enabled: bool) -> bool:
        sub = await self.get_subscription(subscription_id, chat_id)
        if not sub:
            return False
        next_run = _digest_next_run(sub["schedule_key"]) if enabled else 0
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "UPDATE group_digest_subscriptions SET enabled=?,next_run_at=?,updated_at=? WHERE id=? AND chat_id=?",
                (1 if enabled else 0, float(next_run), _utc_now().isoformat(), int(subscription_id), int(chat_id)),
            )
            await conn.commit()
            return int(cur.rowcount or 0) > 0

    async def update_schedule(self, subscription_id: int, chat_id: int, schedule_key: str) -> bool:
        """Change weekdays/time and recalculate the next publication."""
        next_run = _digest_next_run(schedule_key)
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "UPDATE group_digest_subscriptions SET schedule_key=?,next_run_at=?,updated_at=? WHERE id=? AND chat_id=?",
                (
                    str(schedule_key), float(next_run), _utc_now().isoformat(),
                    int(subscription_id), int(chat_id),
                ),
            )
            await conn.commit()
            return int(cur.rowcount or 0) > 0

    async def toggle_enabled(self, subscription_id: int, chat_id: int) -> Optional[bool]:
        sub = await self.get_subscription(subscription_id, chat_id)
        if not sub:
            return None
        enabled = not bool(sub["enabled"])
        next_run = _digest_next_run(sub["schedule_key"]) if enabled else 0
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute(
                "UPDATE group_digest_subscriptions SET enabled=?,next_run_at=?,updated_at=? WHERE id=? AND chat_id=?",
                (1 if enabled else 0, float(next_run), _utc_now().isoformat(), int(subscription_id), int(chat_id)),
            )
            await conn.commit()
        return enabled

    async def delete_subscription(self, subscription_id: int, chat_id: int) -> bool:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "DELETE FROM group_digest_subscriptions WHERE id=? AND chat_id=?",
                (int(subscription_id), int(chat_id)),
            )
            await conn.execute("DELETE FROM group_digest_sent_tracks WHERE subscription_id=?", (int(subscription_id),))
            await conn.commit()
            return int(cur.rowcount or 0) > 0

    async def get_due_subscriptions(self, now_ts: float, limit: int = 20) -> List[Dict[str, Any]]:
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT id,chat_id,thread_id,genre_id,genre_name,schedule_key,track_count,enabled,created_by,next_run_at,last_run_at "
                "FROM group_digest_subscriptions WHERE enabled=1 AND next_run_at<=? ORDER BY next_run_at LIMIT ?",
                (float(now_ts), int(limit)),
            )
            rows = await cur.fetchall()
            await cur.close()
        return [self._row_to_subscription(r) for r in rows]

    async def mark_run(self, subscription_id: int, schedule_key: str, last_run_at: Optional[float] = None) -> None:
        run_ts = float(last_run_at or time.time())
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute(
                "UPDATE group_digest_subscriptions SET last_run_at=?,next_run_at=?,updated_at=? WHERE id=?",
                (run_ts, _digest_next_run(schedule_key, run_ts + 1), _utc_now().isoformat(), int(subscription_id)),
            )
            await conn.commit()

    async def get_recent_sent_keys(self, subscription_id: int, days: Optional[int] = None) -> Set[str]:
        """Return all identities blocked by the no-repeat window."""
        keep_days = _digest_no_repeat_days() if days is None else max(1, int(days))
        since = time.time() - keep_days * 86400
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT track_key FROM group_digest_sent_tracks WHERE subscription_id=? AND sent_at>=?",
                (int(subscription_id), float(since)),
            )
            rows = await cur.fetchall()
            await cur.close()
        return {str(r[0]) for r in rows if r and r[0]}

    async def record_sent_many(self, subscription_id: int, track_keys: List[str]) -> None:
        """Remember every available identity of a sent track in one transaction."""
        keys = list(dict.fromkeys(str(key) for key in (track_keys or []) if key))
        if not keys:
            return
        sent_at = time.time()
        async with _sqlite_connection(self.db_path) as conn:
            await conn.executemany(
                "INSERT INTO group_digest_sent_tracks(subscription_id,track_key,sent_at) VALUES(?,?,?) "
                "ON CONFLICT(subscription_id,track_key) DO UPDATE SET sent_at=excluded.sent_at",
                [(int(subscription_id), key, sent_at) for key in keys],
            )
            await conn.commit()

    async def record_sent(self, subscription_id: int, track_key: str) -> None:
        """Backward-compatible single-key history writer."""
        await self.record_sent_many(subscription_id, [track_key])

    async def remove_cached_audio(self, query_key: str) -> None:
        if not query_key:
            return
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute("DELETE FROM tg_audio_cache WHERE query_key=?", (str(query_key),))
            await conn.commit()
        self._audio_index_ts = 0.0

    async def cached_audio_index(self, force: bool = False) -> Dict[str, Dict[str, Any]]:
        ttl = int(getattr(config, "DIGEST_AUDIO_INDEX_TTL", 300) or 300)
        if not force and self._audio_index and time.time() - self._audio_index_ts < ttl:
            return self._audio_index
        async with self._audio_index_lock:
            if not force and self._audio_index and time.time() - self._audio_index_ts < ttl:
                return self._audio_index
            async with _sqlite_connection(self.db_path) as conn:
                try:
                    cur = await conn.execute(
                        "SELECT query_key,file_id,artist,title,duration,created_at FROM tg_audio_cache "
                        "WHERE file_id IS NOT NULL AND artist IS NOT NULL AND title IS NOT NULL "
                        "ORDER BY created_at DESC LIMIT ?",
                        (int(getattr(config, "DIGEST_AUDIO_INDEX_LIMIT", 100000) or 100000),),
                    )
                    rows = await cur.fetchall()
                    await cur.close()
                except Exception:
                    rows = []
            index: Dict[str, Dict[str, Any]] = {}
            for query_key, file_id, artist, title, duration, created_at in rows:
                item = {
                    "query_key": query_key, "file_id": file_id, "artist": artist or "",
                    "title": title or "", "duration": int(duration or 0), "created_at": int(created_at or 0),
                }
                for key in _digest_track_keys(item["artist"], item["title"]):
                    index.setdefault(key, item)
            self._audio_index = index
            self._audio_index_ts = time.time()
            return index

async def _admin_db_path() -> str:
    # используем ту же БД, что и file_id cache (по умолчанию bot_stats.db)
    try:
        return await bot_instance._tg_db_path()
    except Exception:
        return getattr(config, "STATS_DB_PATH", "bot_stats.db")


# ==================== USER DATA (history/preferences) ====================
class UserStore:
    """SQLite-backed store for user preferences and history.

    Uses the same DB file as stats/file_id cache (STATS_DB_PATH by default).
    """

    def __init__(self, db_path: str):
        self.db_path = db_path
        self._init_lock = asyncio.Lock()
        self._initialized = False

    async def init(self) -> None:
        async with self._init_lock:
            if self._initialized:
                return
            await self._ensure_tables()
            self._initialized = True

    async def _ensure_tables(self) -> None:
        async with _sqlite_connection(self.db_path) as conn:
            try:
                if getattr(config, "DB_WAL_MODE", True):
                    await conn.execute("PRAGMA journal_mode=WAL")
                await conn.execute("PRAGMA synchronous=NORMAL")
            except Exception as exc:
                logger.warning("UserStore SQLite pragma setup failed: %s", exc)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS user_preferences (
                    user_id INTEGER PRIMARY KEY,
                    prefer_source TEXT,
                    artist_only_default INTEGER,
                    prefer_bitrate_kbps INTEGER,
                    updated_at REAL
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS user_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    uid TEXT,
                    track_json TEXT,
                    ts REAL
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS user_last (
                    user_id INTEGER PRIMARY KEY,
                    uid TEXT,
                    track_json TEXT,
                    ts REAL
                )
            """)
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS user_favorites (
                    user_id INTEGER NOT NULL,
                    uid TEXT NOT NULL,
                    track_json TEXT NOT NULL,
                    ts REAL NOT NULL,
                    PRIMARY KEY(user_id, uid)
                )
            """)
            # Fallback for Telegram clients that open the bot from a t.me
            # deep-link but deliver only plain /start without the payload.
            # The pending item is short-lived and contains no secret data.
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS user_pending_favorite (
                    user_id INTEGER PRIMARY KEY,
                    uid TEXT NOT NULL,
                    track_json TEXT NOT NULL,
                    ts REAL NOT NULL
                )
            """)
            await conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_user_favorites_user_ts "
                "ON user_favorites(user_id, ts DESC)"
            )
            await conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_user_history_user_ts "
                "ON user_history(user_id, ts DESC)"
            )
            await conn.execute(
                "CREATE INDEX IF NOT EXISTS idx_user_history_user_uid_ts "
                "ON user_history(user_id, uid, ts DESC)"
            )
            await conn.commit()

    @staticmethod
    def _dump_track(track: Any) -> str:
        try:
            if isinstance(track, TrackInfo):
                payload = track.to_dict()
            elif isinstance(track, dict):
                payload = dict(track)
            else:
                payload = dict(getattr(track, '__dict__', {}))
            payload['uid'] = _track_uid_from_any(payload) or payload.get('uid')
            return json.dumps(payload, ensure_ascii=False)
        except Exception:
            return json.dumps({'uid': _track_uid_from_any(track)}, ensure_ascii=False)

    @staticmethod
    def _load_track(track_json: str) -> Dict[str, Any]:
        try:
            d = json.loads(track_json or "{}")
            if isinstance(d, dict):
                return d
        except Exception:
            pass
        return {}

    async def get_preferences(self, user_id: int) -> Dict[str, Any]:
        await self.init()
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT prefer_source, artist_only_default, prefer_bitrate_kbps FROM user_preferences WHERE user_id=?",
                (int(user_id),)
            )
            row = await cur.fetchone()
            await cur.close()
        if not row:
            return {
                'prefer_source': None,
                'artist_only_default': None,
                'prefer_bitrate_kbps': None,
            }
        return {
            'prefer_source': row[0],
            'artist_only_default': row[1],
            'prefer_bitrate_kbps': row[2],
        }

    async def set_preferences(self, user_id: int, **fields) -> None:
        await self.init()
        prefer_source = fields.get('prefer_source')
        artist_only_default = fields.get('artist_only_default')
        prefer_bitrate_kbps = fields.get('prefer_bitrate_kbps')
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute(
                """
                INSERT INTO user_preferences(user_id, prefer_source, artist_only_default, prefer_bitrate_kbps, updated_at)
                VALUES(?,?,?,?,?)
                ON CONFLICT(user_id) DO UPDATE SET
                    prefer_source=COALESCE(excluded.prefer_source, user_preferences.prefer_source),
                    artist_only_default=COALESCE(excluded.artist_only_default, user_preferences.artist_only_default),
                    prefer_bitrate_kbps=COALESCE(excluded.prefer_bitrate_kbps, user_preferences.prefer_bitrate_kbps),
                    updated_at=excluded.updated_at
                """,
                (int(user_id), prefer_source, artist_only_default, prefer_bitrate_kbps, time.time())
            )
            await conn.commit()

    async def add_history(self, user_id: int, uid: str, track: Any) -> None:
        """Append a successful delivery to history without altering old records."""
        await self.init()
        uid = str(uid or _track_uid_from_any(track) or "")
        if not uid:
            return
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute(
                "INSERT INTO user_history(user_id, uid, track_json, ts) VALUES(?,?,?,?)",
                (int(user_id), uid, self._dump_track(track), time.time()),
            )
            await conn.commit()

    async def list_history(self, user_id: int, limit: int = 30) -> List[Dict[str, Any]]:
        await self.init()
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT track_json FROM user_history WHERE user_id=? ORDER BY ts DESC LIMIT ?",
                (int(user_id), int(limit))
            )
            rows = await cur.fetchall()
            await cur.close()
        return [self._load_track(r[0]) for r in (rows or [])]

    async def get_history_track(self, user_id: int, uid: str) -> Optional[Dict[str, Any]]:
        await self.init()
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT track_json FROM user_history WHERE user_id=? AND uid=? ORDER BY ts DESC LIMIT 1",
                (int(user_id), str(uid))
            )
            row = await cur.fetchone()
            await cur.close()
        return self._load_track(row[0]) if row else None

    async def add_favorite(self, user_id: int, uid: str, track: Any) -> None:
        await self.init()
        uid = str(uid or _track_uid_from_any(track) or "")
        if not uid:
            return
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute(
                """
                INSERT INTO user_favorites(user_id, uid, track_json, ts) VALUES(?,?,?,?)
                ON CONFLICT(user_id, uid) DO UPDATE SET track_json=excluded.track_json, ts=excluded.ts
                """,
                (int(user_id), uid, self._dump_track(track), time.time()),
            )
            await conn.commit()

    async def set_pending_favorite(self, user_id: int, uid: str, track: Any) -> None:
        """Remember a recent download for the t.me favorite-link fallback."""
        await self.init()
        uid = str(uid or _track_uid_from_any(track) or "")
        if not uid:
            return
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute(
                """
                INSERT INTO user_pending_favorite(user_id, uid, track_json, ts)
                VALUES(?,?,?,?)
                ON CONFLICT(user_id) DO UPDATE SET
                    uid=excluded.uid, track_json=excluded.track_json, ts=excluded.ts
                """,
                (int(user_id), uid, self._dump_track(track), time.time()),
            )
            await conn.commit()

    async def get_pending_favorite(self, user_id: int, max_age_seconds: int = 300) -> Optional[Dict[str, Any]]:
        """Consume a recent pending favorite created by a successful download."""
        await self.init()
        now = time.time()
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT uid, track_json, ts FROM user_pending_favorite WHERE user_id=? LIMIT 1",
                (int(user_id),),
            )
            row = await cur.fetchone()
            await cur.close()
            await conn.execute("DELETE FROM user_pending_favorite WHERE user_id=?", (int(user_id),))
            await conn.commit()
        if not row or (now - float(row[2] or 0)) > max(1, int(max_age_seconds)):
            return None
        track = self._load_track(row[1])
        if track:
            track['uid'] = str(row[0])
        return track or None

    async def clear_pending_favorite(self, user_id: int) -> None:
        await self.init()
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute("DELETE FROM user_pending_favorite WHERE user_id=?", (int(user_id),))
            await conn.commit()

    async def remove_favorite(self, user_id: int, uid: str) -> None:
        await self.init()
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute("DELETE FROM user_favorites WHERE user_id=? AND uid=?", (int(user_id), str(uid)))
            await conn.commit()

    async def is_favorite(self, user_id: int, uid: str) -> bool:
        await self.init()
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT 1 FROM user_favorites WHERE user_id=? AND uid=? LIMIT 1",
                (int(user_id), str(uid)),
            )
            row = await cur.fetchone()
            await cur.close()
        return bool(row)

    async def list_favorites(self, user_id: int, limit: int = 30) -> List[Dict[str, Any]]:
        await self.init()
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT track_json FROM user_favorites WHERE user_id=? ORDER BY ts DESC LIMIT ?",
                (int(user_id), int(limit)),
            )
            rows = await cur.fetchall()
            await cur.close()
        return [self._load_track(r[0]) for r in (rows or [])]

    async def get_favorite(self, user_id: int, uid: str) -> Optional[Dict[str, Any]]:
        await self.init()
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT track_json FROM user_favorites WHERE user_id=? AND uid=? LIMIT 1",
                (int(user_id), str(uid)),
            )
            row = await cur.fetchone()
            await cur.close()
        return self._load_track(row[0]) if row else None

    async def set_last(self, user_id: int, uid: str, track: Any) -> None:
        await self.init()
        uid = str(uid or "")
        if not uid:
            return
        async with _sqlite_connection(self.db_path) as conn:
            await conn.execute(
                "INSERT OR REPLACE INTO user_last(user_id, uid, track_json, ts) VALUES(?,?,?,?)",
                (int(user_id), uid, self._dump_track(track), time.time())
            )
            await conn.commit()

    async def get_last(self, user_id: int) -> Optional[Dict[str, Any]]:
        await self.init()
        async with _sqlite_connection(self.db_path) as conn:
            cur = await conn.execute(
                "SELECT track_json FROM user_last WHERE user_id=?",
                (int(user_id),)
            )
            row = await cur.fetchone()
            await cur.close()
        if not row:
            return None
        return self._load_track(row[0])


# ==================== АДМИН-ПАНЕЛЬ ====================
def _is_admin(user_id: int) -> bool:
    """Проверка админ-доступа.
    Поддерживает:
      - config.ADMIN_IDS = [123, 456]
      - config.OWNER_ID = 123 (запасной вариант)
    """
    try:
        uid = int(user_id)
    except Exception:
        return False
    try:
        admin_ids = getattr(config, "ADMIN_IDS", None) or []
        if uid in set(int(x) for x in admin_ids):
            return True
    except Exception:
        pass
    try:
        owner_id = getattr(config, "OWNER_ID", None)
        if owner_id is not None and uid == int(owner_id):
            return True
    except Exception:
        pass
    return False

def _admin_keyboard():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔑 VK токены", callback_data="tokens_menu")],
        [InlineKeyboardButton("📊 Статистика", callback_data="admin_stats")],
        [InlineKeyboardButton("👥 Группы с ботом", callback_data="admin_groups:0")],
        [InlineKeyboardButton("🩺 Состояние", callback_data="admin_monitor")],
        [InlineKeyboardButton("🧹 Кэш / Очистка", callback_data="admin_cache")],
        [InlineKeyboardButton("⚙️ Источники / Настройки", callback_data="admin_settings")],
        [InlineKeyboardButton("🚫 Модерация", callback_data="admin_moderation")],
        [InlineKeyboardButton("📣 Рассылка", callback_data="admin_broadcast")],
        [InlineKeyboardButton("📎 Приём файлов → file_id", callback_data="admin_fileid")],
        [InlineKeyboardButton("❌ Закрыть", callback_data="admin_close")]
    ])


async def admin_command(update: Update, context: CallbackContext) -> None:
    """Команда /admin — админка (видна только админам)"""
    user_id = update.effective_user.id if update.effective_user else 0
    if not _is_admin(user_id):
        return  # молча игнорируем, чтобы "админка" не светилась
    await update.message.reply_text(
        "🛠️ <b>VLMB Admin</b>\n\nВыберите раздел:",
        parse_mode='HTML',
        reply_markup=_admin_keyboard()
    )

async def _render_tokens_menu() -> tuple[str, InlineKeyboardMarkup]:
    tokens = await bot_instance.vk_token_manager.list_tokens()
    if not tokens:
        text = "🔑 <b>VK токены</b>\n\n❌ Токены не найдены. Добавьте новый токен."
    else:
        text = "🔑 <b>VK токены</b>\n\n"
        for i, t in enumerate(tokens, 1):
            masked = t.get('token_masked', '***')
            status = "🟢 Активен" if t.get('is_active') and not t.get('is_expired') else "🔴 Неактивен"
            if t.get('is_expired'):
                status = "⚠️ Истёк"
            text += f"{i}. <code>{masked}</code>\n"
            text += f"   {status} | Добавил: {t.get('added_by')}\n"
            if t.get('description'):
                text += f"   Описание: {t.get('description')}\n"
            text += "\n"

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Добавить токен", callback_data="add_token")],
        [InlineKeyboardButton("➖ Удалить токен", callback_data="remove_token_menu")],
        [InlineKeyboardButton("🔄 Обновить", callback_data="refresh_tokens")],
        [InlineKeyboardButton("🔙 Назад", callback_data="admin_menu")]
    ])
    return text, kb

async def show_tokens_menu(update: Update, context: CallbackContext) -> None:
    """Показать меню токенов (callback)"""
    query = update.callback_query
    await safe_answer_callback(query)
    if not _is_admin(query.from_user.id):
        return

    text, kb = await _render_tokens_menu()
    await safe_edit_message_text(query, text, parse_mode='HTML', reply_markup=kb, disable_web_page_preview=True)

async def show_remove_token_menu(update: Update, context: CallbackContext) -> None:
    """Показать список токенов для удаления"""
    query = update.callback_query
    await safe_answer_callback(query)
    if not _is_admin(query.from_user.id):
        return

    tokens = await bot_instance.vk_token_manager.list_tokens()
    full_tokens = [t.get('token_full') for t in tokens if t.get('token_full')]

    # сохраним полный список токенов в state, чтобы по индексу удалять без передачи токена в callback_data
    await bot_instance.set_user_state(query.from_user.id, "remove_token", {"tokens_full": full_tokens})

    if not tokens:
        await safe_edit_message_text(query, "🔑 <b>Удаление токенов</b>\n\n❌ Токены не найдены.", parse_mode='HTML',
                                     reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Назад", callback_data="tokens_menu")]]))
        return

    rows = []
    row = []
    for i, t in enumerate(tokens, 1):
        row.append(InlineKeyboardButton(f"🗑️ {i}", callback_data=f"del_token:{i-1}"))
        if len(row) == 5:
            rows.append(row); row = []
    if row: rows.append(row)

    text = "🔑 <b>Удаление VK токенов</b>\n\nВыберите номер токена для удаления:\n"
    for i, t in enumerate(tokens, 1):
        text += f"{i}. <code>{t.get('token_masked','***')}</code>\n"

    rows.append([InlineKeyboardButton("🔙 Назад", callback_data="tokens_menu")])
    await safe_edit_message_text(query, text, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(rows), disable_web_page_preview=True)

async def handle_delete_token_callback(update: Update, context: CallbackContext, idx: int) -> None:
    query = update.callback_query
    await safe_answer_callback(query)
    if not _is_admin(query.from_user.id):
        return

    state = await bot_instance.get_user_state(query.from_user.id)
    tokens_full = (state or {}).get('data', {}).get('tokens_full', [])
    if not tokens_full or idx < 0 or idx >= len(tokens_full):
        await safe_answer_callback(query, "❌ Список токенов устарел. Обновите.", show_alert=True)
        return

    token_full = tokens_full[idx]
    ok = await bot_instance.vk_token_manager.remove_token(token_full)
    if ok:
        await safe_answer_callback(query, "✅ Токен удалён", show_alert=True)
    else:
        await safe_answer_callback(query, "❌ Не удалось удалить токен", show_alert=True)

    # после удаления вернёмся в меню токенов
    await bot_instance.clear_user_state(query.from_user.id)
    text, kb = await _render_tokens_menu()
    await safe_edit_message_text(query, text, parse_mode='HTML', reply_markup=kb, disable_web_page_preview=True)

async def admin_tokens(update: Update, context: CallbackContext):
    """Команда /tokens — быстрый доступ к токенам (только админы)"""
    user_id = update.effective_user.id if update.effective_user else 0
    if not _is_admin(user_id):
        return

    text, kb = await _render_tokens_menu()
    await update.message.reply_text(text=text, parse_mode='HTML', reply_markup=kb, disable_web_page_preview=True)




# ==================== АДМИН-ПАНЕЛЬ: РАЗДЕЛЫ ====================

def _kb_back_to_admin():
    return InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Назад", callback_data="admin_menu")]])

def _fmt_onoff(v: Any) -> str:
    return "✅ ВКЛ" if bool(v) else "❌ ВЫКЛ"

def _safe_int(s: str) -> Optional[int]:
    try:
        return int(str(s).strip())
    except Exception:
        return None

async def _show_admin_stats(query, days: int = 7):
    await safe_answer_callback(query)
    try:
        rows: List[Dict[str, Any]] = []
        if hasattr(bot_instance, "admin_db") and bot_instance.admin_db:
            rows = await bot_instance.admin_db.get_stats(days=days)

        total_searches = sum(r.get("searches", 0) for r in rows)
        total_downloads = sum(r.get("downloads", 0) for r in rows)
        total_ym = sum(r.get("ym", 0) for r in rows)
        total_vk = sum(r.get("vk", 0) for r in rows)
        total_yt = sum(r.get("yt", 0) for r in rows)
        total_errors = sum(r.get("errors", 0) for r in rows)
        broadcast_stats = {"runs": 0, "recipients": 0, "ok": 0, "fail": 0}

        unique_private = 0
        unique_groups = 0
        unique_total = 0
        active_group_chats = 0
        admin_group_chats = 0
        try:
            if hasattr(bot_instance, "admin_db") and bot_instance.admin_db:
                unique_private, unique_groups, unique_total = await bot_instance.admin_db.get_unique_user_counts()
                active_group_chats = await bot_instance.admin_db.count_group_chats(active_only=True)
                admin_group_chats = await bot_instance.admin_db.count_admin_group_chats(active_only=True)
                broadcast_stats = await bot_instance.admin_db.get_broadcast_stats(days=days)
        except Exception:
            logger.exception("Failed to include unique user stats in admin report")

        file_id_cached = 0
        file_id_manual = 0
        file_id_total = 0
        try:
            if hasattr(bot_instance, "admin_db") and bot_instance.admin_db:
                file_id_cached = await bot_instance.admin_db.count_cached_audio_file_ids()
                file_id_manual = await bot_instance.admin_db.count_forwarded_file_ids()
                file_id_total = file_id_cached + file_id_manual
        except Exception:
            pass

        lines = []
        for r in rows[-min(len(rows), 10):]:
            lines.append(
                f"• <code>{r.get('day','')}</code> — 🔎 {r.get('searches',0)} | ⬇️ {r.get('downloads',0)} | "
                f"🎵 {r.get('ym',0)} | 🎶 {r.get('vk',0)} | ▶️ {r.get('yt',0)} | ❗ {r.get('errors',0)}"
            )

        text_msg = (
            f"📊 <b>Статистика (последние {days} дн.)</b>\n\n"
            f"🔎 Поисков: <b>{total_searches}</b>\n"
            f"⬇️ Скачиваний: <b>{total_downloads}</b>\n"
            f"💾 file_id в базе: <b>{file_id_total}</b> (кэш: {file_id_cached} + ручные: {file_id_manual})\n"
            f"🎵 YM поисков: <b>{total_ym}</b>\n"
            f"🎶 VK поисков: <b>{total_vk}</b>\n"
            f"▶️ YouTube поисков: <b>{total_yt}</b>\n"
            f"❗ Ошибок: <b>{total_errors}</b>\n\n"
            f"📣 <b>Рассылки за период</b>\n"
            f"• Запусков: <b>{broadcast_stats.get('runs', 0)}</b>\n"
            f"• Получателей: <b>{broadcast_stats.get('recipients', 0)}</b>\n"
            f"• Доставлено: <b>{broadcast_stats.get('ok', 0)}</b>\n"
            f"• Ошибок доставки: <b>{broadcast_stats.get('fail', 0)}</b>\n\n"
            f"👥 <b>Уникальные пользователи за всё время</b>\n"
            f"• В личке: <b>{unique_private}</b>\n"
            f"• В группах: <b>{unique_groups}</b>\n"
            f"• Итого: <b>{unique_total}</b>\n\n"
            f"💬 <b>Группы с ботом</b>\n"
            f"• Активных групп: <b>{active_group_chats}</b>\n"
            f"• Где бот администратор: <b>{admin_group_chats}</b>\n\n"
            f"<b>Дни:</b>\n" + ("\n".join(lines) if lines else "— нет данных —")
        )

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("📊 7 дн.", callback_data="admin_stats:7"),
             InlineKeyboardButton("📊 30 дн.", callback_data="admin_stats:30")],
            [InlineKeyboardButton("📤 Выгрузить пользователей", callback_data="admin_export:users")],
            [InlineKeyboardButton("📤 Выгрузить чаты/каналы", callback_data="admin_export:admin_chats")],
            [InlineKeyboardButton("👥 Показать группы", callback_data="admin_groups:0")],
            [InlineKeyboardButton("📤 Действия (7 дн) → Excel", callback_data="admin_export:actions_xlsx:7"),
             InlineKeyboardButton("📤 Действия (30 дн) → Excel", callback_data="admin_export:actions_xlsx:30")],
            [InlineKeyboardButton("📥 Excel → прогрев file_id", callback_data="admin_settings:excel_warmup")],
            [InlineKeyboardButton("🔙 Назад", callback_data="admin_menu")]
        ])
        await safe_edit_message_text(query, text_msg, parse_mode="HTML", reply_markup=kb, disable_web_page_preview=True)
    except Exception as e:
        logger.error(f"Admin stats error: {e}", exc_info=True)
        await safe_answer_callback(query, "❌ Не удалось показать статистику", show_alert=True)



async def _handle_admin_export(update: Update, context: CallbackContext, kind: str):
    """Экспорт пользователей/чатов в CSV и отправка администратору."""
    query = update.callback_query
    await safe_answer_callback(query)
    if not _is_admin(query.from_user.id):
        return
    if not hasattr(bot_instance, "admin_db") or not bot_instance.admin_db:
        await safe_answer_callback(query, "❌ AdminDB не инициализирован", show_alert=True)
        return

    try:
        if kind == "users":
            rows = await bot_instance.admin_db.export_users_rows()
            filename = f"vlmb_users_{_utc_now().date().isoformat()}.csv"
        elif kind == "admin_chats":
            rows = await bot_instance.admin_db.export_admin_chats_rows()
            filename = f"vlmb_chats_{_utc_now().date().isoformat()}.csv"
        elif kind.startswith("actions_xlsx"):
            # kind format: actions_xlsx:7 or actions_xlsx:30
            days = 7
            try:
                parts = kind.split(":")
                if len(parts) >= 2:
                    days = int(parts[1])
            except Exception:
                days = 7

            rows = await bot_instance.admin_db.export_user_actions_rows(days=days)
            filename = f"vlmb_actions_{days}d_{_utc_now().date().isoformat()}.xlsx"
        else:
            await safe_answer_callback(query, "❌ Неизвестный тип выгрузки", show_alert=True)
            return


        if filename.lower().endswith(".xlsx"):
            # XLSX (openpyxl)
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "actions"

            headers = list(rows[0].keys()) if rows else ["ts", "user_id", "username", "first_name", "last_name", "chat_id", "action", "payload"]
            ws.append(headers)

            for r in rows:
                ws.append([r.get(h, "") for h in headers])

            # лёгкая авто-ширина (без тяжёлых подсчётов)
            try:
                for i, h in enumerate(headers, 1):
                    ws.column_dimensions[get_column_letter(i)].width = min(60, max(10, len(str(h)) + 2))
            except Exception:
                pass

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            bio.name = filename
        else:
            # CSV (UTF-8 with BOM for Excel)
            output = io.StringIO()
            if rows:
                import csv as _csv
                writer = _csv.DictWriter(output, fieldnames=list(rows[0].keys()), delimiter=';')
                writer.writeheader()
                writer.writerows(rows)
            csv_text = output.getvalue()
            data = ("﻿" + csv_text).encode("utf-8", errors="ignore")

            bio = io.BytesIO(data)
            bio.name = filename

        await safe_send_document(context.bot, 
            chat_id=query.message.chat.id,
            document=bio,
            caption=(f"📤 Выгрузка: <b>{'Пользователи' if kind=='users' else ('Чаты/каналы' if kind=='admin_chats' else 'Действия пользователей')}</b>\n"
                     f"Записей: <b>{len(rows)}</b>"),
            parse_mode="HTML"
        )
    except Exception as e:
        logger.error(f"Admin export error: {e}", exc_info=True)
        await safe_answer_callback(query, "❌ Не удалось выгрузить", show_alert=True)
async def _show_admin_monitor(query):
    await safe_answer_callback(query)
    try:
        cache_stats: Dict[str, Any] = {}
        try:
            cache_stats = bot_instance.cache.get_cache_stats() if hasattr(bot_instance, "cache") else {}
        except Exception:
            cache_stats = {}

        active_sessions = 0
        try:
            active_sessions = len(getattr(bot_instance.session_manager, "sessions", {}))
        except Exception:
            pass

        bg_tasks = 0
        try:
            bg_tasks = len(getattr(bot_instance, "background_tasks", []))
        except Exception:
            pass

        mem = _process_memory_percent()

        redis_ok = "⚠️ отключен"
        try:
            if hasattr(bot_instance, "cache") and bot_instance.cache and bot_instance.cache.redis_client:
                await bot_instance.cache.redis_client.ping()
                redis_ok = "✅ ok"
        except Exception:
            redis_ok = "❌ ошибка"

        ym_ok = "✅ ok" if getattr(getattr(bot_instance, "yandex_music", None), "_initialized", False) else "⚠️ не инициализирован"
        vk_ok = "✅ ok" if getattr(config, "ENABLE_VK_MUSIC", True) else "❌ выключен"
        if not getattr(config, "ENABLE_YOUTUBE_MUSIC", True):
            yt_ok = "❌ выключен"
        elif getattr(getattr(bot_instance, "youtube_music", None), "_initialized", False):
            yt_ok = "✅ ok"
        else:
            yt_ok = "⚠️ yt-dlp недоступен"

        text_msg = (
            "🩺 <b>Состояние бота</b>\n\n"
            f"👥 Активных сессий: <b>{active_sessions}</b>\n"
            f"🧵 Фоновых задач: <b>{bg_tasks}</b>\n"
        )
        if mem is not None:
            text_msg += f"🧠 Память: <b>{mem:.1f}%</b>\n"
        else:
            text_msg += "🧠 Память: <b>—</b>\n"

        if cache_stats:
            text_msg += (
                "\n<b>Кэш:</b>\n"
                f"• hits/misses: <b>{cache_stats.get('hits',0)}</b>/<b>{cache_stats.get('misses',0)}</b>\n"
                f"• redis hits/misses: <b>{cache_stats.get('redis_hits',0)}</b>/<b>{cache_stats.get('redis_misses',0)}</b>\n"
                f"• hit-rate: <b>{cache_stats.get('hit_rate',0):.1%}</b>, redis hit-rate: <b>{cache_stats.get('redis_hit_rate',0):.1%}</b>\n"
            )

        try:
            m = await bot_instance.metrics.snapshot()
            queue_state = bot_instance.download_queue.snapshot()
            text_msg += (
                "\n<b>Метрики:</b>\n"
                f"• requests/search: <b>{m['counters'].get('search.total', 0)}</b>\n"
                f"• downloads: <b>{m['counters'].get('download.total', 0)}</b>\n"
                f"• download errors: <b>{m['counters'].get('download.errors', 0)}</b>\n"
                f"• queue: <b>{queue_state.get('queued', 0)}</b> / workers: <b>{queue_state.get('workers', 0)}</b>\n"
            )
            search_lat = m.get('latency', {}).get('search', {})
            if search_lat:
                text_msg += f"• search P95: <b>{search_lat.get('p95', 0):.2f}s</b>\n"
        except Exception:
            pass

        text_msg += (
            "\n<b>Сервисы:</b>\n"
            f"• Redis: <b>{redis_ok}</b>\n"
            f"• Yandex Music: <b>{ym_ok}</b>\n"
            f"• VK: <b>{vk_ok}</b>\n"
            f"• YouTube: <b>{yt_ok}</b>\n"
        )

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Обновить", callback_data="admin_monitor")],
            [InlineKeyboardButton("🔙 Назад", callback_data="admin_menu")]
        ])
        await safe_edit_message_text(query, text_msg, parse_mode="HTML", reply_markup=kb, disable_web_page_preview=True)
    except Exception as e:
        logger.error(f"Admin monitor error: {e}", exc_info=True)
        await safe_answer_callback(query, "❌ Не удалось показать состояние", show_alert=True)

async def _show_admin_cache_menu(query):
    await safe_answer_callback(query)
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🧹 Очистить локальный кэш", callback_data="admin_cache:local")],
        [InlineKeyboardButton("🧹 Очистить Redis", callback_data="admin_cache:redis")],
        [InlineKeyboardButton("🧹 Очистить сессии", callback_data="admin_cache:sessions")],
        [InlineKeyboardButton("🧹 Очистить temp файлы", callback_data="admin_cache:temp")],
        [InlineKeyboardButton("🔙 Назад", callback_data="admin_menu")]
    ])
    await safe_edit_message_text(query, "🧹 <b>Кэш / Очистка</b>\n\nВыберите действие:", parse_mode="HTML", reply_markup=kb)

async def _handle_admin_cache_action(query, action: str):
    await safe_answer_callback(query)
    try:
        if action == "local":
            try:
                if hasattr(bot_instance, "cache") and bot_instance.cache:
                    bot_instance.cache.local_cache.clear()
            except Exception:
                pass
            await safe_answer_callback(query, "✅ Локальный кэш очищен", show_alert=True)

        elif action == "redis":
            if hasattr(bot_instance, "cache") and bot_instance.cache and bot_instance.cache.redis_client:
                try:
                    await bot_instance.cache.redis_client.flushdb()
                    await safe_answer_callback(query, "✅ Redis очищен", show_alert=True)
                except Exception as e:
                    await safe_answer_callback(query, f"❌ Redis: {e}", show_alert=True)
            else:
                await safe_answer_callback(query, "⚠️ Redis отключен", show_alert=True)

        elif action == "sessions":
            try:
                if hasattr(bot_instance, "session_manager") and bot_instance.session_manager:
                    bot_instance.session_manager.sessions.clear()
                if hasattr(bot_instance, "user_sessions"):
                    bot_instance.user_sessions.clear()
            except Exception:
                pass
            await safe_answer_callback(query, "✅ Сессии очищены", show_alert=True)

        elif action == "temp":
            try:
                if hasattr(bot_instance, "file_manager") and bot_instance.file_manager:
                    await bot_instance.file_manager.cleanup_old_files()
            except Exception:
                pass
            await safe_answer_callback(query, "✅ Запрошена очистка temp", show_alert=True)
    finally:
        await _show_admin_cache_menu(query)

async def _show_admin_settings(update: Update, context: CallbackContext):
    """Админка: источники/настройки"""
    query = update.callback_query
    await safe_answer_callback(query, "")

    try:
        # Подтягиваем настройки из БД (если есть) в runtime config
        try:
            if hasattr(bot_instance, "admin_db") and bot_instance.admin_db:
                await bot_instance.admin_db.apply_runtime_settings()
        except Exception:
            pass

        enable_ym = getattr(config, "ENABLE_YANDEX_MUSIC", True)
        enable_vk = getattr(config, "ENABLE_VK_MUSIC", True)
        enable_yt = getattr(config, "ENABLE_YOUTUBE_MUSIC", True)
        priority = getattr(config, "SOURCE_PRIORITY", "fastest")

        text_msg = (
            "⚙️ <b>Источники / Настройки</b>\n\n"
            f"🎵 Yandex Music: <b>{_fmt_onoff(enable_ym)}</b>\n"
            f"🎶 VK: <b>{_fmt_onoff(enable_vk)}</b>\n"
            f"▶️ YouTube: <b>{_fmt_onoff(enable_yt)}</b>\n"
            f"⚡ Приоритет: <b>{html.escape(str(priority))}</b>\n\n"
            "📥 <b>Excel → прогрев file_id</b>: загрузите .xlsx (лист <code>Artists</code>, колонка <code>Artist</code>) "
            "— бот будет по очереди искать артистов и скачивать найденные треки, чтобы максимально набрать Telegram <code>file_id</code>."
        )

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton(f"🎵 YM: {_fmt_onoff(enable_ym)}", callback_data="admin_settings:toggle_ym")],
            [InlineKeyboardButton(f"🎶 VK: {_fmt_onoff(enable_vk)}", callback_data="admin_settings:toggle_vk")],
            [InlineKeyboardButton(f"▶️ YouTube: {_fmt_onoff(enable_yt)}", callback_data="admin_settings:toggle_yt")],
            [
                InlineKeyboardButton("⚡ fastest", callback_data="admin_settings:priority:fastest"),
                InlineKeyboardButton("🎵 ym", callback_data="admin_settings:priority:ym"),
                InlineKeyboardButton("🎶 vk", callback_data="admin_settings:priority:vk"),
                InlineKeyboardButton("▶️ yt", callback_data="admin_settings:priority:yt"),
            ],
            [InlineKeyboardButton("📥 Excel → прогрев file_id", callback_data="admin_settings:excel_warmup")],
            [InlineKeyboardButton("🛑 Остановить прогрев", callback_data="admin_settings:excel_warmup_cancel")],
            [InlineKeyboardButton("🔙 Назад", callback_data="admin_menu")],
        ])

        await safe_edit_message_text(query, text_msg, parse_mode="HTML", reply_markup=kb, disable_web_page_preview=True)
    except Exception as e:
        logger.error(f"Admin settings error: {e}", exc_info=True)
        await safe_answer_callback(query, "❌ Не удалось показать настройки", show_alert=True)


async def _handle_admin_settings_action(update: Update, context: CallbackContext, callback_data: str):
    """Обработчик кнопок админки настроек"""
    query = update.callback_query
    await safe_answer_callback(query, "")

    if not hasattr(bot_instance, "admin_db") or not bot_instance.admin_db:
        await safe_answer_callback(query, "❌ AdminDB не инициализирован", show_alert=True)
        return

    # callback_data вида: admin_settings:<action>
    action = callback_data.split(":", 1)[1] if ":" in callback_data else callback_data

    try:
        if action == "toggle_ym":
            cur = getattr(config, "ENABLE_YANDEX_MUSIC", True)
            await bot_instance.admin_db.set_setting("enable_ym", (not cur))

        elif action == "toggle_vk":
            cur = getattr(config, "ENABLE_VK_MUSIC", True)
            await bot_instance.admin_db.set_setting("enable_vk", (not cur))

        elif action == "toggle_yt":
            cur = getattr(config, "ENABLE_YOUTUBE_MUSIC", True)
            await bot_instance.admin_db.set_setting("enable_yt", (not cur))

        elif action.startswith("priority:"):
            val = action.split(":", 1)[1]
            if val not in ("fastest", "ym", "vk", "yt"):
                val = "fastest"
            await bot_instance.admin_db.set_setting("source_priority", val)

        elif action == "excel_warmup":
            # Переводим админа в режим ожидания excel-файла
            user_id = query.from_user.id
            await bot_instance.set_user_state(user_id, "await_excel_warmup", {"ts": time.time()})
            await safe_edit_message_text(
                query,
                "📥 <b>Отправь сюда Excel (.xlsx)</b>\n\n"
                "Требования:\n"
                "• Лист <code>Artists</code> (или первый лист)\n"
                "• Колонка <code>Artist</code> (или первая колонка)\n\n"
                "После загрузки бот начнёт прогрев: поиск артистов по очереди → скачивание треков → сохранение file_id.",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Назад", callback_data="admin_settings")]]),
                disable_web_page_preview=True
            )
            return

        elif action == "excel_warmup_cancel":
            user_id = query.from_user.id
            bot_instance.excel_warmup_cancel.add(user_id)
            await bot_instance.set_user_state(user_id, "idle", {})
            await safe_answer_callback(query, "🛑 Ок, останавливаю (если было запущено).", show_alert=True)
            await _show_admin_settings(update, context)
            return

        # Обновим runtime и перерисуем меню
        try:
            await bot_instance.admin_db.apply_runtime_settings()
        except Exception:
            pass
        await _show_admin_settings(update, context)

    except Exception as e:
        logger.error(f"Admin settings action error: {e}", exc_info=True)
        await safe_answer_callback(query, "❌ Ошибка при изменении настроек", show_alert=True)


async def _show_admin_moderation(query):
    await safe_answer_callback(query)
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🚫 Забанить по ID", callback_data="admin_moderation:ban")],
        [InlineKeyboardButton("✅ Разбанить по ID", callback_data="admin_moderation:unban")],
        [InlineKeyboardButton("📃 Список банов", callback_data="admin_moderation:list")],
        [InlineKeyboardButton("🔙 Назад", callback_data="admin_menu")]
    ])
    await safe_edit_message_text(query, "🚫 <b>Модерация</b>\n\nВыберите действие:", parse_mode="HTML", reply_markup=kb)

async def _handle_admin_moderation_action(query, action: str):
    await safe_answer_callback(query)
    if not hasattr(bot_instance, "admin_db") or not bot_instance.admin_db:
        await safe_answer_callback(query, "❌ AdminDB не инициализирован", show_alert=True)
        return

    if action == "ban":
        await bot_instance.set_user_state(query.from_user.id, "ban_user")
        await safe_edit_message_text(
            query,
            "🚫 <b>Бан пользователя</b>\n\nОтправьте <code>user_id</code> (только число).\n\nПример: <code>123456789</code>",
            parse_mode="HTML",
            reply_markup=_kb_back_to_admin()
        )

    elif action == "unban":
        await bot_instance.set_user_state(query.from_user.id, "unban_user")
        await safe_edit_message_text(
            query,
            "✅ <b>Разбан пользователя</b>\n\nОтправьте <code>user_id</code> (только число).",
            parse_mode="HTML",
            reply_markup=_kb_back_to_admin()
        )

    elif action == "list":
        bans = await bot_instance.admin_db.list_bans(limit=50)
        if not bans:
            await safe_edit_message_text(query, "📃 <b>Баны</b>\n\n— список пуст —", parse_mode="HTML", reply_markup=_kb_back_to_admin())
            return
        text_lines = ["📃 <b>Баны (последние 50)</b>\n"]
        for b in bans[:50]:
            uid = b.get("user_id")
            reason = b.get("reason") or ""
            banned_at = b.get("banned_at") or ""
            text_lines.append(f"• <code>{uid}</code> — {banned_at}{(' — ' + reason) if reason else ''}")
        await safe_edit_message_text(query, "\n".join(text_lines), parse_mode="HTML", reply_markup=_kb_back_to_admin(), disable_web_page_preview=True)

    else:
        await _show_admin_moderation(query)

async def _handle_ban_input(update: Update, context: CallbackContext, state_data: Dict[str, Any]):
    user_id = update.effective_user.id
    if not _is_admin(user_id):
        return
    txt = update.effective_message.text.strip()
    uid = _safe_int(txt.split()[0]) if txt else None
    if not uid:
        await update.effective_message.reply_text("❌ Нужен числовой <code>user_id</code>.", parse_mode="HTML")
        return
    reason = " ".join(txt.split()[1:]).strip() if len(txt.split()) > 1 else ""
    try:
        await bot_instance.admin_db.ban_user(uid, reason=reason, banned_by=user_id)
        bot_instance.user_manager.banned_users.add(int(uid))
        bot_instance.user_manager._last_ban_refresh = time.monotonic()
        await update.effective_message.reply_text(f"✅ Забанен: <code>{uid}</code>", parse_mode="HTML")
    except Exception as e:
        await update.effective_message.reply_text(f"❌ Ошибка: {e}", parse_mode="HTML")
    finally:
        await bot_instance.clear_user_state(user_id)

async def _handle_unban_input(update: Update, context: CallbackContext, state_data: Dict[str, Any]):
    user_id = update.effective_user.id
    if not _is_admin(user_id):
        return
    txt = update.effective_message.text.strip()
    uid = _safe_int(txt.split()[0]) if txt else None
    if not uid:
        await update.effective_message.reply_text("❌ Нужен числовой <code>user_id</code>.", parse_mode="HTML")
        return
    try:
        ok = await bot_instance.admin_db.unban_user(uid)
        if ok:
            bot_instance.user_manager.banned_users.discard(int(uid))
            bot_instance.user_manager._last_ban_refresh = time.monotonic()
            await update.effective_message.reply_text(f"✅ Разбанен: <code>{uid}</code>", parse_mode="HTML")
        else:
            await update.effective_message.reply_text(f"⚠️ Не был в бане: <code>{uid}</code>", parse_mode="HTML")
    except Exception as e:
        await update.effective_message.reply_text(f"❌ Ошибка: {e}", parse_mode="HTML")
    finally:
        await bot_instance.clear_user_state(user_id)


async def _show_admin_groups(query, page: int = 0):
    await safe_answer_callback(query)
    if not getattr(bot_instance, "admin_db", None):
        await safe_answer_callback(query, "❌ AdminDB не инициализирован", show_alert=True)
        return

    page_size = max(1, min(int(getattr(config, "ADMIN_GROUPS_PAGE_SIZE", 8) or 8), 20))
    total = await bot_instance.admin_db.count_group_chats(active_only=True)
    max_page = max(0, math.ceil(total / page_size) - 1)
    page = max(0, min(int(page or 0), max_page))
    groups = await bot_instance.admin_db.list_group_chats(
        active_only=True,
        limit=page_size,
        offset=page * page_size,
    )

    lines = [
        "👥 <b>Группы, в которые добавлен бот</b>",
        "",
        f"Всего активных групп: <b>{total}</b>",
        "Нажмите на группу, чтобы подготовить сообщение.",
        "",
    ]
    rows: List[List[InlineKeyboardButton]] = []
    for index, group in enumerate(groups, start=page * page_size + 1):
        chat_id = int(group.get("chat_id") or 0)
        title = str(group.get("title") or group.get("username") or chat_id)
        username = str(group.get("username") or "")
        status = str(group.get("bot_status") or "member")
        admin_mark = "🛡" if group.get("bot_is_admin") else "👤"
        send_mark = "✅" if group.get("can_send_messages") or status in ("member", "administrator", "creator", "") else "⚠️"
        username_part = f" · @{_esc(username)}" if username else ""
        last_seen = str(group.get("last_seen") or "").replace("T", " ")[:19]
        lines.append(
            f"{index}. {admin_mark} <b>{_esc(title)}</b>{username_part}\n"
            f"   ID: <code>{chat_id}</code> · {send_mark} <code>{_esc(status)}</code>"
            + (f" · {_esc(last_seen)}" if last_seen else "")
        )
        button_title = title if len(title) <= 38 else title[:35] + "…"
        rows.append([
            InlineKeyboardButton(
                f"📨 {button_title}",
                callback_data=f"admin_broadcast:group:{chat_id}",
            )
        ])

    if not groups:
        lines.append("— активные группы пока не зафиксированы —")

    nav: List[InlineKeyboardButton] = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️", callback_data=f"admin_groups:{page - 1}"))
    if page < max_page:
        nav.append(InlineKeyboardButton("➡️", callback_data=f"admin_groups:{page + 1}"))
    if nav:
        rows.append(nav)
    rows.append([InlineKeyboardButton("📣 Рассылка", callback_data="admin_broadcast")])
    rows.append([InlineKeyboardButton("🔙 В админку", callback_data="admin_menu")])

    await safe_edit_message_text(
        query,
        "\n".join(lines),
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(rows),
        disable_web_page_preview=True,
    )


async def _show_admin_broadcast(query, context: CallbackContext = None):
    await safe_answer_callback(query)
    if not hasattr(bot_instance, "admin_db") or not bot_instance.admin_db:
        await safe_answer_callback(query, "❌ AdminDB не инициализирован", show_alert=True)
        return
    try:
        await bot_instance.clear_user_state(query.from_user.id)
        users_cnt = await bot_instance.admin_db.count_users()
        groups_cnt = await bot_instance.admin_db.count_group_chats(active_only=True)
        admin_chats_cnt = len(await bot_instance.admin_db.get_admin_chat_ids())

        text_msg = (
            "📣 <b>Рассылка</b>\n\n"
            f"👤 Пользователей в базе: <b>{users_cnt}</b>\n"
            f"👥 Активных групп с ботом: <b>{groups_cnt}</b>\n"
            f"🛡 Чатов/каналов, где бот админ: <b>{admin_chats_cnt}</b>\n\n"
            "Выберите получателя:"
        )

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("👤 Конкретному пользователю", callback_data="admin_broadcast:to_user")],
            [InlineKeyboardButton("👥 Всем пользователям", callback_data="admin_broadcast:to_users")],
            [InlineKeyboardButton("💬 Конкретной группе", callback_data="admin_broadcast:to_group")],
            [InlineKeyboardButton("📋 Выбрать группу из списка", callback_data="admin_groups:0")],
            [InlineKeyboardButton("📣 Во все активные группы", callback_data="admin_broadcast:to_groups")],
            [InlineKeyboardButton("🛡 Во все чаты/каналы, где бот админ", callback_data="admin_broadcast:to_admin_chats")],
            [InlineKeyboardButton("🔙 Назад", callback_data="admin_menu")]
        ])
        await safe_edit_message_text(query, text_msg, parse_mode="HTML", reply_markup=kb, disable_web_page_preview=True)
    except Exception:
        logger.exception("Admin broadcast show error")
        await safe_answer_callback(query, "❌ Не удалось открыть рассылку", show_alert=True)


async def _handle_admin_broadcast_action(query, action: str, context: CallbackContext = None):
    await safe_answer_callback(query)
    if not getattr(bot_instance, "admin_db", None):
        await safe_answer_callback(query, "❌ AdminDB не инициализирован", show_alert=True)
        return

    if action in ("to_user", "to_group"):
        target_kind = "user" if action == "to_user" else "group"
        await bot_instance.set_user_state(
            query.from_user.id,
            "broadcast_target_id",
            {"target_kind": target_kind},
        )
        prompt = (
            "👤 <b>Сообщение конкретному пользователю</b>\n\n"
            "Отправьте числовой <code>user_id</code>. Пользователь должен быть зарегистрирован в базе бота."
            if target_kind == "user"
            else
            "💬 <b>Сообщение конкретной группе</b>\n\n"
            "Отправьте числовой <code>chat_id</code> группы или выберите её из списка."
        )
        kb_rows = []
        if target_kind == "group":
            kb_rows.append([InlineKeyboardButton("📋 Показать группы", callback_data="admin_groups:0")])
        kb_rows.append([InlineKeyboardButton("🔙 Назад", callback_data="admin_broadcast")])
        await safe_edit_message_text(
            query,
            prompt,
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup(kb_rows),
        )
        return

    if action.startswith("group:"):
        chat_id = _safe_int(action.split(":", 1)[1])
        group = await bot_instance.admin_db.get_group_chat(chat_id) if chat_id is not None else None
        if not group or str(group.get("bot_status") or "") in ("left", "kicked"):
            await safe_answer_callback(query, "❌ Группа не найдена или бот уже удалён", show_alert=True)
            return
        title = str(group.get("title") or group.get("username") or chat_id)
        await bot_instance.set_user_state(
            query.from_user.id,
            "broadcast_text",
            {"target": "single_chat", "chat_id": int(chat_id), "label": title},
        )
        await safe_edit_message_text(
            query,
            "📣 <b>Сообщение в группу</b>\n\n"
            f"Получатель: <b>{_esc(title)}</b>\n"
            f"ID: <code>{chat_id}</code>\n\n"
            "Отправьте текст одним сообщением.",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 К группам", callback_data="admin_groups:0")]]),
        )
        return

    target_labels = {
        "to_users": ("users", "всем пользователям в личные сообщения"),
        "to_groups": ("groups", "во все активные группы"),
        "to_admin_chats": ("admin_chats", "во все чаты/каналы, где бот администратор"),
        # Совместимость со старой кнопкой.
        "start": ("admin_chats", "во все чаты/каналы, где бот администратор"),
    }
    target_info = target_labels.get(action)
    if target_info:
        target, label = target_info
        await bot_instance.set_user_state(query.from_user.id, "broadcast_text", {"target": target})
        await safe_edit_message_text(
            query,
            "📣 <b>Рассылка</b>\n\n"
            f"Цель: <b>{_esc(label)}</b>\n\n"
            "Отправьте текст рассылки одним сообщением.",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Назад", callback_data="admin_broadcast")]])
        )
        return

    await _show_admin_broadcast(query, context=context)


async def _handle_broadcast_target_input(
    update: Update,
    context: CallbackContext,
    state_data: Dict[str, Any],
) -> None:
    user_id = update.effective_user.id
    if not _is_admin(user_id):
        return
    raw = str(update.effective_message.text or "").strip()
    target_id = _safe_int(raw.split()[0]) if raw else None
    target_kind = str((state_data or {}).get("data", {}).get("target_kind") or "")
    if target_id is None:
        await update.effective_message.reply_text("❌ Нужен числовой ID.", parse_mode="HTML")
        return

    if target_kind == "user":
        if target_id <= 0:
            await update.effective_message.reply_text("❌ <code>user_id</code> должен быть положительным числом.", parse_mode="HTML")
            return
        user = await bot_instance.admin_db.get_user(target_id)
        if not user:
            await update.effective_message.reply_text(
                "❌ Пользователь не найден в базе. Бот может написать пользователю только после того, как пользователь уже взаимодействовал с ним.",
                parse_mode="HTML",
            )
            return
        display_name = " ".join(filter(None, (user.get("first_name"), user.get("last_name")))).strip()
        if user.get("username"):
            display_name = f"{display_name} (@{user['username']})".strip()
        display_name = display_name or str(target_id)
        target = "single_user"
        label = display_name
    elif target_kind == "group":
        group = await bot_instance.admin_db.get_group_chat(target_id)
        if not group or str(group.get("bot_status") or "") in ("left", "kicked"):
            await update.effective_message.reply_text(
                "❌ Активная группа с таким <code>chat_id</code> не найдена. Откройте список групп в админке.",
                parse_mode="HTML",
            )
            return
        target = "single_chat"
        label = str(group.get("title") or group.get("username") or target_id)
    else:
        await bot_instance.clear_user_state(user_id)
        await update.effective_message.reply_text("❌ Состояние рассылки устарело. Откройте админку заново.")
        return

    await bot_instance.set_user_state(
        user_id,
        "broadcast_text",
        {"target": target, "chat_id": int(target_id), "label": label},
    )
    await update.effective_message.reply_text(
        "✅ Получатель выбран.\n\n"
        f"<b>{_esc(label)}</b>\nID: <code>{target_id}</code>\n\n"
        "Теперь отправьте текст сообщения.",
        parse_mode="HTML",
    )


async def _send_broadcast_text(bot: Any, chat_id: int, html_text: str, plain_text: str) -> None:
    try:
        await safe_send_message(
            bot,
            chat_id=chat_id,
            text=html_text,
            parse_mode="HTML",
            disable_web_page_preview=True,
        )
    except BadRequest as exc:
        error_text = str(exc).lower()
        if "parse" not in error_text and "entity" not in error_text:
            raise
        await safe_send_message(
            bot,
            chat_id=chat_id,
            text=plain_text,
            disable_web_page_preview=True,
        )


async def _handle_broadcast_text(update: Update, context: CallbackContext, state_data: Dict[str, Any]):
    user_id = update.effective_user.id
    if not _is_admin(user_id):
        return
    if not getattr(bot_instance, "admin_db", None):
        await update.effective_message.reply_text("❌ AdminDB не инициализирован", parse_mode="HTML")
        await bot_instance.clear_user_state(user_id)
        return

    message = update.effective_message
    plain_text = str(message.text or "")
    html_text = str(getattr(message, "text_html", None) or html.escape(plain_text))
    state_payload = (state_data or {}).get("data", {})
    target = str(state_payload.get("target") or "admin_chats")
    await bot_instance.clear_user_state(user_id)

    if target == "users":
        targets = await bot_instance.admin_db.get_user_ids()
        label = "пользователям в личные сообщения"
    elif target == "groups":
        targets = await bot_instance.admin_db.get_group_chat_ids(active_only=True)
        label = "активным группам"
    elif target == "single_user":
        targets = [int(state_payload.get("chat_id"))]
        label = f"пользователю {_esc(state_payload.get('label') or targets[0])}"
    elif target == "single_chat":
        targets = [int(state_payload.get("chat_id"))]
        label = f"группе {_esc(state_payload.get('label') or targets[0])}"
    else:
        targets = await bot_instance.admin_db.get_admin_chat_ids()
        label = "чатам/каналам, где бот администратор"

    # Remove duplicates without changing the existing recency/order from SQLite.
    targets = list(dict.fromkeys(int(chat_id) for chat_id in targets if chat_id is not None))
    if not targets:
        await message.reply_text("⚠️ Подходящих получателей не найдено.", parse_mode="HTML")
        return

    ok = 0
    fail = 0
    failed_ids: List[int] = []
    await message.reply_text(
        f"📣 Начинаю отправку: <b>{len(targets)}</b> получателей ({label})…",
        parse_mode="HTML",
    )
    delay = max(0.0, float(getattr(config, "BROADCAST_SEND_DELAY", 0.05) or 0.0))
    for chat_id in targets:
        try:
            await _send_broadcast_text(context.bot, chat_id, html_text, plain_text)
            ok += 1
        except Exception as exc:
            fail += 1
            if len(failed_ids) < 20:
                failed_ids.append(chat_id)
            logger.warning("Broadcast delivery failed for chat_id=%s: %s", chat_id, exc)
        if delay:
            await asyncio.sleep(delay)

    try:
        await bot_instance.admin_db.log_event(
            "broadcast",
            user_id=user_id,
            extra={"target": target, "recipients": len(targets), "ok": ok, "fail": fail},
        )
    except Exception:
        logger.exception("Failed to store broadcast statistics")

    failure_details = ""
    if failed_ids:
        failure_details = "\nНе доставлено: " + ", ".join(f"<code>{chat_id}</code>" for chat_id in failed_ids)
        if fail > len(failed_ids):
            failure_details += f" и ещё {fail - len(failed_ids)}"
    await message.reply_text(
        f"✅ Отправка завершена.\nУспешно: <b>{ok}</b>\nОшибки: <b>{fail}</b>{failure_details}",
        parse_mode="HTML",
    )


# ==================== ГРУППОВЫЕ РАССЫЛКИ ПОДБОРОК ====================
async def _is_group_admin(update: Update, context: CallbackContext, notify: bool = True) -> bool:
    chat = update.effective_chat
    user = update.effective_user
    message = update.effective_message
    if not chat or chat.type not in ("group", "supergroup"):
        if notify and message:
            await message.reply_text("Эта настройка доступна только внутри группы.")
        return False

    # Anonymous administrators appear as the group sender_chat.
    sender_chat = getattr(message, "sender_chat", None) if message else None
    if sender_chat and int(getattr(sender_chat, "id", 0) or 0) == int(chat.id):
        return True
    if not user:
        return False
    try:
        member = await context.bot.get_chat_member(chat.id, user.id)
        return getattr(member, "status", "") in ("administrator", "creator")
    except Exception:
        if notify and message:
            await message.reply_text("Не удалось проверить права администратора группы.")
        return False


async def observe_group_topic(update: Update, context: CallbackContext) -> None:
    """Remember forum topics seen by the bot; Bot API cannot enumerate all topics."""
    try:
        if not bot_instance or not getattr(bot_instance, "digest_store", None):
            return
        chat = update.effective_chat
        message = update.effective_message
        if not chat or chat.type not in ("group", "supergroup") or not message:
            return
        thread_id = getattr(message, "message_thread_id", None)
        if not thread_id:
            return
        title = None
        created = getattr(message, "forum_topic_created", None)
        edited = getattr(message, "forum_topic_edited", None)
        if created:
            title = getattr(created, "name", None)
        elif edited:
            title = getattr(edited, "name", None)
        await bot_instance.digest_store.remember_topic(chat.id, thread_id, title)
    except Exception:
        logger.debug("Topic observation failed", exc_info=True)


def _digest_wizard_key(user_id: int, chat_id: int) -> Tuple[int, int]:
    return int(user_id), int(chat_id)


def _digest_destination_label(thread_id: int, topics: Optional[List[Dict[str, Any]]] = None) -> str:
    if not thread_id:
        return "Основной чат"
    for topic in topics or []:
        if int(topic.get("thread_id") or 0) == int(thread_id):
            return str(topic.get("title") or f"Раздел #{thread_id}")
    return f"Раздел #{thread_id}"


def _digest_schedule_label(schedule_key: str) -> str:
    legacy = (DIGEST_SCHEDULES.get(str(schedule_key)) or {}).get("label")
    if legacy:
        return str(legacy)
    days, hour, minute = _digest_parse_schedule(schedule_key)
    short_names = [short for day, short, _ in DIGEST_WEEKDAYS if day in days]
    if days == set(range(7)):
        day_text = "Каждый день"
    elif days == set(range(5)):
        day_text = "По будням"
    elif days == {5, 6}:
        day_text = "По выходным"
    else:
        day_text = ", ".join(short_names)
    return f"{day_text} в {hour:02d}:{minute:02d}"



def _user_mix_genre_keyboard() -> InlineKeyboardMarkup:
    rows: List[List[InlineKeyboardButton]] = []
    row: List[InlineKeyboardButton] = []
    for genre_id, genre_name in DEEZER_GENRES.items():
        row.append(InlineKeyboardButton(genre_name, callback_data=f"mix:genre:{genre_id}"))
        if len(row) == 2:
            rows.append(row)
            row = []
    if row:
        rows.append(row)
    rows.append([InlineKeyboardButton("❌ Закрыть", callback_data="mix:close")])
    return InlineKeyboardMarkup(rows)


def _user_mix_count_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("3 трека", callback_data="mix:count:3"),
            InlineKeyboardButton("5 треков", callback_data="mix:count:5"),
        ],
        [
            InlineKeyboardButton("7 треков", callback_data="mix:count:7"),
            InlineKeyboardButton("10 треков", callback_data="mix:count:10"),
        ],
        [InlineKeyboardButton("🔙 Другой жанр", callback_data="mix:new")],
        [InlineKeyboardButton("❌ Закрыть", callback_data="mix:close")],
    ])


async def _render_user_mix_genres(target: Any, edit: bool = False) -> None:
    text = (
        "🎲 <b>Произвольная музыкальная подборка</b>\n\n"
        "Выберите жанр. Бот возьмёт популярные позиции Deezer, а аудио "
        "отправит из Telegram-кэша или автоматически скачает через VK, Яндекс.Музыку или YouTube."
    )
    if edit:
        await safe_edit_message_text(
            target,
            text,
            parse_mode="HTML",
            reply_markup=_user_mix_genre_keyboard(),
        )
    else:
        await target.effective_message.reply_text(
            text,
            parse_mode="HTML",
            reply_markup=_user_mix_genre_keyboard(),
        )


async def _open_user_mix(
    update: Update,
    context: CallbackContext,
    check_access: bool = True,
) -> None:
    if check_access and not await user_check(update):
        return
    message = update.effective_message
    chat = update.effective_chat
    user = update.effective_user
    if not message or not chat or not user:
        return

    try:
        if getattr(bot_instance, "admin_db", None):
            await bot_instance.admin_db.track_chat(update, bot=context.bot)
            await bot_instance.admin_db.track_user(user)
    except Exception:
        pass
    if chat.type in ("group", "supergroup"):
        await observe_group_topic(update, context)

    key = _user_mix_key(user.id, chat.id)
    _USER_MIX_WIZARDS[key] = {
        "user_id": int(user.id),
        "chat_id": int(chat.id),
        "thread_id": int(getattr(message, "message_thread_id", 0) or 0),
        "created_at": time.time(),
    }
    await _render_user_mix_genres(update, edit=False)


async def user_mix_command(update: Update, context: CallbackContext) -> None:
    """Start an on-demand genre mix for any non-banned user."""
    await _open_user_mix(update, context, check_access=True)


async def _run_user_mix(application: Application, subscription: Dict[str, Any]) -> None:
    history_id = int(subscription["id"])
    chat_id = int(subscription["chat_id"])
    try:
        result = await send_group_digest(application, subscription, manual=False)
        if not result.get("ok") and not result.get("notified"):
            kwargs: Dict[str, Any] = {
                "chat_id": int(subscription["chat_id"]),
                "text": "⚠️ Подборка не отправлена: " + str(
                    result.get("reason") or "нет подходящих треков"
                ),
            }
            thread_id = int(subscription.get("thread_id") or 0)
            if thread_id:
                kwargs["message_thread_id"] = thread_id
            await safe_send_message(application.bot, **kwargs)
    except asyncio.CancelledError:
        raise
    except Exception as exc:
        logger.exception("User mix failed: %s", exc)
        kwargs = {
            "chat_id": int(subscription["chat_id"]),
            "text": "⚠️ Не удалось сформировать подборку. Попробуйте ещё раз позже.",
        }
        thread_id = int(subscription.get("thread_id") or 0)
        if thread_id:
            kwargs["message_thread_id"] = thread_id
        try:
            await safe_send_message(application.bot, **kwargs)
        except Exception:
            pass
    finally:
        _USER_MIX_RUNNING.discard(history_id)
        _USER_MIX_CHAT_RUNNING.discard(chat_id)


async def handle_user_mix_callback(update: Update, context: CallbackContext) -> None:
    query = update.callback_query
    if not query or not query.message or not query.from_user:
        return

    action = str(query.data or "")
    user_id = int(query.from_user.id)
    chat_id = int(query.message.chat.id)
    key = _user_mix_key(user_id, chat_id)

    if action == "mix:close":
        _USER_MIX_WIZARDS.pop(key, None)
        await safe_answer_callback(query)
        await safe_edit_message_text(query, "✅ Подборка закрыта.")
        return

    if action == "mix:new":
        _USER_MIX_WIZARDS[key] = {
            "user_id": user_id,
            "chat_id": chat_id,
            "thread_id": int(getattr(query.message, "message_thread_id", 0) or 0),
            "created_at": time.time(),
        }
        await safe_answer_callback(query)
        await _render_user_mix_genres(query, edit=True)
        return

    state = _USER_MIX_WIZARDS.get(key)
    if not state or time.time() - float(state.get("created_at") or 0) > 900:
        _USER_MIX_WIZARDS.pop(key, None)
        await safe_answer_callback(
            query,
            "Настройка устарела. Запустите /mix ещё раз.",
            show_alert=True,
        )
        return

    expected_thread = int(state.get("thread_id") or 0)
    actual_thread = int(getattr(query.message, "message_thread_id", 0) or 0)
    if expected_thread != actual_thread:
        await safe_answer_callback(
            query,
            "Используйте кнопки в исходном разделе.",
            show_alert=True,
        )
        return

    if action.startswith("mix:genre:"):
        try:
            genre_id = int(action.rsplit(":", 1)[1])
        except Exception:
            genre_id = -1
        genre_name = DEEZER_GENRES.get(genre_id)
        if genre_name is None:
            await safe_answer_callback(query, "Неизвестный жанр.", show_alert=True)
            return
        state["genre_id"] = genre_id
        state["genre_name"] = genre_name
        state["created_at"] = time.time()
        await safe_answer_callback(query)
        await safe_edit_message_text(
            query,
            f"🎲 <b>{_esc(genre_name)}</b>\n\nВыберите количество треков:",
            parse_mode="HTML",
            reply_markup=_user_mix_count_keyboard(),
        )
        return

    if action.startswith("mix:count:"):
        if "genre_id" not in state:
            await safe_answer_callback(query, "Сначала выберите жанр.", show_alert=True)
            return
        try:
            track_count = int(action.rsplit(":", 1)[1])
        except Exception:
            track_count = 0
        if track_count not in {3, 5, 7, 10}:
            await safe_answer_callback(query, "Недопустимое количество треков.", show_alert=True)
            return

        history_id = _user_mix_history_id(user_id, chat_id)
        if history_id in _USER_MIX_RUNNING or _digest_send_lock(history_id).locked():
            await safe_answer_callback(
                query,
                "Ваша предыдущая подборка ещё формируется.",
                show_alert=True,
            )
            return
        if chat_id in _USER_MIX_CHAT_RUNNING:
            await safe_answer_callback(
                query,
                "В этом чате уже формируется другая подборка.",
                show_alert=True,
            )
            return

        # Count the whole mix as one download action for the normal user rate
        # limiter. The actual per-track downloads are additionally protected by
        # the bot's global and per-chat semaphores.
        if not await bot_instance.user_manager.check_rate_limit(user_id, "download"):
            await safe_answer_callback(
                query,
                "Слишком много подборок. Попробуйте немного позже.",
                show_alert=True,
            )
            return

        _USER_MIX_RUNNING.add(history_id)
        _USER_MIX_CHAT_RUNNING.add(chat_id)
        _USER_MIX_WIZARDS.pop(key, None)
        await safe_answer_callback(query, "Формирую подборку…")
        await safe_edit_message_text(
            query,
            f"🎲 <b>{_esc(state['genre_name'])}</b>\n\n"
            f"Запрошено треков: <b>{track_count}</b>. Подготовка началась…",
            parse_mode="HTML",
        )

        subscription = {
            "id": history_id,
            "chat_id": chat_id,
            "thread_id": expected_thread,
            "genre_id": int(state["genre_id"]),
            "genre_name": str(state["genre_name"]),
            "track_count": track_count,
            "no_repeat_days": _user_mix_no_repeat_days(),
            "is_user_mix": True,
            "created_by": user_id,
        }
        try:
            task = bot_instance.create_background_task(
                _run_user_mix(context.application, subscription),
                f"user-mix:{history_id}",
            )
        except Exception:
            _USER_MIX_RUNNING.discard(history_id)
            _USER_MIX_CHAT_RUNNING.discard(chat_id)
            raise
        return

    await safe_answer_callback(query, "Неизвестное действие подборки.", show_alert=True)


async def _digest_chat_is_forum(context: CallbackContext, chat_id: int, fallback_chat: Any = None) -> bool:
    if bool(getattr(fallback_chat, "is_forum", False)):
        return True
    try:
        chat = await context.bot.get_chat(chat_id)
        return bool(getattr(chat, "is_forum", False))
    except Exception:
        return False


async def _render_digest_home(target: Any, context: CallbackContext, edit: bool = False) -> None:
    chat = target.message.chat if edit else target.effective_chat
    subscriptions = await bot_instance.digest_store.list_subscriptions(chat.id)
    is_forum = await _digest_chat_is_forum(context, chat.id, chat)
    enabled = sum(1 for sub in subscriptions if sub.get("enabled"))
    text = (
        "🎧 <b>Рассылки музыкальных подборок</b>\n\n"
        "Настраивать их могут только администраторы этой группы. "
        "Популярность и жанры берутся из Deezer. Сначала бот использует Telegram file_id-кэш, "
        "а отсутствующие композиции сам находит и скачивает через VK, Яндекс.Музыку или YouTube.\n\n"
        f"Активных рассылок: <b>{enabled}</b> из <b>{len(subscriptions)}</b>\n"
        f"Разделы Telegram: <b>{'включены' if is_forum else 'не используются'}</b>\n"
        f"Часовой пояс: <code>{_esc(str(getattr(config, 'DIGEST_TIMEZONE', 'Europe/Amsterdam')))}</code>"
    )
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Новая рассылка", callback_data="digest:new")],
        [InlineKeyboardButton("📋 Мои рассылки", callback_data="digest:list")],
        [InlineKeyboardButton("❌ Закрыть", callback_data="digest:close")],
    ])
    if edit:
        await safe_answer_callback(target)
        await safe_edit_message_text(target, text, parse_mode="HTML", reply_markup=kb)
    else:
        await target.effective_message.reply_text(text, parse_mode="HTML", reply_markup=kb)


async def digest_command(update: Update, context: CallbackContext) -> None:
    if not await _is_group_admin(update, context):
        return
    try:
        await bot_instance.admin_db.track_chat(update, bot=context.bot)
        await bot_instance.admin_db.track_user(update.effective_user)
    except Exception:
        pass
    await observe_group_topic(update, context)
    await _render_digest_home(update, context, edit=False)


async def _render_digest_destinations(query: Any, context: CallbackContext, state: Dict[str, Any]) -> None:
    chat_id = int(state["chat_id"])
    topics = await bot_instance.digest_store.list_topics(chat_id, limit=12)
    current_thread = int(state.get("current_thread_id") or 0)
    rows: List[List[InlineKeyboardButton]] = [
        [InlineKeyboardButton("💬 Основной чат", callback_data="digest:dest:0")]
    ]
    if current_thread:
        rows.append([InlineKeyboardButton("📌 Текущий раздел", callback_data=f"digest:dest:{current_thread}")])
    for topic in topics:
        tid = int(topic.get("thread_id") or 0)
        if not tid or tid == current_thread:
            continue
        title = str(topic.get("title") or f"Раздел #{tid}")
        if len(title) > 42:
            title = title[:41] + "…"
        rows.append([InlineKeyboardButton(f"🧵 {title}", callback_data=f"digest:dest:{tid}")])
        if len(rows) >= 11:
            break
    rows.append([InlineKeyboardButton("🔙 Назад", callback_data="digest:home")])
    await safe_edit_message_text(
        query,
        "📍 <b>Куда отправлять подборку?</b>\n\n"
        "Telegram не отдаёт ботам полный список тем. Здесь показаны текущий раздел и темы, которые бот уже видел.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(rows),
    )


async def _render_digest_genres(query: Any) -> None:
    rows: List[List[InlineKeyboardButton]] = []
    items = list(DEEZER_GENRES.items())
    for i in range(0, len(items), 2):
        row = []
        for genre_id, name in items[i:i + 2]:
            row.append(InlineKeyboardButton(name, callback_data=f"digest:genre:{genre_id}"))
        rows.append(row)
    rows.append([InlineKeyboardButton("🔙 Назад", callback_data="digest:new")])
    await safe_edit_message_text(
        query,
        "🎼 <b>Выберите тип подборки</b>\n\nТреки будут взяты из актуального чарта Deezer выбранного жанра.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(rows),
    )


async def _render_digest_schedules(query: Any, state: Dict[str, Any]) -> None:
    selected = {int(day) for day in state.get("selected_days", set())}
    rows: List[List[InlineKeyboardButton]] = []
    weekday_buttons: List[InlineKeyboardButton] = []
    for day, short_name, _ in DIGEST_WEEKDAYS:
        mark = "✅" if day in selected else "▫️"
        weekday_buttons.append(
            InlineKeyboardButton(f"{mark} {short_name}", callback_data=f"digest:day:{day}")
        )
    for index in range(0, len(weekday_buttons), 4):
        rows.append(weekday_buttons[index:index + 4])
    rows.append([InlineKeyboardButton("Каждый день", callback_data="digest:days:all")])
    rows.append([
        InlineKeyboardButton("Будни", callback_data="digest:days:weekdays"),
        InlineKeyboardButton("Выходные", callback_data="digest:days:weekend"),
    ])
    rows.append([InlineKeyboardButton("➡️ Указать время", callback_data="digest:days:done")])
    rows.append([InlineKeyboardButton("🔙 Назад", callback_data="digest:genres")])
    selected_text = ", ".join(
        short for day, short, _ in DIGEST_WEEKDAYS if day in selected
    ) or "не выбраны"
    await safe_edit_message_text(
        query,
        "📅 <b>Выберите дни публикации</b>\n\n"
        f"Сейчас: <b>{_esc(selected_text)}</b>\n"
        "Можно отметить любые дни недели.",
        parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(rows),
    )


async def _render_digest_counts(query: Any) -> None:
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("3 трека", callback_data="digest:count:3"), InlineKeyboardButton("5 треков", callback_data="digest:count:5")],
        [InlineKeyboardButton("7 треков", callback_data="digest:count:7"), InlineKeyboardButton("10 треков", callback_data="digest:count:10")],
        [InlineKeyboardButton("🔙 Назад", callback_data="digest:schedules")],
    ])
    await safe_edit_message_text(
        query,
        "🔢 <b>Сколько треков отправлять?</b>\n\nТреки из кэша отправляются сразу, остальные бот автоматически найдёт и скачает.",
        parse_mode="HTML",
        reply_markup=kb,
    )


async def _render_digest_list(query: Any, context: CallbackContext) -> None:
    chat_id = int(query.message.chat.id)
    topics = await bot_instance.digest_store.list_topics(chat_id, limit=30)
    subscriptions = await bot_instance.digest_store.list_subscriptions(chat_id)
    if not subscriptions:
        await safe_edit_message_text(
            query,
            "📋 <b>Рассылок пока нет.</b>",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("➕ Создать", callback_data="digest:new")],
                [InlineKeyboardButton("🔙 Назад", callback_data="digest:home")],
            ]),
        )
        return

    lines = ["📋 <b>Рассылки этой группы</b>\n"]
    rows: List[List[InlineKeyboardButton]] = []
    for sub in subscriptions[:12]:
        status = "✅" if sub["enabled"] else "⏸"
        dest = _digest_destination_label(sub["thread_id"], topics)
        lines.append(
            f"<b>#{sub['id']}</b> {status} {_esc(sub['genre_name'])} · {sub['track_count']} тр.\n"
            f"   {_esc(_digest_schedule_label(sub['schedule_key']))} · {_esc(dest)}"
        )
        rows.append([
            InlineKeyboardButton("▶️ Тест", callback_data=f"digest:test:{sub['id']}"),
            InlineKeyboardButton("🕒 Время", callback_data=f"digest:edit_schedule:{sub['id']}"),
        ])
        rows.append([
            InlineKeyboardButton("⏯ Вкл/выкл", callback_data=f"digest:toggle:{sub['id']}"),
            InlineKeyboardButton("🗑 Удалить", callback_data=f"digest:delete:{sub['id']}"),
        ])
    rows.append([InlineKeyboardButton("➕ Новая", callback_data="digest:new")])
    rows.append([InlineKeyboardButton("🔙 Назад", callback_data="digest:home")])
    await safe_edit_message_text(query, "\n\n".join(lines), parse_mode="HTML", reply_markup=InlineKeyboardMarkup(rows))


def _digest_source_track_score(candidate: Dict[str, Any], track: Dict[str, Any]) -> int:
    """Score a search result only for resolving a concrete Deezer chart item.

    This is not user-facing search ranking. A result is accepted only when both
    performer and title match closely enough, so an automatic digest does not
    download an unrelated recording with a similar word in its title.
    """
    candidate_artist = str(candidate.get("artist") or "")
    candidate_title = str(candidate.get("name") or candidate.get("title") or "")
    track_artist = str(track.get("artist") or "")
    track_title = str(track.get("title") or "")
    if not candidate_artist or not candidate_title or not track_artist or not track_title:
        return -1

    candidate_keys = set(_digest_track_keys(candidate_artist, candidate_title))
    track_keys = set(_digest_track_keys(track_artist, track_title))
    if candidate_keys.intersection(track_keys):
        score = 1000
    else:
        ca = _digest_normalize(candidate_artist)
        ta = _digest_normalize(track_artist)
        ct = _digest_normalize(candidate_title)
        tt = _digest_normalize(track_title)
        ct_base = _digest_normalize(re.sub(r"\s*[\(\[].*?[\)\]]\s*$", "", candidate_title))
        tt_base = _digest_normalize(re.sub(r"\s*[\(\[].*?[\)\]]\s*$", "", track_title))

        title_match = bool(ct and tt and (ct == tt or (ct_base and ct_base == tt_base)))
        artist_match = bool(
            ca and ta and (
                ca == ta
                or ca in ta
                or ta in ca
                or _digest_normalize(candidate_artist.split(",", 1)[0])
                == _digest_normalize(track_artist.split(",", 1)[0])
            )
        )
        if not title_match or not artist_match:
            return -1
        score = 700

    # Prefer complete/full tracks and preserve configured provider ordering.
    duration = int(track.get("duration") or 0)
    if duration >= 90:
        score += 20
    source = str(track.get("source") or "")
    # For unattended digests prefer Yandex by default: VK direct links expire
    # more often and can otherwise consume the whole digest timeout. This does
    # not affect ordinary user search ranking.
    priority = str(getattr(config, "DIGEST_SOURCE_PRIORITY", "yandex_first") or "yandex_first")
    preferred_source = "ym" if priority in ("yandex_first", "ym") else (
        "yt" if priority in ("youtube_first", "youtube", "yt") else "vk"
    )
    if source == preferred_source:
        score += 25
    return score


async def _resolve_digest_source_track(candidate: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Find the actual downloadable VK/Yandex/YouTube track for a Deezer chart row."""
    artist = str(candidate.get("artist") or "").strip()
    title = str(candidate.get("name") or candidate.get("title") or "").strip()
    if not artist or not title:
        return None

    search_limit = max(5, int(getattr(config, "DIGEST_SOURCE_SEARCH_LIMIT", 16) or 16))
    try:
        results = await bot_instance.search_all_sources(f"{artist} {title}", limit=search_limit)
    except Exception as exc:
        logger.warning("Digest source search failed for %s - %s: %s", artist, title, exc)
        return None

    ranked = sorted(
        ((int(_digest_source_track_score(candidate, track)), track) for track in (results or [])),
        key=lambda item: item[0],
        reverse=True,
    )
    if not ranked or ranked[0][0] < 0:
        logger.info("Digest track not found in enabled download sources: %s - %s", artist, title)
        return None
    return ranked[0][1]


def _digest_file_history_key(file_id: str) -> str:
    if not file_id:
        return ""
    return "tg:" + hashlib.sha256(str(file_id).encode("utf-8", errors="ignore")).hexdigest()


def _digest_invalid_file_id_error(exc: Exception) -> bool:
    message = str(exc).lower()
    return any(token in message for token in (
        "wrong file identifier",
        "file_id_invalid",
        "file id invalid",
        "file identifier",
        "failed to get http url content",
    ))


def _digest_fatal_destination_error(exc: Exception) -> bool:
    message = str(exc).lower()
    return any(token in message for token in (
        "chat not found",
        "topic_closed",
        "topic closed",
        "message thread not found",
        "forbidden",
        "bot was kicked",
    ))


async def _select_digest_candidates(subscription: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Return fresh Deezer chart candidates, whether cached or not.

    More candidates than the requested count are returned so the sender can skip
    unavailable provider results and still fill the digest without repeating old
    tracks.
    """
    requested_count = max(1, int(subscription.get("track_count", 5) or 5))
    fetch_limit = max(100, requested_count * 25)
    chart = await bot_instance.deezer.get_chart_tracks(
        bot_instance.session,
        int(subscription.get("genre_id", 0)),
        fetch_limit,
    )
    if not chart:
        return []

    cache_index = await bot_instance.digest_store.cached_audio_index()
    history_days = max(
        1,
        int(subscription.get("no_repeat_days") or _digest_no_repeat_days()),
    )
    recent = await bot_instance.digest_store.get_recent_sent_keys(
        int(subscription["id"]),
        days=history_days,
    )
    eligible: List[Dict[str, Any]] = []
    seen_track_keys: Set[str] = set()

    for position, candidate in enumerate(chart):
        artist = str(candidate.get("artist") or "").strip()
        title = str(candidate.get("name") or "").strip()
        candidate_keys = _digest_track_keys(artist, title)
        deezer_id = candidate.get("deezer_id")
        deezer_key = f"dz:{deezer_id}" if deezer_id not in (None, "") else ""
        canonical = candidate_keys[0] if candidate_keys else deezer_key
        if not canonical or canonical in seen_track_keys:
            continue

        cached = None
        for key in candidate_keys:
            cached = cache_index.get(key)
            if cached:
                break

        history_keys = list(candidate_keys)
        if deezer_key:
            history_keys.append(deezer_key)
        file_id = str((cached or {}).get("file_id") or "")
        file_history_key = _digest_file_history_key(file_id)
        if file_history_key:
            history_keys.append(file_history_key)
        history_keys = list(dict.fromkeys(key for key in history_keys if key))

        if recent.intersection(history_keys):
            continue

        eligible.append({
            "file_id": file_id,
            "query_key": (cached or {}).get("query_key") or "",
            "artist": artist,
            "title": title,
            "duration": int((cached or {}).get("duration") or 0),
            "deezer_id": deezer_id,
            "track_key": canonical,
            "history_keys": history_keys,
            "chart_position": position,
        })
        seen_track_keys.add(canonical)

    if not eligible:
        return []

    pool_limit = max(requested_count * 12, 60)
    attempt_limit = max(
        requested_count,
        int(getattr(config, "DIGEST_MAX_CANDIDATE_ATTEMPTS", requested_count * 6) or requested_count * 6),
    )
    pool = eligible[:pool_limit]
    selected: List[Dict[str, Any]] = []
    rng = random.SystemRandom()
    while pool and len(selected) < attempt_limit:
        weights = [1.0 / (1.0 + float(item.get("chart_position", 0)) * 0.08) for item in pool]
        chosen = rng.choices(pool, weights=weights, k=1)[0]
        selected.append(chosen)
        pool.remove(chosen)
    return selected


async def _send_digest_candidate(
    application: Application,
    subscription: Dict[str, Any],
    candidate: Dict[str, Any],
    blocked_history_keys: Optional[Set[str]] = None,
) -> Optional[Dict[str, Any]]:
    """Send one candidate from cache or download it and populate the cache."""
    chat_id = int(subscription["chat_id"])
    thread_id = int(subscription.get("thread_id") or 0)
    subscription_id = int(subscription["id"])
    synthetic_user_id = -1_000_000_000 - subscription_id

    async def send_audio_value(audio_value: Any, artist: str, title: str) -> Any:
        kwargs: Dict[str, Any] = {
            "chat_id": chat_id,
            "audio": audio_value,
            "title": str(title or candidate.get("title") or "")[:64],
            "performer": str(artist or candidate.get("artist") or "")[:64],
        }
        caption = getattr(config, "TRACK_CREDIT_CAPTION", None)
        if caption:
            kwargs["caption"] = caption
            kwargs["parse_mode"] = "HTML"
        if thread_id:
            kwargs["message_thread_id"] = thread_id
        return await safe_send_audio(application.bot, **kwargs)

    actual_track: Optional[Dict[str, Any]] = None
    sent_message = None
    used_cache = False
    original_file_id = str(candidate.get("file_id") or "")
    blocked = blocked_history_keys or set()
    initial_keys = set(candidate.get("history_keys") or [])
    if blocked.intersection(initial_keys):
        return None

    if original_file_id:
        try:
            sent_message = await send_audio_value(original_file_id, candidate.get("artist", ""), candidate.get("title", ""))
            used_cache = True
        except BadRequest as exc:
            if not _digest_invalid_file_id_error(exc):
                raise
            await bot_instance.digest_store.remove_cached_audio(candidate.get("query_key") or "")
            candidate["file_id"] = ""

    if sent_message is None:
        actual_track = await _resolve_digest_source_track(candidate)
        if not actual_track:
            return None

        actual_identity_keys = set(_digest_track_keys(
            str(actual_track.get("artist") or ""),
            str(actual_track.get("title") or ""),
        ))
        if blocked.intersection(actual_identity_keys):
            return None

        # A source-specific metadata variant may already be cached even when the
        # Deezer artist/title alias was not found in the digest index.
        actual_cached_id = await bot_instance.get_tg_file_id(
            str(actual_track.get("artist") or ""),
            str(actual_track.get("title") or ""),
            actual_track.get("vk_key"),
        )
        if actual_cached_id:
            actual_file_key = _digest_file_history_key(actual_cached_id)
            if actual_file_key and actual_file_key in blocked:
                return None
            try:
                sent_message = await send_audio_value(
                    actual_cached_id,
                    str(actual_track.get("artist") or candidate.get("artist") or ""),
                    str(actual_track.get("title") or candidate.get("title") or ""),
                )
                used_cache = True
            except BadRequest as exc:
                if not _digest_invalid_file_id_error(exc):
                    raise
                await bot_instance.delete_tg_file_id(
                    str(actual_track.get("artist") or ""),
                    str(actual_track.get("title") or ""),
                    actual_track.get("vk_key"),
                )
                sent_message = None

        if sent_message is None:
            await bot_instance.audio_memory_semaphore.acquire()
            try:
                try:
                    audio_data = await bot_instance.download_audio_any_source(
                        actual_track,
                        synthetic_user_id,
                        fast=True,
                    )
                except Exception as exc:
                    logger.warning(
                        "Digest download failed for %s - %s: %s",
                        candidate.get("artist"), candidate.get("title"), exc,
                    )
                    return None
                if not audio_data:
                    return None

                audio_file = BytesIO(audio_data)
                del audio_data
                audio_file.name = bot_instance.create_safe_filename(
                    str(actual_track.get("artist") or candidate.get("artist") or ""),
                    str(actual_track.get("title") or candidate.get("title") or ""),
                    str(actual_track.get("audio_ext") or "mp3"),
                )
                try:
                    sent_message = await send_audio_value(
                        audio_file,
                        str(actual_track.get("artist") or candidate.get("artist") or ""),
                        str(actual_track.get("title") or candidate.get("title") or ""),
                    )
                finally:
                    audio_file.close()
            finally:
                bot_instance.audio_memory_semaphore.release()

    if sent_message is None:
        return None

    sent_audio = getattr(sent_message, "audio", None)
    final_file_id = str(getattr(sent_audio, "file_id", None) or candidate.get("file_id") or "")
    unique_id = getattr(sent_audio, "file_unique_id", None)
    duration = int(getattr(sent_audio, "duration", None) or candidate.get("duration") or (actual_track or {}).get("duration") or 0)

    # Save both the provider metadata and the Deezer alias. Future digests can
    # therefore use the Telegram file_id immediately without another search.
    if final_file_id:
        if actual_track:
            await bot_instance.set_tg_file_id(
                str(actual_track.get("artist") or candidate.get("artist") or ""),
                str(actual_track.get("title") or candidate.get("title") or ""),
                final_file_id,
                unique_id,
                duration,
                actual_track.get("vk_key"),
            )
        await bot_instance.set_tg_file_id(
            str(candidate.get("artist") or ""),
            str(candidate.get("title") or ""),
            final_file_id,
            unique_id,
            duration,
            None,
        )
        # Refresh the digest's in-memory DB index after adding a new alias.
        try:
            bot_instance.digest_store._audio_index_ts = 0.0
        except Exception:
            pass

    history_keys = list(candidate.get("history_keys") or [])
    if actual_track:
        history_keys.extend(_digest_track_keys(
            str(actual_track.get("artist") or ""),
            str(actual_track.get("title") or ""),
        ))
    file_history_key = _digest_file_history_key(final_file_id)
    if file_history_key:
        history_keys.append(file_history_key)

    return {
        "history_keys": list(dict.fromkeys(key for key in history_keys if key)),
        "used_cache": used_cache,
        "downloaded": not used_cache,
    }


async def send_group_digest(application: Application, subscription: Dict[str, Any], manual: bool = False) -> Dict[str, Any]:
    subscription_id = int(subscription["id"])
    async with _digest_send_lock(subscription_id):
        candidates = await _select_digest_candidates(subscription)
        if not candidates:
            days = max(1, int(subscription.get("no_repeat_days") or _digest_no_repeat_days()))
            return {
                "ok": False,
                "sent": 0,
                "notified": False,
                "reason": (
                    f"В актуальном чарте Deezer нет новых композиций: треки, опубликованные "
                    f"за последние {days} дн., не повторяются."
                ),
            }

        chat_id = int(subscription["chat_id"])
        thread_id = int(subscription.get("thread_id") or 0)
        requested_count = max(1, int(subscription.get("track_count", 5) or 5))
        preparing_text = (
            f"🎧 <b>{_esc(subscription.get('genre_name') or 'Музыкальная подборка')}</b>\n\n"
            f"Формирую до <b>{requested_count}</b> популярных треков по данным Deezer. "
            "Отсутствующие в кэше композиции бот скачает автоматически."
        )
        header_kwargs: Dict[str, Any] = {"chat_id": chat_id, "text": preparing_text, "parse_mode": "HTML"}
        if thread_id:
            header_kwargs["message_thread_id"] = thread_id

        try:
            header_message = await safe_send_message(application.bot, **header_kwargs)
        except TelegramError as exc:
            if _digest_fatal_destination_error(exc):
                await bot_instance.digest_store.set_enabled(subscription_id, chat_id, False)
            raise

        sent = 0
        cached_count = 0
        downloaded_count = 0
        failed_count = 0
        sent_this_run: Set[str] = set()
        count_manual = bool(getattr(config, "DIGEST_TEST_COUNTS_AS_SENT", True))

        concurrency = max(
            1,
            min(4, int(getattr(config, "DIGEST_DOWNLOAD_CONCURRENCY", 3) or 3)),
        )
        cursor = 0

        # Process a few tracks in parallel. Each batch is capped by the number
        # of still-needed tracks, so the digest never sends more than requested.
        while sent < requested_count and cursor < len(candidates):
            remaining = requested_count - sent
            batch_size = min(concurrency, remaining, len(candidates) - cursor)
            batch = candidates[cursor:cursor + batch_size]
            cursor += batch_size

            batch_started = time.monotonic()
            batch_results = await asyncio.gather(
                *[
                    _send_digest_candidate(
                        application,
                        subscription,
                        candidate,
                        blocked_history_keys=set(sent_this_run),
                    )
                    for candidate in batch
                ],
                return_exceptions=True,
            )

            for candidate, result in zip(batch, batch_results):
                if isinstance(result, TelegramError):
                    if _digest_fatal_destination_error(result):
                        await bot_instance.digest_store.set_enabled(subscription_id, chat_id, False)
                        raise result
                    failed_count += 1
                    logger.warning(
                        "Digest Telegram send failed for %s - %s: %s",
                        candidate.get("artist"), candidate.get("title"), result,
                    )
                    continue
                if isinstance(result, Exception):
                    failed_count += 1
                    logger.warning(
                        "Digest candidate failed for %s - %s: %s",
                        candidate.get("artist"), candidate.get("title"), result,
                    )
                    continue
                if not result:
                    failed_count += 1
                    continue

                sent += 1
                cached_count += 1 if result.get("used_cache") else 0
                downloaded_count += 1 if result.get("downloaded") else 0
                result_history_keys = list(
                    result.get("history_keys")
                    or candidate.get("history_keys")
                    or [candidate.get("track_key")]
                )
                sent_this_run.update(key for key in result_history_keys if key)
                if not manual or count_manual:
                    await bot_instance.digest_store.record_sent_many(
                        subscription_id,
                        result_history_keys,
                    )

            logger.info(
                "Digest #%s batch completed in %.1fs: sent=%s/%s failed=%s",
                subscription_id,
                time.monotonic() - batch_started,
                sent,
                requested_count,
                failed_count,
            )

            # A visible progress update avoids the impression that the bot has
            # stopped while provider downloads are still being prepared.
            if sent < requested_count and cursor < len(candidates):
                progress_text = (
                    f"🎧 <b>{_esc(subscription.get('genre_name') or 'Музыкальная подборка')}</b>\n\n"
                    f"Отправлено <b>{sent}</b> из {requested_count}. "
                    f"Из кэша: <b>{cached_count}</b> · скачано: <b>{downloaded_count}</b>.\n"
                    "Подготавливаю следующие композиции…"
                )
                await safe_edit_text(header_message, progress_text, parse_mode="HTML")

            if sent < requested_count and cursor < len(candidates):
                await asyncio.sleep(float(getattr(config, "DIGEST_BATCH_DELAY", 0.25) or 0.25))

        if sent:
            shortage = ""
            if sent < requested_count:
                shortage = (
                    f"\nОтправлено <b>{sent}</b> из {requested_count}: для остальных позиций "
                    "не удалось найти или скачать точное совпадение, а старые треки не повторяются."
                )
            final_text = (
                f"🎧 <b>{_esc(subscription.get('genre_name') or 'Музыкальная подборка')}</b>\n\n"
                f"Популярные треки по данным Deezer. Отправлено: <b>{sent}</b>.\n"
                f"Из Telegram-кэша: <b>{cached_count}</b> · скачано сейчас: <b>{downloaded_count}</b>."
                f"{shortage}"
            )
            await safe_edit_text(header_message, final_text, parse_mode="HTML")
            return {
                "ok": True,
                "sent": sent,
                "cached": cached_count,
                "downloaded": downloaded_count,
                "failed": failed_count,
                "notified": True,
                "reason": "",
            }

        reason = (
            "Не удалось найти и скачать точные версии новых треков в доступных источниках "
            "VK, Яндекс.Музыка и YouTube."
        )
        await safe_edit_text(
            header_message,
            f"⚠️ <b>Подборка не отправлена</b>\n\n{_esc(reason)}",
            parse_mode="HTML",
        )
        return {"ok": False, "sent": 0, "notified": True, "reason": reason}


async def group_digest_scheduler(application: Application) -> None:
    """Minute-level scheduler for enabled group digests."""
    await asyncio.sleep(10)
    while True:
        try:
            due = await bot_instance.digest_store.get_due_subscriptions(time.time(), limit=20)
            for subscription in due:
                try:
                    result = await send_group_digest(application, subscription, manual=False)
                    logger.info(
                        "Group digest #%s completed: sent=%s reason=%s",
                        subscription.get("id"), result.get("sent"), result.get("reason"),
                    )
                except asyncio.CancelledError:
                    raise
                except Exception as exc:
                    logger.error("Group digest #%s failed: %s", subscription.get("id"), exc)
                finally:
                    await bot_instance.digest_store.mark_run(
                        int(subscription["id"]), str(subscription.get("schedule_key") or "d18")
                    )
        except asyncio.CancelledError:
            return
        except Exception:
            logger.exception("Group digest scheduler iteration failed")
        await asyncio.sleep(60)


async def handle_digest_callback(update: Update, context: CallbackContext) -> None:
    query = update.callback_query
    if not query or not query.message:
        return
    if not await _is_group_admin(update, context, notify=False):
        await safe_answer_callback(query, "Только администратор этой группы может менять рассылки.", show_alert=True)
        return

    chat_id = int(query.message.chat.id)
    user_id = int(query.from_user.id)
    action = str(query.data or "")
    key = _digest_wizard_key(user_id, chat_id)
    state = _DIGEST_WIZARDS.get(key)

    if action == "digest:home":
        _DIGEST_WIZARDS.pop(key, None)
        await _render_digest_home(query, context, edit=True)
        return
    if action == "digest:close":
        _DIGEST_WIZARDS.pop(key, None)
        await safe_answer_callback(query)
        await safe_edit_message_text(query, "✅ Настройка рассылок закрыта.")
        return
    if action == "digest:list":
        _DIGEST_WIZARDS.pop(key, None)
        await safe_answer_callback(query)
        await _render_digest_list(query, context)
        return
    if action.startswith("digest:edit_schedule:"):
        try:
            subscription_id = int(action.rsplit(":", 1)[1])
        except Exception:
            subscription_id = 0
        subscription = await bot_instance.digest_store.get_subscription(subscription_id, chat_id)
        if not subscription:
            await safe_answer_callback(query, "Рассылка не найдена.", show_alert=True)
            return
        days, _, _ = _digest_parse_schedule(str(subscription.get("schedule_key") or "d18"))
        current_thread = int(getattr(query.message, "message_thread_id", 0) or 0)
        state = {
            "chat_id": chat_id,
            "user_id": user_id,
            "wizard_thread_id": current_thread,
            "selected_days": set(days),
            "awaiting_time": False,
            "edit_subscription_id": subscription_id,
            "created_at": time.time(),
        }
        _DIGEST_WIZARDS[key] = state
        await safe_answer_callback(query)
        await _render_digest_schedules(query, state)
        return

    if action == "digest:new":
        await safe_answer_callback(query)
        current_thread = int(getattr(query.message, "message_thread_id", 0) or 0)
        is_forum = await _digest_chat_is_forum(context, chat_id, query.message.chat)
        state = {
            "chat_id": chat_id, "user_id": user_id, "current_thread_id": current_thread,
            "wizard_thread_id": current_thread, "thread_id": 0, "is_forum": is_forum,
            "selected_days": set(range(7)), "awaiting_time": False, "created_at": time.time(),
        }
        _DIGEST_WIZARDS[key] = state
        if current_thread:
            await bot_instance.digest_store.remember_topic(chat_id, current_thread, None)
        if is_forum:
            await _render_digest_destinations(query, context, state)
        else:
            await _render_digest_genres(query)
        return
    if action == "digest:genres":
        await safe_answer_callback(query)
        await _render_digest_genres(query)
        return
    if action == "digest:schedules":
        if not state:
            await safe_answer_callback(query, "Настройка устарела. Начните заново.", show_alert=True)
            return
        state["awaiting_time"] = False
        await safe_answer_callback(query)
        await _render_digest_schedules(query, state)
        return

    if action.startswith("digest:dest:"):
        if not state:
            await safe_answer_callback(query, "Настройка устарела. Начните заново.", show_alert=True)
            return
        try:
            state["thread_id"] = max(0, int(action.rsplit(":", 1)[1]))
        except Exception:
            await safe_answer_callback(query, "Некорректный раздел.", show_alert=True)
            return
        await safe_answer_callback(query)
        await _render_digest_genres(query)
        return

    if action.startswith("digest:genre:"):
        if not state:
            await safe_answer_callback(query, "Настройка устарела. Начните заново.", show_alert=True)
            return
        try:
            genre_id = int(action.rsplit(":", 1)[1])
        except Exception:
            genre_id = -1
        if genre_id not in DEEZER_GENRES:
            await safe_answer_callback(query, "Неизвестный жанр.", show_alert=True)
            return
        state["genre_id"] = genre_id
        state["genre_name"] = DEEZER_GENRES[genre_id]
        await safe_answer_callback(query)
        await _render_digest_schedules(query, state)
        return

    if action.startswith("digest:day:"):
        if not state:
            await safe_answer_callback(query, "Настройка устарела. Начните заново.", show_alert=True)
            return
        try:
            day = int(action.rsplit(":", 1)[1])
        except Exception:
            day = -1
        if day < 0 or day > 6:
            await safe_answer_callback(query, "Некорректный день.", show_alert=True)
            return
        selected = {int(value) for value in state.get("selected_days", set())}
        if day in selected:
            selected.remove(day)
        else:
            selected.add(day)
        state["selected_days"] = selected
        await safe_answer_callback(query)
        await _render_digest_schedules(query, state)
        return

    if action.startswith("digest:days:"):
        if not state:
            await safe_answer_callback(query, "Настройка устарела. Начните заново.", show_alert=True)
            return
        mode = action.rsplit(":", 1)[1]
        if mode == "all":
            state["selected_days"] = set(range(7))
            await safe_answer_callback(query)
            await _render_digest_schedules(query, state)
            return
        if mode == "weekdays":
            state["selected_days"] = set(range(5))
            await safe_answer_callback(query)
            await _render_digest_schedules(query, state)
            return
        if mode == "weekend":
            state["selected_days"] = {5, 6}
            await safe_answer_callback(query)
            await _render_digest_schedules(query, state)
            return
        if mode == "done":
            selected = {int(value) for value in state.get("selected_days", set())}
            if not selected:
                await safe_answer_callback(query, "Выберите хотя бы один день.", show_alert=True)
                return
            state["awaiting_time"] = True
            state["wizard_message_id"] = int(query.message.message_id)
            await safe_answer_callback(query)
            await safe_edit_message_text(
                query,
                "🕒 <b>Введите время публикации</b>\n\n"
                "Отправьте время сообщением в формате <code>ЧЧ:ММ</code>, "
                "например <code>08:30</code> или <code>21:05</code>.\n\n"
                f"Часовой пояс: <code>{_esc(str(getattr(config, 'DIGEST_TIMEZONE', 'Europe/Amsterdam')))}</code>",
                parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔙 Изменить дни", callback_data="digest:schedules")
                ]]),
            )
            return

    if action.startswith("digest:count:"):
        if not state or "genre_id" not in state or "schedule_key" not in state:
            await safe_answer_callback(query, "Настройка устарела. Начните заново.", show_alert=True)
            return
        try:
            track_count = int(action.rsplit(":", 1)[1])
        except Exception:
            track_count = 5
        existing = await bot_instance.digest_store.list_subscriptions(chat_id)
        max_subscriptions = int(getattr(config, "DIGEST_MAX_PER_GROUP", 12) or 12)
        if len(existing) >= max_subscriptions:
            await safe_answer_callback(query, f"В группе уже достигнут лимит: {max_subscriptions} рассылок.", show_alert=True)
            return
        subscription_id = await bot_instance.digest_store.create_subscription(
            chat_id=chat_id,
            thread_id=int(state.get("thread_id") or 0),
            genre_id=int(state["genre_id"]),
            genre_name=str(state["genre_name"]),
            schedule_key=str(state["schedule_key"]),
            track_count=track_count,
            created_by=user_id,
        )
        _DIGEST_WIZARDS.pop(key, None)
        await safe_answer_callback(query, "Рассылка создана")
        await safe_edit_message_text(
            query,
            f"✅ <b>Рассылка #{subscription_id} создана</b>\n\n"
            f"Жанр: <b>{_esc(state['genre_name'])}</b>\n"
            f"Расписание: <b>{_esc(_digest_schedule_label(state['schedule_key']))}</b>\n"
            f"Треков: <b>{track_count}</b>\n"
            f"Куда: <b>{_esc(_digest_destination_label(int(state.get('thread_id') or 0)))}</b>",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("▶️ Проверить сейчас", callback_data=f"digest:test:{subscription_id}")],
                [InlineKeyboardButton("📋 Все рассылки", callback_data="digest:list")],
            ]),
        )
        return

    if action.startswith("digest:test:"):
        try:
            subscription_id = int(action.rsplit(":", 1)[1])
        except Exception:
            subscription_id = 0
        subscription = await bot_instance.digest_store.get_subscription(subscription_id, chat_id)
        if not subscription:
            await safe_answer_callback(query, "Рассылка не найдена.", show_alert=True)
            return
        await safe_answer_callback(query, "Формирую тестовую подборку…")
        result = await send_group_digest(context.application, subscription, manual=True)
        if not result.get("ok") and not result.get("notified"):
            kwargs: Dict[str, Any] = {
                "chat_id": chat_id,
                "text": "⚠️ Тестовая подборка не отправлена: " + str(result.get("reason") or "нет подходящих треков"),
            }
            thread_id = int(getattr(query.message, "message_thread_id", 0) or 0)
            if thread_id:
                kwargs["message_thread_id"] = thread_id
            await safe_send_message(context.bot, **kwargs)
        return

    if action.startswith("digest:toggle:"):
        try:
            subscription_id = int(action.rsplit(":", 1)[1])
        except Exception:
            subscription_id = 0
        enabled = await bot_instance.digest_store.toggle_enabled(subscription_id, chat_id)
        if enabled is None:
            await safe_answer_callback(query, "Рассылка не найдена.", show_alert=True)
            return
        await safe_answer_callback(query, "Рассылка включена" if enabled else "Рассылка приостановлена")
        await _render_digest_list(query, context)
        return

    if action.startswith("digest:delete:"):
        try:
            subscription_id = int(action.rsplit(":", 1)[1])
        except Exception:
            subscription_id = 0
        deleted = await bot_instance.digest_store.delete_subscription(subscription_id, chat_id)
        await safe_answer_callback(query, "Рассылка удалена" if deleted else "Рассылка не найдена")
        await _render_digest_list(query, context)
        return

    await safe_answer_callback(query, "Неизвестное действие рассылки.", show_alert=True)

# ==================== ОСНОВНЫЕ КОМАНДЫ ====================
async def start(update: Update, context: CallbackContext) -> None:
    """Обработчик команды /start с информацией о Яндекс.Музыке"""
    if not await user_check(update):
        return

    user = update.effective_user

    # Deep link from a downloaded audio: /start fav_<track_uid>.
    # The track is resolved from the user's download history, so no track metadata
    # or secrets are exposed in the URL.
    start_payload = ""
    try:
        if getattr(context, "args", None):
            start_payload = str(context.args[0] or "").strip()
    except Exception:
        start_payload = ""

    # Preferred path: Telegram supplies the deep-link payload.
    # Fallback path: some Telegram desktop/client flows open the bot and emit
    # plain /start without preserving the payload. Consume the short-lived
    # pending favorite created when this user received the downloaded audio.
    pending_track = None
    if getattr(bot_instance, "user_store", None):
        if start_payload.startswith(("fav_", "unfav_")):
            uid = start_payload[4:].strip()
            if uid:
                if not uid.startswith("h:"):
                    uid = "h:" + uid
                pending_track = await bot_instance.user_store.get_history_track(user.id, uid)
                # A payload was received successfully, so discard any fallback
                # record for the same user to avoid a later accidental /start.
                if pending_track:
                    await bot_instance.user_store.clear_pending_favorite(user.id)
        elif not start_payload:
            pending_track = await bot_instance.user_store.get_pending_favorite(user.id, max_age_seconds=300)

    if pending_track and getattr(bot_instance, "user_store", None):
        uid = _track_uid_from_any(pending_track)
        if uid:
            is_unfavorite = start_payload.startswith("unfav_")
            is_favorite = await bot_instance.user_store.is_favorite(user.id, uid)
            if is_unfavorite:
                if is_favorite:
                    await bot_instance.user_store.remove_favorite(user.id, uid)
                text = "💔 <b>Убрано из избранного</b>\n\n" + _esc(f"{pending_track.get('artist','')} — {pending_track.get('title','')}")
            else:
                if not is_favorite:
                    await bot_instance.user_store.add_favorite(user.id, uid, pending_track)
                    text = "❤️ <b>Добавлено в избранное</b>\n\n" + _esc(f"{pending_track.get('artist','')} — {pending_track.get('title','')}")
                else:
                    text = "❤️ <b>Трек уже в избранном</b>\n\n" + _esc(f"{pending_track.get('artist','')} — {pending_track.get('title','')}")
            await update.message.reply_text(text, parse_mode="HTML", reply_markup=private_main_keyboard())
            return


    # Регистрируем чат/пользователя (для статистики/экспорта/рассылки)
    try:
        if hasattr(bot_instance, "admin_db") and bot_instance.admin_db:
            await bot_instance.admin_db.track_chat(update, bot=context.bot)
            await bot_instance.admin_db.track_user(user)
            try:
                await bot_instance.admin_db.log_user_action(
                    user.id,
                    update.effective_chat.id if update.effective_chat else 0,
                    "start",
                    {"chat_type": str(update.effective_chat.type) if update.effective_chat else ""}
                )
            except Exception:
                pass
    except Exception:
        pass

    chat_type = update.effective_chat.type
    
    await bot_instance.clear_user_state(user.id)
    
    sources = []
    if bot_instance.yandex_music._initialized:
        sources.append("🎵 Яндекс.Музыка")
    if bot_instance and bot_instance.vk_enabled():
        sources.append("🎶 VK")
    if (
        bot_instance
        and getattr(config, "ENABLE_YOUTUBE_MUSIC", True)
        and bot_instance.youtube_music._initialized
    ):
        sources.append("▶️ YouTube")

    sources_text = " + ".join(sources) if sources else "недоступны"
    
    if chat_type in ["group", "supergroup"]:
        bot_username = await get_bot_username(context)
        welcome_text = f"""🎧 <b>Музыкальный бот добавлен в чат!</b>

<b>Доступные источники:</b> {sources_text}

<b>Как использовать:</b>
• Напишите <code>@{bot_username} название трека</code>
• Или <code>/search название трека</code>
• <code>/mix</code> — произвольная подборка с выбором жанра

<b>Бот покажет топ-{config.GROUP_CHAT_RESULTS_COUNT} треков с приоритетом исполнителя!</b>"""
        
        keyboard = [[InlineKeyboardButton("🔍 Начать поиск", url=f"https://t.me/{bot_username}?start=search")]]
        
        await update.message.reply_text(
            welcome_text, 
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        welcome_text = f"""🎧 <b>Добро пожаловать в VLMB!</b>

Привет, {user.first_name}! Я помогу найти и скачать музыку.

<b>Доступные источники:</b> {sources_text}

<b>Основные возможности:</b>
• 🔍 <b>Поиск музыки</b> - найдём в Яндекс.Музыке, VK и YouTube
• 🎤 <b>Приоритет исполнителя</b> - точное совпадение исполнителя показывается первым
• 🎲 <b>Подборка по жанру</b> - произвольный набор популярных треков
• 📊 <b>Топ-чарты</b> - популярные треки со всего мира
• 👥 <b>Похожие исполнители</b> - откройте для себя новую музыку

Просто отправьте название трека или исполнителя 👇"""
        
        keyboard = [
            [InlineKeyboardButton("🔍 Поиск музыки", callback_data="new_search")],
            [InlineKeyboardButton("📊 Чарты", callback_data="charts_menu")],
            [InlineKeyboardButton("👥 Похожие исполнители", callback_data="similar_menu")],
            [InlineKeyboardButton("🆘 Помощь", callback_data="help_menu")]
        ]
        
        await update.message.reply_text(
            welcome_text, 
            parse_mode='HTML',
            reply_markup=private_main_keyboard()
        )

async def search_command(update: Update, context: CallbackContext) -> None:
    """Handle /search with any non-empty query."""
    if not await user_check(update):
        return

    try:
        if getattr(bot_instance, "admin_db", None):
            await bot_instance.admin_db.track_chat(update, bot=context.bot)
            await bot_instance.admin_db.track_user(update.effective_user)
    except Exception:
        pass

    query = " ".join(context.args).strip()
    if not query:
        await update.effective_message.reply_text(
            "❌ <b>Укажите запрос для поиска</b>", parse_mode="HTML"
        )
        return
    await process_search(update, context, query)

async def help_command(update: Update, context: CallbackContext) -> None:
    """Показать краткую справку."""
    if not await user_check(update):
        return
    try:
        if getattr(bot_instance, "admin_db", None):
            await bot_instance.admin_db.track_chat(update, bot=context.bot)
            await bot_instance.admin_db.track_user(update.effective_user)
    except Exception:
        pass

    help_text = """
🎧 <b>VLMB Music Bot — помощь</b>

<b>Команды:</b>
/search [запрос] — поиск музыки
/history — история загрузок
/favorites — избранные треки
/favorite — добавить/убрать последний трек из избранного
/repeat — повторить последний трек
/settings — настройки источника и качества
/mix — произвольная подборка с выбором жанра
/digest — настроить подборки в группе (для администраторов)
/help — эта справка

<b>Как искать:</b>
• в личке просто отправьте исполнителя или название;
• в группе: <code>@бот запрос</code> или <code>Найти запрос</code>;
• в теме форума бот отвечает и отправляет аудио в ту же тему.

Точное совпадение по исполнителю показывается первым. Другие фильтры качества и релевантности не применяются.
"""
    await update.effective_message.reply_text(help_text, parse_mode="HTML")



async def playlist_command(update: Update, context: CallbackContext) -> None:
    """Extract a YouTube playlist/album URL and expose batch download actions."""
    if not await user_check(update):
        return
    url = " ".join(context.args).strip()
    if not url:
        await user_mix_command(update, context)
        return
    if not getattr(bot_instance, "playlist_manager", None):
        await update.effective_message.reply_text("📚 Модуль playlist временно недоступен.")
        return
    try:
        rows = await bot_instance.playlist_manager.extract(url, getattr(config, "MAX_PLAYLIST_TRACKS", 100))
    except Exception as exc:
        logger.warning("Playlist extraction failed: %s", exc)
        await update.effective_message.reply_text("❌ Не удалось прочитать playlist/album. Проверьте ссылку и доступность списка.")
        return
    if not rows:
        await update.effective_message.reply_text("❌ В playlist/album не найдено доступных треков.")
        return
    sid = await bot_instance.session_manager.create_session(update.effective_user.id, url, rows)
    sess = await bot_instance.session_manager.get_session(sid)
    if sess:
        sess.chat_id = update.effective_chat.id
        sess.original_message_id = update.effective_message.message_id
        sess.search_message_id = update.effective_message.message_id
    lines = [f"📚 <b>Playlist / Album</b>\n", f"Треков: <b>{len(rows)}</b>\n"]
    for i, tr in enumerate(rows[:20], 1):
        lines.append(f"{i}. {_esc(tr.get('artist',''))} — {_esc(tr.get('title',''))}")
    if len(rows) > 20:
        lines.append(f"… ещё {len(rows)-20}")
    kb = [
        [InlineKeyboardButton(f"⬇️ Скачать всё ({len(rows)})", callback_data=f"plall:{sid}")],
    ]
    for i, tr in enumerate(rows[:15]):
        uid = _track_uid_from_any(tr)
        kb.append([InlineKeyboardButton(f"{i+1}. {_source_badge('yt')} {str(tr.get('artist',''))[:18]} — {str(tr.get('title',''))[:22]}", callback_data=f"pldl:{sid}:{uid}")])
    await update.effective_message.reply_text("\n".join(lines), parse_mode="HTML", reply_markup=InlineKeyboardMarkup(kb), disable_web_page_preview=True)


async def settings_command(update: Update, context: CallbackContext) -> None:
    """User preferences for provider and bitrate."""
    if not await user_check(update):
        return
    prefs = await bot_instance.user_store.get_preferences(update.effective_user.id) if bot_instance.user_store else {}
    source = prefs.get("prefer_source") or "auto"
    bitrate = prefs.get("prefer_bitrate_kbps") or getattr(config, "YM_PREFERRED_MAX_BITRATE_KBPS", 192)
    text = (
        "⚙️ <b>Настройки</b>\n\n"
        f"🎵 Источник: <b>{_esc(str(source))}</b>\n"
        f"🎚 Качество: <b>{int(bitrate)} kbps</b>\n\n"
        "Настройки применяются к новым поискам и загрузкам."
    )
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🎵 Авто", callback_data="user_settings:source:auto"), InlineKeyboardButton("🎶 VK", callback_data="user_settings:source:vk")],
        [InlineKeyboardButton("🎧 Yandex", callback_data="user_settings:source:ym"), InlineKeyboardButton("▶️ YouTube", callback_data="user_settings:source:yt")],
        [InlineKeyboardButton("128 kbps", callback_data="user_settings:bitrate:128"), InlineKeyboardButton("192 kbps", callback_data="user_settings:bitrate:192"), InlineKeyboardButton("320 kbps", callback_data="user_settings:bitrate:320")],
    ])
    await update.effective_message.reply_text(text, parse_mode="HTML", reply_markup=kb)


async def history_command(update: Update, context: CallbackContext) -> None:
    """Показать историю загрузок пользователя."""
    if not await user_check(update):
        return
    user_id = update.effective_user.id
    if not getattr(bot_instance, 'user_store', None):
        await update.message.reply_text("❌ История недоступна")
        return
    hist = await bot_instance.user_store.list_history(user_id, limit=int(getattr(config, 'HISTORY_LIMIT', 20) or 20))
    if not hist:
        await update.message.reply_text("🕘 История пока пустая. Скачайте любой трек — и он появится здесь.")
        return
    rows = []
    btn_max = int(getattr(config, "BUTTON_TEXT_MAX_LENGTH", 48) or 48)
    def _t(text: str) -> str:
        text = (text or "").strip()
        return text if len(text) <= btn_max else text[:btn_max-1] + "…"
    for tr in hist[:15]:
        uid = _track_uid_from_any(tr) or tr.get('uid')
        src = (tr.get('source') or 'vk')
        badge = _source_badge(src)
        label = _t(f"{_track_duration_text(tr)} ⬇️ [{badge}] {tr.get('artist','')} — {tr.get('title','')}")
        rows.append([InlineKeyboardButton(label, callback_data=f"histdl:{uid}")])
    await update.message.reply_text("🕘 <b>История</b>", parse_mode='HTML', reply_markup=InlineKeyboardMarkup(rows))


async def favorites_command(update: Update, context: CallbackContext) -> None:
    """Показать избранные треки пользователя."""
    if not await user_check(update):
        return
    user_id = update.effective_user.id
    if not getattr(bot_instance, 'user_store', None):
        await update.message.reply_text("❌ Избранное недоступно")
        return
    favorites = await bot_instance.user_store.list_favorites(
        user_id, limit=int(getattr(config, 'HISTORY_LIMIT', 20) or 20)
    )
    if not favorites:
        await update.message.reply_text(
            "❤️ <b>Избранное пока пустое.</b>\n\n"
            "• Нажмите ❤️ рядом с найденным треком — это самый простой способ добавить его.\n"
            "• Или после скачивания используйте /favorite, чтобы добавить последний трек.",
            parse_mode="HTML",
            reply_markup=private_main_keyboard(),
        )
        return
    rows = []
    btn_max = int(getattr(config, "BUTTON_TEXT_MAX_LENGTH", 48) or 48)
    for tr in favorites[:15]:
        uid = _track_uid_from_any(tr) or tr.get('uid')
        label = f"❤️ {_source_badge(tr.get('source') or 'vk')} {tr.get('artist','')} — {tr.get('title','')}"
        if len(label) > btn_max:
            label = label[:btn_max - 1] + "…"
        rows.append([
            InlineKeyboardButton(label, callback_data=f"favdl:{uid}"),
            InlineKeyboardButton("✕", callback_data=f"favrm:{uid}"),
        ])
    await update.message.reply_text(
        "❤️ <b>Избранное</b>", parse_mode='HTML', reply_markup=InlineKeyboardMarkup(rows)
    )


async def favorite_last_command(update: Update, context: CallbackContext) -> None:
    """Toggle the last downloaded track in favorites."""
    if not await user_check(update):
        return
    user_id = update.effective_user.id
    if not getattr(bot_instance, 'user_store', None):
        await update.message.reply_text("❌ Избранное недоступно")
        return
    last = await bot_instance.user_store.get_last(user_id)
    if not last:
        await update.message.reply_text("❤️ Сначала скачайте трек, который хотите добавить в избранное.")
        return
    uid = _track_uid_from_any(last)
    if not uid:
        await update.message.reply_text("❌ Не удалось определить трек.")
        return
    if await bot_instance.user_store.is_favorite(user_id, uid):
        await bot_instance.user_store.remove_favorite(user_id, uid)
        await update.message.reply_text("💔 Убрано из избранного.")
    else:
        await bot_instance.user_store.add_favorite(user_id, uid, last)
        await update.message.reply_text(f"❤️ Добавлено: <b>{last.get('artist','')} — {last.get('title','')}</b>", parse_mode='HTML')


async def repeat_last_command(update: Update, context: CallbackContext) -> None:
    """Повторить последнее (скачать последний отправленный трек)."""
    if not await user_check(update):
        return
    user_id = update.effective_user.id
    if not getattr(bot_instance, 'user_store', None):
        await update.message.reply_text("❌ Функция недоступна")
        return
    last = await bot_instance.user_store.get_last(user_id)
    if not last:
        await update.message.reply_text("↩️ Пока нечего повторять. Сначала скачайте любой трек.")
        return
    await download_and_send_audio(update, context, update.effective_message, last, update.effective_message.message_id)

async def _extract_artists_from_xlsx(file_path: str, max_artists: int = None) -> List[str]:
    """Читает .xlsx и возвращает список артистов (лист Artists/первая колонка)."""
    loop = asyncio.get_running_loop()

    def _read() -> List[str]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = wb["Artists"] if "Artists" in wb.sheetnames else wb[wb.sheetnames[0]]
            # Попробуем найти колонку 'Artist' по заголовку
            rows = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
            headers = next(rows, None) or ()
            artist_col = 0
            if headers:
                for i, h in enumerate(headers):
                    if h is None:
                        continue
                    if str(h).strip().lower() in ("artist", "исполнитель", "артист"):
                        artist_col = i
                        break
            artists = []
            seen = set()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                val = row[artist_col] if artist_col < len(row) else None
                if not val:
                    continue
                name = str(val).strip()
                if not name:
                    continue
                key = name.lower()
                if key in seen:
                    continue
                seen.add(key)
                artists.append(name)
                if max_artists and len(artists) >= max_artists:
                    break
            return artists
        finally:
            wb.close()

    return await loop.run_in_executor(None, _read)


async def _run_excel_warmup_for_admin(update: Update, context: CallbackContext, artists: List[str]):
    """Прогрев: по каждому артисту ищем и скачиваем пачку треков, чтобы набрать file_id."""
    chat_id = update.effective_chat.id
    user_id = update.effective_user.id

    max_artists = int(getattr(config, "EXCEL_WARMUP_MAX_ARTISTS", 1000000))
    tracks_per_artist = int(getattr(config, "EXCEL_WARMUP_TRACKS_PER_ARTIST", 10))
    min_tracks_to_send = int(getattr(config, "EXCEL_WARMUP_MIN_TRACKS_PER_ARTIST", 1))

    artists = artists[:max_artists]

    bot_instance.excel_warmup_cancel.discard(user_id)

    progress_msg = await safe_send_message(
        context.bot,
        chat_id,
        f"🔥 <b>Прогрев file_id запущен</b>\n"
        f"Артистов: <b>{len(artists)}</b>\n"
        f"Треков/артист: <b>{tracks_per_artist}</b>\n\n"
        f"Чтобы остановить: /admin → ⚙️ Настройки → 🛑 Остановить прогрев",
        parse_mode="HTML",
        disable_web_page_preview=True
    )

    sent_total = 0
    processed = 0
    failed = 0

    for artist in artists:
        if user_id in bot_instance.excel_warmup_cancel:
            break

        processed += 1
        try:
            # Ищем по всем источникам
            results = await bot_instance.search_all_sources(artist, limit=max(30, tracks_per_artist))
            if not results:
                continue

            # Берём первые результаты после сортировки по исполнителю
            try:
                best = await bot_instance.find_best_tracks(results, artist, count=tracks_per_artist)
            except Exception:
                best = results[:tracks_per_artist]

            if len(best) < min_tracks_to_send:
                continue

            for t in best[:tracks_per_artist]:
                if user_id in bot_instance.excel_warmup_cancel:
                    break
                # Важно: отправка аудио = получение Telegram file_id и сохранение в кэш
                try:
                    async with bot_instance._excel_warmup_sem:
                        sent_ok = await download_and_send_audio(
                            update, context, update.effective_message, t, update.effective_message.message_id
                        )
                        # gentle pacing to avoid Telegram/YM/VK throttling
                        delay = float(getattr(config, 'EXCEL_WARMUP_SEND_DELAY', 0.15) or 0)
                        if delay > 0:
                            await asyncio.sleep(delay)
                    if sent_ok:
                        sent_total += 1
                    else:
                        failed += 1
                except Exception as e:
                    logger.warning(f"Warmup send failed for {artist}: {e}")
                    failed += 1

        except Exception as e:
            logger.warning(f"Warmup failed for {artist}: {e}")
            failed += 1

        # апдейт прогресса каждые 5 артистов
        if progress_msg and (processed % 5 == 0):
            try:
                await safe_edit_text(
                    progress_msg,
                    f"🔥 <b>Прогрев file_id</b>\n"
                    f"Готово артистов: <b>{processed}/{len(artists)}</b>\n"
                    f"Отправлено треков: <b>{sent_total}</b>\n"
                    f"Ошибок: <b>{failed}</b>",
                    parse_mode="HTML",
                    disable_web_page_preview=True
                )
            except Exception:
                pass

    stopped = (user_id in bot_instance.excel_warmup_cancel)
    bot_instance.excel_warmup_cancel.discard(user_id)

    final_text = (
        "🛑 <b>Прогрев остановлен</b>\n" if stopped else "✅ <b>Прогрев завершён</b>\n"
    ) + (
        f"Артистов обработано: <b>{processed}</b>\n"
        f"Отправлено треков: <b>{sent_total}</b>\n"
        f"Ошибок: <b>{failed}</b>"
    )

    if progress_msg:
        try:
            await safe_edit_text(progress_msg, final_text, parse_mode="HTML", disable_web_page_preview=True)
        except Exception:
            try:
                await safe_send_message(context.bot, chat_id, final_text, parse_mode="HTML")
            except Exception:
                pass


async def handle_document_upload(update: Update, context: CallbackContext) -> None:
    """Приём Excel для прогрева (только админ, только когда включён режим ожидания)."""
    if not update.message or not update.message.document:
        return

    user_id = update.effective_user.id
    if not _is_admin(user_id):
        return

    state = bot_instance.user_states.get(user_id, {}).get("state")
    # The generic media handler (group=1) stores documents in file_id mode.
    if state == "admin_collect_fileid":
        return

    if state != "await_excel_warmup":
        return

    doc = update.message.document
    filename = (doc.file_name or "").lower()
    if not filename.endswith(".xlsx"):
        await safe_send_message(context.bot, update.effective_chat.id, "❌ Нужен файл .xlsx", parse_mode="HTML")
        return

    await bot_instance.set_user_state(user_id, "excel_warmup_processing", {"file": doc.file_id, "name": doc.file_name})

    # скачиваем файл
    local_path = None
    try:
        tg_file = await context.bot.get_file(doc.file_id)
        tmp_dir = getattr(getattr(bot_instance, "file_manager", None), "temp_dir", None) or tempfile.gettempdir()
        local_path = os.path.join(tmp_dir, f"excel_warmup_{user_id}_{int(time.time())}.xlsx")
        await tg_file.download_to_drive(custom_path=local_path)

        await safe_send_message(context.bot, update.effective_chat.id, "📄 Excel получен. Читаю список артистов…", parse_mode="HTML")
        artists = await _extract_artists_from_xlsx(local_path, max_artists=int(getattr(config, "EXCEL_WARMUP_MAX_ARTISTS", 1000000)))

        if not artists:
            await bot_instance.set_user_state(user_id, "idle", {})
            await safe_send_message(context.bot, update.effective_chat.id, "❌ В файле не найден список артистов (проверь лист Artists и колонку Artist).", parse_mode="HTML")
            return

        await safe_send_message(context.bot, update.effective_chat.id, f"✅ Нашёл артистов: <b>{len(artists)}</b>. Запускаю прогрев…", parse_mode="HTML")

        # Запускаем как отдельную управляемую задачу, чтобы бот не зависал на одном апдейте.
        bot_instance.create_background_task(
            _run_excel_warmup_for_admin(update, context, artists),
            f"excel-warmup:{user_id}",
        )

        await bot_instance.set_user_state(user_id, "idle", {})
    except Exception as e:
        logger.error(f"Excel warmup upload error: {e}", exc_info=True)
        await bot_instance.set_user_state(user_id, "idle", {})
        await safe_send_message(context.bot, update.effective_chat.id, f"❌ Ошибка обработки Excel: {html.escape(str(e))}", parse_mode="HTML")
    finally:
        if local_path:
            try:
                await async_os.remove(local_path)
            except FileNotFoundError:
                pass
            except Exception as exc:
                logger.warning("Failed to remove uploaded Excel temp file %s: %s", local_path, exc)


async def handle_media_upload(update: Update, context: CallbackContext) -> None:
    """Ловим любые медиа от админов в режиме admin_collect_fileid и сохраняем file_id."""
    if not update.message:
        return
    user_id = update.effective_user.id if update.effective_user else 0
    if not _is_admin(user_id):
        return

    state = bot_instance.user_states.get(user_id, {}).get("state")
    if state != "admin_collect_fileid":
        return

    msg = update.message

    file_id = None
    unique_id = None
    file_type = None
    file_name = None
    mime_type = None
    file_size = None

    # For mapping to instant-send cache (artist + title)
    mapped_artist = None
    mapped_title = None
    mapped_duration = None

    def _parse_artist_title(s: str):
        if not s:
            return (None, None)
        s = s.strip()
        # common separators: "Artist - Title" or "Artist — Title"
        for sep in [" - ", " — ", " – ", "-", "—", "–"]:
            if sep in s:
                parts = [p.strip() for p in s.split(sep, 1)]
                if len(parts) == 2 and parts[0] and parts[1]:
                    return parts[0], parts[1]
        return (None, None)

    try:
        if msg.audio:
            file_type = "audio"
            file_id = msg.audio.file_id
            unique_id = getattr(msg.audio, "file_unique_id", None)
            file_name = getattr(msg.audio, "file_name", None)
            mime_type = getattr(msg.audio, "mime_type", None)
            file_size = getattr(msg.audio, "file_size", None)
            mapped_artist = (getattr(msg.audio, "performer", None) or "").strip() or None
            mapped_title = (getattr(msg.audio, "title", None) or "").strip() or None
            mapped_duration = getattr(msg.audio, "duration", None)
            # fallback: caption "Artist - Title"
            if (not mapped_artist or not mapped_title) and getattr(msg, "caption", None):
                a, t = _parse_artist_title(msg.caption)
                mapped_artist = mapped_artist or a
                mapped_title = mapped_title or t
        elif msg.voice:
            file_type = "voice"
            file_id = msg.voice.file_id
            unique_id = getattr(msg.voice, "file_unique_id", None)
            mime_type = getattr(msg.voice, "mime_type", None)
            file_size = getattr(msg.voice, "file_size", None)
        elif msg.video:
            file_type = "video"
            file_id = msg.video.file_id
            unique_id = getattr(msg.video, "file_unique_id", None)
            file_name = getattr(msg.video, "file_name", None)
            mime_type = getattr(msg.video, "mime_type", None)
            file_size = getattr(msg.video, "file_size", None)
        elif msg.video_note:
            file_type = "video_note"
            file_id = msg.video_note.file_id
            unique_id = getattr(msg.video_note, "file_unique_id", None)
            file_size = getattr(msg.video_note, "file_size", None)
        elif msg.animation:
            file_type = "animation"
            file_id = msg.animation.file_id
            unique_id = getattr(msg.animation, "file_unique_id", None)
            file_name = getattr(msg.animation, "file_name", None)
            mime_type = getattr(msg.animation, "mime_type", None)
            file_size = getattr(msg.animation, "file_size", None)
        elif msg.document:
            file_type = "document"
            file_id = msg.document.file_id
            unique_id = getattr(msg.document, "file_unique_id", None)
            file_name = getattr(msg.document, "file_name", None)
            mime_type = getattr(msg.document, "mime_type", None)
            file_size = getattr(msg.document, "file_size", None)
            candidate = getattr(msg, "caption", None) or os.path.splitext(file_name or "")[0]
            mapped_artist, mapped_title = _parse_artist_title(candidate)
        elif msg.sticker:
            file_type = "sticker"
            file_id = msg.sticker.file_id
            unique_id = getattr(msg.sticker, "file_unique_id", None)
            file_size = getattr(msg.sticker, "file_size", None)
        elif msg.photo:
            # берём самый большой размер
            ph = msg.photo[-1]
            file_type = "photo"
            file_id = ph.file_id
            unique_id = getattr(ph, "file_unique_id", None)
            file_size = getattr(ph, "file_size", None)

        if not file_id:
            return

        if hasattr(bot_instance, "admin_db") and bot_instance.admin_db:
            await bot_instance.admin_db.add_forwarded_file_id(
                file_id=file_id,
                file_unique_id=unique_id,
                file_type=file_type,
                file_name=file_name,
                mime_type=mime_type,
                file_size=file_size,
                added_by=user_id,
            )
            mapped_note = ""
            if mapped_artist and mapped_title:
                try:
                    await bot_instance.set_tg_file_id(
                        mapped_artist,
                        mapped_title,
                        file_id,
                        unique_id=unique_id,
                        duration=mapped_duration,
                        vk_key=None,
                    )
                    mapped_note = (
                        f"\nПривязка: <code>{html.escape(mapped_artist)}</code> — "
                        f"<code>{html.escape(mapped_title)}</code>"
                    )
                except Exception as exc:
                    logger.debug(f"Manual file_id mapping failed: {exc}")

            cached_cnt = await bot_instance.admin_db.count_cached_audio_file_ids()
            manual_cnt = await bot_instance.admin_db.count_forwarded_file_ids()
            total_cnt = cached_cnt + manual_cnt

            await safe_send_message(
                context.bot,
                update.effective_chat.id,
                "✅ <b>file_id записан</b>\n"
                f"Тип: <code>{html.escape(str(file_type or 'file'))}</code>\n"
                f"Всего file_id в базе: <b>{total_cnt}</b> "
                f"(кэш: {cached_cnt} + ручные: {manual_cnt}){mapped_note}",
                parse_mode="HTML",
                message_thread_id=getattr(update.effective_message, "message_thread_id", None),
            )
        else:
            await safe_send_message(context.bot, update.effective_chat.id, "❌ AdminDB не инициализирован", parse_mode="HTML")

    except Exception as e:
        logger.error(f"Admin file_id capture error (media): {e}", exc_info=True)
        await safe_send_message(context.bot, update.effective_chat.id, "❌ Не удалось записать file_id", parse_mode="HTML")




async def handle_contact(update: Update, context: CallbackContext) -> None:
    """Сохраняем телефон, если пользователь сам поделился контактом.
    Важно: Telegram не отдаёт телефон автоматически — только через share contact.
    """
    try:
        msg = update.effective_message
        if not msg or not getattr(msg, "contact", None):
            return
        contact = msg.contact
        from_user = update.effective_user
        if not from_user:
            return
        # сохраняем только если это контакт самого пользователя (иначе спам/чужие номера)
        if getattr(contact, "user_id", None) and int(contact.user_id) != int(from_user.id):
            return
        phone = getattr(contact, "phone_number", None) or ""
        phone = str(phone).strip()
        if not phone:
            return
        if hasattr(bot_instance, "admin_db") and bot_instance.admin_db:
            await bot_instance.admin_db.track_user(from_user)
            await bot_instance.admin_db.set_user_phone(from_user.id, phone)
        # Не спамим в чате, просто тихо подтверждаем
        try:
            await msg.reply_text("✅ Телефон сохранён (будет виден только в статистике/экспорте).", quote=True)
        except Exception:
            pass
    except Exception:
        pass
async def handle_digest_time_input(update: Update, context: CallbackContext) -> bool:
    """Consume a custom digest time entered by the group administrator."""
    message = update.effective_message
    chat = update.effective_chat
    user = update.effective_user
    if not message or not chat or not user or not getattr(message, "text", None):
        return False
    if chat.type not in ("group", "supergroup"):
        return False

    key = _digest_wizard_key(user.id, chat.id)
    state = _DIGEST_WIZARDS.get(key)
    if not state or not state.get("awaiting_time"):
        return False
    if time.time() - float(state.get("created_at") or 0) > 1800:
        _DIGEST_WIZARDS.pop(key, None)
        await message.reply_text("⌛ Настройка устарела. Запустите /digest заново.")
        return True

    expected_thread = int(state.get("wizard_thread_id") or 0)
    actual_thread = int(getattr(message, "message_thread_id", 0) or 0)
    if expected_thread != actual_thread:
        await message.reply_text("Введите время в том же разделе, где открыт мастер /digest.")
        return True
    if not await _is_group_admin(update, context, notify=False):
        _DIGEST_WIZARDS.pop(key, None)
        await message.reply_text(
            "Только администратор этой группы может завершить настройку рассылки."
        )
        return True

    match = re.fullmatch(r"\s*([01]?\d|2[0-3]):([0-5]\d)\s*", message.text)
    if not match:
        await message.reply_text(
            "Некорректное время. Используйте формат <code>ЧЧ:ММ</code>, "
            "например <code>09:30</code>.",
            parse_mode="HTML",
        )
        return True

    selected_days = {int(value) for value in state.get("selected_days", set())}
    if not selected_days:
        state["awaiting_time"] = False
        await message.reply_text("Сначала выберите хотя бы один день через /digest.")
        return True

    hour, minute = int(match.group(1)), int(match.group(2))
    state["schedule_key"] = _digest_encode_schedule(selected_days, hour, minute)
    state["awaiting_time"] = False

    edit_subscription_id = int(state.get("edit_subscription_id") or 0)
    if edit_subscription_id:
        updated = await bot_instance.digest_store.update_schedule(
            edit_subscription_id,
            int(chat.id),
            str(state["schedule_key"]),
        )
        _DIGEST_WIZARDS.pop(key, None)
        if not updated:
            await message.reply_text("Рассылка не найдена или уже удалена.")
            return True
        await message.reply_text(
            f"✅ <b>Расписание #{edit_subscription_id} обновлено</b>\n\n"
            f"{_esc(_digest_schedule_label(state['schedule_key']))}\n"
            f"Часовой пояс: <code>{_esc(str(getattr(config, 'DIGEST_TIMEZONE', 'Europe/Amsterdam')))}</code>",
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("📋 Все рассылки", callback_data="digest:list")
            ]]),
        )
        return True

    kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("3 трека", callback_data="digest:count:3"),
            InlineKeyboardButton("5 треков", callback_data="digest:count:5"),
        ],
        [
            InlineKeyboardButton("7 треков", callback_data="digest:count:7"),
            InlineKeyboardButton("10 треков", callback_data="digest:count:10"),
        ],
        [InlineKeyboardButton("🔙 Изменить дни и время", callback_data="digest:schedules")],
    ])
    await message.reply_text(
        "✅ <b>Расписание задано</b>\n\n"
        f"{_esc(_digest_schedule_label(state['schedule_key']))}\n"
        f"Часовой пояс: <code>{_esc(str(getattr(config, 'DIGEST_TIMEZONE', 'Europe/Amsterdam')))}</code>\n\n"
        "Теперь выберите количество треков:",
        parse_mode="HTML",
        reply_markup=kb,
    )
    return True


async def handle_message(update: Update, context: CallbackContext) -> None:
    """Handle text messages in private chats, groups, and forum topics."""
    message = update.effective_message
    if not message or not getattr(message, "text", None):
        return
    message_text = message.text.strip()
    if not message_text:
        return
    chat_type = update.effective_chat.type
    user_id = update.effective_user.id

    if await handle_digest_time_input(update, context):
        return

    try:
        if getattr(bot_instance, "admin_db", None):
            await bot_instance.admin_db.track_chat(update, bot=context.bot)
            await bot_instance.admin_db.track_user(update.effective_user)
            if await bot_instance.user_manager.is_user_banned(user_id):
                return
    except Exception:
        pass

    if _is_admin(user_id):
        state = await bot_instance.get_user_state(user_id)
        if state and state.get("state") == "add_token":
            await handle_token_input(update, context); return
        if state and state.get("state") == "broadcast_text":
            await _handle_broadcast_text(update, context, state); return
        if state and state.get("state") == "broadcast_target_id":
            await _handle_broadcast_target_input(update, context, state); return
        if state and state.get("state") == "ban_user":
            await _handle_ban_input(update, context, state); return
        if state and state.get("state") == "unban_user":
            await _handle_unban_input(update, context, state); return

    bot_instance.mark_user_activity()
    if message_text.startswith("/"):
        return

    # Persistent private-menu buttons always win over transient input states.
    # This prevents the "similar_artists" state from treating "История",
    # "Чарты", "Настройки", etc. as artist names.
    if chat_type == "private":
        menu_action = _private_menu_action(message_text)
        if menu_action:
            # Navigation is intentionally exempt from the short anti-spam cooldown.
            if not await user_check(update, enforce_cooldown=False):
                return
        elif not await user_check(update):
            return

        if menu_action == "search":
            await bot_instance.clear_user_state(user_id)
            await message.reply_text(
                "🔍 <b>Введите название трека или исполнителя:</b>",
                parse_mode="HTML", reply_markup=private_main_keyboard(),
            )
            return
        if menu_action == "mix":
            await bot_instance.clear_user_state(user_id)
            await _open_user_mix(update, context, check_access=False)
            return
        if menu_action == "charts":
            await bot_instance.clear_user_state(user_id)
            await message.reply_text(
                "📊 <b>Выберите тип чартов:</b>", parse_mode="HTML",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("🌍 Глобальные хиты", callback_data="charts:global")],
                    [InlineKeyboardButton("📅 Недельные хиты", callback_data="charts:weekly")],
                    [InlineKeyboardButton("🎸 Рок", callback_data="charts:rock")],
                    [InlineKeyboardButton("🎵 Поп", callback_data="charts:pop")],
                    [InlineKeyboardButton("🎤 Хип-хоп", callback_data="charts:hiphop")],
                    [InlineKeyboardButton("🎧 Электроника", callback_data="charts:electronic")],
                    [InlineKeyboardButton("❌ Закрыть", callback_data="close_search")],
                ]),
            )
            return
        if menu_action == "similar":
            await bot_instance.set_user_state(user_id, "similar_artists")
            await message.reply_text(
                "🎧 <b>Введите имя исполнителя, чтобы найти похожих:</b>\n\n"
                "Например: <i>Miyagi</i>",
                parse_mode="HTML", reply_markup=private_main_keyboard(),
            )
            return
        if menu_action == "favorites":
            await bot_instance.clear_user_state(user_id)
            await favorites_command(update, context)
            return
        if menu_action == "history":
            await bot_instance.clear_user_state(user_id)
            await history_command(update, context)
            return
        if menu_action == "settings":
            await bot_instance.clear_user_state(user_id)
            await settings_command(update, context)
            return
        if menu_action == "help":
            await bot_instance.clear_user_state(user_id)
            await help_command(update, context)
            return

        state = await bot_instance.get_user_state(user_id)
        if state and state.get("state") == "similar_artists":
            await handle_similar_artists_input(update, context)
            return

        await bot_instance.clear_user_state(user_id)
        query = message_text[6:].strip() if message_text.lower().startswith("найти ") else message_text
        username = await get_bot_username(context)
        if username:
            query = query.replace(f"@{username}", "").strip()
        if query:
            await process_search(update, context, query)
        return

    if chat_type in ("group", "supergroup"):
        username = await get_bot_username(context)
        query = None
        if username and f"@{username}" in message_text:
            query = message_text.replace(f"@{username}", "").strip()
        elif message_text.lower().startswith("найти "):
            query = message_text[6:].strip()
        if query and await user_check(update):
            await process_search(update, context, query)

# ==================== КЛАСС ЗАПУСКА БОТА ====================
class BotRunner:
    """Класс для запуска и управления ботом"""
    
    def __init__(self):
        self.application = None
        self._stop_event = asyncio.Event()
        self._shutdown_started = False
    
    async def setup(self):
        """Настройка бота"""
        global bot_instance
        
        bot_instance = AsyncMusicBot()
        # Compatibility across forks: init/initialize/init_session
        if hasattr(bot_instance, 'initialize'):
            await bot_instance.initialize()
        elif hasattr(bot_instance, 'init_session'):
            await bot_instance.init_session()
        elif hasattr(bot_instance, 'init'):
            await bot_instance.init()

        # Admin datastore (stats/settings/bans/broadcast registry)
        bot_instance.admin_db = AdminDB(await _admin_db_path())
        await bot_instance.admin_db.init()
        await bot_instance.admin_db.apply_runtime_settings()
        await bot_instance.user_manager.refresh_bans(bot_instance.admin_db, force=True)
        bot_instance.digest_store = GroupDigestStore(await _admin_db_path())
        await bot_instance.digest_store.init()
        
        self.application = (
            Application.builder()
            .token(config.TELEGRAM_BOT_TOKEN)
            .request(
                HTTPXRequest(
                    connect_timeout=getattr(config, 'TELEGRAM_CONNECT_TIMEOUT', 10),
                    read_timeout=getattr(config, 'TELEGRAM_READ_TIMEOUT', 25),
                    write_timeout=getattr(config, 'TELEGRAM_WRITE_TIMEOUT', 25),
                    pool_timeout=getattr(config, 'TELEGRAM_POOL_TIMEOUT', 10),
                    connection_pool_size=getattr(config, 'TELEGRAM_CONNECTION_POOL_SIZE', 32),
                )
            )
            .concurrent_updates(int(getattr(config, 'TELEGRAM_CONCURRENT_UPDATES', 8) or 8))
            .build()
        )

        # Cache bot username once (avoid frequent get_me() calls in hot paths)
        try:
            me = await self.application.bot.get_me()
            if bot_instance:
                bot_instance.bot_username = me.username or ""
        except Exception as exc:
            logger.debug("Unable to cache bot username during setup: %s", exc)

        handlers = [
            CommandHandler("start", start),
            CommandHandler("search", search_command),
            CommandHandler("help", help_command),
            CommandHandler("history", history_command),
            CommandHandler("favorites", favorites_command),
            CommandHandler("favorite", favorite_last_command),
            CommandHandler("repeat", repeat_last_command),
            CommandHandler("settings", settings_command),
            CommandHandler("admin", admin_command),
            CommandHandler("tokens", admin_tokens),
            CommandHandler("mix", user_mix_command),
            CommandHandler("playlist", playlist_command),
            CommandHandler("album", playlist_command),
            CommandHandler("digest", digest_command),
            MessageHandler(filters.Document.ALL, handle_document_upload),
            MessageHandler(filters.CONTACT, handle_contact),
            MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message),
            CallbackQueryHandler(button_handler)
        ]
        
        # Separate early group: PTB continues to groups -1 and 0, so accounting
        # never intercepts command, message or callback handlers.
        self.application.add_handler(
            ChatMemberHandler(observe_bot_chat_membership, ChatMemberHandler.MY_CHAT_MEMBER),
            group=-3,
        )
        self.application.add_handler(MessageHandler(filters.ALL, track_unique_user_usage), group=-2)
        self.application.add_handler(CallbackQueryHandler(track_unique_user_usage), group=-2)
        self.application.add_handler(MessageHandler(filters.ALL, observe_group_topic), group=-1)
        for handler in handlers:
            self.application.add_handler(handler)
        self.application.add_error_handler(error_handler)

        # Media ловушка для админского режима приёма file_id (принимаем все сообщения,
        # а внутри handler-а фильтруем только сообщения с медиа). Это максимально совместимо
        # с разными версиями python-telegram-bot.
        self.application.add_handler(MessageHandler(filters.ALL, handle_media_upload), group=1)
        
        logger.info("✅ Бот настроен и готов к работе")
    
    async def run(self):
        """Запуск бота"""
        try:
            await self.application.initialize()
            await self.application.start()
            await self.application.updater.start_polling()
            bot_instance.create_background_task(
                group_digest_scheduler(self.application),
                "group-digest-scheduler",
            )
            
            logger.info("🎵 VLMB Bot запущен и готов к работе!")
            logger.info("🚀 100% асинхронная архитектура")
            logger.info(
                "⚙️ Параллелизм: updates=%s, downloads=%s",
                int(getattr(config, "TELEGRAM_CONCURRENT_UPDATES", 8) or 8),
                bot_instance._download_concurrency,
            )
            logger.info("🔴 Redis кэширование активировано" if bot_instance.cache.redis_client else "⚠️ Redis отключен")
            logger.info("🎵 Яндекс.Музыка активирована" if bot_instance.yandex_music._initialized else "⚠️ Яндекс.Музыка отключена")
            logger.info("🎶 VK Music активирован" if (bot_instance and bot_instance.vk_enabled()) else "⚠️ VK отключен")
            
            await self._stop_event.wait()
            
        except asyncio.CancelledError:
            raise
        except Exception:
            logger.exception("Критическая ошибка при запуске")
            raise
        finally:
            await self.shutdown()
    
    async def shutdown(self):
        """Корректное завершение работы"""
        if self._shutdown_started:
            return
        self._shutdown_started = True
        logger.info("🔄 Завершение работы бота...")

        try:
            if self.application:
                # Updater может быть не запущен, если упали до start_polling()
                upd = getattr(self.application, "updater", None)
                if upd:
                    is_running = bool(getattr(upd, "running", False) or getattr(upd, "_running", False))
                    if is_running:
                        await upd.stop()
                # stop/shutdown тоже делаем максимально безопасно
                try:
                    await self.application.stop()
                except Exception:
                    logger.exception("Ошибка при остановке Telegram application")
                try:
                    await self.application.shutdown()
                except Exception:
                    logger.exception("Ошибка при shutdown Telegram application")
        except Exception:
            logger.exception("Ошибка при завершении приложения")
        
        if bot_instance:
            try:
                if hasattr(bot_instance, "download_queue"):
                    await bot_instance.download_queue.close()
                if hasattr(bot_instance, "playlist_manager"):
                    await bot_instance.playlist_manager.close()
            except Exception:
                logger.exception("Ошибка при закрытии download queue")
            try:
                if hasattr(bot_instance, 'close_session'):
                    await bot_instance.close_session()
                elif hasattr(bot_instance, 'shutdown'):
                    await bot_instance.shutdown()
                elif hasattr(bot_instance, 'session') and getattr(bot_instance, 'session'):
                    await bot_instance.session.close()
            except Exception:
                logger.exception("Ошибка при закрытии ресурсов музыкального бота")

        logger.info("✅ Бот успешно остановлен")
    
    def stop(self):
        """Остановка бота"""
        self._stop_event.set()


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Global PTB error handler with traceback and persistent error metrics."""
    error = getattr(context, "error", None)
    if error is None:
        logger.error("Telegram update handling error without exception details")
        return

    logger.error(
        "Telegram update handling error",
        exc_info=(type(error), error, error.__traceback__),
    )

    try:
        instance = globals().get("bot_instance")
        admin_db = getattr(instance, "admin_db", None) if instance else None
        if admin_db is not None:
            effective_user = getattr(update, "effective_user", None)
            effective_chat = getattr(update, "effective_chat", None)
            await admin_db.log_error(
                getattr(effective_user, "id", None),
                getattr(effective_chat, "id", None),
                "telegram_update",
                str(_secret_filter._redact(str(error))),
            )
    except Exception:
        logger.exception("Failed to persist Telegram error metrics")

async def main():
    """Основная функция"""
    if not config.TELEGRAM_BOT_TOKEN:
        raise RuntimeError("Не установлен TELEGRAM_BOT_TOKEN. Задайте его в окружении или в systemd EnvironmentFile.")
    if not acquire_single_instance_lock():
        raise RuntimeError("Другой экземпляр бота уже запущен (instance lock занят)")

    runner = BotRunner()

    def signal_handler(signum, frame):
        logger.info(f"🛑 Получен сигнал {signum}, завершаем работу...")
        runner.stop()

    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTERM, signal_handler)

    try:
        await runner.setup()
        await runner.run()
    finally:
        await runner.shutdown()
        release_single_instance_lock()

if __name__ == '__main__':
    if not YANDEX_MUSIC_AVAILABLE and getattr(config, 'ENABLE_YANDEX_MUSIC', True):
        logger.warning("⚠️ Для работы с Яндекс.Музыкой установите: pip install yandex-music")
        logger.info("🤖 Бот запустится без поддержки Яндекс.Музыки")
    if not YT_DLP_AVAILABLE and getattr(config, 'ENABLE_YOUTUBE_MUSIC', True):
        logger.warning('⚠️ Для работы с YouTube установите: pip install -U "yt-dlp[default]"')
        logger.info("🤖 Бот запустится без поддержки YouTube")
    
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        pass
    except Exception:
        logger.exception("Бот завершён из-за критической ошибки")
        raise